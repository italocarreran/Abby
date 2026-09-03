# -*- coding: utf-8 -*-
"""
Comparador_Tabulado.py
======================

Compara, por central y hora, el SOBRECOSTO junto con las tres variables que lo
explican -- Generacion, CV y CMg -- entre las tres etapas del proceso
(Definitivo, Reliquidacion preliminar, Reliquidacion definitiva), leyendo el
archivo "02 Consolidado_Tabulado" de cada etapa.

Es un programa INDEPENDIENTE de Comparador_Etapas.py (el de los .mdb). Comparte
la misma carpeta base y el mismo estilo de ventana, pero su almacenamiento y
sus Excel son propios, asi que se pueden correr por separado sin pisarse.

De la hoja "Sobrecostos" se leen las columnas A:E, G, I:J y W. La fila 2 es el
encabezado, los datos parten en la 3.

Como encuentra los archivos
---------------------------
Se apoya en donde esta el .mdb de SSCC de cada etapa, porque "Detalles diarios"
siempre cuelga de ahi:

    Definitivo   <CMgReales>/AAMM/Sobrecostos/02 Definitivo/Auxiliares/   <- el .mdb
                 <CMgReales>/AAMM/Sobrecostos/02 Definitivo/Detalles diarios/  <- hermana
    Rpre/Rdef    <caso>/01 Sobrecostos/                          <- el .mdb
                 <caso>/01 Sobrecostos/Detalles diarios/         <- dentro

Dentro de "Detalles diarios" toma el unico .xlsm/.xlsx que haya, SIN exigir un
nombre exacto. Si hay varios, descarta los que terminen en "- copia" y se queda
con el mas reciente.

Almacenamiento incremental
--------------------------
    00_Salidas/AAAA/_comparador_tabulado/
        estado.json
        rutas.json
        parquet_variables/aamm=2501/etapa=def/datos.parquet
        vistas_variables/vista_2501.parquet
    00_Salidas/AAAA/MM Mes/Comparacion_Variables_AAMM.xlsx
    00_Salidas/AAAA/_comparador_tabulado/Comparacion_Variables_AAAA.xlsx

Al abrir la ventana NO se lee ningun archivo: solo se listan carpetas y se
compara la fecha de modificacion. El contenido se lee unicamente al consolidar.

Ubicacion: este .py va en ``Comparadores/``, carpeta hermana de la carpeta
del Revisor.

Requiere: pandas, pyarrow, duckdb, xlsxwriter, openpyxl
Opcional (muy recomendado): python-calamine  -> lectura varias veces mas rapida
"""

import importlib
import importlib.util
import json
import os
import re
import socket
import subprocess
import sys
import threading
import time
import traceback
import unicodedata
from datetime import datetime
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# --------------------------------------------------------------------------
# Dependencias externas
# --------------------------------------------------------------------------
# Se comprueban antes de importarlas: si falta alguna se avisa en una ventana y
# se sale, en vez de reventar con un ModuleNotFoundError pelado. Despues se
# importan de forma normal, sin `modulo = None` en el except: ese patron deja a
# cada modulo con tipo `Module | None` y el analizador de VS Code marca despues
# TODOS los usos ("no se puede acceder al atributo en None").
REQUISITOS = ["pandas", "pyarrow", "duckdb", "xlsxwriter", "openpyxl"]


def _falta(nombre):
    """find_spec puede lanzar excepcion en vez de devolver None."""
    if nombre in sys.modules:
        return False
    try:
        return importlib.util.find_spec(nombre) is None
    except (ImportError, ValueError):
        return True


FALTAN = [n for n in REQUISITOS if _falta(n)]

if FALTAN:
    _msg = ("Faltan librerias para ejecutar este programa:\n\n  "
            + ", ".join(FALTAN)
            + "\n\nInstalalas con:\n\n  python -m pip install "
            + " ".join(FALTAN))
    try:
        _r = tk.Tk()
        _r.withdraw()
        messagebox.showerror("Comparador de etapas — faltan librerias", _msg)
        _r.destroy()
    except Exception:
        print(_msg)
    sys.exit(1)

class _Perezoso:
    """Carga un modulo la PRIMERA vez que se usa, no al arrancar.

    pandas, pyarrow y duckdb tardan varios segundos en importarse cuando el
    antivirus corporativo revisa cada archivo del paquete, y ese costo se
    pagaba antes de dibujar la ventana. Con esto la ventana aparece de
    inmediato y el costo se paga recien al consolidar, donde ademas se ve
    en la barra de progreso.

    Se comporta como el modulo: pd.DataFrame, duckdb.connect, etc.
    """

    __slots__ = ("_nombre", "_modulo")

    def __init__(self, nombre):
        object.__setattr__(self, "_nombre", nombre)
        object.__setattr__(self, "_modulo", None)

    def _cargar(self):
        m = object.__getattribute__(self, "_modulo")
        if m is None:
            m = importlib.import_module(object.__getattribute__(self, "_nombre"))
            object.__setattr__(self, "_modulo", m)
        return m

    def __getattr__(self, attr):
        return getattr(self._cargar(), attr)

    def __dir__(self):
        return dir(self._cargar())


duckdb = _Perezoso("duckdb")
openpyxl = _Perezoso("openpyxl")
pd = _Perezoso("pandas")
pa = _Perezoso("pyarrow")
pq = _Perezoso("pyarrow.parquet")
xlsxwriter = _Perezoso("xlsxwriter")

# python_calamine es OPCIONAL: acelera varias veces la lectura de los .xlsm.
# Se comprueba aqui y se importa con importlib donde se usa, para que el editor
# no marque "import no resuelto" a quien no lo tenga instalado.
TIENE_CALAMINE = not _falta("python_calamine")


# ==========================================================================
# Constantes
# ==========================================================================
APP_TITULO = "Comparador de consolidados tabulados — Def / Rpre / Rdef"

BASE = Path(__file__).resolve().parent


def _morir(titulo, mensaje):
    """Muestra un error util incluso cuando el script corre con pythonw."""
    try:
        raiz = tk.Tk()
        raiz.withdraw()
        messagebox.showerror(titulo, mensaje)
        raiz.destroy()
    except Exception:
        print(f"{titulo}: {mensaje}")
    raise SystemExit(1)


def _hallar_revisor(raiz):
    """Carpeta hermana que contiene Revisor_Reliquidacion.py."""
    preferida = raiz / "Revisor_Relq"
    if (preferida / "Revisor_Reliquidacion.py").is_file():
        return preferida
    for carpeta in sorted(p for p in raiz.iterdir() if p.is_dir()):
        if (carpeta / "Revisor_Reliquidacion.py").is_file():
            return carpeta
    return None


DIR_REVISOR = _hallar_revisor(BASE.parent)
if DIR_REVISOR is None:
    _morir(
        "No se encontro la carpeta del Revisor",
        "Este comparador tiene que estar en una carpeta hermana de la del\n"
        "Revisor (la que contiene Revisor_Reliquidacion.py).\n\n"
        f"Se busco en: {BASE.parent}",
    )

sys.path.insert(0, str(DIR_REVISOR))
from comun import salidas as _sal

CONFIG_PATH = DIR_REVISOR / "config.json"
SALIDAS = _sal.raiz_salidas(BASE)


def _anio_de(aamm):
    """Devuelve el anio de cuatro digitos de un AAMM valido."""
    partes = _sal.partir_aamm(aamm)
    if partes is None:
        raise ValueError(f"AAMM invalido: {aamm!r}")
    return partes[0]


def cdir(anio):
    return _sal.carpeta_comparador(SALIDAS, anio, "_comparador_tabulado")


def dir_parquet(anio):
    return cdir(anio) / "parquet_variables"


def dir_vistas(anio):
    return cdir(anio) / "vistas_variables"


def estado_path(anio):
    return cdir(anio) / "estado.json"


def rutas_path(anio):
    return cdir(anio) / "rutas.json"


# Carpeta del comparador de .mdb: se consulta en modo solo lectura.
def cdir_mdb(anio):
    return _sal.carpeta_comparador(SALIDAS, anio, "_comparador")


def rutas_mdb_path(anio):
    return cdir_mdb(anio) / "rutas.json"

NOMBRE_JSON_MES = "_traspaso_actualizador.json"
CLAVE_JSON_PROPIA = "comparador_etapas"

ETAPAS = ["def", "rpre", "rdef"]
ETIQUETA = {"def": "Def", "rpre": "Rpre", "rdef": "Rdef"}
SLOTS = ["sscc", "sob"]
ETIQUETA_SLOT = {"sscc": "SOB_SSCC", "sob": "SOB"}

PAT_SSCC = re.compile(r"ENTRADA[\s_]*SOB[\s_]*SSCC", re.IGNORECASE)
PAT_SOB = re.compile(r"ENTRADA[\s_]*SOB(?![\s_]*SSCC)", re.IGNORECASE)
PAT_COPIA = re.compile(r"(-\s*copia|-\s*copy|\(\d+\))\s*$", re.IGNORECASE)

COLOR_ROJO = "#c0392b"
COLOR_AMARILLO = "#b8860b"
COLOR_VERDE = "#1e7a1e"
COLOR_GRIS = "#7f8c8d"

LIMITE_FILAS_HOJA = 1_048_000

# --- lectura del Consolidado_Tabulado ---
NOMBRE_CARPETA_DETALLES = "Detalles diarios"
EXT_CONSOL = (".xlsm", ".xlsx", ".xlsb")

# Columnas de la hoja "Sobrecostos", por letra. B (Hora) es la hora DENTRO DEL
# DIA; la hora mensual se calcula acumulando los dias previos.
COLS_CONSOL = {
    "A": "fecha", "B": "hora_dia", "C": "tipo", "D": "central",
    "E": "sobrecosto", "G": "gen", "I": "cv", "J": "cmg", "W": "usd",
}
FILA_ENCABEZADO_CONSOL = 2   # la fila 2 son titulos; los datos parten en la 3
# ==========================================================================
# Utilidades
# ==========================================================================
def normalizar(texto):
    """Sin tildes, sin espacios/guiones bajos, en minusculas."""
    if texto is None:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(texto))
    s = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"[\s_]+", "", s).strip().lower()


def normalizar_suave(texto):
    """Sin tildes, espacios colapsados, minusculas (para nombres de carpeta)."""
    if texto is None:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(texto))
    s = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


# --------------------------------------------------------------------------
# Acceso a disco con cache (clave para que esto no tarde minutos en el NAS)
# --------------------------------------------------------------------------
# En una carpeta de red cada consulta es un viaje por la red. Listar con
# iterdir() y despues preguntar is_dir()/stat() archivo por archivo son
# decenas de viajes por carpeta; os.scandir() trae nombre, tipo y fecha de
# una sola vez, y ademas se guarda el resultado por si la misma carpeta se
# consulta otra vez en el mismo refresco.
_CACHE_DIR = {}


def limpiar_cache():
    """Se llama al empezar cada refresco, para no mostrar datos viejos."""
    _CACHE_DIR.clear()


class _Entrada:
    """Un archivo o carpeta, con lo que hace falta ya leido."""

    __slots__ = ("nombre", "ruta", "es_dir", "mtime", "size")

    def __init__(self, nombre, ruta, es_dir, mtime, size):
        self.nombre = nombre
        self.ruta = ruta
        self.es_dir = es_dir
        self.mtime = mtime
        self.size = size


def listar(carpeta):
    """Contenido de una carpeta en UNA sola consulta al disco (con cache)."""
    if carpeta is None:
        return []
    clave = str(carpeta)
    if clave in _CACHE_DIR:
        return _CACHE_DIR[clave]
    out = []
    try:
        with os.scandir(clave) as it:
            for e in it:
                try:
                    st = e.stat()
                    out.append(_Entrada(e.name, Path(e.path), e.is_dir(),
                                        int(st.st_mtime), st.st_size))
                except OSError:
                    continue
    except (OSError, ValueError):
        out = []
    _CACHE_DIR[clave] = out
    return out


def huella_entrada(ruta):
    """mtime+tamano de un archivo, aprovechando el listado ya leido."""
    p = Path(ruta)
    for e in listar(p.parent):
        if e.nombre == p.name and not e.es_dir:
            return f"{e.mtime}_{e.size}"
    return None


def get_usuario():
    usuario = os.environ.get("USERNAME") or os.environ.get("USER") or "desconocido"
    return f"{socket.gethostname()}_{usuario}"


def leer_json(path, defecto=None):
    try:
        if Path(path).exists():
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        return None  # existe pero no se pudo interpretar -> None (distinto de vacio)
    return {} if defecto is None else defecto


def escribir_json_atomico(path, data):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def leer_config():
    todo = leer_json(CONFIG_PATH, {})
    if not isinstance(todo, dict):
        return {}
    return todo.get(get_usuario(), {}) or {}


def guardar_config(data):
    todo = leer_json(CONFIG_PATH, {})
    if todo is None:
        return  # config.json ilegible: mejor perder un ajuste que el archivo
    if not isinstance(todo, dict):
        todo = {}
    todo.setdefault(get_usuario(), {}).update(data)
    escribir_json_atomico(CONFIG_PATH, todo)


def abrir_en_explorador(ruta, es_archivo=True):
    if not ruta:
        return
    p = Path(ruta)
    if not p.exists():
        p = p.parent
        if not p.exists():
            return
    carpeta = p.parent if (es_archivo and p.is_file()) else p
    try:
        if sys.platform == "win32":
            subprocess.Popen(["explorer", str(carpeta)])
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(carpeta)])
        else:
            subprocess.Popen(["xdg-open", str(carpeta)])
    except Exception:
        pass


def huella(ruta):
    """Identidad barata de un archivo en disco de red: mtime + tamanio.

    Se apoya en el listado ya leido de la carpeta, para no hacer un viaje
    extra por cada archivo.
    """
    h = huella_entrada(ruta)
    if h is not None:
        return h
    try:
        st = Path(ruta).stat()
        return f"{int(st.st_mtime)}_{st.st_size}"
    except Exception:
        return None


def es_copia(nombre):
    tallo = Path(nombre).stem
    return bool(PAT_COPIA.search(tallo))



def subcarpeta(padre, nombre_buscado):
    """Subcarpeta por nombre normalizado (tolera tildes y mayusculas)."""
    if padre is None:
        return None
    objetivo = normalizar_suave(nombre_buscado)
    entradas = [e for e in listar(padre) if e.es_dir]
    for e in entradas:
        if normalizar_suave(e.nombre) == objetivo:
            return e.ruta
    for e in entradas:                      # segunda pasada: que contenga
        if objetivo in normalizar_suave(e.nombre):
            return e.ruta
    return None


def buscar_mdb(carpeta, patron):
    """Mas reciente que calce con el patron, descartando copias de Windows."""
    if carpeta is None:
        return None
    cands = [e for e in listar(carpeta)
             if not e.es_dir
             and Path(e.nombre).suffix.lower() in (".mdb", ".accdb")
             and patron.search(e.nombre)
             and not es_copia(e.nombre)
             and not e.nombre.startswith("~$")]
    if not cands:
        return None
    cands.sort(key=lambda e: e.mtime, reverse=True)
    return cands[0].ruta


def meses_del_anio(anio):
    """['2401', '2402', ... '2412'] a partir de 2024 o de '24'."""
    a = str(anio).strip()
    if len(a) == 4:
        aa = a[2:]
    elif len(a) == 2:
        aa = a
    else:
        return []
    if not aa.isdigit():
        return []
    return [f"{aa}{m:02d}" for m in range(1, 13)]


def fmt_tiempo(seg):
    m, s = divmod(int(seg), 60)
    h, m = divmod(m, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def ahora():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# Hojas que genera este programa; cualquier otra es de alguien mas.
HOJAS_PROPIAS_FIJAS = ["RESUMEN"]

def una_fila(cur, defecto=None):
    """fetchone() que nunca devuelve None: un COUNT/SUM siempre trae fila,
    pero el tipo declarado es Optional y hay que desempaquetarlo con cuidado."""
    fila = cur.fetchone()
    return fila if fila is not None else defecto


def es_hoja_propia(nombre, meses=None):
    """True si la hoja la genera este programa (y por lo tanto se puede pisar).

    Una hoja de mes se reconoce por su nombre (AAMM o AAMM_2), no por estar en
    la lista de meses que se exporta ahora: si un mes se saca del consolidado,
    su hoja tiene que desaparecer, no quedar congelada como si fuera de otro.
    """
    if nombre in HOJAS_PROPIAS_FIJAS or nombre == "RESUMEN":
        return True
    base, _, resto = nombre.partition("_")
    if not (len(base) == 4 and base.isdigit() and 1 <= int(base[2:]) <= 12):
        return False
    return resto == "" or resto.isdigit()


def hojas_ajenas(destino, log=print):
    """Hojas del archivo que NO genera este programa."""
    try:
        wb = openpyxl.load_workbook(str(destino), read_only=True)
        try:
            return [n for n in wb.sheetnames if not es_hoja_propia(n)]
        finally:
            wb.close()
    except Exception as e:
        log(f"  ! No se pudo revisar {destino.name} ({e}); se reescribe entero.")
        return []


def respaldar(destino, anio, log=print):
    """Copia el archivo antes de reescribirlo. Deja las ultimas 5."""
    try:
        import shutil
        carpeta = cdir(anio) / "respaldos"
        carpeta.mkdir(parents=True, exist_ok=True)
        marca = datetime.now().strftime("%Y%m%d_%H%M%S")
        copia = carpeta / f"{destino.stem}_{marca}{destino.suffix}"
        shutil.copy2(destino, copia)
        previas = sorted(carpeta.glob(f"{destino.stem}_*{destino.suffix}"))
        for vieja in previas[:-5]:
            try:
                vieja.unlink()
            except Exception:
                pass
        log(f"  Respaldo: {copia.name}")
        return copia
    except Exception as e:
        log(f"  ! No se pudo respaldar {destino.name}: {e}")
        return None


def cargar_estado(anio):
    est = leer_json(estado_path(anio), {})
    return est if isinstance(est, dict) else {}


def guardar_estado(anio, est):
    escribir_json_atomico(estado_path(anio), est)


def mes_incluido(est, aamm):
    """Si el mes entra al consolidado anual. Por defecto si."""
    reg = est.get(aamm) or {}
    return bool(reg.get("incluir", True))


def fijar_incluido(est, aamm, valor):
    est.setdefault(aamm, {})["incluir"] = bool(valor)


def firma_vistas(meses):
    """Que meses entran al consolidado y con que version de su vista.

    Si esta firma no cambio, el consolidado anual ya esta al dia y no hay para
    que reescribirlo.
    """
    f = {}
    for m in meses:
        v = path_vista(m)
        if v.exists():
            f[m] = int(v.stat().st_mtime)
    return f


def color_de(estado):
    return {
        "falta": COLOR_ROJO,
        "pendiente": COLOR_AMARILLO,
        "desactualizado": COLOR_AMARILLO,
        "ok": COLOR_VERDE,
    }.get(estado, COLOR_GRIS)


# Tramos de <CMgReales>\AAMM\Sobrecostos\02 Definitivo\Auxiliares.
# Cada tramo trae alternativas por si alguien renombro la carpeta.
TRAMOS_DEF = [
    ("Sobrecostos", ("Sobrecosto",)),
    ("02 Definitivo", ("Definitivo",)),
    ("Auxiliares", ("Auxiliar", "Auxiliares Definitivo")),
]


def carpeta_definitivo(raiz_cmg, aamm):
    """Devuelve (carpeta_auxiliares | None, hasta_donde_se_llego).

    El segundo valor es para mostrarlo en la ventana: si la estructura cambio,
    conviene ver en que tramo se corto en lugar de un "sin ruta" pelado.
    """
    if not raiz_cmg:
        return None, "falta la carpeta CMgReales"
    raiz = Path(raiz_cmg)
    if not raiz.is_dir():
        return None, f"no se ve la carpeta {raiz}"
    p = subcarpeta(raiz, aamm)
    if p is None:
        cand = raiz / aamm
        p = cand if cand.is_dir() else None
    if p is None:
        return None, f"llego hasta {raiz} — no hay carpeta {aamm}"
    for tramo, alternativas in TRAMOS_DEF:
        sig = subcarpeta(p, tramo)
        if sig is None:
            for alt in alternativas:
                sig = subcarpeta(p, alt)
                if sig is not None:
                    break
        if sig is None:
            return None, f"llego hasta {p} — falta '{tramo}'"
        p = sig
    return p, str(p)


# ==========================================================================
# Resolucion de rutas del Consolidado_Tabulado
# ==========================================================================
def ruta_json_mes(aamm):
    return _sal.carpeta_mes(SALIDAS, aamm) / NOMBRE_JSON_MES


def rutas_desde_json_mes(aamm):
    """(rdef, rpre) con las rutas de los .mdb que dejo el otro programa."""
    data = leer_json(ruta_json_mes(aamm), {})
    if not isinstance(data, dict):
        return {}, {}
    r = data.get("rutas") or {}
    rdef = {}
    if isinstance(r, dict):
        if r.get("mdb_sscc"):
            rdef["sscc"] = r["mdb_sscc"]
        if r.get("mdb_sob"):
            rdef["sob"] = r["mdb_sob"]
    propio = data.get(CLAVE_JSON_PROPIA) or {}
    rpre = {}
    if isinstance(propio, dict):
        blo = propio.get("rpre") or {}
        if isinstance(blo, dict):
            if blo.get("mdb_sscc"):
                rpre["sscc"] = blo["mdb_sscc"]
            if blo.get("mdb_sob"):
                rpre["sob"] = blo["mdb_sob"]
    return rdef, rpre


def rutas_manuales_mdb(anio):
    """Las carpetas que el usuario ya eligio en el comparador de .mdb.

    Se leen en modo solo lectura: este programa nunca escribe ese archivo.
    """
    d = leer_json(rutas_mdb_path(anio), {})
    return d if isinstance(d, dict) else {}


def excel_en_detalles(carpeta, log=print):
    """El Consolidado_Tabulado de una carpeta 'Detalles diarios'.

    NO se exige un nombre exacto: se toma el unico Excel que haya. Si hay
    varios, se descartan las copias ("- copia", "- Copy", "(2)") y los
    temporales de Excel, y se prefiere el que mencione "Consolidado" o
    "Tabulado"; a igualdad, el mas reciente.
    """
    if carpeta is None:
        return None, "sin carpeta"
    carpeta = Path(carpeta)
    cands = [e for e in listar(carpeta)
             if not e.es_dir and Path(e.nombre).suffix.lower() in EXT_CONSOL
             and not e.nombre.startswith("~$")]
    if not cands:
        return None, f"en {carpeta} no hay ningun Excel"
    sin_copias = [e for e in cands if not es_copia(e.nombre)]
    if sin_copias:
        cands = sin_copias
    if len(cands) > 1:
        preferidos = [e for e in cands
                      if "consolidado" in normalizar_suave(e.nombre)
                      or "tabulado" in normalizar_suave(e.nombre)]
        if preferidos:
            cands = preferidos
    cands.sort(key=lambda e: e.mtime, reverse=True)
    aviso = ""
    if len(cands) > 1:
        aviso = (f"habia {len(cands)} archivos, se tomo el mas reciente: "
                 f"{cands[0].nombre}")
    return cands[0].ruta, aviso


def detalles_junto_al_mdb(ruta_mdb):
    """La carpeta 'Detalles diarios' que corresponde a un .mdb de SSCC.

    Se busca primero DENTRO de la carpeta del .mdb (caso Rpre/Rdef, donde
    cuelga de '01 Sobrecostos') y despues como HERMANA (caso Definitivo,
    donde el .mdb esta en 'Auxiliares' y 'Detalles diarios' esta al lado).
    Asi la misma funcion sirve para las tres etapas sin ramificar.
    """
    if not ruta_mdb:
        return None, "no hay ruta del .mdb de SSCC"
    p = Path(ruta_mdb)
    carpeta = p.parent if p.suffix else p
    if not carpeta.is_dir():
        return None, f"no existe la carpeta del .mdb: {carpeta}"
    dentro = subcarpeta(carpeta, NOMBRE_CARPETA_DETALLES)
    if dentro is not None:
        return dentro, ""
    hermana = subcarpeta(carpeta.parent, NOMBRE_CARPETA_DETALLES)
    if hermana is not None:
        return hermana, ""
    return None, (f"no hay '{NOMBRE_CARPETA_DETALLES}' ni dentro de {carpeta} "
                  f"ni junto a ella")


def mdb_sscc_de_etapa(aamm, etapa, raiz_cmg, manual_mdb):
    """Ubica el .mdb de SSCC de una etapa, que es el ancla de la busqueda."""
    # 1) carpeta elegida a mano en el comparador de .mdb
    blo = (manual_mdb.get(aamm, {}) or {}).get(etapa, {}) or {}
    if isinstance(blo, dict):
        if blo.get("sscc"):
            return blo["sscc"], "archivo elegido a mano"
        if blo.get("carpeta"):
            c = Path(blo["carpeta"])
            sub = subcarpeta(c, "01 Sobrecostos")
            f = buscar_mdb(sub if sub is not None else c, PAT_SSCC)
            if f:
                return str(f), "carpeta elegida a mano"
            return str(sub or c), "carpeta elegida a mano (sin .mdb SSCC)"
    # 2) JSON del mes (lo deja el otro programa)
    rdef, rpre = rutas_desde_json_mes(aamm)
    origen = {"rdef": rdef, "rpre": rpre}.get(etapa, {})
    if origen.get("sscc"):
        return origen["sscc"], f"desde {NOMBRE_JSON_MES}"
    # 3) autodeteccion del Definitivo
    if etapa == "def":
        carp, hasta = carpeta_definitivo(raiz_cmg, aamm)
        if carp is not None:
            f = buscar_mdb(carp, PAT_SSCC)
            return (str(f) if f else str(carp)), "autodetectado"
        return None, hasta
    return None, (f"falta la ruta de {ETIQUETA[etapa]}: se toma del comparador "
                  f"de .mdb (JSON del mes o carpeta elegida alla)")


def resolver_rutas(aamm, raiz_cmg, manuales, manual_mdb):
    """({etapa: ruta|None}, {etapa: diagnostico}) del Consolidado_Tabulado."""
    res, diag = {}, {}
    for etapa in ETAPAS:
        # lo elegido a mano AQUI manda sobre todo
        m = (manuales.get(aamm, {}) or {}).get(etapa)
        if m:
            res[etapa] = m
            diag[etapa] = "archivo elegido a mano"
            continue
        mdb, origen = mdb_sscc_de_etapa(aamm, etapa, raiz_cmg, manual_mdb)
        if not mdb:
            res[etapa], diag[etapa] = None, origen
            continue
        carpeta, msg = detalles_junto_al_mdb(mdb)
        if carpeta is None:
            res[etapa], diag[etapa] = None, msg
            continue
        f, aviso = excel_en_detalles(carpeta)
        res[etapa] = str(f) if f else None
        diag[etapa] = aviso or (str(carpeta) if f else f"en {carpeta} no hay Excel")
    return res, diag



def tabla_por_nombre(nombres, candidatos):
    """Primer candidato presente, comparando normalizado."""
    mapa = {normalizar(n): n for n in nombres}
    for c in candidatos:
        real = mapa.get(normalizar(c))
        if real:
            return real
    return None


def _col_a_indice(letra):
    """'A'->0, 'B'->1, ... 'W'->22 (0-index para listas de calamine/openpyxl)."""
    n = 0
    for c in letra:
        n = n * 26 + (ord(c.upper()) - ord("A") + 1)
    return n - 1


IDX_CONSOL = {_col_a_indice(l): nom for l, nom in COLS_CONSOL.items()}





def leer_consolidado_tabulado(ruta, log=print):
    """DataFrame: central, tipo, dia, hora_dia, sobrecosto, gen, cv, cmg, usd.

    Usa python_calamine si esta disponible (varias veces mas rapido que
    openpyxl en archivos grandes, que es el "truquito" para que esto no se
    demore como los .mdb por red); si no esta instalado, cae a openpyxl.
    """
    ruta = Path(ruta)
    filas = None
    if TIENE_CALAMINE:
        try:
            # Carga dinamica a proposito: python_calamine es OPCIONAL, y un
            # `import` normal lo marca como "no se pudo resolver" en el editor
            # de quien no lo tenga instalado. importlib lo pide en tiempo de
            # ejecucion, solo cuando TIENE_CALAMINE confirmo que esta.
            calamine = importlib.import_module("python_calamine")
            wb = calamine.CalamineWorkbook.from_path(str(ruta))
            nombre = tabla_por_nombre(wb.sheet_names, ["Sobrecostos"]) or wb.sheet_names[0]
            filas = wb.get_sheet_by_name(nombre).to_python()
        except Exception as e:
            log(f"    ! calamine fallo con {ruta.name} ({e}); se usa openpyxl.")
            filas = None
    if filas is None:
        wb = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
        try:
            nombre = tabla_por_nombre(wb.sheetnames, ["Sobrecostos"]) or wb.sheetnames[0]
            filas = [list(r) for r in wb[nombre].iter_rows(values_only=True)]
        finally:
            wb.close()

    if len(filas) <= FILA_ENCABEZADO_CONSOL:
        raise RuntimeError(f"{ruta.name}: no hay filas de datos.")
    maxi = max(IDX_CONSOL)
    datos = []
    for r in filas[FILA_ENCABEZADO_CONSOL:]:
        if len(r) <= maxi or r[_col_a_indice("D")] is None:
            continue
        fila = {nom: r[i] for i, nom in IDX_CONSOL.items()}
        datos.append(fila)
    if not datos:
        raise RuntimeError(f"{ruta.name}: 0 filas de datos utilizables.")
    df = pd.DataFrame(datos)

    def a_fecha(v):
        if hasattr(v, "date"):
            return v.date() if hasattr(v, "hour") else v
        return v
    df["fecha"] = df["fecha"].map(a_fecha)
    df["dia"] = pd.to_datetime(df["fecha"], errors="coerce").dt.day
    malas = int(df["dia"].isna().sum())
    if malas > 0:
        log(f"    ! {ruta.name}: {malas} fila(s) con fecha invalida, se descartan.")
        df = df[df["dia"].notna()]
    df["dia"] = df["dia"].astype(int)
    df["hora_dia"] = pd.to_numeric(df["hora_dia"], errors="coerce")
    df = df[df["hora_dia"].notna()]
    df["hora_dia"] = df["hora_dia"].astype(int)

    for c in ("sobrecosto", "gen", "cv", "cmg", "usd"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    df["central"] = df["central"].astype(str).str.strip()
    df["tipo"] = df["tipo"].astype(str).str.strip()
    df = df[(df["central"].str.len() > 0) & (df["central"].str.lower() != "nan")]
    return df[["central", "tipo", "dia", "hora_dia",
              "sobrecosto", "gen", "cv", "cmg", "usd"]]


def hora_mensual(dia, hora_dia, acumulado_por_dia_dict):
    """hora_mes = horas de los dias 1..dia-1 + hora_dia.

    acumulado_por_dia_dict: {dia: horas_acumuladas_antes_de_ese_dia}, calculado
    sobre las horas que realmente trae el archivo (asi un dia de 23 o 25
    horas por cambio de hora no desalinea los dias siguientes: el archivo
    manda, no el calendario).
    """
    return acumulado_por_dia_dict[dia] + hora_dia


def acumulado_por_dia(df):
    """{dia: horas_acumuladas_antes_de_ese_dia} a partir de lo que trae el df."""
    por_dia = df.groupby("dia")["hora_dia"].max().to_dict()
    dias = sorted(por_dia)
    acc, total = {}, 0
    for d in dias:
        acc[d] = total
        total += por_dia[d]
    return acc


# ---- almacenamiento (parquet), separado del de sobrecostos por tipo ----


def dir_datos(aamm, etapa):
    return dir_parquet(_anio_de(aamm)) / f"aamm={aamm}" / f"etapa={etapa}"


def path_vista(aamm):
    return dir_vistas(_anio_de(aamm)) / f"vista_{aamm}.parquet"


def path_excel_mes(aamm):
    return _sal.carpeta_mes(SALIDAS, aamm) / f"Comparacion_Variables_{aamm}.xlsx"


def path_excel_anual(aa):
    return cdir(aa) / f"Comparacion_Variables_{_sal.normalizar_anio(aa)}.xlsx"


def estado_etapa(est, aamm, etapa, rutas):
    r = rutas.get(etapa)
    if not (r and Path(r).is_file()):
        return "falta"
    reg = ((est.get(aamm) or {}).get(etapa) or {})
    if not reg.get("consolidado"):
        return "pendiente"
    if reg.get("huella") != huella(r):
        return "desactualizado"
    return "ok"


def consolidar_etapa(aamm, etapa, ruta, est, log=print):
    log(f"  {aamm} {ETIQUETA[etapa]} · {Path(ruta).name}")
    df = leer_consolidado_tabulado(ruta, log=log)
    acc = acumulado_por_dia(df)
    df["hora_mes"] = [hora_mensual(d, h, acc) for d, h in zip(df["dia"], df["hora_dia"])]
    df["etapa"] = etapa
    d = dir_datos(aamm, etapa)
    d.mkdir(parents=True, exist_ok=True)
    pq.write_table(pa.Table.from_pandas(
        df[["central", "tipo", "hora_mes", "sobrecosto", "gen", "cv", "cmg", "usd", "etapa"]],
        preserve_index=False), d / "datos.parquet", compression="snappy")
    est.setdefault(aamm, {})[etapa] = {
        "consolidado": ahora(), "huella": huella(ruta),
        "filas": int(len(df)), "horas": int(df["hora_mes"].max()) if len(df) else 0,
    }
    est[aamm].pop("vista", None)
    log(f"  {aamm} {ETIQUETA[etapa]}: {len(df):,} filas, hasta hora {df['hora_mes'].max() if len(df) else 0}.")
    return True


COLUMNAS_VISTA = [
    "aamm", "central", "tipo", "hora_mes",
    "sc_def", "sc_rpre", "sc_rdef", "d_rpre_def", "d_rdef_rpre", "d_rdef_def",
    "gen_def", "gen_rpre", "gen_rdef", "cv_def", "cv_rpre", "cv_rdef",
    "cmg_def", "cmg_rpre", "cmg_rdef", "usd",
    "cambia_gen", "cambia_cv", "cambia_cmg", "formula_mismatch", "detalle",
]


def etapas_consolidadas(aamm):
    return [e for e in ETAPAS if (dir_datos(aamm, e) / "datos.parquet").exists()]


def vista_completa(aamm):
    v = path_vista(aamm)
    if not v.exists():
        return False
    try:
        return not (set(COLUMNAS_VISTA) - set(pq.read_schema(v).names))
    except Exception:
        return False


def construir_vista(aamm, tol=0.01, log=print):
    """Pivotea SC/Gen/CV/CMg de las etapas disponibles y arma las banderas.

    'cambia_X' es cierto si esa variable difiere en mas de `tol` (proporcion,
    ej. 0.01 = 1%) entre alguna pareja de etapas que si estan consolidadas.
    'formula_mismatch' es cierto si en alguna etapa Sobrecosto no calza con
    (CV-CMg)*Gen*USD mas alla de la tolerancia.
    """
    disp = etapas_consolidadas(aamm)
    if not disp:
        log(f"  {aamm}: nada consolidado en variables, sin vista.")
        return None
    dir_vistas(_anio_de(aamm)).mkdir(parents=True, exist_ok=True)
    con = duckdb.connect()
    try:
        base = str((dir_parquet(_anio_de(aamm)) / f"aamm={aamm}").as_posix()) + "/**/*.parquet"
        sel = ",\n".join(
            f"MAX(CASE WHEN etapa='{e}' THEN {c} END) AS {c}_{e}"
            for e in ETAPAS for c in ("sobrecosto", "gen", "cv", "cmg", "usd"))
        sql_base = f"""
        WITH s AS (
            SELECT etapa, central, tipo, hora_mes,
                   SUM(sobrecosto) AS sobrecosto, SUM(gen) AS gen,
                   AVG(cv) AS cv, AVG(cmg) AS cmg, AVG(usd) AS usd
            FROM read_parquet('{base}', hive_partitioning=1)
            GROUP BY 1,2,3,4
        )
        SELECT central, tipo, hora_mes, {sel}
        FROM s GROUP BY 1,2,3
        """

        def monto(pref, e):
            return f"{pref}_{e}" if e in disp else "NULL::DOUBLE"

        def cambia(pref):
            pares = [(a, b) for i, a in enumerate(disp) for b in disp[i + 1:]]
            if not pares:
                return "FALSE"
            cond = " OR ".join(
                f"(ABS(COALESCE({pref}_{a},0) - COALESCE({pref}_{b},0)) "
                f"> {tol} * GREATEST(ABS(COALESCE({pref}_{a},0)), ABS(COALESCE({pref}_{b},0)), 1))"
                for a, b in pares)
            return f"({cond})"

        def formula(e):
            if e not in disp:
                return "FALSE"
            return (f"(ABS(COALESCE(sobrecosto_{e},0) "
                    f"- (COALESCE(cv_{e},0)-COALESCE(cmg_{e},0))*COALESCE(gen_{e},0)*COALESCE(usd_{e},0)) "
                    f"> {tol} * GREATEST(ABS(COALESCE(sobrecosto_{e},0)), 1))")

        detalle_partes = []
        for pref, etq in (("gen", "Gen"), ("cv", "CV"), ("cmg", "CMg")):
            for i, a in enumerate(disp):
                for b in disp[i + 1:]:
                    detalle_partes.append(
                        f"CASE WHEN ABS(COALESCE({pref}_{a},0)-COALESCE({pref}_{b},0)) "
                        f"> {tol} * GREATEST(ABS(COALESCE({pref}_{a},0)),ABS(COALESCE({pref}_{b},0)),1) "
                        f"THEN '{etq} {ETIQUETA[a]}->{ETIQUETA[b]}: ' || "
                        f"printf('%.3f', COALESCE({pref}_{a},0)) || ' -> ' || "
                        f"printf('%.3f', COALESCE({pref}_{b},0)) END")
        for e in disp:
            detalle_partes.append(
                f"CASE WHEN {formula(e)} THEN 'Formula no calza en {ETIQUETA[e]} "
                f"(SC {ETIQUETA[e]} vs (CV-CMg)*Gen*USD)' END")
        # Cada CASE WHEN da NULL si no aplica; se arma una lista, se botan los
        # NULL y se unen con salto de linea. Asi el detalle no queda con
        # lineas en blanco quemando espacio quando solo cambio una variable.
        lista_sql = "[" + ", ".join(detalle_partes) + "]" if detalle_partes else "[]"
        detalle_sql = (f"array_to_string(list_filter({lista_sql}, "
                       f"x -> x IS NOT NULL), CHR(10))")

        sql = f"""
        WITH p AS ({sql_base})
        SELECT
            '{aamm}' AS aamm, central, tipo, hora_mes,
            {monto("sobrecosto","def")}  AS sc_def,
            {monto("sobrecosto","rpre")} AS sc_rpre,
            {monto("sobrecosto","rdef")} AS sc_rdef,
            {(f"{monto('sobrecosto','rpre')}-{monto('sobrecosto','def')}"
              if 'def' in disp and 'rpre' in disp else 'NULL::DOUBLE')}  AS d_rpre_def,
            {(f"{monto('sobrecosto','rdef')}-{monto('sobrecosto','rpre')}"
              if 'rpre' in disp and 'rdef' in disp else 'NULL::DOUBLE')} AS d_rdef_rpre,
            {(f"{monto('sobrecosto','rdef')}-{monto('sobrecosto','def')}"
              if 'def' in disp and 'rdef' in disp else 'NULL::DOUBLE')}  AS d_rdef_def,
            {monto("gen","def")} AS gen_def, {monto("gen","rpre")} AS gen_rpre, {monto("gen","rdef")} AS gen_rdef,
            {monto("cv","def")} AS cv_def, {monto("cv","rpre")} AS cv_rpre, {monto("cv","rdef")} AS cv_rdef,
            {monto("cmg","def")} AS cmg_def, {monto("cmg","rpre")} AS cmg_rpre, {monto("cmg","rdef")} AS cmg_rdef,
            COALESCE({monto("usd","rdef")}, {monto("usd","rpre")}, {monto("usd","def")}) AS usd,
            {cambia("gen")} AS cambia_gen,
            {cambia("cv")} AS cambia_cv,
            {cambia("cmg")} AS cambia_cmg,
            ({" OR ".join(formula(e) for e in disp)}) AS formula_mismatch,
            NULLIF({detalle_sql}, '') AS detalle
        FROM p
        ORDER BY central, tipo, hora_mes
        """
        con.execute(f"COPY ({sql}) TO '{path_vista(aamm).as_posix()}' "
                   f"(FORMAT PARQUET, COMPRESSION SNAPPY)")
        n = una_fila(con.execute(
            f"SELECT COUNT(*) FROM read_parquet('{path_vista(aamm).as_posix()}')"),
            (0,))[0]
    finally:
        con.close()
    log(f"  {aamm}: vista de variables rearmada ({n:,} filas, etapas "
       f"{', '.join(ETIQUETA[e] for e in disp)}).")
    return n


def asegurar_vista(aamm, tol=0.01, log=print):
    if vista_completa(aamm):
        return True
    if not etapas_consolidadas(aamm):
        return False
    if path_vista(aamm).exists():
        log(f"  {aamm}: la vista de variables es de una version anterior, se rearma.")
    return construir_vista(aamm, tol=tol, log=log) is not None


CAB = ["Central", "Tipo", "Hora Mensual",
          "SC Def", "SC Rpre", "SC Rdef", "Rpre−Def", "Rdef−Rpre", "Rdef−Def",
          "Gen Def", "Gen Rpre", "Gen Rdef",
          "CV Def", "CV Rpre", "CV Rdef",
          "CMg Def", "CMg Rpre", "CMg Rdef", "USD",
          "Cambia Gen", "Cambia CV", "Cambia CMg", "Formula no calza", "Detalle"]


def exportar_excel(destino, meses, solo_dif, tolerancia, log=print,
                             preservar=True):
    """Igual espiritu que exportar_excel: solo pisa sus propias hojas."""
    for m in meses:
        asegurar_vista(m, tol=tolerancia, log=log)
    vistas = [(m, path_vista(m).as_posix()) for m in meses if path_vista(m).exists()]
    if not vistas:
        raise RuntimeError("No hay ninguna vista de variables construida para exportar.")
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    ajenas = hojas_ajenas(destino, log=log) if (preservar and destino.exists()) else []
    if ajenas:
        _escribir_preservando(destino, vistas, solo_dif, tolerancia, ajenas, log)
    else:
        _escribir_desde_cero(destino, vistas, solo_dif, tolerancia, log)
    log(f"Excel de variables listo: {destino}")


def _filtro(solo_dif, tol):
    if not solo_dif:
        return ""
    return (f"WHERE ABS(COALESCE(d_rpre_def,0)) > {tol} "
            f"OR ABS(COALESCE(d_rdef_rpre,0)) > {tol} "
            f"OR ABS(COALESCE(d_rdef_def,0)) > {tol} "
            f"OR cambia_gen OR cambia_cv OR cambia_cmg OR formula_mismatch")


def _escribir_desde_cero(destino, vistas, solo_dif, tolerancia, log):
    con = duckdb.connect()
    wb = None
    try:
        wb = xlsxwriter.Workbook(str(destino), {"constant_memory": True})
        f_head = wb.add_format({"bold": True, "bg_color": "#1f3864", "font_color": "white",
                                "border": 1, "text_wrap": True, "valign": "vcenter"})
        f_num = wb.add_format({"num_format": "#,##0.00"})
        f_dif = wb.add_format({"num_format": "#,##0.00", "bg_color": "#fde9d9"})
        f_alerta = wb.add_format({"bg_color": "#f8cbad"})
        for aamm, pvs in vistas:
            filtro = _filtro(solo_dif, tolerancia)
            cur = con.execute(f"SELECT {', '.join(COLUMNAS_VISTA[1:])} "
                              f"FROM read_parquet('{pvs}') {filtro} "
                              f"ORDER BY central, tipo, hora_mes")
            h = wb.add_worksheet(aamm)
            h.write_row(0, 0, CAB, f_head)
            h.set_row(0, 26)
            h.freeze_panes(1, 3)
            h.set_column(0, 1, 26)
            h.set_column(3, 18, 13, f_num)
            h.set_column(22, 22, 70)
            h.autofilter(0, 0, 0, len(CAB) - 1)
            i_cg, i_cv, i_cm, i_fm = 19, 20, 21, 22
            fila = 1
            while True:
                lote = cur.fetchmany(50_000)
                if not lote:
                    break
                for reg in lote:
                    marca = any(reg[i] for i in (i_cg, i_cv, i_cm, i_fm))
                    for j, v in enumerate(reg):
                        if isinstance(v, bool):
                            h.write(fila, j, "SI" if v else "",
                                   f_alerta if v else None)
                        elif v is None:
                            h.write_blank(fila, j, None)
                        elif isinstance(v, (int, float)):
                            h.write_number(fila, j, float(v),
                                          f_dif if j in (6, 7, 8) and
                                          abs(float(v)) > tolerancia else f_num)
                        else:
                            h.write(fila, j, str(v), f_alerta if marca and j == 22 else None)
                    fila += 1
            log(f"  {aamm}: {fila - 1:,} filas.")
    finally:
        if wb is not None:
            wb.close()
        con.close()


def _escribir_preservando(destino, vistas, solo_dif, tolerancia, ajenas, log):
    log(f"  Se conservan {len(ajenas)} hoja(s) de otra persona: "
       f"{', '.join(ajenas[:6])}{' ...' if len(ajenas) > 6 else ''}")
    respaldar(destino, _anio_de(vistas[0][0]), log)
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    azul = PatternFill("solid", fgColor="1F3864")
    naranjo = PatternFill("solid", fgColor="F8CBAD")
    crema = PatternFill("solid", fgColor="FDE9D9")
    blanco_negrita = Font(bold=True, color="FFFFFF")
    con = duckdb.connect()
    try:
        wb = openpyxl.load_workbook(str(destino))
        for n in list(wb.sheetnames):
            if es_hoja_propia(n):
                del wb[n]
        i_cg, i_cv, i_cm, i_fm = 19, 20, 21, 22
        for aamm, pvs in vistas:
            filtro = _filtro(solo_dif, tolerancia)
            h = wb.create_sheet(aamm)
            for j, tx in enumerate(CAB, start=1):
                c = h.cell(row=1, column=j, value=tx)
                c.fill, c.font = azul, blanco_negrita
                c.alignment = Alignment(wrap_text=True, vertical="center")
            h.freeze_panes = "D2"
            fila = 2
            cur = con.execute(f"SELECT {', '.join(COLUMNAS_VISTA[1:])} "
                              f"FROM read_parquet('{pvs}') {filtro} "
                              f"ORDER BY central, tipo, hora_mes")
            while True:
                lote = cur.fetchmany(50_000)
                if not lote:
                    break
                for reg in lote:
                    if fila > LIMITE_FILAS_HOJA:
                        log(f"  ! {aamm} paso el limite de filas; usa 'Solo con diferencia'.")
                        break
                    marca = any(reg[i] for i in (i_cg, i_cv, i_cm, i_fm))
                    for j, v in enumerate(reg, start=1):
                        if v is None:
                            continue
                        if isinstance(v, bool):
                            if v:
                                h.cell(row=fila, column=j, value="SI").fill = naranjo
                        elif isinstance(v, (int, float)):
                            c = h.cell(row=fila, column=j, value=float(v))
                            c.number_format = "#,##0.00"
                            if j - 1 in (6, 7, 8) and abs(float(v)) > tolerancia:
                                c.fill = crema
                        else:
                            c = h.cell(row=fila, column=j, value=str(v))
                            if marca and j - 1 == 22:
                                c.fill = naranjo
                    fila += 1
            h.auto_filter.ref = f"A1:{get_column_letter(len(CAB))}1"
            log(f"  {aamm}: {fila - 2:,} filas.")
        wb.save(str(destino))
    finally:
        con.close()


# ==========================================================================
# Ventana
# ==========================================================================
class App:
    def __init__(self, root):
        self.root = root
        self.cfg = leer_config()
        anio_inicial = self.cfg.get("tab_anio", "")
        self.est = cargar_estado(anio_inicial) if meses_del_anio(anio_inicial) else {}
        rj = leer_json(rutas_path(anio_inicial), {}) if meses_del_anio(anio_inicial) else {}
        self.manuales = rj if isinstance(rj, dict) else {}
        self.rutas = {}       # aamm -> {etapa: ruta|None}
        self.diag = {}        # aamm -> {etapa: diagnostico}
        self.filas = {}
        self.expandido = {}
        self.trabajando = False
        self.timer = {"on": False, "t0": 0.0}

        root.title(APP_TITULO)
        root.geometry("1120x780")
        root.minsize(900, 600)

        self.var_anio = tk.StringVar(value=self.cfg.get("tab_anio", ""))
        self.var_cmg = tk.StringVar(value=self.cfg.get("comp_cmgreales", ""))
        self.var_estado = tk.StringVar(value="Listo")
        self.var_tiempo = tk.StringVar(value="00:00:00")
        self.var_tol = tk.StringVar(value=self.cfg.get("tab_tolerancia", "0.01"))
        self.var_solo_dif = tk.BooleanVar(value=False)
        self.var_forzar = tk.BooleanVar(value=False)
        self.var_preservar = tk.BooleanVar(value=True)
        self.var_anual = tk.StringVar(value=str(cdir(anio_inicial)) if meses_del_anio(anio_inicial) else "")

        self._construir()
        self.log(f"Carpeta base: {BASE}")
        self.log("Lectura rapida de Excel: "
                 + ("python-calamine activo" if TIENE_CALAMINE
                    else "python-calamine NO instalado (se usa openpyxl, mas lento). "
                         "Instalalo con: pip install python-calamine"))
        if self.var_anio.get():
            self.refrescar()

    # ---------------- construccion ----------------
    def _construir(self):
        root = self.root
        pie = tk.Frame(root)
        pie.pack(side="bottom", fill="x", pady=6)
        self.btn_refrescar = tk.Button(pie, text="ACTUALIZAR estado",
                                       command=self.refrescar, width=18)
        self.btn_refrescar.pack(side="left", padx=6)
        self.btn_cons = tk.Button(pie, text="Consolidar pendientes", bg="#b8860b",
                                  fg="white", width=20,
                                  command=lambda: self.lanzar(self.consolidar, forzar=False))
        self.btn_cons.pack(side="left", padx=6)
        self.btn_todo = tk.Button(pie, text="Reconsolidar TODO", bg="#7f8c8d",
                                  fg="white", width=18, command=self.confirmar_todo)
        self.btn_todo.pack(side="left", padx=6)
        self.btn_excel = tk.Button(pie, text="Exportar Excel", bg="#1e7a1e",
                                   fg="white", width=16,
                                   command=lambda: self.lanzar(self.exportar))
        self.btn_excel.pack(side="left", padx=6)
        tk.Label(pie, textvariable=self.var_tiempo, font=("Consolas", 11, "bold"),
                 fg="#1e7a1e").pack(side="right", padx=10)

        marco_log = tk.LabelFrame(root, text="Bitacora")
        marco_log.pack(side="bottom", fill="x", padx=8, pady=(0, 4))
        self.txt = tk.Text(marco_log, height=9, font=("Consolas", 9), wrap="none")
        sb = tk.Scrollbar(marco_log, command=self.txt.yview)
        self.txt.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.txt.pack(fill="x", expand=True)
        self.barra = ttk.Progressbar(root, mode="determinate")
        self.barra.pack(side="bottom", fill="x", padx=8)
        tk.Label(root, textvariable=self.var_estado, anchor="w",
                 fg="#1f3864").pack(side="bottom", fill="x", padx=10)

        canvas = tk.Canvas(root, borderwidth=0, highlightthickness=0)
        scroll = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        cont = tk.Frame(canvas)
        win = canvas.create_window((0, 0), window=cont, anchor="nw")

        def ajustar(_=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(win, width=canvas.winfo_width())
        cont.bind("<Configure>", ajustar)
        canvas.bind("<Configure>", ajustar)
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"))

        f1 = tk.LabelFrame(cont, text="Anio a comparar (AAAA o AA)", padx=8, pady=6)
        f1.pack(fill="x", padx=12, pady=6)
        e = tk.Entry(f1, textvariable=self.var_anio, width=10,
                     font=("Consolas", 12, "bold"), justify="center")
        e.pack(side="left")
        e.bind("<Return>", lambda _: self.refrescar())
        tk.Button(f1, text="Cargar los 12 meses",
                  command=self.refrescar).pack(side="left", padx=8)

        f2 = tk.LabelFrame(cont, text="Carpeta CMgReales (para ubicar el Definitivo)",
                           padx=8, pady=6)
        f2.pack(fill="x", padx=12, pady=4)
        self.lbl_cmg = tk.Label(f2, textvariable=self.var_cmg, cursor="hand2",
                                anchor="w", font=("Segoe UI", 9))
        self.lbl_cmg.pack(fill="x")
        self.lbl_cmg.bind("<Button-1>",
                          lambda _: abrir_en_explorador(self.var_cmg.get(), False))
        tk.Button(f2, text="Examinar", command=self.elegir_cmg).pack(pady=(4, 0))

        f3 = tk.Frame(cont)
        f3.pack(fill="x", padx=12, pady=(2, 6))
        tk.Label(f3, anchor="w", justify="left", fg=COLOR_GRIS, font=("Segoe UI", 8),
                 text="Las rutas de Rpre y Rdef salen de donde esta el .mdb de SSCC "
                      "(las que ya elegiste en el comparador de Access).\n"
                      "Dentro de esa carpeta se busca 'Detalles diarios' y se toma el "
                      "Excel que haya ahi, sea cual sea su nombre."
                 ).pack(anchor="w")
        for txt, col in (("rojo = falta el archivo", COLOR_ROJO),
                         ("amarillo = esta pero sin consolidar / cambio despues", COLOR_AMARILLO),
                         ("verde = consolidado y al dia", COLOR_VERDE)):
            tk.Label(f3, text="  " + txt, fg=col,
                     font=("Segoe UI", 8, "bold")).pack(side="left")

        self.marco_meses = tk.Frame(cont)
        self.marco_meses.pack(fill="x", padx=12, pady=4)

        f4 = tk.LabelFrame(cont, text="Paso 2 — parquet a Excel", padx=8, pady=6)
        f4.pack(fill="x", padx=12, pady=8)
        tk.Label(f4, anchor="w", justify="left", fg=COLOR_GRIS, font=("Segoe UI", 8),
                 text="Cada mes deja su Excel en Salidas\\AAMM\\ y el consolidado del "
                      "anio en Salidas\\_comparador_tabulado\\.\nSolo se rehace lo que "
                      "quedo viejo."
                 ).pack(fill="x")
        self.lbl_anual = tk.Label(f4, textvariable=self.var_anual, cursor="hand2",
                                  anchor="w", font=("Segoe UI", 9), fg="#1f3864")
        self.lbl_anual.pack(fill="x", pady=(4, 0))
        self.lbl_anual.bind("<Button-1>",
                            lambda _: abrir_en_explorador(self.var_anual.get(), True))
        fb = tk.Frame(f4)
        fb.pack(fill="x", pady=(4, 0))
        tk.Checkbutton(fb, text="Rehacer todos los Excel aunque esten al dia",
                       variable=self.var_forzar).pack(side="left")
        tk.Checkbutton(fb, text="Conservar hojas agregadas por otros",
                       variable=self.var_preservar).pack(side="left", padx=10)
        tk.Checkbutton(fb, text="Solo filas con diferencia o cambio de variable",
                       variable=self.var_solo_dif).pack(side="left", padx=10)
        tk.Label(fb, text="tolerancia:").pack(side="left", padx=(10, 2))
        tk.Entry(fb, textvariable=self.var_tol, width=8).pack(side="left")
        tk.Label(fb, text="(proporcion: 0.01 = 1%)", fg=COLOR_GRIS,
                 font=("Segoe UI", 8)).pack(side="left", padx=4)

    # ---------------- log / estado ----------------
    def log(self, msg):
        self.txt.insert("end", str(msg) + "\n")
        self.txt.see("end")
        try:
            self.root.update_idletasks()
        except Exception:
            pass

    def set_estado(self, txt):
        self.var_estado.set(txt)
        try:
            self.root.update_idletasks()
        except Exception:
            pass

    def tick(self):
        if self.timer["on"]:
            self.var_tiempo.set(fmt_tiempo(time.time() - self.timer["t0"]))
            self.root.after(500, self.tick)

    def botones(self, activos):
        estado = "normal" if activos else "disabled"
        for b in (self.btn_refrescar, self.btn_cons, self.btn_todo, self.btn_excel):
            b.config(state=estado)

    def lanzar(self, funcion, **kw):
        if self.trabajando:
            return
        self.trabajando = True
        self.botones(False)
        self.timer.update({"on": True, "t0": time.time()})
        self.tick()

        def correr():
            try:
                funcion(**kw)
            except Exception as e:
                self.log(f"ERROR: {e}")
                self.log(traceback.format_exc())
            finally:
                self.timer["on"] = False
                self.trabajando = False
                self.root.after(0, lambda: self.botones(True))
                self.root.after(0, self.pintar)
                self.root.after(0, lambda: self.set_estado("Listo"))

        threading.Thread(target=correr, daemon=True).start()

    # ---------------- selectores ----------------
    def elegir_cmg(self):
        ini = self.var_cmg.get() if Path(self.var_cmg.get() or "x").is_dir() else ""
        r = filedialog.askdirectory(title="Carpeta CMgReales", initialdir=ini)
        if r:
            self.var_cmg.set(r)
            guardar_config({"comp_cmgreales": r})
            self.refrescar()

    def elegir_archivo(self, aamm, etapa):
        act = self.rutas.get(aamm, {}).get(etapa)
        ini = str(Path(act).parent) if act and Path(act).exists() else \
            (self.var_cmg.get() or str(BASE))
        r = filedialog.askopenfilename(
            title=f"{aamm} · {ETIQUETA[etapa]} · Consolidado tabulado",
            initialdir=ini,
            filetypes=[("Excel", "*.xlsm *.xlsx *.xlsb"), ("Todos", "*.*")])
        if not r:
            return
        self.manuales.setdefault(aamm, {})[etapa] = r
        escribir_json_atomico(rutas_path(self.var_anio.get().strip()), self.manuales)
        self.refrescar(solo=aamm)

    def elegir_carpeta(self, aamm, etapa):
        """Apuntar directo a una carpeta 'Detalles diarios'."""
        r = filedialog.askdirectory(
            title=f"{aamm} · {ETIQUETA[etapa]} · carpeta con el consolidado",
            initialdir=self.var_cmg.get() or str(BASE))
        if not r:
            return
        f, aviso = excel_en_detalles(r, log=self.log)
        if f is None:
            self.log(f"{aamm} {ETIQUETA[etapa]}: {aviso}")
            return
        self.log(f"{aamm} {ETIQUETA[etapa]} · {f.name}"
                 + (f"   ({aviso})" if aviso else ""))
        self.manuales.setdefault(aamm, {})[etapa] = str(f)
        escribir_json_atomico(rutas_path(self.var_anio.get().strip()), self.manuales)
        self.refrescar(solo=aamm)

    def olvidar_manual(self, aamm, etapa):
        if (self.manuales.get(aamm) or {}).pop(etapa, None) is not None:
            escribir_json_atomico(rutas_path(self.var_anio.get().strip()), self.manuales)
            self.log(f"{aamm} {ETIQUETA[etapa]}: se vuelve a la deteccion automatica.")
            self.refrescar(solo=aamm)

    # ---------------- filas de mes ----------------
    def refrescar(self, solo=None):
        """Busca los archivos en segundo plano, para no congelar la ventana.

        Recorrer las carpetas del NAS toma varios segundos; hacerlo en el hilo
        principal dejaba la ventana en blanco y sin responder. Aqui el trabajo
        de disco va en un hilo y solo el pintado vuelve al hilo de tkinter,
        que es el unico que puede tocar los widgets.
        """
        anio = self.var_anio.get().strip()
        meses = meses_del_anio(anio)
        if not meses:
            self.log("Indica el anio (por ejemplo 2025 o 25).")
            return
        if self.trabajando:
            return
        guardar_config({"tab_anio": anio})
        self.est = cargar_estado(anio)
        rj = leer_json(rutas_path(anio), {})
        self.manuales = rj if isinstance(rj, dict) else {}
        self.var_anual.set(str(cdir(anio)))
        raiz = self.var_cmg.get().strip()
        self.trabajando = True
        self.botones(False)
        objetivo = [solo] if solo else meses
        self.barra.config(maximum=len(objetivo), value=0)

        def buscar():
            t0 = time.time()
            try:
                limpiar_cache()      # el disco pudo cambiar desde el refresco anterior
                man_mdb = rutas_manuales_mdb(self.var_anio.get().strip())
                for i, m in enumerate(objetivo, 1):
                    self.set_estado(f"Buscando archivos... {m}  ({i}/{len(objetivo)})")
                    self.rutas[m], self.diag[m] = resolver_rutas(
                        m, raiz, self.manuales, man_mdb)
                    self.barra.config(value=i)
            except Exception as e:
                self.log(f"ERROR buscando archivos: {e}")
                self.log(traceback.format_exc())
            finally:
                seg = time.time() - t0
                self.trabajando = False

                def terminar():
                    self.botones(True)
                    if solo is None or not self.filas:
                        self.construir_filas(meses)
                    self.pintar()
                    self.set_estado("Listo")
                    if seg > 3:
                        self.log(f"Busqueda de archivos: {seg:.1f} s "
                                 f"({len(objetivo)} mes(es)).")
                self.root.after(0, terminar)

        threading.Thread(target=buscar, daemon=True).start()

    def construir_filas(self, meses):
        for w in self.marco_meses.winfo_children():
            w.destroy()
        self.filas = {}
        for m in meses:
            self.expandido.setdefault(m, False)
            caja = tk.Frame(self.marco_meses, relief="groove", bd=1)
            caja.pack(fill="x", pady=2)
            cab = tk.Frame(caja)
            cab.pack(fill="x")
            var_inc = tk.BooleanVar(value=mes_incluido(self.est, m))
            tk.Checkbutton(cab, variable=var_inc,
                           command=lambda mm=m: self.cambiar_inclusion(mm)).pack(side="left")
            btn = tk.Button(cab, text=f"▶  {m}", width=10, anchor="w",
                            font=("Consolas", 11, "bold"), relief="flat",
                            command=lambda mm=m: self.alternar(mm))
            btn.pack(side="left", padx=4, pady=2)
            chips = {}
            for etapa in ETAPAS:
                l = tk.Label(cab, text=ETIQUETA[etapa], width=7,
                             font=("Segoe UI", 9, "bold"), relief="ridge", padx=4, pady=1)
                l.pack(side="left", padx=3)
                chips[etapa] = l
            info = tk.Label(cab, text="", anchor="w", fg=COLOR_GRIS, font=("Segoe UI", 8))
            info.pack(side="left", padx=8)
            tk.Button(cab, text="Consolidar mes", font=("Segoe UI", 8),
                      command=lambda mm=m: self.lanzar(self.consolidar, meses=[mm],
                                                       forzar=False)).pack(side="right", padx=4)
            tk.Button(cab, text="Rearmar vista", font=("Segoe UI", 8),
                      command=lambda mm=m: self.lanzar(self.rearmar_vista,
                                                       meses=[mm])).pack(side="right", padx=2)

            det = tk.Frame(caja)
            lineas = {}
            for etapa in ETAPAS:
                fe = tk.Frame(det)
                fe.pack(fill="x", padx=16, pady=1)
                tk.Label(fe, text=ETIQUETA[etapa], width=6, anchor="w",
                         font=("Consolas", 8, "bold")).pack(side="left")
                lr = tk.Label(fe, text="", anchor="w", cursor="hand2", font=("Segoe UI", 8))
                lr.pack(side="left", fill="x", expand=True)
                lr.bind("<Button-1>", lambda e, mm=m, et=etapa:
                        abrir_en_explorador(self.rutas.get(mm, {}).get(et), True))
                tk.Button(fe, text="Auto", font=("Segoe UI", 8),
                          command=lambda mm=m, et=etapa:
                          self.olvidar_manual(mm, et)).pack(side="right", padx=2)
                tk.Button(fe, text="Archivo", font=("Segoe UI", 8),
                          command=lambda mm=m, et=etapa:
                          self.elegir_archivo(mm, et)).pack(side="right", padx=2)
                tk.Button(fe, text="Carpeta", font=("Segoe UI", 8),
                          command=lambda mm=m, et=etapa:
                          self.elegir_carpeta(mm, et)).pack(side="right", padx=2)
                lineas[etapa] = lr
            self.filas[m] = {"btn": btn, "chips": chips, "info": info,
                             "det": det, "lineas": lineas, "inc": var_inc}

    def cambiar_inclusion(self, m):
        f = self.filas.get(m)
        if not f:
            return
        self.est = cargar_estado(self.var_anio.get().strip())
        fijar_incluido(self.est, m, f["inc"].get())
        guardar_estado(self.var_anio.get().strip(), self.est)
        self.pintar()

    def alternar(self, m):
        f = self.filas.get(m)
        if not f:
            return
        self.expandido[m] = not self.expandido[m]
        if self.expandido[m]:
            f["det"].pack(fill="x", pady=(0, 4))
            f["btn"].config(text=f"▼  {m}")
        else:
            f["det"].pack_forget()
            f["btn"].config(text=f"▶  {m}")

    def pintar(self):
        for m, f in self.filas.items():
            rutas = self.rutas.get(m) or {e: None for e in ETAPAS}
            dg = self.diag.get(m) or {}
            ests = []
            for etapa in ETAPAS:
                st = estado_etapa(self.est, m, etapa, rutas)
                ests.append(st)
                f["chips"][etapa].config(fg="white", bg=color_de(st))
                r = rutas.get(etapa)
                lr = f["lineas"][etapa]
                msg = dg.get(etapa) or ""
                if r and Path(r).is_file():
                    lr.config(text=str(r) + (f"   ({msg})" if msg and not
                                             msg.startswith(str(Path(r).parent)) else ""),
                              fg="#1f3864")
                else:
                    lr.config(text=f"[sin archivo] {msg}".strip(), fg=COLOR_ROJO)
            reg = self.est.get(m) or {}
            hechas = [ETIQUETA[e] for e in ETAPAS if (reg.get(e) or {}).get("consolidado")]
            info = (f"consolidado: {', '.join(hechas) or '—'}"
                    f"   |   vista: {'si' if path_vista(m).exists() else 'no'}")
            if not mes_incluido(self.est, m):
                info += "   |   FUERA del consolidado anual"
            f["inc"].set(mes_incluido(self.est, m))
            f["info"].config(text=info)
            todos_ok = all(x == "ok" for x in ests)
            falta = any(x == "falta" for x in ests)
            f["btn"].config(fg=color_de("ok" if todos_ok else
                                        ("falta" if falta else "pendiente")))

    # ---------------- procesos ----------------
    def meses_visibles(self):
        return meses_del_anio(self.var_anio.get().strip())

    def tolerancia(self):
        try:
            return float(str(self.var_tol.get()).replace(",", "."))
        except Exception:
            return 0.01

    def confirmar_todo(self):
        if messagebox.askyesno(
                "Reconsolidar todo",
                "Vuelve a leer TODOS los consolidados tabulados de los 12 meses.\n\n"
                "¿Seguir?"):
            self.lanzar(self.consolidar, forzar=True)

    def consolidar(self, meses=None, forzar=False):
        if not meses_del_anio(self.var_anio.get().strip()):
            # Sin esto, cargar_estado()/rutas_manuales_mdb() revientan con
            # ValueError (via comun.salidas.normalizar_anio): antes de la
            # Tarea 2 estas rutas no dependian del anio y nunca fallaban.
            self.log("Elegi un anio valido antes de consolidar.")
            return
        meses = meses or self.meses_visibles()
        raiz = self.var_cmg.get().strip()
        man_mdb = rutas_manuales_mdb(self.var_anio.get().strip())
        self.est = cargar_estado(self.var_anio.get().strip())
        tareas = []
        for m in meses:
            self.rutas[m], self.diag[m] = resolver_rutas(m, raiz, self.manuales, man_mdb)
            for etapa in ETAPAS:
                st = estado_etapa(self.est, m, etapa, self.rutas[m])
                if st != "falta" and (forzar or st in ("pendiente", "desactualizado")):
                    tareas.append((m, etapa, st))
        if not tareas:
            self.log("Nada por consolidar: todo lo que tiene archivo ya esta al dia.")
            return
        self.log(f"\n=== Consolidando {len(tareas)} etapa(s) ===")
        self.barra.config(maximum=len(tareas), value=0)
        tocados = set()
        for i, (m, etapa, st) in enumerate(tareas, 1):
            self.set_estado(f"[{i}/{len(tareas)}] {m} {ETIQUETA[etapa]} ({st})")
            try:
                consolidar_etapa(m, etapa, self.rutas[m][etapa], self.est, log=self.log)
                tocados.add(m)
                guardar_estado(self.var_anio.get().strip(), self.est)
            except Exception as e:
                self.log(f"  ERROR en {m} {ETIQUETA[etapa]}: {e}")
                self.log(traceback.format_exc())
            self.barra.config(value=i)
            self.root.after(0, self.pintar)
        tol = self.tolerancia()
        for m in sorted(tocados):
            self.set_estado(f"Rearmando vista {m}")
            try:
                construir_vista(m, tol=tol, log=self.log)
                self.est.setdefault(m, {})["vista"] = ahora()
                guardar_estado(self.var_anio.get().strip(), self.est)
            except Exception as e:
                self.log(f"  ERROR armando vista {m}: {e}")
        self.log("=== Consolidacion terminada ===\n")

    def rearmar_vista(self, meses=None):
        self.est = cargar_estado(self.var_anio.get().strip())
        tol = self.tolerancia()
        for m in (meses or self.meses_visibles()):
            if etapas_consolidadas(m):
                construir_vista(m, tol=tol, log=self.log)
                self.est.setdefault(m, {})["vista"] = ahora()
        guardar_estado(self.var_anio.get().strip(), self.est)

    def exportar(self):
        tol = self.tolerancia()
        guardar_config({"tab_tolerancia": str(tol)})
        self.est = cargar_estado(self.var_anio.get().strip())
        anio = self.var_anio.get().strip()
        aa = anio[2:] if len(anio) == 4 else anio
        meses = self.meses_visibles()
        disponibles = [m for m in meses if path_vista(m).exists()]
        if not disponibles:
            self.log("No hay ningun mes consolidado todavia.")
            return
        for m in disponibles:
            asegurar_vista(m, tol=tol, log=self.log)
        forzar = self.var_forzar.get()
        solo_dif = self.var_solo_dif.get()

        pend = []
        for m in disponibles:
            dm = path_excel_mes(m)
            if (forzar or not dm.exists()
                    or dm.stat().st_mtime < path_vista(m).stat().st_mtime):
                pend.append(m)
        if pend:
            self.log(f"\n=== Excel por mes: {', '.join(pend)} ===")
            self.barra.config(maximum=len(pend), value=0)
            for i, m in enumerate(pend, 1):
                self.set_estado(f"[{i}/{len(pend)}] Excel de {m}")
                dm = path_excel_mes(m)
                dm.parent.mkdir(parents=True, exist_ok=True)
                exportar_excel(dm, [m], solo_dif, tol, log=self.log,
                               preservar=self.var_preservar.get())
                self.barra.config(value=i)
        else:
            self.log("\nExcel por mes: todos al dia, no se rehace ninguno.")

        incluidos = [m for m in disponibles if mes_incluido(self.est, m)]
        fuera = [m for m in disponibles if m not in incluidos]
        if fuera:
            self.log(f"\nFuera del consolidado por eleccion tuya: {', '.join(fuera)}")
        if not incluidos:
            self.log("Ningun mes marcado para el consolidado anual.")
        else:
            anual = path_excel_anual(aa)
            firma = firma_vistas(incluidos)
            previa = (self.est.get("_excel_anual") or {}).get("firma") or {}
            igual = ({str(k): v for k, v in firma.items()}
                     == {str(k): v for k, v in previa.items()})
            if anual.exists() and igual and not forzar and not pend:
                self.log(f"Consolidado anual ya al dia: {anual.name}")
            else:
                self.set_estado("Escribiendo el consolidado anual...")
                self.log(f"\n=== Consolidado anual ({len(incluidos)} meses: "
                         f"{', '.join(incluidos)}) ===")
                t0 = time.time()
                exportar_excel(anual, incluidos, solo_dif, tol, log=self.log,
                               preservar=self.var_preservar.get())
                self.est["_excel_anual"] = {
                    "archivo": str(anual), "meses": incluidos, "firma": firma,
                    "escrito": ahora(), "segundos": round(time.time() - t0, 1),
                }
                guardar_estado(self.var_anio.get().strip(), self.est)
                self.log(f"  ({self.est['_excel_anual']['segundos']} s)")
            self.var_anual.set(str(anual))
        self.barra.config(value=self.barra["maximum"])
        self.log("=== Listo ===\n")


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
