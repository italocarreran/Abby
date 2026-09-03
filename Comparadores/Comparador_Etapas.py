# -*- coding: utf-8 -*-
"""
Comparador_Etapas.py
====================

Consolida los sobrecostos horarios de los .mdb de las tres etapas del proceso
(Definitivo, Reliquidacion Preliminar, Reliquidacion Definitiva) para los 12
meses de un anio, y exporta un Excel de comparacion con el detalle horario por
central (una hoja por mes).

Por cada mes y etapa hay DOS Access:
    slot "sscc" -> 03b ENTRADA_SOB_SSCC_AAMM_*.mdb   (tipos SCCF, CO, CCA)
    slot "sob"  -> 03b ENTRADA_SOB_AAMM_*.mdb        (tipos SCMT, SCPC)

De cada .mdb se leen:
    - tabla Sobrecostos  (Clave Anio_Mes, Tipo_sobrecosto, Central, Hora Mensual, Sobrecosto)
    - tabla Central_Empresa_Actualizada, y si no existe, Central_Empresa

Almacenamiento incremental (nunca se reescribe lo ya consolidado):

    __config__/AAAA/_comparador/
        estado.json                     <- huella (mtime+tamanio) y fecha de cada consolidacion
        rutas.json                      <- rutas elegidas a mano (respaldo propio)
        parquet/
            sobrecostos/aamm=2401/etapa=def/datos.parquet
            centrales/  aamm=2401/etapa=def/datos.parquet
        vistas/
            vista_2401.parquet          <- tabla ya comparada del mes (cache para el Excel)
    00_Salidas/AAAA/MM Mes/Comparacion_AAMM.xlsx
    00_Salidas/AAAA/Comparacion_Etapas_AAAA.xlsx

Agregar un mes escribe solo sus archivos. El Excel se arma desde las vistas, asi
que reexportar un anio completo no vuelve a tocar ningun Access.

Ubicacion: este .py va en ``Comparadores/``, carpeta hermana de la carpeta
del Revisor.

Requiere: pyodbc, pandas, pyarrow, duckdb, xlsxwriter
          + driver "Microsoft Access Driver (*.mdb, *.accdb)" de la MISMA
            arquitectura (32/64 bits) que el Python que lo ejecuta.
"""

import json
import importlib
import importlib.util
import os
import queue
import re
import socket
import subprocess
import sys
import threading
import time
import traceback
import unicodedata
from datetime import datetime
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# --------------------------------------------------------------------------
# Dependencias externas
# --------------------------------------------------------------------------
# Se comprueban antes de importarlas: si falta alguna se avisa en una ventana y
# se sale, en vez de reventar con un ModuleNotFoundError pelado. Despues se
# importan de forma normal, sin `modulo = None` en el except: ese patron deja a
# cada modulo con tipo `Module | None` y el analizador de VS Code marca despues
# TODOS los usos ("no se puede acceder al atributo en None").
REQUISITOS = ["pyodbc", "pandas", "pyarrow", "duckdb", "xlsxwriter", "openpyxl"]


def _falta(nombre):
    """find_spec puede lanzar excepcion en vez de devolver None."""
    if nombre in sys.modules:
        return False
    try:
        return importlib.util.find_spec(nombre) is None
    except (ImportError, ValueError):
        return True


FALTAN = [n for n in REQUISITOS if _falta(n)]

if FALTAN:
    _msg = ("Faltan librerias para ejecutar este programa:\n\n  "
            + ", ".join(FALTAN)
            + "\n\nInstalalas con:\n\n  python -m pip install "
            + " ".join(FALTAN))
    try:
        _r = tk.Tk()
        _r.withdraw()
        messagebox.showerror("Comparador de etapas — faltan librerias", _msg)
        _r.destroy()
    except Exception:
        print(_msg)
    sys.exit(1)

class _Perezoso:
    """Carga un modulo la PRIMERA vez que se usa, no al arrancar.

    pandas, pyarrow y duckdb tardan varios segundos en importarse cuando el
    antivirus corporativo revisa cada archivo del paquete, y ese costo se
    pagaba antes de dibujar la ventana. Con esto la ventana aparece de
    inmediato y el costo se paga recien al consolidar, donde ademas se ve
    en la barra de progreso.

    Se comporta como el modulo: pd.DataFrame, duckdb.connect, etc.
    """

    __slots__ = ("_nombre", "_modulo")

    def __init__(self, nombre):
        object.__setattr__(self, "_nombre", nombre)
        object.__setattr__(self, "_modulo", None)

    def _cargar(self):
        m = object.__getattribute__(self, "_modulo")
        if m is None:
            m = importlib.import_module(object.__getattribute__(self, "_nombre"))
            object.__setattr__(self, "_modulo", m)
        return m

    def __getattr__(self, attr):
        return getattr(self._cargar(), attr)

    def __dir__(self):
        return dir(self._cargar())


duckdb = _Perezoso("duckdb")
openpyxl = _Perezoso("openpyxl")
pd = _Perezoso("pandas")
pa = _Perezoso("pyarrow")
pq = _Perezoso("pyarrow.parquet")
xlsxwriter = _Perezoso("xlsxwriter")
pyodbc = _Perezoso("pyodbc")


# ==========================================================================
# Constantes
# ==========================================================================
APP_TITULO = "Comparador de etapas — Def / Rpre / Rdef"

BASE = Path(__file__).resolve().parent


def _morir(titulo, mensaje):
    """Muestra un error util incluso cuando el script corre con pythonw."""
    try:
        raiz = tk.Tk()
        raiz.withdraw()
        messagebox.showerror(titulo, mensaje)
        raiz.destroy()
    except Exception:
        print(f"{titulo}: {mensaje}")
    raise SystemExit(1)


def _hallar_revisor(raiz):
    """Carpeta hermana que contiene Revisor_Reliquidacion.py."""
    preferida = raiz / "Revisor_Relq"
    if (preferida / "Revisor_Reliquidacion.py").is_file():
        return preferida
    for carpeta in sorted(p for p in raiz.iterdir() if p.is_dir()):
        if (carpeta / "Revisor_Reliquidacion.py").is_file():
            return carpeta
    return None


DIR_REVISOR = _hallar_revisor(BASE.parent)
if DIR_REVISOR is None:
    _morir(
        "No se encontro la carpeta del Revisor",
        "Este comparador tiene que estar en una carpeta hermana de la del\n"
        "Revisor (la que contiene Revisor_Reliquidacion.py).\n\n"
        f"Se busco en: {BASE.parent}",
    )
assert DIR_REVISOR is not None

DIR_RAIZ = DIR_REVISOR.parent
sys.path.insert(0, str(DIR_RAIZ))
try:
    from __comun__ import salidas as _sal
    from __comun__ import tema as _tema
except ImportError as e:
    _morir(
        "Falta la carpeta __comun__/",
        "Tiene que estar la carpeta '__comun__' hermana de Revisor_Relq y de\n"
        "Comparadores. Baja el repositorio completo, no los .py sueltos.\n\n"
        f"Carpeta actual: {DIR_RAIZ}\n\nDetalle: {e}",
    )

CONFIG_RAIZ = _sal.raiz_config(BASE)
CONFIG_PATH = CONFIG_RAIZ / "config.json"
SALIDAS = _sal.raiz_salidas(BASE)


def _anio_de(aamm):
    """Devuelve el anio de cuatro digitos de un AAMM valido."""
    partes = _sal.partir_aamm(aamm)
    if partes is None:
        raise ValueError(f"AAMM invalido: {aamm!r}")
    return partes[0]


def cdir(anio):
    return _sal.carpeta_comparador(CONFIG_RAIZ, anio, "_comparador")


def dir_parquet(anio):
    return cdir(anio) / "parquet"


def dir_sob_raiz(anio):
    return dir_parquet(anio) / "sobrecostos"


def dir_cen_raiz(anio):
    return dir_parquet(anio) / "centrales"


def dir_vistas(anio):
    return cdir(anio) / "vistas"


def dir_actual(anio):
    return cdir(anio) / "actual"


def actual_parquet(anio):
    return dir_actual(anio) / "central_empresa_actual.parquet"


def estado_path(anio):
    return cdir(anio) / "estado.json"


def rutas_path(anio):
    return cdir(anio) / "rutas.json"

# Clave nueva y propia dentro del JSON de traspaso del revisor. Se agrega sin
# tocar nada de lo que ya hay, para no romper a los otros scripts.
CLAVE_JSON_PROPIA = "comparador_etapas"
NOMBRE_JSON_MES = "_traspaso_actualizador.json"

ETAPAS = ["def", "rpre", "rdef"]
ETIQUETA = {"def": "Def", "rpre": "Rpre", "rdef": "Rdef"}
SLOTS = ["sscc", "sob"]
ETIQUETA_SLOT = {"sscc": "SOB_SSCC", "sob": "SOB"}

# Patrones de nombre de archivo
# Sin anclar al inicio: en el Definitivo los archivos NO empiezan con "03b",
# asi que lo que identifica a cada uno es el trozo ENTRADA_SOB(_SSCC), donde sea
# que aparezca en el nombre. El lookahead negativo evita que el patron del de
# energia se coma tambien al de SSCC.
PAT_SSCC = re.compile(r"ENTRADA[\s_]*SOB[\s_]*SSCC", re.IGNORECASE)
PAT_SOB = re.compile(r"ENTRADA[\s_]*SOB(?![\s_]*SSCC)", re.IGNORECASE)
PAT_COPIA = re.compile(r"(-\s*copia|-\s*copy|\(\d+\))\s*$", re.IGNORECASE)

# Maestro del propietario vigente:
#   T:\Facturacion\<AAAA>\<MM Mes>\{02 Definitivo|01 Preliminar}\SSCC\
#   4_REMUNERACION_SC_CO_AAMM_*.xlsx|xlsm  ->  hoja "Configuracion Empresa", A y B
PAT_REMUN = re.compile(r"^4[\s_]*REMUNERACI[OÓ]N[\s_]*SC[\s_]*CO", re.IGNORECASE)
PAT_ANIO_DIR = re.compile(r"^(20\d{2})$")
PAT_MES_DIR = re.compile(r"^(\d{1,2})\b")
RAMAS_FACT = ["02 Definitivo", "01 Preliminar"]
HOJA_CONFIG_EMPRESA = "Configuracion Empresa"

# La paleta clara conserva exactamente los colores historicos. La ventana
# reemplaza este dict solo si el tema pedido se pudo aplicar completo.
COLORES = _tema.paleta("claro")

# Nombres de columna que se buscan en la tabla Sobrecostos (normalizados)
MAPA_SOB = {
    "claveaniomes": "clave_aamm",
    "claveanomes": "clave_aamm",
    "clavea_omes": "clave_aamm",
    "tiposobrecosto": "tipo",
    "central": "central",
    "horamensual": "hora_mes",
    "sobrecosto": "sobrecosto",
}
MAPA_CEN = {"central": "central", "empresa": "empresa"}


# ==========================================================================
# Utilidades
# ==========================================================================
def normalizar(texto):
    """Sin tildes, sin espacios/guiones bajos, en minusculas."""
    if texto is None:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(texto))
    s = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"[\s_]+", "", s).strip().lower()


def normalizar_suave(texto):
    """Sin tildes, espacios colapsados, minusculas (para nombres de carpeta)."""
    if texto is None:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(texto))
    s = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


# --------------------------------------------------------------------------
# Acceso a disco con cache (clave para que esto no tarde minutos en el NAS)
# --------------------------------------------------------------------------
# En una carpeta de red cada consulta es un viaje por la red. Listar con
# iterdir() y despues preguntar is_dir()/stat() archivo por archivo son
# decenas de viajes por carpeta; os.scandir() trae nombre, tipo y fecha de
# una sola vez, y ademas se guarda el resultado por si la misma carpeta se
# consulta otra vez en el mismo refresco.
_CACHE_DIR = {}


def limpiar_cache():
    """Se llama al empezar cada refresco, para no mostrar datos viejos."""
    _CACHE_DIR.clear()


class _Entrada:
    """Un archivo o carpeta, con lo que hace falta ya leido."""

    __slots__ = ("nombre", "ruta", "es_dir", "mtime", "size")

    def __init__(self, nombre, ruta, es_dir, mtime, size):
        self.nombre = nombre
        self.ruta = ruta
        self.es_dir = es_dir
        self.mtime = mtime
        self.size = size


def listar(carpeta):
    """Contenido de una carpeta en UNA sola consulta al disco (con cache)."""
    if carpeta is None:
        return []
    clave = str(carpeta)
    if clave in _CACHE_DIR:
        return _CACHE_DIR[clave]
    out = []
    try:
        with os.scandir(clave) as it:
            for e in it:
                try:
                    st = e.stat()
                    out.append(_Entrada(e.name, Path(e.path), e.is_dir(),
                                        int(st.st_mtime), st.st_size))
                except OSError:
                    continue
    except (OSError, ValueError):
        out = []
    _CACHE_DIR[clave] = out
    return out


def huella_entrada(ruta):
    """mtime+tamano de un archivo, aprovechando el listado ya leido."""
    p = Path(ruta)
    for e in listar(p.parent):
        if e.nombre == p.name and not e.es_dir:
            return f"{e.mtime}_{e.size}"
    return None


def get_usuario():
    usuario = os.environ.get("USERNAME") or os.environ.get("USER") or "desconocido"
    return f"{socket.gethostname()}_{usuario}"


def leer_json(path, defecto=None):
    try:
        if Path(path).exists():
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        return None  # existe pero no se pudo interpretar -> None (distinto de vacio)
    return {} if defecto is None else defecto


def escribir_json_atomico(path, data):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def leer_config():
    todo = leer_json(CONFIG_PATH, {})
    if not isinstance(todo, dict):
        return {}
    return todo.get(get_usuario(), {}) or {}


def guardar_config(data):
    todo = leer_json(CONFIG_PATH, {})
    if todo is None:
        return  # config.json ilegible: mejor perder un ajuste que el archivo
    if not isinstance(todo, dict):
        todo = {}
    todo.setdefault(get_usuario(), {}).update(data)
    escribir_json_atomico(CONFIG_PATH, todo)


def abrir_en_explorador(ruta, es_archivo=True):
    if not ruta:
        return
    p = Path(ruta)
    if not p.exists():
        p = p.parent
        if not p.exists():
            return
    carpeta = p.parent if (es_archivo and p.is_file()) else p
    try:
        if sys.platform == "win32":
            subprocess.Popen(["explorer", str(carpeta)])
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(carpeta)])
        else:
            subprocess.Popen(["xdg-open", str(carpeta)])
    except Exception:
        pass


def huella(ruta):
    """Identidad barata de un archivo en disco de red: mtime + tamanio.

    Se apoya en el listado ya leido de la carpeta, para no hacer un viaje
    extra por cada archivo.
    """
    h = huella_entrada(ruta)
    if h is not None:
        return h
    try:
        st = Path(ruta).stat()
        return f"{int(st.st_mtime)}_{st.st_size}"
    except Exception:
        return None


def es_copia(nombre):
    tallo = Path(nombre).stem
    return bool(PAT_COPIA.search(tallo))



def subcarpeta(padre, nombre_buscado):
    """Subcarpeta por nombre normalizado (tolera tildes y mayusculas)."""
    if padre is None:
        return None
    objetivo = normalizar_suave(nombre_buscado)
    entradas = [e for e in listar(padre) if e.es_dir]
    for e in entradas:
        if normalizar_suave(e.nombre) == objetivo:
            return e.ruta
    for e in entradas:                      # segunda pasada: que contenga
        if objetivo in normalizar_suave(e.nombre):
            return e.ruta
    return None


def buscar_mdb(carpeta, patron):
    """Mas reciente que calce con el patron, descartando copias de Windows."""
    if carpeta is None:
        return None
    cands = [e for e in listar(carpeta)
             if not e.es_dir
             and Path(e.nombre).suffix.lower() in (".mdb", ".accdb")
             and patron.search(e.nombre)
             and not es_copia(e.nombre)
             and not e.nombre.startswith("~$")]
    if not cands:
        return None
    cands.sort(key=lambda e: e.mtime, reverse=True)
    return cands[0].ruta


def meses_del_anio(anio):
    """['2401', '2402', ... '2412'] a partir de 2024 o de '24'."""
    a = str(anio).strip()
    if len(a) == 4:
        aa = a[2:]
    elif len(a) == 2:
        aa = a
    else:
        return []
    if not aa.isdigit():
        return []
    return [f"{aa}{m:02d}" for m in range(1, 13)]


def fmt_tiempo(seg):
    m, s = divmod(int(seg), 60)
    h, m = divmod(m, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def ahora():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ==========================================================================
# Resolucion de rutas de los .mdb
# ==========================================================================
def ruta_json_mes(aamm):
    return _sal.carpeta_mes(CONFIG_RAIZ, aamm) / NOMBRE_JSON_MES


def rutas_desde_json_mes(aamm):
    """Lee el JSON de traspaso del mes. Devuelve (rdef, rpre) como dicts."""
    data = leer_json(ruta_json_mes(aamm), {})
    if not isinstance(data, dict):
        return {}, {}
    rutas = data.get("rutas") or {}
    rdef = {}
    if isinstance(rutas, dict):
        if rutas.get("mdb_sscc"):
            rdef["sscc"] = rutas["mdb_sscc"]
        if rutas.get("mdb_sob"):
            rdef["sob"] = rutas["mdb_sob"]
    propio = data.get(CLAVE_JSON_PROPIA) or {}
    rpre = {}
    if isinstance(propio, dict):
        blo = propio.get("rpre") or {}
        if isinstance(blo, dict):
            if blo.get("mdb_sscc"):
                rpre["sscc"] = blo["mdb_sscc"]
            if blo.get("mdb_sob"):
                rpre["sob"] = blo["mdb_sob"]
    return rdef, rpre


def guardar_rpre_en_json_mes(aamm, slot, ruta, log=print):
    """Agrega la ruta Rpre al JSON del mes bajo una clave propia.

    Solo agrega/actualiza `comparador_etapas`; nunca toca `rutas`, `planilla`
    ni ninguna otra clave, asi que los demas scripts siguen leyendo lo mismo.
    """
    p = ruta_json_mes(aamm)
    data = leer_json(p, {})
    if data is None:
        log(f"  ! {p.name} de {aamm} existe pero no se pudo interpretar: no se escribe.")
        return False
    if not isinstance(data, dict):
        log(f"  ! {p.name} de {aamm} no es un objeto JSON: no se escribe.")
        return False
    if not p.exists():
        # No inventamos un JSON de traspaso completo: solo el bloque propio.
        data = {"origen": "Comparador_Etapas", "version": 1, "aamm": aamm}
    bloque = data.setdefault(CLAVE_JSON_PROPIA, {})
    if not isinstance(bloque, dict):
        bloque = {}
        data[CLAVE_JSON_PROPIA] = bloque
    rpre = bloque.setdefault("rpre", {})
    if not isinstance(rpre, dict):
        rpre = {}
        bloque["rpre"] = rpre
    rpre["mdb_sscc" if slot == "sscc" else "mdb_sob"] = str(ruta)
    bloque["actualizado"] = ahora()
    try:
        escribir_json_atomico(p, data)
        return True
    except Exception as e:
        log(f"  ! No se pudo escribir {p}: {e}")
        return False


def buscar_mdb_arbol(base, patron, profundidad=4):
    """Ultimo recurso: baja por el arbol buscando el .mdb que calce.

    Se limita la profundidad y se saltan las carpetas de detalles diarios, que
    son las caras en un disco de red.
    """
    try:
        base = Path(base)
        if not base.is_dir():
            return None
        cands = []
        raiz_partes = len(base.parts)
        for actual, dirs, archivos in os.walk(base):
            pa = Path(actual)
            if len(pa.parts) - raiz_partes >= profundidad:
                dirs[:] = []
            dirs[:] = [d for d in dirs
                       if "detalle" not in normalizar_suave(d)
                       and not d.startswith(".")]
            for nom in archivos:
                if (Path(nom).suffix.lower() in (".mdb", ".accdb")
                        and patron.search(nom) and not es_copia(nom)
                        and not nom.startswith("~$")):
                    cands.append(pa / nom)
        if not cands:
            return None
        cands.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        return cands[0]
    except Exception:
        return None


# Tramos de <CMgReales>\AAMM\Sobrecostos\02 Definitivo\Auxiliares.
# Cada tramo trae alternativas por si alguien renombro la carpeta.
TRAMOS_DEF = [
    ("Sobrecostos", ("Sobrecosto",)),
    ("02 Definitivo", ("Definitivo",)),
    ("Auxiliares", ("Auxiliar", "Auxiliares Definitivo")),
]


def carpeta_definitivo(raiz_cmg, aamm):
    """Devuelve (carpeta_auxiliares | None, hasta_donde_se_llego).

    El segundo valor es para mostrarlo en la ventana: si la estructura cambio,
    conviene ver en que tramo se corto en lugar de un "sin ruta" pelado.
    """
    if not raiz_cmg:
        return None, "falta la carpeta CMgReales"
    raiz = Path(raiz_cmg)
    if not raiz.is_dir():
        return None, f"no se ve la carpeta {raiz}"
    p = subcarpeta(raiz, aamm)
    if p is None:
        cand = raiz / aamm
        p = cand if cand.is_dir() else None
    if p is None:
        return None, f"llego hasta {raiz} — no hay carpeta {aamm}"
    for tramo, alternativas in TRAMOS_DEF:
        sig = subcarpeta(p, tramo)
        if sig is None:
            for alt in alternativas:
                sig = subcarpeta(p, alt)
                if sig is not None:
                    break
        if sig is None:
            return None, f"llego hasta {p} — falta '{tramo}'"
        p = sig
    return p, str(p)


# Subcarpetas del arbol de reliquidacion donde vive cada .mdb
SUB_POR_SLOT = {
    "sscc": ("01 Sobrecostos", ("Sobrecostos",)),
    "sob": ("01.a Sobrecostos de Energia", ("Sobrecostos de Energia",
                                            "01a Sobrecostos de Energia")),
}


def mdb_presentes(base, tope=6, profundidad=3):
    """Nombres de los .mdb que hay bajo una carpeta, para poder diagnosticar.

    Cuando el patron no calza con nada, saber que archivos hay ahi es lo unico
    que permite entender por que.
    """
    vistos = []
    try:
        base = Path(base)
        raiz_partes = len(base.parts)
        for actual, dirs, archivos in os.walk(base):
            if len(Path(actual).parts) - raiz_partes >= profundidad:
                dirs[:] = []
            dirs[:] = [d for d in dirs if "detalle" not in normalizar_suave(d)]
            for nom in archivos:
                if (Path(nom).suffix.lower() in (".mdb", ".accdb")
                        and not nom.startswith("~$")):
                    vistos.append(nom)
                    if len(vistos) >= tope:
                        return vistos
    except Exception:
        pass
    return vistos


def mdb_en_carpeta(carpeta, log=print):
    """Encuentra los dos .mdb a partir de UNA carpeta.

    Sirve igual para la carpeta `02 CASO RELIQUIDACION` (busca en
    `01 Sobrecostos` y `01.a Sobrecostos de Energia`) y para una carpeta
    `Auxiliares` donde los dos estan juntos. Si no los ve donde deberian,
    baja por el arbol como ultimo recurso.

    Devuelve ({slot: ruta}, diagnostico).
    """
    base = Path(carpeta)
    if not base.is_dir():
        return {}, f"no se ve la carpeta {base}"
    # Si apuntaron a la carpeta del mes, bajar sola al caso.
    if normalizar_suave(base.name) != normalizar_suave("02 CASO RELIQUIDACION"):
        cr = subcarpeta(base, "02 CASO RELIQUIDACION")
        if cr is not None:
            base = cr
    out, faltan = {}, []
    for slot in SLOTS:
        patron = PAT_SSCC if slot == "sscc" else PAT_SOB
        sub, alternativas = SUB_POR_SLOT[slot]
        c = subcarpeta(base, sub)
        if c is None:
            for alt in alternativas:
                c = subcarpeta(base, alt)
                if c is not None:
                    break
        f = buscar_mdb(c, patron) if c is not None else None
        if f is None:
            f = buscar_mdb(base, patron)          # todos juntos en la carpeta
        if f is None:
            f = buscar_mdb_arbol(base, patron)    # ultimo recurso
        if f is not None:
            out[slot] = str(f)
        else:
            faltan.append(ETIQUETA_SLOT[slot])
    diag = ""
    if faltan:
        diag = f"en {base} no aparece: {', '.join(faltan)}"
        hay = mdb_presentes(base)
        if hay:
            diag += f"  ·  si hay: {', '.join(hay)}"
    return out, diag


def resolver_rutas(aamm, raiz_cmg, rutas_manuales):
    """({etapa: {slot: ruta|None}}, {etapa: diagnostico}) para un mes.

    Prioridad, de menor a mayor: autodeteccion > JSON del mes >
    carpeta elegida a mano > archivo elegido a mano.
    """
    res = {e: {s: None for s in SLOTS} for e in ETAPAS}
    diag = {e: "" for e in ETAPAS}

    # 1) Definitivo por autodeteccion en CMgReales
    carp, hasta = carpeta_definitivo(raiz_cmg, aamm)
    if carp is None:
        diag["def"] = hasta
    else:
        encontrados, d = mdb_en_carpeta(carp)
        res["def"].update(encontrados)
        diag["def"] = d or hasta

    # 2) Rdef y Rpre desde el JSON del mes
    rdef, rpre = rutas_desde_json_mes(aamm)
    for slot, r in rdef.items():
        res["rdef"][slot] = r
    for slot, r in rpre.items():
        res["rpre"][slot] = r

    man = (rutas_manuales or {}).get(aamm, {})
    if not isinstance(man, dict):
        man = {}

    # 3) Carpeta elegida a mano: se resuelve cada vez, asi una revision nueva
    #    (R01E donde habia R01D) se toma sola sin volver a examinar.
    for etapa in ETAPAS:
        blo = man.get(etapa) or {}
        if isinstance(blo, dict) and blo.get("carpeta"):
            encontrados, d = mdb_en_carpeta(blo["carpeta"])
            res[etapa].update(encontrados)
            if d:
                diag[etapa] = d
            elif etapa != "def":
                diag[etapa] = str(blo["carpeta"])

    # 4) Archivo elegido a mano: manda sobre todo lo demas
    for etapa in ETAPAS:
        blo = man.get(etapa) or {}
        if isinstance(blo, dict):
            for slot in SLOTS:
                if blo.get(slot):
                    res[etapa][slot] = blo[slot]
    return res, diag


# ==========================================================================
# Lectura de los .mdb
# ==========================================================================
def driver_access():
    for d in pyodbc.drivers():
        if "Microsoft Access Driver" in d:
            return d
    return None


def conectar(ruta_mdb):
    drv = driver_access()
    if not drv:
        raise RuntimeError(
            "No hay driver 'Microsoft Access Driver (*.mdb, *.accdb)' instalado, "
            "o es de otra arquitectura (32/64 bits) que este Python."
        )
    cs = f"DRIVER={{{drv}}};DBQ={ruta_mdb};ReadOnly=1;"
    return pyodbc.connect(cs, autocommit=True)


def tablas(con):
    cur = con.cursor()
    return [row.table_name for row in cur.tables(tableType="TABLE")]


def tabla_por_nombre(nombres, candidatos):
    """Primer candidato presente, comparando normalizado."""
    mapa = {normalizar(n): n for n in nombres}
    for c in candidatos:
        real = mapa.get(normalizar(c))
        if real:
            return real
    return None


def mapear_columnas(cols, mapa):
    """{nombre_real: nombre_canonico} segun el mapa normalizado."""
    out = {}
    for c in cols:
        can = mapa.get(normalizar(c))
        if can and can not in out.values():
            out[c] = can
    return out


def leer_sobrecostos(ruta_mdb, aamm, log=print):
    """DataFrame con clave_aamm, tipo, central, hora_mes, sobrecosto."""
    with conectar(ruta_mdb) as con:
        tn = tabla_por_nombre(tablas(con), ["Sobrecostos", "Sobrecosto"])
        if not tn:
            raise RuntimeError(f"No hay tabla Sobrecostos en {Path(ruta_mdb).name}")
        cur = con.cursor()
        cur.execute(f"SELECT * FROM [{tn}] WHERE 1=0")
        cols = [d[0] for d in cur.description]
        ren = mapear_columnas(cols, MAPA_SOB)
        faltan = {"tipo", "central", "hora_mes", "sobrecosto"} - set(ren.values())
        if faltan:
            raise RuntimeError(
                f"{Path(ruta_mdb).name}: tabla [{tn}] sin columnas {sorted(faltan)}. "
                f"Encontradas: {cols}")
        sel = ", ".join(f"[{c}]" for c in ren)
        partes = []
        for chunk in pd.read_sql(f"SELECT {sel} FROM [{tn}]", con, chunksize=200_000):
            partes.append(chunk.rename(columns=ren))
        df = (pd.concat(partes, ignore_index=True) if partes
              else pd.DataFrame(columns=list(ren.values())))

    if "clave_aamm" not in df.columns:
        df["clave_aamm"] = aamm
    df["clave_aamm"] = df["clave_aamm"].astype(str).str.strip()

    # El .mdb deberia traer solo el mes; si trae mas, se filtra y se avisa.
    if len(df):
        propios = df["clave_aamm"] == aamm
        if propios.any() and not propios.all():
            log(f"      · se descartan {int((~propios).sum())} filas de otras "
                f"claves Anio_Mes.")
            df = df[propios]
        elif not propios.any():
            log(f"      ! ninguna fila con clave {aamm} (claves vistas: "
                f"{sorted(df['clave_aamm'].unique())[:5]}). Se consolidan bajo {aamm}.")
            df["clave_aamm"] = aamm

    df["central"] = df["central"].astype(str).str.strip()
    df["tipo"] = df["tipo"].astype(str).str.strip().str.upper()
    df["hora_mes"] = pd.to_numeric(df["hora_mes"], errors="coerce").astype("Int32")
    df["sobrecosto"] = pd.to_numeric(df["sobrecosto"], errors="coerce").fillna(0.0)
    df = df[df["central"].str.len() > 0]
    df = df[df["central"].str.lower() != "nan"]
    return df[["clave_aamm", "tipo", "central", "hora_mes", "sobrecosto"]]


def leer_centrales(ruta_mdb, log=print):
    """DataFrame con central, empresa, tabla_origen.

    Busca Central_Empresa_Actualizada; si no esta, Central_Empresa.
    """
    vacio = pd.DataFrame(columns=["central", "empresa", "tabla_origen"])
    with conectar(ruta_mdb) as con:
        nombres = tablas(con)
        tn = tabla_por_nombre(nombres, ["Central_Empresa_Actualizada"])
        if not tn:
            tn = tabla_por_nombre(nombres, ["Central_Empresa"])
        if not tn:
            log(f"      ! {Path(ruta_mdb).name}: sin tabla de propietarios.")
            return vacio
        cur = con.cursor()
        cur.execute(f"SELECT * FROM [{tn}] WHERE 1=0")
        cols = [d[0] for d in cur.description]
        ren = mapear_columnas(cols, MAPA_CEN)
        if "central" not in ren.values() or "empresa" not in ren.values():
            log(f"      ! [{tn}] sin Central/Empresa ({cols}).")
            return vacio
        sel = ", ".join(f"[{c}]" for c in ren)
        df = pd.read_sql(f"SELECT {sel} FROM [{tn}]", con).rename(columns=ren)

    df["central"] = df["central"].astype(str).str.strip()
    df["empresa"] = df["empresa"].astype(str).str.strip()
    df = df[(df["central"].str.len() > 0) & (df["central"].str.lower() != "nan")]
    df = df.drop_duplicates(subset=["central"], keep="first")
    df["tabla_origen"] = tn
    return df[["central", "empresa", "tabla_origen"]]


# ==========================================================================
# Propietario vigente ("Actual") desde T:\Facturacion
# ==========================================================================
def carpetas_facturacion(raiz):
    """[(anio, mes_num, carpeta_mes)] de la mas nueva a la mas vieja."""
    out = []
    try:
        raiz = Path(raiz)
        if not raiz.is_dir():
            return out
        anios = sorted(
            [e for e in listar(raiz)
             if e.es_dir and PAT_ANIO_DIR.match(e.nombre.strip())],
            key=lambda e: int(e.nombre.strip()), reverse=True)
        for da in anios:
            meses = []
            for e in listar(da.ruta):
                if not e.es_dir:
                    continue
                mm = PAT_MES_DIR.match(e.nombre.strip())
                if mm and 1 <= int(mm.group(1)) <= 12:
                    meses.append((int(mm.group(1)), e.ruta))
            meses.sort(key=lambda x: x[0], reverse=True)
            for num, dm in meses:
                out.append((int(da.nombre.strip()), num, dm))
    except Exception:
        pass
    return out


def buscar_maestro_actual(raiz_fact, log=print):
    """El 4_REMUNERACION_SC_CO mas nuevo disponible.

    Recorre anio y mes de mas nuevo a mas viejo; dentro de cada mes prueba
    primero `02 Definitivo\\SSCC` y despues `01 Preliminar\\SSCC`. El primero
    que encuentre gana. Devuelve (ruta, etiqueta) o (None, None).
    """
    for anio, num, dmes in carpetas_facturacion(raiz_fact):
        for rama in RAMAS_FACT:
            cr = subcarpeta(dmes, rama)
            if cr is None:
                continue
            cs = subcarpeta(cr, "SSCC")
            if cs is None:
                continue
            f = buscar_archivo(cs, PAT_REMUN, (".xlsx", ".xlsm", ".xlsb"))
            if f:
                etiqueta = f"{anio}-{num:02d} · {rama}"
                log(f"  Propietario actual: {etiqueta} · {f.name}")
                return f, etiqueta
    return None, None


def buscar_archivo(carpeta, patron, extensiones):
    """Mas reciente que calce, descartando copias de Windows."""
    try:
        carpeta = Path(carpeta)
        if not carpeta.is_dir():
            return None
        cands = [
            e for e in listar(carpeta)
            if not e.es_dir
            and Path(e.nombre).suffix.lower() in extensiones
            and patron.search(e.nombre)
            and not es_copia(e.nombre)
            and not e.nombre.startswith("~$")
        ]
        if not cands:
            return None
        cands.sort(key=lambda e: e.mtime, reverse=True)
        return cands[0].ruta
    except Exception:
        return None


def leer_config_empresa(ruta, log=print):
    """Hoja 'Configuracion Empresa', columnas A y B, datos desde la fila 2."""
    wb = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
    try:
        objetivo = normalizar(HOJA_CONFIG_EMPRESA)
        nombre = None
        for n in wb.sheetnames:
            if normalizar(n) == objetivo:
                nombre = n
                break
        if nombre is None:
            for n in wb.sheetnames:
                if objetivo in normalizar(n):
                    nombre = n
                    break
        if nombre is None:
            raise RuntimeError(
                f"{Path(ruta).name}: no hay hoja '{HOJA_CONFIG_EMPRESA}' "
                f"(hojas: {wb.sheetnames})")
        sh = wb[nombre]
        filas = []
        for i, fila in enumerate(sh.iter_rows(min_row=2, min_col=1, max_col=2,
                                              values_only=True), start=2):
            central = str(fila[0]).strip() if fila[0] is not None else ""
            empresa = str(fila[1]).strip() if len(fila) > 1 and fila[1] is not None else ""
            if not central or central.lower() in ("nan", "0"):
                continue
            filas.append((central, empresa))
        log(f"  Hoja '{nombre}': {len(filas)} centrales.")
    finally:
        wb.close()

    df = pd.DataFrame(filas, columns=["central", "empresa_actual"])
    df["central_norm"] = df["central"].map(normalizar)
    dup = df["central_norm"].duplicated(keep=False)
    if dup.any():
        rep = sorted(df.loc[dup, "central"].unique().tolist())
        log(f"  ! Centrales repetidas en la configuracion: {', '.join(rep[:10])}"
            + (" ..." if len(rep) > 10 else ""))
        df = df.drop_duplicates(subset=["central_norm"], keep="first")
    return df


def consolidar_actual(anio, raiz_fact, est, forzar=False, log=print):
    """Deja el propietario vigente en parquet. Idempotente por huella."""
    ruta, etiqueta = buscar_maestro_actual(raiz_fact, log=log)
    if ruta is None:
        log("  ! No se encontro ningun 4_REMUNERACION_SC_CO bajo Facturacion.")
        return False
    reg = est.get("_actual") or {}
    if (not forzar and reg.get("huella") == huella(ruta)
            and actual_parquet(anio).exists()):
        log(f"  Propietario actual ya al dia ({Path(reg.get('archivo', '')).name}).")
        return False
    df = leer_config_empresa(ruta, log=log)
    dir_actual(anio).mkdir(parents=True, exist_ok=True)
    pq.write_table(pa.Table.from_pandas(df, preserve_index=False),
                   actual_parquet(anio), compression="snappy")
    est["_actual"] = {
        "archivo": str(ruta),
        "etiqueta": etiqueta,
        "huella": huella(ruta),
        "centrales": int(len(df)),
        "leido": ahora(),
    }
    # Cambia la referencia: todas las vistas quedan obsoletas
    for k, v in est.items():
        if isinstance(v, dict) and not k.startswith("_"):
            v.pop("vista", None)
    log(f"  Propietario actual consolidado: {len(df)} centrales ({etiqueta}).")
    return True


# ==========================================================================
# Estado de consolidacion
# ==========================================================================
def mes_incluido(est, aamm):
    """Si el mes entra al consolidado anual. Por defecto si."""
    reg = est.get(aamm) or {}
    return bool(reg.get("incluir", True))


def fijar_incluido(est, aamm, valor):
    est.setdefault(aamm, {})["incluir"] = bool(valor)


def cargar_estado(anio):
    est = leer_json(estado_path(anio), {})
    return est if isinstance(est, dict) else {}


def guardar_estado(anio, est):
    escribir_json_atomico(estado_path(anio), est)


def estado_etapa(est, aamm, etapa, rutas):
    """'falta' | 'pendiente' | 'desactualizado' | 'ok'"""
    r = rutas[etapa]
    presentes = [s for s in SLOTS if r.get(s) and Path(r[s]).is_file()]
    if len(presentes) < len(SLOTS):
        return "falta"
    reg = ((est.get(aamm) or {}).get(etapa) or {})
    if not reg.get("consolidado"):
        return "pendiente"
    for s in SLOTS:
        if reg.get("huellas", {}).get(s) != huella(r[s]):
            return "desactualizado"
    return "ok"


def color_de(estado):
    return {
        "falta": COLORES["rojo"],
        "pendiente": COLORES["amarillo"],
        "desactualizado": COLORES["amarillo"],
        "ok": COLORES["verde"],
    }.get(estado, COLORES["gris"])


def estado_mes(est, aamm, rutas):
    ests = [estado_etapa(est, aamm, e, rutas) for e in ETAPAS]
    if all(x == "ok" for x in ests):
        return "ok"
    if any(x == "falta" for x in ests):
        return "falta"
    return "pendiente"


# ==========================================================================
# Consolidacion -> parquet
# ==========================================================================
def dir_sob(aamm, etapa):
    return dir_sob_raiz(_anio_de(aamm)) / f"aamm={aamm}" / f"etapa={etapa}"


def dir_cen(aamm, etapa):
    return dir_cen_raiz(_anio_de(aamm)) / f"aamm={aamm}" / f"etapa={etapa}"


def path_vista(aamm):
    return dir_vistas(_anio_de(aamm)) / f"vista_{aamm}.parquet"


def dir_resultados_anuales(anio):
    """00_Salidas/AAAA, directo bajo el anio: 00_Salidas solo tiene resultados,
    el estado y los datos intermedios del comparador viven en __config__."""
    normal = _sal.normalizar_anio(anio)
    if normal is None:
        raise ValueError(f"anio no reconocido: {anio!r}")
    return SALIDAS / normal


def path_excel_mes(aamm):
    """El Excel propio del mes, en su carpeta 00_Salidas/AAAA/MM Mes."""
    return _sal.carpeta_mes(SALIDAS, aamm) / f"Comparacion_{aamm}.xlsx"


def path_excel_anual(anio_aa):
    """El consolidado con todos los meses disponibles, directo bajo el anio."""
    return dir_resultados_anuales(anio_aa) / f"Comparacion_Etapas_{_sal.normalizar_anio(anio_aa)}.xlsx"


def firma_vistas(meses):
    """Que meses entran al consolidado y con que version de su vista.

    Si esta firma no cambio, el consolidado anual ya esta al dia y no hay para
    que reescribirlo.
    """
    f = {}
    for m in meses:
        v = path_vista(m)
        if v.exists():
            f[m] = int(v.stat().st_mtime)
    return f


def consolidar_etapa(aamm, etapa, rutas, est, log=print):
    """Lee los 2 .mdb de una etapa y reescribe SOLO su particion."""
    r = rutas[etapa]
    for s in SLOTS:
        if not (r.get(s) and Path(r[s]).is_file()):
            log(f"  {aamm} {ETIQUETA[etapa]}: falta el .mdb {ETIQUETA_SLOT[s]}, se salta.")
            return False

    dfs_sob, dfs_cen, huellas, detalle = [], [], {}, {}
    for s in SLOTS:
        ruta = r[s]
        log(f"  {aamm} {ETIQUETA[etapa]} · {Path(ruta).name}")
        d = leer_sobrecostos(ruta, aamm, log=log)
        d["slot"] = s
        dfs_sob.append(d)
        c = leer_centrales(ruta, log=log)
        c["slot"] = s
        dfs_cen.append(c)
        huellas[s] = huella(ruta)
        tipos = sorted(d["tipo"].unique().tolist())
        detalle[s] = {
            "archivo": Path(ruta).name,
            "filas": int(len(d)),
            "tipos": tipos,
            "total": float(d["sobrecosto"].sum()),
            "tabla_propietarios": (c["tabla_origen"].iloc[0] if len(c) else None),
        }
        log(f"      {len(d):,} filas · tipos {', '.join(tipos) or '-'} · "
            f"total {d['sobrecosto'].sum():,.2f}")

    df = pd.concat(dfs_sob, ignore_index=True)
    df["etapa"] = etapa
    cen = pd.concat(dfs_cen, ignore_index=True)
    # Un propietario por central por etapa: gana el que venga de la tabla
    # Actualizada, y a igualdad el del SSCC.
    cen["prio"] = cen["tabla_origen"].map(
        lambda t: 0 if normalizar(t) == normalizar("Central_Empresa_Actualizada") else 1
    )
    cen["prio2"] = cen["slot"].map({"sscc": 0, "sob": 1}).fillna(9)
    cen = (cen.sort_values(["central", "prio", "prio2"])
              .drop_duplicates(subset=["central"], keep="first"))
    cen["etapa"] = etapa

    d1, d2 = dir_sob(aamm, etapa), dir_cen(aamm, etapa)
    d1.mkdir(parents=True, exist_ok=True)
    d2.mkdir(parents=True, exist_ok=True)
    pq.write_table(pa.Table.from_pandas(
        df[["clave_aamm", "tipo", "central", "hora_mes", "sobrecosto", "slot"]],
        preserve_index=False), d1 / "datos.parquet", compression="snappy")
    pq.write_table(pa.Table.from_pandas(
        cen[["central", "empresa", "tabla_origen", "slot"]],
        preserve_index=False), d2 / "datos.parquet", compression="snappy")

    est.setdefault(aamm, {})[etapa] = {
        "consolidado": ahora(),
        "huellas": huellas,
        "filas": int(len(df)),
        "total": float(df["sobrecosto"].sum()),
        "centrales": int(len(cen)),
        "detalle": detalle,
    }
    # La vista del mes queda obsoleta
    est[aamm].pop("vista", None)
    log(f"  {aamm} {ETIQUETA[etapa]}: consolidado, {len(df):,} filas.")
    return True


# ==========================================================================
# Vista comparada del mes (parquet cache) y Excel
# ==========================================================================
def una_fila(cur, defecto=None):
    """fetchone() que nunca devuelve None: un COUNT/SUM siempre trae fila,
    pero el tipo declarado es Optional y hay que desempaquetarlo con cuidado."""
    fila = cur.fetchone()
    return fila if fila is not None else defecto


def etapas_consolidadas(aamm):
    return [e for e in ETAPAS if (dir_sob(aamm, e) / "datos.parquet").exists()]


def construir_vista(aamm, log=print):
    """Pivotea las etapas del mes en una tabla ancha y la deja en cache."""
    disp = etapas_consolidadas(aamm)
    if not disp:
        log(f"  {aamm}: nada consolidado, sin vista.")
        return None
    dir_vistas(_anio_de(aamm)).mkdir(parents=True, exist_ok=True)
    con = duckdb.connect()
    try:
        sob = str((dir_sob_raiz(_anio_de(aamm)) / f"aamm={aamm}").as_posix()) + "/**/*.parquet"
        cen = str((dir_cen_raiz(_anio_de(aamm)) / f"aamm={aamm}").as_posix()) + "/**/*.parquet"
        sel_m = ",\n".join(
            f"SUM(CASE WHEN etapa = '{e}' THEN sobrecosto END) AS m_{e}" for e in ETAPAS
        )
        sel_e = ",\n".join(
            f"MAX(CASE WHEN etapa = '{e}' THEN empresa END) AS emp_{e}" for e in ETAPAS
        )
        # El propietario vigente es opcional: si no se ha leido, va en NULL.
        if actual_parquet(_anio_de(aamm)).exists():
            cte_actual = (f"SELECT central_norm, empresa_actual "
                          f"FROM read_parquet('{actual_parquet(_anio_de(aamm)).as_posix()}')")
        else:
            cte_actual = ("SELECT NULL::VARCHAR AS central_norm, "
                          "NULL::VARCHAR AS empresa_actual WHERE FALSE")
        def monto(e):
            return (f"COALESCE(m_{e}, 0.0)" if e in disp else "NULL::DOUBLE")

        def delta(a, b):
            # Si a alguna de las dos etapas todavia no la tengo, la diferencia
            # no significa nada: va en blanco en vez de un numero enorme.
            if a in disp and b in disp:
                return f"COALESCE(m_{a}, 0.0) - COALESCE(m_{b}, 0.0)"
            return "NULL::DOUBLE"

        etapas_txt = ", ".join(ETIQUETA[e] for e in disp)
        detalle = " || ' | ' || ".join(
            [f"'{ETIQUETA[e]}: ' || COALESCE(emp_{e}, '(no esta)')" for e in ETAPAS]
            + ["'Actual: ' || COALESCE(empresa_actual, '(no esta)')"])

        sql = f"""
        WITH s AS (
            SELECT etapa, central, tipo, hora_mes, SUM(sobrecosto) AS sobrecosto
            FROM read_parquet('{sob}', hive_partitioning = 1)
            GROUP BY 1, 2, 3, 4
        ),
        p AS (
            SELECT central, tipo, hora_mes, {sel_m}
            FROM s GROUP BY 1, 2, 3
        ),
        c AS (
            SELECT central, {sel_e}
            FROM read_parquet('{cen}', hive_partitioning = 1)
            GROUP BY 1
        ),
        act AS ({cte_actual}),
        base AS (
            SELECT
                p.central, p.tipo, p.hora_mes, p.m_def, p.m_rpre, p.m_rdef,
                {", ".join(f"c.emp_{e}" for e in ETAPAS)},
                a.empresa_actual
            FROM p
            LEFT JOIN c ON c.central = p.central
            LEFT JOIN act a
                ON a.central_norm = lower(replace(replace(
                       strip_accents(p.central), ' ', ''), '_', ''))
        )
        SELECT
            '{aamm}'                                   AS aamm,
            central,
            COALESCE(empresa_actual, {", ".join(f"emp_{e}" for e in reversed(ETAPAS))},
                     '(sin propietario)')              AS empresa,
            {", ".join(f"emp_{e}" for e in ETAPAS)},
            empresa_actual,
            len(list_distinct(list_filter(
                [{", ".join(f"emp_{e}" for e in ETAPAS)}, empresa_actual],
                x -> x IS NOT NULL AND x <> ''))) > 1  AS empresa_cambia,
            (empresa_actual IS NOT NULL AND EXISTS (
                SELECT 1 FROM (SELECT UNNEST(list_filter(
                    [{", ".join(f"emp_{e}" for e in ETAPAS)}],
                    x -> x IS NOT NULL AND x <> '')) AS v) t
                WHERE t.v <> empresa_actual))          AS dif_vs_actual,
            (empresa_actual IS NULL)                   AS sin_actual,
            CASE WHEN len(list_distinct(list_filter(
                [{", ".join(f"emp_{e}" for e in ETAPAS)}, empresa_actual],
                x -> x IS NOT NULL AND x <> ''))) > 1
                 THEN {detalle} END                    AS detalle_propietario,
            tipo,
            hora_mes,
            {monto("def")}                             AS def,
            {monto("rpre")}                            AS rpre,
            {monto("rdef")}                            AS rdef,
            {delta("rpre", "def")}                     AS d_rpre_def,
            {delta("rdef", "rpre")}                    AS d_rdef_rpre,
            {delta("rdef", "def")}                     AS d_rdef_def,
            '{etapas_txt}'                             AS etapas,
            (m_def  IS NOT NULL)                       AS en_def,
            (m_rpre IS NOT NULL)                       AS en_rpre,
            (m_rdef IS NOT NULL)                       AS en_rdef
        FROM base
        ORDER BY central, tipo, hora_mes
        """
        con.execute(
            f"COPY ({sql}) TO '{path_vista(aamm).as_posix()}' "
            f"(FORMAT PARQUET, COMPRESSION SNAPPY)"
        )
        n = una_fila(con.execute(
            f"SELECT COUNT(*) FROM read_parquet('{path_vista(aamm).as_posix()}')"),
            (0,))[0]
    finally:
        con.close()
    log(f"  {aamm}: vista rearmada ({n:,} filas, etapas {', '.join(ETIQUETA[e] for e in disp)}).")
    return n


# Si se agrega una columna a la vista, las vistas ya escritas en disco quedan
# viejas. Antes de exportar se compara el esquema del parquet contra esta lista
# y las que no calzan se rearman solas: es barato (no toca ningun Access) y
# evita un "columna no encontrada" al exportar.
COLUMNAS_VISTA = [
    "aamm", "central", "empresa", "emp_def", "emp_rpre", "emp_rdef",
    "empresa_actual", "empresa_cambia", "dif_vs_actual", "sin_actual",
    "detalle_propietario", "tipo", "hora_mes", "def", "rpre", "rdef",
    "d_rpre_def", "d_rdef_rpre", "d_rdef_def", "etapas",
    "en_def", "en_rpre", "en_rdef",
]


def vista_completa(aamm):
    """True si la vista existe y trae todas las columnas que se esperan hoy."""
    v = path_vista(aamm)
    if not v.exists():
        return False
    try:
        presentes = set(pq.read_schema(v).names)
    except Exception:
        return False
    return not (set(COLUMNAS_VISTA) - presentes)


def asegurar_vista(aamm, log=print):
    """Rearma la vista del mes si falta o si quedo con el formato viejo."""
    if vista_completa(aamm):
        return True
    if not etapas_consolidadas(aamm):
        return False
    if path_vista(aamm).exists():
        log(f"  {aamm}: la vista es de una version anterior, se rearma.")
    return construir_vista(aamm, log=log) is not None


COLUMNAS_EXCEL = [
    ("central", "Central"),
    ("empresa", "Empresa"),
    ("tipo", "Tipo"),
    ("hora_mes", "Hora Mensual"),
    ("def", "Def"),
    ("rpre", "Rpre"),
    ("rdef", "Rdef"),
    ("d_rpre_def", "Rpre - Def"),
    ("d_rdef_rpre", "Rdef - Rpre"),
    ("d_rdef_def", "Rdef - Def"),
    ("emp_def", "Emp Def"),
    ("emp_rpre", "Emp Rpre"),
    ("emp_rdef", "Emp Rdef"),
    ("empresa_actual", "Emp Actual"),
    ("empresa_cambia", "Cambio propietario"),
    ("dif_vs_actual", "Difiere del actual"),
    ("sin_actual", "No esta en config. actual"),
    ("detalle_propietario", "Detalle propietario"),
    ("en_def", "En Def"),
    ("en_rpre", "En Rpre"),
    ("en_rdef", "En Rdef"),
]
COL_DELTAS = (7, 8, 9)          # columnas de diferencia, se resaltan
COL_NUM = (4, 5, 6, 7, 8, 9)    # columnas de monto
COL_HORA = 3

LIMITE_FILAS_HOJA = 1_048_000


# Tipos de sobrecosto de cada familia, en el orden en que se muestran
FAMILIAS = [
    ("SSCC", ["CCA", "CO", "SCCF"]),
    ("Energia", ["SCMT", "SCPC"]),
]
# Hojas que genera este programa. Cualquier otra hoja del archivo es de alguien
# mas y no se toca (ver escribir_preservando).
HOJAS_PROPIAS_FIJAS = ["RESUMEN SSCC", "RESUMEN ENERGIA", "PROPIETARIOS"]


def es_hoja_propia(nombre, meses=None):
    """True si la hoja la genera este programa (y por lo tanto se puede pisar).

    Una hoja de mes se reconoce por su nombre (AAMM o AAMM_2), no por estar en
    la lista de meses que se exporta ahora: si un mes se saca del consolidado,
    su hoja tiene que desaparecer, no quedar congelada como si fuera de otro.
    """
    if nombre in HOJAS_PROPIAS_FIJAS or nombre == "RESUMEN":
        return True
    base, _, resto = nombre.partition("_")
    if not (len(base) == 4 and base.isdigit() and 1 <= int(base[2:]) <= 12):
        return False
    return resto == "" or resto.isdigit()


def datos_resumen(con, pvs_por_mes, familia_tipos):
    """Bloques del resumen: por cada mes, una fila por tipo mas el total.

    Devuelve [(aamm, [(etiqueta, def, rpre, rdef, d1, d2, d3), ...]), ...]
    """
    bloques = []
    for aamm, pvs in pvs_por_mes:
        lista = ", ".join(f"'{x}'" for x in familia_tipos)
        filas = con.execute(f"""
            SELECT tipo, SUM(def), SUM(rpre), SUM(rdef),
                   SUM(d_rpre_def), SUM(d_rdef_rpre), SUM(d_rdef_def)
            FROM read_parquet('{pvs}')
            WHERE tipo IN ({lista})
            GROUP BY tipo
        """).fetchall()
        por_tipo = {r[0]: r[1:] for r in filas}
        if not por_tipo:
            continue
        detalle = []
        for tipo in familia_tipos:
            detalle.append((tipo,) + tuple(por_tipo.get(tipo, (None,) * 6)))
        tot = con.execute(f"""
            SELECT SUM(def), SUM(rpre), SUM(rdef),
                   SUM(d_rpre_def), SUM(d_rdef_rpre), SUM(d_rdef_def)
            FROM read_parquet('{pvs}')
            WHERE tipo IN ({lista})
        """).fetchone()
        detalle.append(("Total",) + tuple(tot))
        bloques.append((aamm, detalle))
    return bloques


CAB_RESUMEN = ["", "Def", "Rpre", "Rdef", "Rpre - Def", "Rdef - Rpre", "Rdef - Def"]


def datos_hoja_mes(con, pvs, claves, filtro, orden=True):
    """Cursor con las filas del detalle de un mes."""
    return con.execute(
        f"SELECT {', '.join(claves)} FROM read_parquet('{pvs}') {filtro}"
        + (" ORDER BY central, tipo, hora_mes" if orden else ""))


def exportar_excel(destino, meses, solo_dif, tolerancia, log=print,
                   preservar=True):
    """Escribe el libro de comparacion.

    Si el archivo ya existe y alguien agrego hojas propias, se conservan: se
    reescriben unicamente las hojas que genera este programa.
    """
    for m in meses:
        asegurar_vista(m, log=log)
    vistas = [(m, path_vista(m).as_posix()) for m in meses if path_vista(m).exists()]
    if not vistas:
        raise RuntimeError("No hay ninguna vista construida para exportar.")
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)

    ajenas = []
    if preservar and destino.exists():
        ajenas = hojas_ajenas(destino, log=log)
    if ajenas:
        escribir_preservando(destino, vistas, solo_dif, tolerancia, ajenas, log)
    else:
        escribir_desde_cero(destino, vistas, solo_dif, tolerancia, log)
    log(f"Excel listo: {destino}")


def hojas_ajenas(destino, log=print):
    """Hojas del archivo que NO genera este programa."""
    try:
        wb = openpyxl.load_workbook(str(destino), read_only=True)
        try:
            return [n for n in wb.sheetnames if not es_hoja_propia(n)]
        finally:
            wb.close()
    except Exception as e:
        log(f"  ! No se pudo revisar {destino.name} ({e}); se reescribe entero.")
        return []


def escribir_desde_cero(destino, vistas, solo_dif, tolerancia, log):
    """Camino rapido: el archivo no existe o no tiene hojas de nadie mas."""
    claves = [c for c, _ in COLUMNAS_EXCEL]
    encabezados = [t for _, t in COLUMNAS_EXCEL]
    con = duckdb.connect()
    wb = None
    try:
        wb = xlsxwriter.Workbook(str(destino), {"constant_memory": True})
        f_head = wb.add_format({"bold": True, "bg_color": "#1f3864",
                                "font_color": "white", "border": 1,
                                "text_wrap": True, "valign": "vcenter"})
        f_mes = wb.add_format({"bold": True, "bg_color": "#1f3864",
                               "font_color": "white"})
        f_tot = wb.add_format({"bold": True, "num_format": "#,##0.00",
                               "top": 1})
        f_tot_txt = wb.add_format({"bold": True, "top": 1})
        f_num = wb.add_format({"num_format": "#,##0.00"})
        f_int = wb.add_format({"num_format": "0"})
        f_dif = wb.add_format({"num_format": "#,##0.00", "bg_color": "#fde9d9"})
        f_alerta = wb.add_format({"bg_color": "#f8cbad"})

        # --- resumenes, uno por familia
        for etiqueta, tipos in FAMILIAS:
            hoja = wb.add_worksheet(f"RESUMEN {etiqueta.upper()}")
            hoja.write_row(0, 0, CAB_RESUMEN, f_head)
            hoja.set_column(0, 0, 14)
            hoja.set_column(1, 6, 18)
            hoja.freeze_panes(1, 1)
            fila = 1
            for aamm, detalle in datos_resumen(con, vistas, tipos):
                hoja.write(fila, 0, aamm, f_mes)
                for j in range(1, len(CAB_RESUMEN)):
                    hoja.write_blank(fila, j, None, f_mes)
                fila += 1
                for reg in detalle:
                    es_total = reg[0] == "Total"
                    hoja.write(fila, 0, reg[0], f_tot_txt if es_total else None)
                    for j, v in enumerate(reg[1:], start=1):
                        if v is None:
                            hoja.write_blank(fila, j, None)
                        else:
                            hoja.write_number(fila, j, float(v),
                                              f_tot if es_total else f_num)
                    fila += 1

        # --- una hoja por mes con el detalle horario
        for aamm, pvs in vistas:
            filtro = ""
            if solo_dif:
                filtro = (f"WHERE ABS(d_rpre_def) > {tolerancia} "
                          f"OR ABS(d_rdef_rpre) > {tolerancia} "
                          f"OR ABS(d_rdef_def) > {tolerancia} "
                          f"OR empresa_cambia OR sin_actual")
            cur = datos_hoja_mes(con, pvs, claves, filtro)
            i_cambia = claves.index("empresa_cambia")
            i_dif_act = claves.index("dif_vs_actual")
            i_sin_act = claves.index("sin_actual")
            i_det = claves.index("detalle_propietario")

            def nueva(nombre):
                h = wb.add_worksheet(nombre)
                h.write_row(0, 0, encabezados, f_head)
                h.set_row(0, 28)
                h.freeze_panes(1, 3)
                h.set_column(0, 0, 26)
                h.set_column(1, 1, 30)
                h.set_column(2, 2, 8)
                h.set_column(COL_HORA, COL_HORA, 12, f_int)
                h.set_column(4, 9, 15, f_num)
                h.set_column(10, 13, 26)
                h.set_column(14, 16, 13)
                h.set_column(i_det, i_det, 70)
                h.autofilter(0, 0, 0, len(encabezados) - 1)
                return h

            hoja = nueva(aamm)
            fila, extra = 1, 0
            while True:
                lote = cur.fetchmany(50_000)
                if not lote:
                    break
                for reg in lote:
                    if fila > LIMITE_FILAS_HOJA:
                        extra += 1
                        hoja = nueva(f"{aamm}_{extra + 1}")
                        fila = 1
                        log(f"  ! {aamm} paso el limite de filas: sigue en "
                            f"{aamm}_{extra + 1}")
                    marcar = bool(reg[i_dif_act]) or bool(reg[i_sin_act])
                    for j, v in enumerate(reg):
                        if isinstance(v, bool):
                            hoja.write(fila, j, "SI" if v else "",
                                       f_alerta if (v and j in (i_cambia, i_dif_act,
                                                                i_sin_act)) else None)
                        elif v is None:
                            hoja.write_blank(fila, j, None)
                        elif j in COL_DELTAS:
                            hoja.write_number(fila, j, float(v),
                                              f_dif if abs(float(v)) > tolerancia
                                              else f_num)
                        elif j in COL_NUM or j == COL_HORA:
                            hoja.write_number(fila, j, float(v))
                        else:
                            hoja.write(fila, j, str(v),
                                       f_alerta if (marcar and j == i_det) else None)
                    fila += 1
            log(f"  {aamm}: {fila - 1:,} filas.")

        # --- propietarios
        hoja = wb.add_worksheet("PROPIETARIOS")
        hoja.write_row(0, 0, CAB_PROP, f_head)
        hoja.set_column(0, 0, 8)
        hoja.set_column(1, 5, 28)
        hoja.set_column(6, 8, 14)
        hoja.set_column(9, 9, 80)
        fila = 1
        for aamm, pvs in vistas:
            for r in con.execute(SQL_PROP.format(pvs=pvs)).fetchall():
                hoja.write(fila, 0, aamm)
                for j, v in enumerate(r, start=1):
                    if isinstance(v, bool):
                        hoja.write(fila, j, "SI" if v else "",
                                   f_alerta if v else None)
                    elif v is None:
                        hoja.write_blank(fila, j, None)
                    else:
                        hoja.write(fila, j, str(v))
                fila += 1
        hoja.autofilter(0, 0, 0, len(CAB_PROP) - 1)
        hoja.freeze_panes(1, 2)
    finally:
        if wb is not None:
            wb.close()
        con.close()


CAB_PROP = ["Mes", "Central", "Emp Def", "Emp Rpre", "Emp Rdef", "Emp Actual",
            "Cambio", "Difiere del actual", "No esta en config. actual", "Detalle"]

SQL_PROP = """
    SELECT DISTINCT central, emp_def, emp_rpre, emp_rdef, empresa_actual,
           empresa_cambia, dif_vs_actual, sin_actual, detalle_propietario
    FROM read_parquet('{pvs}')
    ORDER BY empresa_cambia DESC, central
"""


def escribir_preservando(destino, vistas, solo_dif, tolerancia, ajenas, log):
    """Reescribe solo las hojas propias, dejando intactas las de los demas.

    Es mas lento que escribir de cero porque hay que cargar el libro entero en
    memoria, pero es la unica forma de no borrar el trabajo de otra persona.
    """
    log(f"  Se conservan {len(ajenas)} hoja(s) de otra persona: "
        f"{', '.join(ajenas[:6])}{' ...' if len(ajenas) > 6 else ''}")
    respaldo = respaldar(destino, _anio_de(vistas[0][0]), log)
    claves = [c for c, _ in COLUMNAS_EXCEL]
    encabezados = [t for _, t in COLUMNAS_EXCEL]

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    azul = PatternFill("solid", fgColor="1F3864")
    naranjo = PatternFill("solid", fgColor="F8CBAD")
    crema = PatternFill("solid", fgColor="FDE9D9")
    blanco_negrita = Font(bold=True, color="FFFFFF")
    negrita = Font(bold=True)

    con = duckdb.connect()
    try:
        wb = openpyxl.load_workbook(str(destino))
        for n in list(wb.sheetnames):
            if es_hoja_propia(n):
                del wb[n]

        def cabecera(hoja, titulos):
            for j, tx in enumerate(titulos, start=1):
                c = hoja.cell(row=1, column=j, value=tx)
                c.fill = azul
                c.font = blanco_negrita
                c.alignment = Alignment(wrap_text=True, vertical="center")
            hoja.freeze_panes = "A2"

        for etiqueta, tipos in FAMILIAS:
            hoja = wb.create_sheet(f"RESUMEN {etiqueta.upper()}")
            cabecera(hoja, CAB_RESUMEN)
            hoja.column_dimensions["A"].width = 14
            for j in range(2, 8):
                hoja.column_dimensions[get_column_letter(j)].width = 18
            fila = 2
            for aamm, detalle in datos_resumen(con, vistas, tipos):
                c = hoja.cell(row=fila, column=1, value=aamm)
                c.fill = azul
                c.font = blanco_negrita
                for j in range(2, 8):
                    hoja.cell(row=fila, column=j).fill = azul
                fila += 1
                for reg in detalle:
                    es_total = reg[0] == "Total"
                    c = hoja.cell(row=fila, column=1, value=reg[0])
                    if es_total:
                        c.font = negrita
                    for j, v in enumerate(reg[1:], start=2):
                        if v is None:
                            continue
                        c = hoja.cell(row=fila, column=j, value=float(v))
                        c.number_format = "#,##0.00"
                        if es_total:
                            c.font = negrita
                    fila += 1

        i_cambia = claves.index("empresa_cambia")
        i_dif_act = claves.index("dif_vs_actual")
        i_sin_act = claves.index("sin_actual")
        i_det = claves.index("detalle_propietario")
        for aamm, pvs in vistas:
            filtro = ""
            if solo_dif:
                filtro = (f"WHERE ABS(d_rpre_def) > {tolerancia} "
                          f"OR ABS(d_rdef_rpre) > {tolerancia} "
                          f"OR ABS(d_rdef_def) > {tolerancia} "
                          f"OR empresa_cambia OR sin_actual")
            hoja = wb.create_sheet(aamm)
            cabecera(hoja, encabezados)
            hoja.freeze_panes = "D2"
            fila = 2
            cur = datos_hoja_mes(con, pvs, claves, filtro)
            while True:
                lote = cur.fetchmany(50_000)
                if not lote:
                    break
                for reg in lote:
                    if fila > LIMITE_FILAS_HOJA:
                        log(f"  ! {aamm} paso el limite de filas de Excel; "
                            f"se corta. Usa 'Solo filas con diferencia'.")
                        break
                    marcar = bool(reg[i_dif_act]) or bool(reg[i_sin_act])
                    for j, v in enumerate(reg, start=1):
                        if v is None:
                            continue
                        if isinstance(v, bool):
                            if not v:
                                continue
                            c = hoja.cell(row=fila, column=j, value="SI")
                            if j - 1 in (i_cambia, i_dif_act, i_sin_act):
                                c.fill = naranjo
                        elif isinstance(v, (int, float)):
                            c = hoja.cell(row=fila, column=j, value=float(v))
                            if j - 1 == COL_HORA:
                                c.number_format = "0"
                            else:
                                c.number_format = "#,##0.00"
                                if (j - 1 in COL_DELTAS
                                        and abs(float(v)) > tolerancia):
                                    c.fill = crema
                        else:
                            c = hoja.cell(row=fila, column=j, value=str(v))
                            if marcar and j - 1 == i_det:
                                c.fill = naranjo
                    fila += 1
            hoja.auto_filter.ref = f"A1:{get_column_letter(len(encabezados))}1"
            log(f"  {aamm}: {fila - 2:,} filas.")

        hoja = wb.create_sheet("PROPIETARIOS")
        cabecera(hoja, CAB_PROP)
        fila = 2
        for aamm, pvs in vistas:
            for r in con.execute(SQL_PROP.format(pvs=pvs)).fetchall():
                hoja.cell(row=fila, column=1, value=aamm)
                for j, v in enumerate(r, start=2):
                    if v is None:
                        continue
                    if isinstance(v, bool):
                        if v:
                            hoja.cell(row=fila, column=j, value="SI").fill = naranjo
                    else:
                        hoja.cell(row=fila, column=j, value=str(v))
                fila += 1
        hoja.auto_filter.ref = f"A1:{get_column_letter(len(CAB_PROP))}1"

        wb.save(str(destino))
    except Exception:
        if respaldo:
            log(f"  ! Fallo la escritura. El archivo previo esta en {respaldo}")
        raise
    finally:
        con.close()


def respaldar(destino, anio, log=print):
    """Copia el archivo antes de reescribirlo. Deja las ultimas 5."""
    try:
        import shutil
        carpeta = dir_resultados_anuales(anio) / "respaldos"
        carpeta.mkdir(parents=True, exist_ok=True)
        marca = datetime.now().strftime("%Y%m%d_%H%M%S")
        copia = carpeta / f"{destino.stem}_{marca}{destino.suffix}"
        shutil.copy2(destino, copia)
        previas = sorted(carpeta.glob(f"{destino.stem}_*{destino.suffix}"))
        for vieja in previas[:-5]:
            try:
                vieja.unlink()
            except Exception:
                pass
        log(f"  Respaldo: {copia.name}")
        return copia
    except Exception as e:
        log(f"  ! No se pudo respaldar {destino.name}: {e}")
        return None


# ==========================================================================
# Ventana
# ==========================================================================
class App:
    def __init__(self, root):
        self.root = root
        self.cola = queue.Queue()
        self.cfg = leer_config()
        anio_inicial = self.cfg.get("comp_anio", "")
        self.est = cargar_estado(anio_inicial) if meses_del_anio(anio_inicial) else {}
        rj = leer_json(rutas_path(anio_inicial), {}) if meses_del_anio(anio_inicial) else {}
        self.rutas_manuales = rj if isinstance(rj, dict) else {}
        self.rutas = {}           # aamm -> {etapa: {slot: ruta}}
        self.diag = {}            # aamm -> {etapa: hasta donde llego la busqueda}
        self.filas = {}           # aamm -> widgets
        self.expandido = {}       # aamm -> bool
        self.trabajando = False
        self.timer = {"on": False, "t0": 0.0}

        root.title(APP_TITULO)
        root.geometry("1120x780")
        root.minsize(900, 600)

        self.var_anio = tk.StringVar(value=self.cfg.get("comp_anio", ""))
        self.var_cmg = tk.StringVar(value=self.cfg.get("comp_cmgreales", ""))
        self.var_fact = tk.StringVar(value=self.cfg.get("comp_facturacion", ""))
        self.var_actual = tk.StringVar(value="[sin leer]")
        self.var_anual = tk.StringVar(value=str(dir_resultados_anuales(anio_inicial)) if meses_del_anio(anio_inicial) else "")
        self.var_estado = tk.StringVar(value="Listo")
        self.var_tiempo = tk.StringVar(value="00:00:00")
        self.var_solo_dif = tk.BooleanVar(value=False)
        self.var_tol = tk.StringVar(value=self.cfg.get("comp_tolerancia", "0.005"))
        self.var_forzar = tk.BooleanVar(value=False)
        self.var_preservar = tk.BooleanVar(value=True)
        self.var_tema_oscuro = tk.BooleanVar(
            value=self.cfg.get("tema", "claro") == "oscuro")

        self._construir()
        self._aplicar_tema()
        self.root.after(100, self._bombear_cola)
        drv = driver_access()
        self.log(f"Driver Access: {drv or 'NO ENCONTRADO'} · Python "
                 f"{'64' if sys.maxsize > 2**32 else '32'} bits, "
                 f"{sys.version.split()[0]}")
        if not drv:
            self.log("  ! Sin driver de Access no se puede leer ningun .mdb. "
                     "Revisa que Python y Office sean de la misma arquitectura.")
        self.log(f"Carpeta base: {BASE}")
        self.pintar_actual()
        if self.var_anio.get():
            self.refrescar()

    # ---------------- construccion de la ventana ----------------
    def _construir(self):
        root = self.root

        # 1) barra inferior fija
        pie = tk.Frame(root)
        pie.pack(side="bottom", fill="x", pady=6)

        self.btn_refrescar = tk.Button(pie, text="ACTUALIZAR estado",
                                       command=self.refrescar, width=18)
        self.btn_refrescar.pack(side="left", padx=6)
        self.btn_cons_pend = tk.Button(pie, text="Consolidar pendientes",
                                       bg=COLORES["amarillo"], fg=COLORES["texto_estado"], width=20,
                                       command=lambda: self.lanzar(self.consolidar,
                                                                  forzar=False))
        self.btn_cons_pend.pack(side="left", padx=6)
        self.btn_cons_todo = tk.Button(pie, text="Reconsolidar TODO",
                                       bg=COLORES["gris"], fg=COLORES["texto_estado"], width=18,
                                       command=self.confirmar_todo)
        self.btn_cons_todo.pack(side="left", padx=6)
        self.btn_excel = tk.Button(pie, text="Exportar Excel", bg=COLORES["verde"],
                                   fg=COLORES["texto_estado"], width=16,
                                   command=lambda: self.lanzar(self.exportar))
        self.btn_excel.pack(side="left", padx=6)
        tk.Checkbutton(pie, text="Tema oscuro", variable=self.var_tema_oscuro,
                       command=self.cambiar_tema).pack(side="left", padx=8)
        tk.Label(pie, textvariable=self.var_tiempo, font=("Consolas", 11, "bold"),
                 fg=COLORES["verde"]).pack(side="right", padx=10)

        # log
        marco_log = tk.LabelFrame(root, text="Bitacora")
        marco_log.pack(side="bottom", fill="x", padx=8, pady=(0, 4))
        self.txt = tk.Text(marco_log, height=9, font=("Consolas", 9), wrap="none")
        sb = tk.Scrollbar(marco_log, command=self.txt.yview)
        self.txt.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.txt.pack(fill="x", expand=True)
        self.barra = ttk.Progressbar(root, mode="determinate")
        self.barra.pack(side="bottom", fill="x", padx=8)
        tk.Label(root, textvariable=self.var_estado, anchor="w",
                 fg=COLORES["enlace"]).pack(side="bottom", fill="x", padx=10)

        # 2) canvas con scroll
        canvas = tk.Canvas(root, borderwidth=0, highlightthickness=0)
        scroll = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        cont = tk.Frame(canvas)
        win = canvas.create_window((0, 0), window=cont, anchor="nw")

        def ajustar(_=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(win, width=canvas.winfo_width())
        cont.bind("<Configure>", ajustar)
        canvas.bind("<Configure>", ajustar)
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(int(-e.delta / 120), "units"))
        self.cont = cont

        # --- encabezado: anio y carpeta CMgReales
        f1 = tk.LabelFrame(cont, text="Anio a comparar (AAAA o AA)", padx=8, pady=6)
        f1.pack(fill="x", padx=12, pady=6)
        e = tk.Entry(f1, textvariable=self.var_anio, width=10,
                     font=("Consolas", 12, "bold"), justify="center")
        e.pack(side="left")
        e.bind("<Return>", lambda _: self.refrescar())
        tk.Button(f1, text="Cargar los 12 meses",
                  command=self.refrescar).pack(side="left", padx=8)
        tk.Label(f1, text="genera las claves AAMM de enero a diciembre",
                 fg=COLORES["gris"]).pack(side="left", padx=6)

        f2 = tk.LabelFrame(cont, text="Carpeta CMgReales (para los Definitivos)",
                           padx=8, pady=6)
        f2.pack(fill="x", padx=12, pady=4)
        self.lbl_cmg = tk.Label(f2, textvariable=self.var_cmg, cursor="hand2",
                                anchor="w", font=("Segoe UI", 9))
        self.lbl_cmg.pack(fill="x")
        self.lbl_cmg.bind("<Button-1>",
                          lambda _: abrir_en_explorador(self.var_cmg.get(), False))
        tk.Button(f2, text="Examinar", command=self.elegir_cmg).pack(pady=(4, 0))

        f2b = tk.LabelFrame(
            cont, text="Carpeta Facturacion (propietario vigente — 'Actual')",
            padx=8, pady=6)
        f2b.pack(fill="x", padx=12, pady=4)
        self.lbl_fact = tk.Label(f2b, textvariable=self.var_fact, cursor="hand2",
                                 anchor="w", font=("Segoe UI", 9))
        self.lbl_fact.pack(fill="x")
        self.lbl_fact.bind("<Button-1>",
                           lambda _: abrir_en_explorador(self.var_fact.get(), False))
        self.lbl_actual = tk.Label(f2b, textvariable=self.var_actual, anchor="w",
                                   font=("Segoe UI", 8), fg=COLORES["gris"])
        self.lbl_actual.pack(fill="x")
        fbb = tk.Frame(f2b)
        fbb.pack(fill="x", pady=(4, 0))
        tk.Button(fbb, text="Examinar", command=self.elegir_fact).pack(side="left")
        self.btn_actual = tk.Button(
            fbb, text="Releer propietario actual",
            command=lambda: self.lanzar(self.actualizar_actual, forzar=True))
        self.btn_actual.pack(side="left", padx=8)
        tk.Label(fbb, text="4_REMUNERACION_SC_CO · hoja 'Configuracion Empresa' (A y B)",
                 fg=COLORES["gris"], font=("Segoe UI", 8)).pack(side="left", padx=6)

        # --- leyenda
        f3 = tk.Frame(cont)
        f3.pack(fill="x", padx=12, pady=(2, 6))
        tk.Label(f3, text="La casilla de cada mes decide si entra al "
                          "consolidado anual.   ",
                 fg=COLORES["enlace"], font=("Segoe UI", 8)).pack(side="left")
        for txt, col in (("rojo = falta la ruta", COLORES["rojo"]),
                         ("amarillo = esta pero sin consolidar / cambio despues", COLORES["amarillo"]),
                         ("verde = consolidado y al dia", COLORES["verde"])):
            tk.Label(f3, text="  " + txt, fg=col,
                     font=("Segoe UI", 8, "bold")).pack(side="left")

        # --- contenedor de los 12 meses
        self.marco_meses = tk.Frame(cont)
        self.marco_meses.pack(fill="x", padx=12, pady=4)

        # --- exportacion
        f4 = tk.LabelFrame(cont, text="Paso 2 — parquet a Excel", padx=8, pady=6)
        f4.pack(fill="x", padx=12, pady=8)
        tk.Label(f4, anchor="w", justify="left", fg=COLORES["gris"],
                 font=("Segoe UI", 8),
                 text="Cada mes deja su propio Excel en Salidas\\AAMM\\ y el "
                      "consolidado con todos los meses disponibles queda en "
                      "Salidas\\_comparador\\.\n"
                      "Se rehace solo lo que quedo viejo: el mes que cambio y, "
                      "por lo tanto, el consolidado."
                 ).pack(fill="x")
        self.lbl_anual = tk.Label(f4, textvariable=self.var_anual, cursor="hand2",
                                  anchor="w", font=("Segoe UI", 9), fg=COLORES["enlace"])
        self.lbl_anual.pack(fill="x", pady=(4, 0))
        self.lbl_anual.bind("<Button-1>",
                            lambda _: abrir_en_explorador(self.var_anual.get(), True))
        fb = tk.Frame(f4)
        fb.pack(fill="x", pady=(4, 0))
        tk.Checkbutton(fb, text="Rehacer todos los Excel aunque esten al dia",
                       variable=self.var_forzar).pack(side="left")
        tk.Checkbutton(fb, text="Conservar hojas agregadas por otros",
                       variable=self.var_preservar).pack(side="left", padx=10)
        tk.Checkbutton(fb, text="Solo filas con diferencia",
                       variable=self.var_solo_dif).pack(side="left", padx=10)
        tk.Label(fb, text="tolerancia:").pack(side="left", padx=(10, 2))
        tk.Entry(fb, textvariable=self.var_tol, width=8).pack(side="left")

    def _aplicar_tema(self):
        """Aplica el tema sin impedir que la herramienta arranque si falla."""
        global COLORES
        modo = "oscuro" if self.var_tema_oscuro.get() else "claro"
        try:
            colores = _tema.aplicar(self.root, modo)
            _tema.pintar_tk(self.root, colores)
        except Exception as exc:
            self.log(f"No se pudo aplicar el tema {modo}; se conserva el aspecto anterior: {exc}")
            return False
        COLORES = colores
        self.btn_cons_pend.config(bg=COLORES["amarillo"],
                                  fg=COLORES["texto_estado"])
        self.btn_cons_todo.config(bg=COLORES["gris"],
                                  fg=COLORES["texto_estado"])
        self.btn_excel.config(bg=COLORES["verde"],
                              fg=COLORES["texto_estado"])
        return True

    def cambiar_tema(self):
        """Guarda y aplica en vivo la preferencia compartida del equipo."""
        modo = "oscuro" if self.var_tema_oscuro.get() else "claro"
        guardar_config({"tema": modo})
        if self._aplicar_tema():
            self.pintar()

    # ---------------- log / estado ----------------
    def log(self, msg):
        self.cola.put(("log", str(msg)))

    def _bombear_cola(self):
        """Aplica en el hilo de tkinter los cambios pedidos por los workers."""
        while True:
            try:
                accion, valor = self.cola.get_nowait()
            except queue.Empty:
                break
            if accion == "log":
                self.txt.insert("end", valor + "\n")
                self.txt.see("end")
            elif accion == "estado":
                self.var_estado.set(valor)
            elif accion == "barra":
                if valor.pop("final", False):
                    self.barra.config(value=self.barra["maximum"])
                else:
                    self.barra.config(**valor)
            elif accion == "llamar":
                funcion, args = valor
                funcion(*args)
        self.root.after(100, self._bombear_cola)

    def _llamar_en_ui(self, funcion, *args):
        self.cola.put(("llamar", (funcion, args)))

    def set_progreso(self, **kw):
        self.cola.put(("barra", kw))

    def set_estado(self, txt):
        self.cola.put(("estado", txt))

    def tick(self):
        if self.timer["on"]:
            self.var_tiempo.set(fmt_tiempo(time.time() - self.timer["t0"]))
            self.root.after(500, self.tick)

    def botones(self, activos):
        estado = "normal" if activos else "disabled"
        for b in (self.btn_refrescar, self.btn_cons_pend,
                  self.btn_cons_todo, self.btn_excel, self.btn_actual):
            b.config(state=estado)

    def lanzar(self, funcion, **kw):
        """Corre en hilo aparte para que la ventana no se congele."""
        if self.trabajando:
            return
        self.trabajando = True
        self.botones(False)
        self.timer.update({"on": True, "t0": time.time()})
        self.tick()

        def correr():
            try:
                funcion(**kw)
            except Exception as e:
                self.log(f"ERROR: {e}")
                self.log(traceback.format_exc())
            finally:
                self.timer["on"] = False
                self.trabajando = False
                self._llamar_en_ui(self.botones, True)
                self._llamar_en_ui(self.pintar)
                self.set_estado("Listo")

        threading.Thread(target=correr, daemon=True).start()

    # ---------------- selectores ----------------
    def elegir_cmg(self):
        ini = self.var_cmg.get() if Path(self.var_cmg.get() or "x").is_dir() else ""
        r = filedialog.askdirectory(title="Carpeta CMgReales", initialdir=ini)
        if r:
            self.var_cmg.set(r)
            guardar_config({"comp_cmgreales": r})
            self.refrescar()

    def elegir_fact(self):
        ini = self.var_fact.get() if Path(self.var_fact.get() or "x").is_dir() else ""
        r = filedialog.askdirectory(title="Carpeta Facturacion", initialdir=ini)
        if r:
            self.var_fact.set(r)
            guardar_config({"comp_facturacion": r})
            self.lanzar(self.actualizar_actual, forzar=False)

    def actualizar_actual(self, forzar=False):
        raiz = self.var_fact.get().strip()
        if not raiz:
            self.log("Falta la carpeta Facturacion para leer el propietario vigente.")
            return
        self.est = cargar_estado(self.var_anio.get().strip())
        self.set_estado("Buscando el 4_REMUNERACION_SC_CO mas nuevo...")
        cambio = consolidar_actual(self.var_anio.get().strip(), raiz, self.est, forzar=forzar, log=self.log)
        guardar_estado(self.var_anio.get().strip(), self.est)
        self._llamar_en_ui(self.pintar_actual)
        if cambio:
            # Cambio la referencia: hay que rearmar las vistas ya existentes
            for m in self.meses_visibles():
                if etapas_consolidadas(m):
                    construir_vista(m, log=self.log)
                    self.est.setdefault(m, {})["vista"] = ahora()
            guardar_estado(self.var_anio.get().strip(), self.est)

    def pintar_actual(self):
        reg = (self.est or {}).get("_actual") or {}
        anio = self.var_anio.get().strip()
        if reg.get("archivo") and meses_del_anio(anio):
            self.var_actual.set(
                f"{reg.get('etiqueta', '')} · {Path(reg['archivo']).name} · "
                f"{reg.get('centrales', 0)} centrales · leido {reg.get('leido', '')}")
            self.lbl_actual.config(fg=COLORES["verde"] if actual_parquet(anio).exists()
                                   else COLORES["amarillo"])
        else:
            self.var_actual.set("[sin leer — el Excel saldra sin la columna Actual]")
            self.lbl_actual.config(fg=COLORES["rojo"])

    def _carpeta_diag(self, aamm, etapa):
        """La carpeta mas profunda a la que llego la busqueda, para abrirla."""
        msg = (self.diag.get(aamm) or {}).get(etapa) or ""
        for marca in ("llego hasta ", "no se ve la carpeta ", "en "):
            if msg.startswith(marca):
                resto = msg[len(marca):]
                for corte in (" — ", " no aparece:"):
                    if corte in resto:
                        resto = resto.split(corte)[0]
                return resto.strip()
        return msg.strip() or None

    def elegir_carpeta(self, aamm, etapa):
        """Una sola carpeta y de ahi salen los dos .mdb."""
        man = self.rutas_manuales.get(aamm, {}).get(etapa, {}) or {}
        ini = man.get("carpeta") or self._carpeta_diag(aamm, etapa) or ""
        if not (ini and Path(ini).is_dir()):
            ini = self.var_cmg.get() or str(BASE)
        titulo = ("Carpeta Auxiliares del Definitivo" if etapa == "def"
                  else "Carpeta 02 CASO RELIQUIDACION")
        r = filedialog.askdirectory(title=f"{aamm} · {ETIQUETA[etapa]} · {titulo}",
                                    initialdir=ini)
        if not r:
            return
        encontrados, d = mdb_en_carpeta(r, log=self.log)
        self.log(f"{aamm} {ETIQUETA[etapa]} · {r}")
        for slot in SLOTS:
            if encontrados.get(slot):
                self.log(f"    {ETIQUETA_SLOT[slot]}: {Path(encontrados[slot]).name}")
        if d:
            self.log(f"    ! {d}")
        blo = self.rutas_manuales.setdefault(aamm, {}).setdefault(etapa, {})
        blo["carpeta"] = r
        # Si antes se habia fijado un archivo suelto, la carpeta lo reemplaza.
        for slot in SLOTS:
            blo.pop(slot, None)
        escribir_json_atomico(rutas_path(self.var_anio.get().strip()), self.rutas_manuales)
        if etapa == "rpre":
            for slot, ruta in encontrados.items():
                guardar_rpre_en_json_mes(aamm, slot, ruta, log=self.log)
        self.refrescar(solo=aamm)

    def elegir_mdb(self, aamm, etapa, slot):
        pat = "03b ENTRADA_SOB_SSCC" if slot == "sscc" else "03b ENTRADA_SOB"
        act = self.rutas.get(aamm, {}).get(etapa, {}).get(slot)
        ini = str(Path(act).parent) if act and Path(act).exists() else \
            (self.var_cmg.get() or str(BASE))
        r = filedialog.askopenfilename(
            title=f"{aamm} · {ETIQUETA[etapa]} · {pat}", initialdir=ini,
            filetypes=[("Access", "*.mdb *.accdb"), ("Todos", "*.*")])
        if not r:
            return
        self.rutas_manuales.setdefault(aamm, {}).setdefault(etapa, {})[slot] = r
        escribir_json_atomico(rutas_path(self.var_anio.get().strip()), self.rutas_manuales)
        if etapa == "rpre":
            if guardar_rpre_en_json_mes(aamm, slot, r, log=self.log):
                self.log(f"  Rpre guardado en Salidas/{aamm}/{NOMBRE_JSON_MES} "
                         f"(clave '{CLAVE_JSON_PROPIA}').")
        self.refrescar(solo=aamm)

    # ---------------- armado de las filas de mes ----------------
    def refrescar(self, solo=None):
        """Busca los archivos en segundo plano, para no congelar la ventana.

        Recorrer las carpetas del NAS toma varios segundos; hacerlo en el hilo
        principal dejaba la ventana en blanco y sin responder. Aqui el trabajo
        de disco va en un hilo y solo el pintado vuelve al hilo de tkinter,
        que es el unico que puede tocar los widgets.
        """
        anio = self.var_anio.get().strip()
        meses = meses_del_anio(anio)
        if not meses:
            self.log("Indica el anio (por ejemplo 2024 o 24).")
            return
        if self.trabajando:
            return
        guardar_config({"comp_anio": anio})
        self.est = cargar_estado(anio)
        rj = leer_json(rutas_path(anio), {})
        self.rutas_manuales = rj if isinstance(rj, dict) else {}
        self.var_anual.set(str(dir_resultados_anuales(anio)))
        raiz = self.var_cmg.get().strip()
        self.trabajando = True
        self.botones(False)
        objetivo = [solo] if solo else meses
        self.set_progreso(maximum=len(objetivo), value=0)

        def buscar():
            t0 = time.time()
            try:
                limpiar_cache()      # el disco pudo cambiar desde el refresco anterior
                for i, m in enumerate(objetivo, 1):
                    self.set_estado(f"Buscando archivos... {m}  ({i}/{len(objetivo)})")
                    self.rutas[m], self.diag[m] = resolver_rutas(
                        m, raiz, self.rutas_manuales)
                    self.set_progreso(value=i)
            except Exception as e:
                self.log(f"ERROR buscando archivos: {e}")
                self.log(traceback.format_exc())
            finally:
                seg = time.time() - t0
                self.trabajando = False

                def terminar():
                    self.botones(True)
                    if solo is None or not self.filas:
                        self.construir_filas(meses)
                    self.pintar()
                    self.set_estado("Listo")
                    if seg > 3:
                        self.log(f"Busqueda de archivos: {seg:.1f} s "
                                 f"({len(objetivo)} mes(es)).")
                self._llamar_en_ui(terminar)

        threading.Thread(target=buscar, daemon=True).start()

    def construir_filas(self, meses):
        for w in self.marco_meses.winfo_children():
            w.destroy()
        self.filas = {}
        for m in meses:
            self.expandido.setdefault(m, False)
            caja = tk.Frame(self.marco_meses, relief="groove", bd=1)
            caja.pack(fill="x", pady=2)

            cab = tk.Frame(caja)
            cab.pack(fill="x")
            var_inc = tk.BooleanVar(value=mes_incluido(self.est, m))
            chk = tk.Checkbutton(
                cab, variable=var_inc,
                command=lambda mm=m: self.cambiar_inclusion(mm))
            chk.pack(side="left")
            btn = tk.Button(cab, text=f"▶  {m}", width=10, anchor="w",
                            font=("Consolas", 11, "bold"), relief="flat",
                            command=lambda mm=m: self.alternar(mm))
            btn.pack(side="left", padx=4, pady=2)

            chips = {}
            for etapa in ETAPAS:
                l = tk.Label(cab, text=ETIQUETA[etapa], width=7,
                             font=("Segoe UI", 9, "bold"), relief="ridge",
                             padx=4, pady=1)
                l.pack(side="left", padx=3)
                chips[etapa] = l
            lbl_info = tk.Label(cab, text="", anchor="w", fg=COLORES["gris"],
                                font=("Segoe UI", 8))
            lbl_info.pack(side="left", padx=8)
            btn_cons = tk.Button(cab, text="Consolidar mes", font=("Segoe UI", 8),
                                 command=lambda mm=m: self.lanzar(
                                     self.consolidar, meses=[mm], forzar=False))
            btn_cons.pack(side="right", padx=4)
            btn_vista = tk.Button(cab, text="Rearmar vista", font=("Segoe UI", 8),
                                  command=lambda mm=m: self.lanzar(
                                      self.rearmar_vista, meses=[mm]))
            btn_vista.pack(side="right", padx=2)

            det = tk.Frame(caja)  # sin pack: se muestra al expandir
            lineas = {}
            for etapa in ETAPAS:
                fe = tk.LabelFrame(det, text=ETIQUETA[etapa], padx=6, pady=2)
                fe.pack(fill="x", padx=16, pady=2)
                lineas[etapa] = {}
                fc = tk.Frame(fe)
                fc.pack(fill="x", pady=(0, 2))
                texto_btn = ("Carpeta Auxiliares del Definitivo"
                             if etapa == "def" else "Carpeta 02 CASO RELIQUIDACION")
                tk.Button(fc, text=texto_btn, font=("Segoe UI", 8),
                          command=lambda mm=m, et=etapa:
                          self.elegir_carpeta(mm, et)).pack(side="left")
                ld = tk.Label(fc, text="", anchor="w", fg=COLORES["gris"],
                              font=("Segoe UI", 8), cursor="hand2")
                ld.pack(side="left", fill="x", expand=True, padx=6)
                ld.bind("<Button-1>", lambda e, mm=m, et=etapa:
                        abrir_en_explorador(self._carpeta_diag(mm, et), False))
                lineas[etapa]["_diag"] = ld
                for slot in SLOTS:
                    fl = tk.Frame(fe)
                    fl.pack(fill="x")
                    tk.Label(fl, text=ETIQUETA_SLOT[slot], width=10, anchor="w",
                             font=("Consolas", 8)).pack(side="left")
                    lr = tk.Label(fl, text="", anchor="w", cursor="hand2",
                                  font=("Segoe UI", 8))
                    lr.pack(side="left", fill="x", expand=True)
                    lr.bind("<Button-1>", lambda e, mm=m, et=etapa, sl=slot:
                            abrir_en_explorador(
                                self.rutas.get(mm, {}).get(et, {}).get(sl), True))
                    tk.Button(fl, text="Examinar", font=("Segoe UI", 8),
                              command=lambda mm=m, et=etapa, sl=slot:
                              self.elegir_mdb(mm, et, sl)).pack(side="right", padx=3)
                    lineas[etapa][slot] = lr
            self.filas[m] = {"caja": caja, "btn": btn, "chips": chips,
                             "inc": var_inc,
                             "info": lbl_info, "det": det, "lineas": lineas,
                             "btn_cons": btn_cons}
        self._aplicar_tema()

    def cambiar_inclusion(self, m):
        """Marcar/desmarcar un mes para el consolidado anual."""
        f = self.filas.get(m)
        if not f:
            return
        self.est = cargar_estado(self.var_anio.get().strip())
        fijar_incluido(self.est, m, f["inc"].get())
        guardar_estado(self.var_anio.get().strip(), self.est)
        self.pintar()

    def alternar(self, m):
        f = self.filas.get(m)
        if not f:
            return
        self.expandido[m] = not self.expandido[m]
        if self.expandido[m]:
            f["det"].pack(fill="x", pady=(0, 4))
            f["btn"].config(text=f"▼  {m}")
        else:
            f["det"].pack_forget()
            f["btn"].config(text=f"▶  {m}")

    def pintar(self):
        for m, f in self.filas.items():
            rutas = self.rutas.get(m) or {e: {s: None for s in SLOTS} for e in ETAPAS}
            dg = (self.diag.get(m) or {})
            for etapa in ETAPAS:
                st = estado_etapa(self.est, m, etapa, rutas)
                col = color_de(st)
                f["chips"][etapa].config(fg=COLORES["texto_estado"], bg=col)
                ld = f["lineas"][etapa].get("_diag")
                if ld is not None:
                    msg = dg.get(etapa) or ""
                    ld.config(text=msg,
                              fg=COLORES["rojo"] if msg.startswith(("llego", "no se ve",
                                                               "falta", "en "))
                              else COLORES["gris"])
                for slot in SLOTS:
                    r = rutas[etapa].get(slot)
                    lr = f["lineas"][etapa][slot]
                    if r and Path(r).is_file():
                        lr.config(text=str(r), fg=COLORES["enlace"])
                    elif r:
                        lr.config(text=f"{r}   ← no existe", fg=COLORES["rojo"])
                    else:
                        msg = dg.get(etapa) or ""
                        lr.config(text=f"[sin ruta] {msg}".strip(), fg=COLORES["rojo"])
            reg = self.est.get(m) or {}
            hechas = [ETIQUETA[e] for e in ETAPAS if (reg.get(e) or {}).get("consolidado")]
            v = path_vista(m)
            info = (f"consolidado: {', '.join(hechas) or '—'}"
                    f"   |   vista: {'si' if v.exists() else 'no'}")
            if not mes_incluido(self.est, m):
                info += "   |   FUERA del consolidado anual"
            f["inc"].set(mes_incluido(self.est, m))
            f["info"].config(text=info)
            f["btn"].config(fg=color_de(estado_mes(self.est, m, rutas)))
        self.pintar_actual()

    # ---------------- procesos ----------------
    def meses_visibles(self):
        return meses_del_anio(self.var_anio.get().strip())

    def confirmar_todo(self):
        if messagebox.askyesno(
                "Reconsolidar todo",
                "Vuelve a leer TODOS los Access de los 12 meses y reescribe el "
                "parquet completo. Puede tomar bastante.\n\n¿Seguir?"):
            self.lanzar(self.consolidar, forzar=True)

    def consolidar(self, meses=None, forzar=False):
        if not meses_del_anio(self.var_anio.get().strip()):
            # Sin esto, cargar_estado()/actual_parquet() revientan con
            # ValueError (via comun.salidas.normalizar_anio): antes de la
            # Tarea 2 estas rutas no dependian del anio y nunca fallaban.
            self.log("Elegi un anio valido antes de consolidar.")
            return
        meses = meses or self.meses_visibles()
        raiz = self.var_cmg.get().strip()
        self.est = cargar_estado(self.var_anio.get().strip())
        if self.var_fact.get().strip() and not actual_parquet(self.var_anio.get().strip()).exists():
            self.log("Leyendo el propietario vigente antes de consolidar...")
            try:
                consolidar_actual(self.var_anio.get().strip(), self.var_fact.get().strip(), self.est, log=self.log)
                guardar_estado(self.var_anio.get().strip(), self.est)
            except Exception as e:
                self.log(f"  ! No se pudo leer el propietario vigente: {e}")
        tareas = []
        for m in meses:
            self.rutas[m], self.diag[m] = resolver_rutas(m, raiz, self.rutas_manuales)
            for etapa in ETAPAS:
                st = estado_etapa(self.est, m, etapa, self.rutas[m])
                if st != "falta" and (forzar or st in ("pendiente", "desactualizado")):
                    tareas.append((m, etapa, st))
        if not tareas:
            self.log("Nada por consolidar: todo lo que tiene rutas ya esta al dia.")
            return

        tocados = set()
        if tareas:
            self.log(f"\n=== Consolidando {len(tareas)} etapa(s) ===")
            self.set_progreso(maximum=len(tareas), value=0)
            for i, (m, etapa, st) in enumerate(tareas, 1):
                self.set_estado(f"[{i}/{len(tareas)}] {m} {ETIQUETA[etapa]} ({st})")
                try:
                    if consolidar_etapa(m, etapa, self.rutas[m], self.est, log=self.log):
                        tocados.add(m)
                        guardar_estado(self.var_anio.get().strip(), self.est)
                except Exception as e:
                    self.log(f"  ERROR en {m} {ETIQUETA[etapa]}: {e}")
                    self.log(traceback.format_exc())
                self.set_progreso(value=i)
                self._llamar_en_ui(self.pintar)
            for m in sorted(tocados):
                self.set_estado(f"Rearmando vista {m}")
                try:
                    construir_vista(m, log=self.log)
                    self.est.setdefault(m, {})["vista"] = ahora()
                    guardar_estado(self.var_anio.get().strip(), self.est)
                except Exception as e:
                    self.log(f"  ERROR armando vista {m}: {e}")

        self.log("=== Consolidacion terminada ===\n")

    def rearmar_vista(self, meses=None):

        self.est = cargar_estado(self.var_anio.get().strip())
        for m in (meses or self.meses_visibles()):
            if etapas_consolidadas(m):
                construir_vista(m, log=self.log)
                self.est.setdefault(m, {})["vista"] = ahora()
        guardar_estado(self.var_anio.get().strip(), self.est)

    def exportar(self):
        """Excel por mes en 00_Salidas/AAAA/MM Mes + consolidado anual directo en AAAA.

        Solo rehace lo que quedo viejo: los meses cuya vista es mas nueva que su
        propio Excel, y el consolidado anual si cambio algun mes.
        """
        try:
            tol = float(str(self.var_tol.get()).replace(",", "."))
        except Exception:
            tol = 0.005
        guardar_config({"comp_tolerancia": str(tol)})
        self.est = cargar_estado(self.var_anio.get().strip())
        anio = self.var_anio.get().strip()
        aa = anio[2:] if len(anio) == 4 else anio
        meses = self.meses_visibles()
        disponibles = [m for m in meses if path_vista(m).exists()]
        if not disponibles:
            self.log("No hay ningun mes consolidado todavia.")
            return
        forzar = self.var_forzar.get()
        solo_dif = self.var_solo_dif.get()
        # Vistas de una version anterior: se rearman ahora, asi su mtime queda
        # nuevo y los Excel de esos meses entran solos a la lista de pendientes.
        for m in disponibles:
            asegurar_vista(m, log=self.log)

        # ---- 1) Comparacion de sobrecostos por tipo (SCCF/CO/CCA/SCMT/SCPC)
        if disponibles:
            pend = []
            for m in disponibles:
                dm = path_excel_mes(m)
                if (forzar or not dm.exists()
                        or dm.stat().st_mtime < path_vista(m).stat().st_mtime):
                    pend.append(m)
            if pend:
                self.log(f"\n=== Excel por mes: {', '.join(pend)} ===")
                self.set_progreso(maximum=len(pend) + 1, value=0)
                for i, m in enumerate(pend, 1):
                    self.set_estado(f"[{i}/{len(pend)}] Excel de {m}")
                    dm = path_excel_mes(m)
                    dm.parent.mkdir(parents=True, exist_ok=True)
                    exportar_excel(dm, [m], solo_dif, tol, log=self.log,
                                   preservar=self.var_preservar.get())
                    self.set_progreso(value=i)
            else:
                self.log("\nExcel por mes: todos al dia, no se rehace ninguno.")

            incluidos = [m for m in disponibles if mes_incluido(self.est, m)]
            fuera = [m for m in disponibles if m not in incluidos]
            if fuera:
                self.log(f"\nFuera del consolidado por eleccion tuya: {', '.join(fuera)}")
            if not incluidos:
                self.log("Ningun mes marcado para el consolidado anual.")
            else:
                anual = path_excel_anual(aa)
                firma = firma_vistas(incluidos)
                firma_previa = (self.est.get("_excel_anual") or {}).get("firma") or {}
                # Las claves del JSON son texto; se comparan como texto en ambos lados.
                igual = ({str(k): v for k, v in firma.items()}
                         == {str(k): v for k, v in firma_previa.items()})
                if anual.exists() and igual and not forzar and not pend:
                    self.log(f"Consolidado anual ya al dia: {anual.name}")
                else:
                    cambio = [m for m in firma if str(m) not in
                              {str(k) for k in firma_previa}] or pend
                    self.set_estado("Escribiendo el consolidado anual...")
                    self.log(f"\n=== Consolidado anual ({len(incluidos)} meses: "
                             f"{', '.join(incluidos)}) ===")
                    if cambio and firma_previa:
                        self.log(f"  se rehace porque cambio: {', '.join(sorted(set(cambio)))}")
                    t0 = time.time()
                    exportar_excel(anual, incluidos, solo_dif, tol, log=self.log,
                                   preservar=self.var_preservar.get())
                    self.est["_excel_anual"] = {
                        "archivo": str(anual), "meses": incluidos, "firma": firma,
                        "escrito": ahora(), "segundos": round(time.time() - t0, 1),
                    }
                    guardar_estado(self.var_anio.get().strip(), self.est)
                    self.log(f"  ({self.est['_excel_anual']['segundos']} s)")
                self.var_anual.set(str(anual))
            self.set_progreso(final=True)

        self.log("=== Listo ===\n")


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
