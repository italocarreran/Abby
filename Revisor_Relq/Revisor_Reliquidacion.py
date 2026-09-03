# -*- coding: utf-8 -*-
"""
Revisor de entregables - CASO RELIQUIDACION
===========================================
Ventana que replica la estructura de carpetas de "02 CASO RELIQUIDACION" y permite:

  * ACTUALIZAR  -> ubica todos los archivos, compara copias contra su maestro por
                   fecha y hora de modificacion (pinta AMARILLO las copias
                   desactualizadas, ROJO lo que falta) y revisa la vigencia de
                   las verificaciones de valores.
  * VERIFICAR   -> boton por archivo que compara sumas de sobrecostos entre
                   archivos (.xlsm / .mdb) y deja registrada la verificacion
                   con su fecha y las fechas de modificacion de sus dependencias.

Requisitos:  pip install xlwings openpyxl pyodbc
Para .mdb se necesita el "Microsoft Access Driver (*.mdb, *.accdb)" con la misma
arquitectura (32/64 bits) que el Python que ejecuta el script.
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from datetime import datetime
import json, subprocess, sys, re, socket, os, traceback, unicodedata, time, csv
import shutil
import threading, queue

DIR_RAIZ_CODIGO = Path(__file__).resolve().parent.parent
if str(DIR_RAIZ_CODIGO) not in sys.path:
    sys.path.insert(0, str(DIR_RAIZ_CODIGO))
from __comun__ import salidas as _sal

# =============================================================================
#  >>> ZONA A AJUSTAR <<<
#  Aca se declara DONDE esta cada valor dentro de cada archivo.
#  Mientras "hoja"/"celda"/"tabla"/"columna" esten vacios, el verificador
#  correspondiente no compara: solo abre el archivo y escribe en el log los
#  nombres de hojas / tablas / columnas disponibles para que puedas completarlos.
#
#  "celda" acepta:   "H120"        -> valor de una celda
#                    "H10:H200"    -> suma de todas las celdas numericas del rango
# =============================================================================

TOLERANCIA = 1.0        # diferencia maxima aceptada al comparar montos (en pesos)

# Residuo maximo aceptado en el descuadre del cuadro de pago (CPRT!I3). No es
# una tolerancia de comparacion: es el redondeo acumulado del reparto
# proporcional, que crece con la cantidad de empresas. En 2312, con 481 pares,
# dio 31,3 pesos.
UMBRAL_DESCUADRE_CPRT = 1000.0

# Diferencia maxima aceptada en UNA fila al recalcular el sobrecosto desde sus
# componentes. No es lo mismo que TOLERANCIA: aca se comparan dos formas de
# calcular el MISMO numero, asi que la unica diferencia legitima es el redondeo
# con que quedo guardada la columna del resultado. Si algun mes sale ruidoso,
# este es el numero a subir.
TOL_SOBRECOSTO_FILA = 1.0

# Diferencia maxima aceptada al comparar el pago por empresa y concepto entre la
# planilla 1 y la 9 (o la 4). Son dos calculos PARALELOS del mismo numero, asi que
# la diferencia legitima es solo arrastre de redondeo. En 2409 la peor fue de 58
# pesos sobre 664 pares.
TOL_PAGO_EMPRESA = 150.0

# Las filas de prorrata suman el 100%, escrito como 1 o como 100 segun la
# planilla. Se aceptan los dos y se avisa si un mismo archivo mezcla ambos.
TOTALES_PRORRATA = (1.0, 100.0)
TOL_PRORRATA = 0.0001         # sobre el total 1; para 100 se escala x100

# Diferencia maxima al comparar la suma por suministrador de la prorrata de una
# planilla contra el Prorrata_Retiros. Son los MISMOS numeros reordenados, asi
# que lo unico legitimo es el redondeo al leerlos.
TOL_PRORRATA_SUMA = 0.0001

# Una central que termina en "-numero" es una unidad, y las unidades son lo que
# tienen los embalses. Si aparece una asi que no esta en la lista, puede ser una
# unidad nueva que nadie agrego todavia.
RE_UNIDAD_CENTRAL = re.compile(r"-\d+\s*$")


def clave_concepto(t):
    """Normaliza un concepto o una empresa para comparar entre planillas.

    Saca tildes, espacios y guiones BAJOS, y pasa a mayusculas. Hace falta porque
    las dos planillas escriben lo mismo distinto: la 1 dice "CO ERNC" con espacio
    y la 9 dice "CO_ERNC" con guion bajo. Sin esto no se cruza ni un concepto.
    Se conservan los parentesis y el signo, que SI distinguen: CSF(+) y CSF(-) son
    conceptos diferentes.
    """
    t = unicodedata.normalize("NFKD", str(t or ""))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"[\s_]+", "", t).upper()


def clave_central(t):
    """Normaliza el nombre de una central para comparar: sin tildes, sin espacios
    ni guiones bajos, en mayusculas. Asi 'El Toro-1' y 'ELTORO-1' son la misma."""
    t = unicodedata.normalize("NFKD", str(t or ""))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"[\s_]+", "", t).upper()


# ---------------------------------------------------------------------------
#  Centrales de embalse
# ---------------------------------------------------------------------------
#  OJO: esta lista esta TAMBIEN en Actualiza_SC_CO.py. Si se agrega o quita una
#  central, hay que cambiarla EN LOS DOS.
#
#  Que pasa si se desincronizan (las dos las caza V10, por suerte):
#    - si alla hay una central que aca no: se pega igual, y la comprobacion
#      "en E9:E solo hay centrales de embalse" la marca.
#    - si aca hay una que alla no: no se pega, y la comprobacion
#      "no quedo fuera ninguna central «-numero»" la marca.
#  O sea que la lista vieja se nota al verificar, no pasa en silencio.
#
#  Al agregar una unidad nueva va con el nombre EXACTO como viene en el origen
#  (columna U de Calculo_SobrecostosSSCC / columna D de Calculo_CO).
CENTRALES_EMBALSE = [
    "CANUTILLAR-1", "CANUTILLAR-2",
    "ELTORO-1", "ELTORO-2", "ELTORO-3", "ELTORO-4",
    "RALCO-1", "RALCO-2",
    "RAPEL-1", "RAPEL-2", "RAPEL-3", "RAPEL-4", "RAPEL-5",
    "PEHUENCHE-1", "PEHUENCHE-2",
    "COLBUN-1", "COLBUN-2",
    "CIPRESES-1", "CIPRESES-2", "CIPRESES-3",
    "PANGUE-1", "PANGUE-2",
    "ANTUCO-1", "ANTUCO-2",
    "ANGOSTURA-1", "ANGOSTURA-2", "ANGOSTURA-3",
]
TOL_MTIME  = 2          # segundos de tolerancia al comparar fechas de modificacion

VALORES = {
    # --- totales que son una COLUMNA completa bajo un encabezado -----------
    #   columna      : letra de la columna de montos
    #   fila_inicio  : primera fila CON DATOS (la del encabezado + 1)
    #   columna_filtro / valores_filtro : para sumar solo ciertos tipos
    "TOTAL_SSCC": {
        "tipo": "excel_col", "archivo": "a_calc_sscc_01",
        "hoja": "SOBRECOSTOS TOTAL", "columna": "F", "fila_inicio": 5,
        "columna_filtro": "", "valores_filtro": [],
        "etiqueta": "Total sobrecosto SSCC (Calculo_SobrecostosSSCC, 01 Sobrecostos)",
    },
    "TOTAL_CO": {
        "tipo": "excel_col", "archivo": "a_calc_co",
        "hoja": "CO TOTAL", "columna": "F", "fila_inicio": 5,
        "columna_filtro": "", "valores_filtro": [],
        "etiqueta": "Total costo de oportunidad (Calculo_CO)",
    },
    "TOTAL_CCA": {
        "tipo": "excel_col", "archivo": "a_cons_cca",
        "hoja": "CCA", "columna": "BC", "fila_inicio": 3,
        "columna_filtro": "", "valores_filtro": [],
        "etiqueta": "Total costo combustible adicional (Consolidado_CCA)",
    },
    "TOTAL_SCMT": {
        "tipo": "excel_col", "archivo": "a_cons_tab",
        "hoja": "Sobrecostos", "columna": "AE", "fila_inicio": 3,
        "columna_filtro": "AB", "valores_filtro": ["SCMT"],
        "etiqueta": "Total sobrecosto SCMT (02 Consolidado_Tabulado)",
    },
    "TOTAL_SCPC": {
        "tipo": "excel_col", "archivo": "a_cons_tab",
        "hoja": "Sobrecostos", "columna": "AE", "fila_inicio": 3,
        "columna_filtro": "AB", "valores_filtro": ["SCPC"],
        "etiqueta": "Total sobrecosto SCPC (02 Consolidado_Tabulado)",
    },
    "TOTAL_SCAGC": {
        "tipo": "excel_col", "archivo": "a_cons_tab",
        "hoja": "Sobrecostos", "columna": "AE", "fila_inicio": 3,
        "columna_filtro": "AB", "valores_filtro": ["SCAGC"],
        "etiqueta": "Total sobrecosto SCAGC (02 Consolidado_Tabulado)",
    },
    "TOTAL_CONSOLIDADO": {
        "tipo": "excel_col", "archivo": "a_consolidado",
        "hoja": "Sobrecostos", "columna": "E", "fila_inicio": 2,
        "columna_filtro": "", "valores_filtro": [],
        "etiqueta": "Total sobrecostos (Consolidado_AAMM)",
    },

    # --- Columnas de control del Consolidado_AAMM contra su origen ------------
    # El Consolidado se arma pegando del 02 Consolidado_Tabulado:
    #     origen A:G + I:J + H   ->   destino A:J
    # o sea que las columnas quedan asi:
    #     origen G (Generacion) -> destino G
    #     origen I (CV)         -> destino H
    #     origen J (CMg)        -> destino I     <- ojo, CV y CMg quedan corridas
    # Comparar las tres sumas caza que el pegado se haya corrido de columna, que
    # es lo que el total de la E sola no detecta.
    #
    # El lado del ORIGEN se filtra por SCMT + SCPC (columna AB), porque el
    # Consolidado solo lleva esos dos tipos. El lado del DESTINO no se filtra: no
    # tiene columna de tipo, justamente porque la AB del origen no se copia.
    #
    # Sumar el CMg es raro (es un precio, no un monto) pero como suma de control
    # sirve igual: si las dos sumas coinciden, la columna es la misma.
    "TAB_GENERACION": {
        "tipo": "excel_col", "archivo": "a_cons_tab",
        "hoja": "Sobrecostos", "columna": "G", "fila_inicio": 3,
        "columna_filtro": "AB", "valores_filtro": ["SCMT", "SCPC"],
        "etiqueta": "Generación, col. G del 02 Consolidado_Tabulado (SCMT+SCPC)",
    },
    "TAB_CV": {
        "tipo": "excel_col", "archivo": "a_cons_tab",
        "hoja": "Sobrecostos", "columna": "I", "fila_inicio": 3,
        "columna_filtro": "AB", "valores_filtro": ["SCMT", "SCPC"],
        "etiqueta": "CV, col. I del 02 Consolidado_Tabulado (SCMT+SCPC)",
    },
    "TAB_CMG": {
        "tipo": "excel_col", "archivo": "a_cons_tab",
        "hoja": "Sobrecostos", "columna": "J", "fila_inicio": 3,
        "columna_filtro": "AB", "valores_filtro": ["SCMT", "SCPC"],
        "etiqueta": "CMg, col. J del 02 Consolidado_Tabulado (SCMT+SCPC)",
    },
    "CONS_GENERACION": {
        "tipo": "excel_col", "archivo": "a_consolidado",
        "hoja": "Sobrecostos", "columna": "G", "fila_inicio": 2,
        "columna_filtro": "", "valores_filtro": [],
        "etiqueta": "Generación, col. G del Consolidado_AAMM",
    },
    "CONS_CV": {
        "tipo": "excel_col", "archivo": "a_consolidado",
        "hoja": "Sobrecostos", "columna": "H", "fila_inicio": 2,
        "columna_filtro": "", "valores_filtro": [],
        "etiqueta": "CV, col. H del Consolidado_AAMM",
    },
    "CONS_CMG": {
        "tipo": "excel_col", "archivo": "a_consolidado",
        "hoja": "Sobrecostos", "columna": "I", "fila_inicio": 2,
        "columna_filtro": "", "valores_filtro": [],
        "etiqueta": "CMg, col. I del Consolidado_AAMM",
    },

    # --- celdas puntuales --------------------------------------------------
    "PAGO_L5": {
        "tipo": "excel", "archivo": "a_pago", "hoja": "VERIFICADORES", "celda": "L5",
        "etiqueta": "Pago_Sobrecostos  VERIFICADORES!L5",
    },
    "PAGO_M5": {
        "tipo": "excel", "archivo": "a_pago", "hoja": "VERIFICADORES", "celda": "M5",
        "etiqueta": "Pago_Sobrecostos  VERIFICADORES!M5",
    },
    "SSCC_EE6": {
        "tipo": "excel", "archivo": "a_calc_sscc_01", "hoja": "SOBRECOSTOS", "celda": "EE6",
        "etiqueta": "Calculo_SobrecostosSSCC  SOBRECOSTOS!EE6",
    },
    "SSCC_H1": {
        "tipo": "excel", "archivo": "a_calc_sscc_01",
        "hoja": "SOBRECOSTOS TOTAL", "celda": "H1",
        "etiqueta": "Calculo_SobrecostosSSCC  'SOBRECOSTOS TOTAL'!H1",
    },

    # 3_REMUNERACION_SUBASTAS_E_ID  (maestro en 04 Planilla 9)
    # El descuadre del cuadro de pago. En la hoja CPRT del cuadro cero:
    #   I1 = SUM(G:G)                                  total del cuadro de pago
    #   I2 = SUM('01.SSCC_Recurso_Tecnico'!K:K)        total de origen
    #   I3 = I2 - I1                                   el residuo
    # Nunca da 0 exacto porque la matriz reparte proporcionalmente y redondea;
    # lo que importa es que sea CHICO. Por eso se compara contra un umbral y no
    # contra cero.
    "DESCUADRE_CPRT": {
        "tipo": "excel", "archivo": "a_0_cuadros", "hoja": "CPRT", "celda": "I3",
        "etiqueta": "Descuadre del cuadro de pago (CPRT!I3 = I2 - I1)",
    },
    "P3_RES_D7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN", "celda": "D7",
                  "etiqueta": "Planilla 3  RESUMEN!D7"},
    "P3_RES_E7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN", "celda": "E7",
                  "etiqueta": "Planilla 3  RESUMEN!E7"},
    "P3_RES_F7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN", "celda": "F7",
                  "etiqueta": "Planilla 3  RESUMEN!F7"},
    "P3_RES_L7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN", "celda": "L7",
                  "etiqueta": "Planilla 3  RESUMEN!L7"},
    "P3_RES_M7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN", "celda": "M7",
                  "etiqueta": "Planilla 3  RESUMEN!M7"},
    "P3_RES_N7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN", "celda": "N7",
                  "etiqueta": "Planilla 3  RESUMEN!N7"},
    "P3_RES_U7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN", "celda": "U7",
                  "etiqueta": "Planilla 3  RESUMEN!U7  (total Subastas)"},
    "P3_RES_V7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN", "celda": "V7",
                  "etiqueta": "Planilla 3  RESUMEN!V7"},
    "P3_RES_W7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN", "celda": "W7",
                  "etiqueta": "Planilla 3  RESUMEN!W7"},
    "P3_DES_D7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN_DESGLOSADO",
                  "celda": "D7", "etiqueta": "Planilla 3  RESUMEN_DESGLOSADO!D7"},
    "P3_DES_E7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN_DESGLOSADO",
                  "celda": "E7", "etiqueta": "Planilla 3  RESUMEN_DESGLOSADO!E7"},
    "P3_DES_F7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN_DESGLOSADO",
                  "celda": "F7", "etiqueta": "Planilla 3  RESUMEN_DESGLOSADO!F7"},
    "P3_DES_K7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN_DESGLOSADO",
                  "celda": "K7", "etiqueta": "Planilla 3  RESUMEN_DESGLOSADO!K7"},
    "P3_DES_L7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN_DESGLOSADO",
                  "celda": "L7", "etiqueta": "Planilla 3  RESUMEN_DESGLOSADO!L7"},
    "P3_DES_M7": {"tipo": "excel", "archivo": "a_3_p9", "hoja": "RESUMEN_DESGLOSADO",
                  "celda": "M7", "etiqueta": "Planilla 3  RESUMEN_DESGLOSADO!M7"},

    # 5_REMUNERACION_CRA  (maestro)
    "P5_D7": {"tipo": "excel", "archivo": "a_5_p9", "hoja": "RESUMEN", "celda": "D7",
              "etiqueta": "Planilla 5  RESUMEN!D7  (total CRA)"},
    "P5_E7": {"tipo": "excel", "archivo": "a_5_p9", "hoja": "RESUMEN", "celda": "E7",
              "etiqueta": "Planilla 5  RESUMEN!E7"},
    "P5_F7": {"tipo": "excel", "archivo": "a_5_p9", "hoja": "RESUMEN", "celda": "F7",
              "etiqueta": "Planilla 5  RESUMEN!F7"},

    # 6_REMUNERACION_REA_Y_CO_ERNC  (maestro)
    "P6_E7": {"tipo": "excel", "archivo": "a_6_p9", "hoja": "RESUMEN", "celda": "E7",
              "etiqueta": "Planilla 6  RESUMEN!E7  (total REA)"},
    "P6_F7": {"tipo": "excel", "archivo": "a_6_p9", "hoja": "RESUMEN", "celda": "F7",
              "etiqueta": "Planilla 6  RESUMEN!F7"},
    "P6_G7": {"tipo": "excel", "archivo": "a_6_p9", "hoja": "RESUMEN", "celda": "G7",
              "etiqueta": "Planilla 6  RESUMEN!G7"},
    "P6_L7": {"tipo": "excel", "archivo": "a_6_p9", "hoja": "RESUMEN", "celda": "L7",
              "etiqueta": "Planilla 6  RESUMEN!L7  (total CO-ERNC)"},
    "P6_M7": {"tipo": "excel", "archivo": "a_6_p9", "hoja": "RESUMEN", "celda": "M7",
              "etiqueta": "Planilla 6  RESUMEN!M7"},
    "P6_N7": {"tipo": "excel", "archivo": "a_6_p9", "hoja": "RESUMEN", "celda": "N7",
              "etiqueta": "Planilla 6  RESUMEN!N7"},

    # 9_Pagos_Retiros y 4_REMUNERACION_SC_CO_CCA  (en 00 Entregables)
    # La tabla de la planilla 9 cambia de largo, asi que el total NO se busca en
    # una celda fija sino en la fila rotulada "Total general".
    "P9_TOTAL_B": {"tipo": "excel_etiqueta", "archivo": "a_9_pagos",
                   "hoja": "VERIFICADORES", "fila_inicio": 1,
                   "columna_etiqueta": "A", "texto_fila": "Total general",
                   "columna_valor": "B",
                   "etiqueta": "Planilla 9  total general (rótulo en A, valor en B)"},
    "P9_TOTAL_E": {"tipo": "excel_etiqueta", "archivo": "a_9_pagos",
                   "hoja": "VERIFICADORES", "fila_inicio": 1,
                   "columna_etiqueta": "D", "texto_fila": "Total general",
                   "columna_valor": "E",
                   "etiqueta": "Planilla 9  total general (rótulo en D, valor en E)"},
    "P4_K6": {"tipo": "excel", "archivo": "a_4_rem", "hoja": "VERIFICADORES",
              "celda": "K6", "etiqueta": "Planilla 4  VERIFICADORES!K6"},
    "P4_L6": {"tipo": "excel", "archivo": "a_4_rem", "hoja": "VERIFICADORES",
              "celda": "L6", "etiqueta": "Planilla 4  VERIFICADORES!L6"},

    # 1_CUADROS_PAGO_SSCC, hoja RESUMEN
    "P1_C10": {"tipo": "excel", "archivo": "a_1_cuadros", "hoja": "RESUMEN", "celda": "C10",
               "etiqueta": "1_CUADROS_PAGO  C10"},
    "P1_C11": {"tipo": "excel", "archivo": "a_1_cuadros", "hoja": "RESUMEN", "celda": "C11",
               "etiqueta": "1_CUADROS_PAGO  C11"},
    "P1_C12": {"tipo": "excel", "archivo": "a_1_cuadros", "hoja": "RESUMEN", "celda": "C12",
               "etiqueta": "1_CUADROS_PAGO  C12"},
    "P1_C13": {"tipo": "excel", "archivo": "a_1_cuadros", "hoja": "RESUMEN", "celda": "C13",
               "etiqueta": "1_CUADROS_PAGO  C13"},
    "P1_C14": {"tipo": "excel", "archivo": "a_1_cuadros", "hoja": "RESUMEN", "celda": "C14",
               "etiqueta": "1_CUADROS_PAGO  C14"},
    "P1_C15": {"tipo": "excel", "archivo": "a_1_cuadros", "hoja": "RESUMEN", "celda": "C15",
               "etiqueta": "1_CUADROS_PAGO  C15"},
    "P1_C17_C25": {"tipo": "excel", "archivo": "a_1_cuadros", "hoja": "RESUMEN",
                   "celda": "C17:C25", "etiqueta": "1_CUADROS_PAGO  suma C17:C25"},

    # --- totales desde bases Access ----------------------------------------
    # columna_tipo: si se indica, el log muestra el desglose por tipo (informativo)
    "MDB_SSCC_TOTAL": {
        "tipo": "mdb", "archivo": "a_mdb_sscc",
        "tabla": "Sobrecostos", "columna": "Sobrecosto", "where": "",
        "columna_tipo": "Tipo_sobrecosto",
        "etiqueta": "Suma sobrecostos (03b ENTRADA_SOB_SSCC.mdb)",
    },
    "MDB_SOB_TOTAL": {
        "tipo": "mdb", "archivo": "a_mdb_sob",
        "tabla": "Sobrecostos", "columna": "Sobrecosto", "where": "",
        "columna_tipo": "Tipo_sobrecosto",
        "etiqueta": "Suma sobrecostos (03b ENTRADA_SOB.mdb)",
    },
    "MDB_OCUPAR_TOTAL": {
        "tipo": "mdb", "archivo": "a_ocupar",
        "tabla": "Sobrecostos", "columna": "Sobrecosto", "where": "",
        "columna_tipo": "Tipo_sobrecosto",
        "etiqueta": "Suma sobrecostos (Ocupar_este_para_Reliquidacion.mdb)",
    },
}


# =============================================================================
#  Verificadores
#  Cada uno se dibuja como boton al lado de su archivo y corre una lista de
#  comprobaciones. Tipos de comprobacion:
#
#   igualdad  izq / der : listas de claves de VALORES; se suma cada lado y se
#                         comparan con TOLERANCIA. Una clave con '-' delante
#                         entra con signo cambiado: der=["-P4_L6"] es K6 = -L6
#                         "absoluto": True compara sin importar el signo. Usarlo
#                         solo donde el signo sea una convencion del archivo, NO
#                         donde el signo sea justamente lo que se verifica.
#   cero      claves    : todas deben valer 0
#   largo     hoja, referencia, bloques, fila_inicio : la ultima fila con datos
#                         de cada bloque de columnas debe coincidir con la del
#                         bloque de referencia (detecta formulas sin extender)
#   marcas    hoja, fila_inicio, reglas : busca valores de error (#N/D, #REF!...)
#                         y textos prohibidos en rangos de columnas
#
#   tabla     compara dos tablas de empresa+montos entre archivos distintos,
#             sin importar el orden, y omite las filas sin empresa
#   pertenencia  todas las empresas de unas columnas deben estar en otras
#   ultimo_igual el ultimo dato util de dos columnas debe coincidir (omite
#             vacios, ceros y errores tipo #REF!)
#
#   Cualquier comprobacion acepta "activa": False para apagarla sin borrarla.
#
#   depende       : archivos cuya fecha de modificacion invalida la verificacion
#   verif_previas : verificadores que deberian estar OK antes de este
# =============================================================================

VERIFICADORES = {
    "V8": {
        # La hoja SOBRECOSTOS de este archivo se llena DESDE el 02 Consolidado_
        # Tabulado ("Traer Consolidado" de Actualiza_datos.py), asi que V17 va
        # primero: si el tabulado tiene el sobrecosto mal calculado, todo lo que
        # sale de aca hereda el error y las comparaciones de totales cuadran igual.
        # Por esta arista V17 entra en la cadena del cuadro cero:
        #     V16 -> V15 -> V14 -> V4 -> V8 -> V17
        "archivo": "a_calc_sscc_01",
        "titulo": "Calculo_SobrecostosSSCC (maestro): SOBRECOSTOS!EE6 = SCAGC, 'SOBRECOSTOS TOTAL'!H1 = 0 y marcas",
        "depende": ["a_cons_tab"],
        "verif_previas": ["V17"],
        "comprobaciones": [
            {"tipo": "igualdad",
             "desc": "SOBRECOSTOS!EE6 = total SCAGC del 02 Consolidado_Tabulado",
             "izq": ["SSCC_EE6"], "der": ["TOTAL_SCAGC"]},
            {"tipo": "cero", "desc": "'SOBRECOSTOS TOTAL'!H1 = 0",
             "claves": ["SSCC_H1"]},
            {"tipo": "marcas", "activa": True,
             "desc": "sin #N/D ni textos REVISAR en las columnas de control",
             "archivo": "a_calc_sscc_01", "hoja": "SOBRECOSTOS", "fila_inicio": 7,
             "reglas": [
                 {"rangos": ["CF:CI"], "errores": True, "textos": ["REVISAR"]},
                 {"rangos": ["CD", "CE", "CS"], "errores": False, "textos": ["REVISAR"]},
             ]},
        ],
    },
    "V9": {
        "archivo": "a_3_p9",
        "titulo": "Planilla 3 (maestro): RESUMEN, RESUMEN_DESGLOSADO y la prorrata",
        "depende": ["a_prorrata"],
        "verif_previas": [],
        "comprobaciones": [
            {"tipo": "igualdad", "desc": "RESUMEN  D7 = E7",
             "izq": ["P3_RES_D7"], "der": ["P3_RES_E7"]},
            {"tipo": "igualdad", "desc": "RESUMEN  L7 = M7",
             "izq": ["P3_RES_L7"], "der": ["P3_RES_M7"]},
            {"tipo": "igualdad", "desc": "RESUMEN  U7 = V7",
             "izq": ["P3_RES_U7"], "der": ["P3_RES_V7"]},
            {"tipo": "igualdad", "desc": "RESUMEN  U7 = D7 + L7",
             "izq": ["P3_RES_U7"], "der": ["P3_RES_D7", "P3_RES_L7"]},
            {"tipo": "igualdad", "desc": "RESUMEN  V7 = E7 + M7",
             "izq": ["P3_RES_V7"], "der": ["P3_RES_E7", "P3_RES_M7"]},
            {"tipo": "cero", "desc": "RESUMEN  F7 = N7 = W7 = 0",
             "claves": ["P3_RES_F7", "P3_RES_N7", "P3_RES_W7"]},
            {"tipo": "igualdad", "desc": "RESUMEN_DESGLOSADO  D7 = E7",
             "izq": ["P3_DES_D7"], "der": ["P3_DES_E7"]},
            {"tipo": "igualdad", "desc": "RESUMEN_DESGLOSADO  K7 = L7",
             "izq": ["P3_DES_K7"], "der": ["P3_DES_L7"]},
            {"tipo": "cero", "desc": "RESUMEN_DESGLOSADO  F7 = M7 = 0",
             "claves": ["P3_DES_F7", "P3_DES_M7"]},
            # La prorrata pegada en esta planilla tiene que ser la del
            # Prorrata_Retiros de ESTE mes. Es facil olvidarse de actualizarla y
            # nada mas la delata: los totales del RESUMEN cuadran igual.
            {"tipo": "prorrata_al_dia",
             "desc": "la prorrata de PRORRATA_RETIROS es la de este mes",
             "archivo": "a_3_p9", "hoja": "PRORRATA_RETIROS",
             "fila_encabezado": 8, "col_inicio": "B",
             "origen": {"archivo": "a_prorrata",
                        "hoja": "PRORRATA_HORARIA_TABULAR", "fila_inicio": 2,
                        "col_hora": "A", "col_suministrador": "B",
                        "col_valor": "C"},
             "tolerancia": TOL_PRORRATA_SUMA},

        ],
    },
    "V10": {
        "archivo": "a_5_p9",
        "titulo": 'Planilla 5 (maestro): RESUMEN, la hoja "SC y CO" y la prorrata',
        "depende": ["a_calc_sscc_01", "a_calc_co", "a_prorrata"],
        "verif_previas": [],
        "comprobaciones": [
            {"tipo": "igualdad", "desc": "RESUMEN  D7 = E7",
             "izq": ["P5_D7"], "der": ["P5_E7"]},
            {"tipo": "cero", "desc": "RESUMEN  F7 = 0", "claves": ["P5_F7"]},

            # --- la hoja "SC y CO" -------------------------------------------
            # Cada fila de prorrata reparte el 100% entre los suministradores, asi
            # que tiene que sumar 1 (o 100). Es la comprobacion que caza una
            # prorrata pegada a medias o corrida de columna.
            # Lo unico que dice si los datos son los de ESTE mes: se compara cada
            # bloque contra su origen, por separado, para poder decir cual traer.
            {"tipo": "bloque_contra_origen",
             "desc": "SC: filas y monto de la planilla 5 = origen (embalses)",
             "destino": {"archivo": "a_5_p9", "hoja": "SC y CO", "fila_inicio": 9,
                         "col_tipo": "D", "tipo": "SCCF", "col_monto": "G",
                         "col_central": "E"},
             "origen": {"archivo": "a_calc_sscc_01", "hoja": "SOBRECOSTOS",
                        "fila_inicio": 7, "col_central": "U", "col_monto": "W"}},
            {"tipo": "bloque_contra_origen",
             "desc": "CO: filas y monto de la planilla 5 = origen (embalses)",
             "destino": {"archivo": "a_5_p9", "hoja": "SC y CO", "fila_inicio": 9,
                         "col_tipo": "D", "tipo": "CO", "col_monto": "G",
                         "col_central": "E"},
             "origen": {"archivo": "a_calc_co", "hoja": "PRORRATA CO",
                        "fila_inicio": 7, "col_central": "D", "col_monto": "G"}},

            {"tipo": "suma_fila",
             "desc": 'SC y CO: cada fila de I:X suma el 100%',
             "archivo": "a_5_p9", "hoja": "SC y CO",
             "rango": "I:X", "fila_inicio": 9, "col_referencia": "E",
             "col_tipo": "D"},

            # Solo pueden estar los embalses de la lista.
            {"tipo": "centrales_en_lista",
             "desc": 'SC y CO: en E9:E solo hay centrales de embalse',
             "archivo": "a_5_p9", "hoja": "SC y CO",
             "columna": "E", "fila_inicio": 9, "exigir": True},

            # Las formulas de Y:AF tienen que cubrir exactamente las filas que hay.
            {"tipo": "formulas_cubren",
             "desc": 'SC y CO: las fórmulas de Y:AB y AD:AF cubren todas las filas',
             "archivo": "a_5_p9", "hoja": "SC y CO",
             # La AC queda AFUERA: no lleva formula.
             "cols": ["Y:AB", "AD:AF"], "fila_inicio": 9,
             "referencia": {"archivo": "a_5_p9", "hoja": "SC y CO",
                            "col": "E", "fila_inicio": 9}},

            # En el ORIGEN: si aparece una central "-numero" que no esta en la
            # lista, puede ser una unidad de embalse nueva y estaria quedando fuera.
            {"tipo": "centrales_en_lista",
             "desc": "origen SC: no quedó fuera ninguna central «-número»",
             "archivo": "a_calc_sscc_01", "hoja": "SOBRECOSTOS",
             "columna": "U", "fila_inicio": 7, "avisar_sufijo": True},
            {"tipo": "centrales_en_lista",
             "desc": "origen CO: no quedó fuera ninguna central «-número»",
             "archivo": "a_calc_co", "hoja": "PRORRATA CO",
             "columna": "D", "fila_inicio": 7, "avisar_sufijo": True},
            # La prorrata pegada en esta planilla tiene que ser la del
            # Prorrata_Retiros de ESTE mes. Es facil olvidarse de actualizarla y
            # nada mas la delata: los totales del RESUMEN cuadran igual.
            {"tipo": "prorrata_al_dia",
             "desc": "la prorrata de PRORRATA_RETIROS es la de este mes",
             "archivo": "a_5_p9", "hoja": "PRORRATA_RETIROS",
             "fila_encabezado": 8, "col_inicio": "B",
             "origen": {"archivo": "a_prorrata",
                        "hoja": "PRORRATA_HORARIA_TABULAR", "fila_inicio": 2,
                        "col_hora": "A", "col_suministrador": "B",
                        "col_valor": "C"},
             "tolerancia": TOL_PRORRATA_SUMA},

        ],
    },
    "V11": {
        "archivo": "a_6_p9",
        "titulo": "Planilla 6 (maestro): RESUMEN y la prorrata",
        "depende": ["a_prorrata"],
        "verif_previas": [],
        "comprobaciones": [
            {"tipo": "igualdad", "desc": "RESUMEN  E7 = F7",
             "izq": ["P6_E7"], "der": ["P6_F7"]},
            {"tipo": "igualdad", "desc": "RESUMEN  L7 = M7",
             "izq": ["P6_L7"], "der": ["P6_M7"]},
            {"tipo": "cero", "desc": "RESUMEN  G7 = N7 = 0",
             "claves": ["P6_G7", "P6_N7"]},
            # La prorrata pegada en esta planilla tiene que ser la del
            # Prorrata_Retiros de ESTE mes. Es facil olvidarse de actualizarla y
            # nada mas la delata: los totales del RESUMEN cuadran igual.
            {"tipo": "prorrata_al_dia",
             "desc": "la prorrata de PRORRATA_RETIROS es la de este mes",
             "archivo": "a_6_p9", "hoja": "PRORRATA_RETIROS",
             "fila_encabezado": 8, "col_inicio": "B",
             "origen": {"archivo": "a_prorrata",
                        "hoja": "PRORRATA_HORARIA_TABULAR", "fila_inicio": 2,
                        "col_hora": "A", "col_suministrador": "B",
                        "col_valor": "C"},
             "tolerancia": TOL_PRORRATA_SUMA},

        ],
    },
    "V4": {
        "archivo": "a_mdb_sscc",
        "titulo": "03b ENTRADA_SOB_SSCC vs Calculo SSCC + Calculo CO + Consolidado CCA",
        "depende": ["a_calc_sscc_01", "a_calc_co", "a_cons_cca"],
        "verif_previas": ["V8"],
        "comprobaciones": [
            {"tipo": "igualdad",
             "desc": "suma del .mdb = SSCC + CO + CCA",
             "izq": ["MDB_SSCC_TOTAL"],
             "der": ["TOTAL_SSCC", "TOTAL_CO", "TOTAL_CCA"]},
        ],
    },
    "V5": {
        "archivo": "a_mdb_sob",
        "titulo": "03b ENTRADA_SOB vs SCMT + SCPC (02 Consolidado_Tabulado)",
        "depende": ["a_cons_tab"],
        # La suma con la que se compara sale del tabulado: si el tabulado esta mal
        # cuadra igual, y las dos puntas quedan mal.
        "verif_previas": ["V17"],
        "comprobaciones": [
            {"tipo": "igualdad", "desc": "suma del .mdb = SCMT + SCPC",
             "izq": ["MDB_SOB_TOTAL"], "der": ["TOTAL_SCMT", "TOTAL_SCPC"]},
        ],
    },
    "V6": {
        "archivo": "a_consolidado",
        "titulo": "Consolidado_AAMM vs SCMT + SCPC (02 Consolidado_Tabulado)",
        "depende": ["a_cons_tab"],
        # Idem V5: las cuatro comprobaciones comparan contra el tabulado.
        "verif_previas": ["V17"],
        "comprobaciones": [
            {"tipo": "igualdad", "desc": "total del Consolidado = SCMT + SCPC",
             "izq": ["TOTAL_CONSOLIDADO"], "der": ["TOTAL_SCMT", "TOTAL_SCPC"]},
            # Las tres columnas de control. Van con el corrimiento del pegado:
            # G->G, I->H, J->I.
            {"tipo": "igualdad",
             "desc": "Generación: col. G del Consolidado = col. G del tabulado",
             "izq": ["CONS_GENERACION"], "der": ["TAB_GENERACION"]},
            {"tipo": "igualdad",
             "desc": "CV: col. H del Consolidado = col. I del tabulado",
             "izq": ["CONS_CV"], "der": ["TAB_CV"]},
            {"tipo": "igualdad",
             "desc": "CMg: col. I del Consolidado = col. J del tabulado",
             "izq": ["CONS_CMG"], "der": ["TAB_CMG"]},
        ],
    },
    "V7": {
        "archivo": "a_pago",
        "titulo": "Pago_Sobrecostos: VERIFICADORES!L5 vs 03b ENTRADA_SOB.mdb y M5 = 0",
        "depende": ["a_mdb_sob"],
        "verif_previas": ["V5"],
        "comprobaciones": [
            {"tipo": "igualdad", "desc": "VERIFICADORES!L5 = suma del ENTRADA_SOB.mdb",
             "izq": ["PAGO_L5"], "der": ["MDB_SOB_TOTAL"]},
            {"tipo": "cero", "desc": "VERIFICADORES!M5 = 0", "claves": ["PAGO_M5"]},
        ],
    },
    "V12": {
        "archivo": "a_ocupar",
        "titulo": "Ocupar_este_para_Reliquidacion vs Subastas + CRA + REA + CO-ERNC",
        "depende": ["a_3_p9", "a_5_p9", "a_6_p9"],
        "verif_previas": ["V9", "V10", "V11"],
        "comprobaciones": [
            {"tipo": "igualdad",
             "desc": "suma del .mdb = U7 (Subastas) + D7 (CRA) + E7 (REA) + L7 (CO-ERNC)",
             "izq": ["MDB_OCUPAR_TOTAL"],
             "der": ["P3_RES_U7", "P5_D7", "P6_E7", "P6_L7"]},

            # Toda central con plata tiene que tener dueño. Una central SIN dueño
            # no es error por si misma (en CONSUMOS_PROPIOS las hay a proposito);
            # lo que no puede pasar es que tenga monto y no se sepa a quien
            # pagarle o cobrarle.
            {"tipo": "centrales_sin_dueno",
             "desc": "toda central con monto tiene dueño en Central_Empresa",
             "archivo": "a_ocupar",
             "tabla_montos": "Sobrecostos", "tabla_duenos": "Central_Empresa"},
        ],
    },
    "V13": {
        "archivo": "a_9_pagos",
        "titulo": "Planilla 9: total general (A/B) = total general (D/E) = suma de Ocupar_este.mdb",
        "depende": ["a_ocupar"],
        "verif_previas": ["V12"],
        "comprobaciones": [
            {"tipo": "igualdad", "desc": "total general de A/B = total general de D/E",
             "izq": ["P9_TOTAL_B"], "der": ["P9_TOTAL_E"]},
            {"tipo": "igualdad", "desc": "total general = suma del Ocupar_este.mdb",
             "izq": ["P9_TOTAL_B"], "der": ["MDB_OCUPAR_TOTAL"]},
        ],
    },
    "V14": {
        "archivo": "a_4_rem",
        "titulo": "Planilla 4: K6 = -L6 y L6 = suma de 03b ENTRADA_SOB_SSCC.mdb",
        "depende": ["a_mdb_sscc"],
        "verif_previas": ["V4"],
        "comprobaciones": [
            {"tipo": "igualdad", "desc": "K6 = -L6",
             "izq": ["P4_K6"], "der": ["-P4_L6"]},
            # La planilla 4 guarda un lado en negativo y el .mdb en positivo,
            # asi que aca el signo es convencion del archivo: se compara en
            # valor absoluto. En la comprobacion de arriba NO, porque ahi el
            # signo opuesto es justamente lo que se verifica.
            {"tipo": "igualdad", "desc": "L6 = suma del ENTRADA_SOB_SSCC.mdb",
             "absoluto": True,
             "izq": ["P4_L6"], "der": ["MDB_SSCC_TOTAL"]},

        ],
    },
    "V15": {
        "archivo": "a_1_cuadros",
        "titulo": "1_CUADROS_PAGO: totales y pago por empresa contra planillas 4 y 9",
        "depende": ["a_4_rem", "a_9_pagos"],
        "verif_previas": ["V14", "V13"],
        "comprobaciones": [
            {"tipo": "igualdad", "absoluto": True,
             "desc": "C10 + C11 + C13 = planilla 4 (L6)",
             "izq": ["P1_C10", "P1_C11", "P1_C13"], "der": ["P4_L6"]},
            {"tipo": "igualdad", "absoluto": True,
             "desc": "C12 + C14 + C15 + suma(C17:C25) = total general de planilla 9",
             "izq": ["P1_C12", "P1_C14", "P1_C15", "P1_C17_C25"], "der": ["P9_TOTAL_B"]},

            # --- el pago por empresa y concepto -------------------------------
            # La planilla 1 y las planillas 9 y 4 son dos calculos PARALELOS del
            # mismo pago: la 1 lo saca por empresa y concepto directo, las otras
            # prorratean por retiros y despues agrupan.
            #
            # Va en V15 y no en V13/V14 porque la planilla 1 se arma DESPUES de
            # la 9: al verificar la 1 tiene sentido compararla contra lo que ya
            # estaba hecho. Al reves, verificar la 9 obligaria a tener lista la 1,
            # que todavia no existe.
            #
            # Las dos comprobaciones ya tienen las dependencias que hacen falta:
            # V15 depende de a_4_rem y a_9_pagos, y arrastra V13 y V14 como
            # previas.
            {"tipo": "pago_por_empresa",
             "desc": "pago por empresa y concepto: planilla 9 = planilla 1",
             "detalle": {"archivo": "a_9_pagos", "hoja": "PAGO_RETIRO",
                         "fila_inicio": 2, "nombre": "P9",
                         "col_concepto": "A", "col_empresa": "G",
                         "col_monto": "H"},
             "resumen": {"archivo": "a_1_cuadros",
                         "hoja": "01.SSCC_Recurso_Técnico",
                         "fila_inicio": 9, "nombre": "P1",
                         "col_concepto": "B", "col_empresa": "C",
                         # La E es PAGA. La D (RECIBE) no corresponde: verificado
                         # en 2409, contra PAGA cuadran los 664 pares y contra
                         # RECIBE solo 41.
                         "col_monto": "E"},
             "tolerancia": TOL_PAGO_EMPRESA},
            {"tipo": "pago_por_empresa",
             "desc": "pago por empresa y concepto: planilla 4 = planilla 1",
             "detalle": {"archivo": "a_4_rem", "hoja": "PAGO_RETIRO",
                         "fila_inicio": 2, "nombre": "P4",
                         "col_concepto": "A", "col_empresa": "G",
                         "col_monto": "H"},
             "resumen": {"archivo": "a_1_cuadros",
                         "hoja": "01.SSCC_Recurso_Técnico",
                         "fila_inicio": 9, "nombre": "P1",
                         "col_concepto": "B", "col_empresa": "C",
                         "col_monto": "E"},
             "tolerancia": TOL_PAGO_EMPRESA},
        ],
    },
    # -------------------------------------------------------------------------
    # V16 - EL CUADRO CERO. Es el que se va a pago: si sale mal hay que
    # refacturarle a todas las empresas. Por eso es el mas detallado.
    #
    # "#3" es la TERCERA hoja del libro. Se pide por posicion y no por nombre,
    # pero ojo: la cuenta incluye las hojas OCULTAS, y este libro tiene dos
    # antes de la que interesa. O sea "#3" es la primera hoja visible.
    # Las dos hojas 01.SSCC_Recurso_Tecnico (la del cuadro 1 y la del cuadro 0)
    # se llaman IGUAL pero son de archivos distintos: ojo al leer.
    # -------------------------------------------------------------------------
    "V16": {
        "archivo": "a_0_cuadros",
        "titulo": "0_CUADROS_RELIQUIDACION: el cuadro que se va a pago",
        "depende": ["a_1_cuadros"],
        "verif_previas": ["V15"],
        "comprobaciones": [
            # 1) La tabla que se copia del cuadro 1 al cuadro 0. Al ser copiar y
            #    pegar tienen que ser identicas; el orden puede cambiar, asi que
            #    se compara indexando por empresa.
            {"tipo": "tabla",
             "desc": "tabla I9:K del 1_CUADROS = tabla A5:C del cuadro 0 (por empresa)",
             "archivo_a": "a_0_cuadros", "hoja_a": "#3",
             "cols_a": ["A", "B", "C"], "fila_a": 5, "nombre_a": "cuadro 0 A5:C",
             "archivo_b": "a_1_cuadros", "hoja_b": "01.SSCC_Recurso_Técnico",
             "cols_b": ["I", "J", "K"], "fila_b": 9, "nombre_b": "1_CUADROS I9:K",
             "nombres_valor": ["paga", "recibe"],
             "sin_duplicados": True},

            # 2) La columna K es la lista consolidada de empresas: tiene que
            #    contener todo lo de A y de F, y no repetir ninguna.
            {"tipo": "pertenencia",
             "desc": "K5:K contiene todas las empresas de A5:A y F5:F, sin repetir",
             "archivo": "a_0_cuadros", "hoja": "#3", "fila_inicio": 5,
             "origen": ["A", "F"], "destino": ["K"],
             "sin_duplicados_destino": True},

            # 3) Las formulas L:U tienen que llegar justo hasta la ultima
            #    empresa de K. La Q queda AFUERA a proposito: esta vacia, es el
            #    espacio que separa dos tablas.
            {"tipo": "formulas_cubren",
             "desc": "las fórmulas de L:P y R:U cubren todas las empresas y no sobran",
             "archivo": "a_0_cuadros", "hoja": "#3",
             "cols": ["L:P", "R:U"], "fila_inicio": 5,
             "referencia": {"archivo": "a_0_cuadros", "hoja": "#3",
                            "col": "K", "fila_inicio": 5}},

            # 3.b) La D es aparte: D5 = B5+C5, o sea que acompaña a la TABLA
            #      PEGADA (A:C), no a la lista K. Y la K suele ser mas larga que
            #      la A, porque junta las empresas de A y de F. Si se comparara
            #      contra K, esta comprobacion fallaria siempre.
            #      Importa porque la D es lo que suma L5 (=SUMAR.SI(A:A;K5;D:D)):
            #      si se queda corta, las ultimas empresas suman 0.
            {"tipo": "formulas_cubren",
             "desc": "la fórmula de D cubre toda la tabla pegada (A:C) y no sobra",
             "archivo": "a_0_cuadros", "hoja": "#3",
             "cols": ["D"], "fila_inicio": 5,
             "referencia": {"archivo": "a_0_cuadros", "hoja": "#3",
                            "col": "A", "fila_inicio": 5}},

            # 4) La columna A de 01.SSCC_Recurso_Tecnico del CUADRO 0 tiene que
            #    tener EXACTAMENTE las mismas empresas que la K de la hoja #3:
            #    ni una de mas ni una de menos. Los 0 y los vacios del final se
            #    descartan, esos pueden estar.
            {"tipo": "mismas_empresas",
             "desc": "las empresas de K5:K (hoja #3) son las mismas de A9:A (01.SSCC del cuadro 0)",
             "lado_a": {"archivo": "a_0_cuadros", "hoja": "#3",
                        "col": "K", "fila_inicio": 5, "nombre": "hoja #3 K5:K"},
             "lado_b": {"archivo": "a_0_cuadros", "hoja": "01.SSCC_Recurso_Técnico",
                        "col": "A", "fila_inicio": 9, "nombre": "01.SSCC A9:A"}},

            # 5) El neto por empresa. En C9:C las empresas vienen REPETIDAS
            #    (una fila por reemplazo), asi que hay que agrupar y sumar G9:G
            #    y comparar contra J+K de la tabla resumen I9:K. Esta misma
            #    comprobacion cubre que las empresas de C y de I sean las mismas,
            #    porque avisa las que estan en un lado y no en el otro.
            {"tipo": "suma_por_empresa",
             "desc": "empresas de C9:C = empresas de I9:I, y suma de G9:G por empresa = J+K",
             "archivo": "a_0_cuadros", "hoja": "01.SSCC_Recurso_Técnico",
             "fila_inicio": 9,
             "col_empresa": "C", "col_monto": "G",
             "cols_resumen": ["I", "J", "K"],
             "nombre_resumen": "tabla I9:K del cuadro 0"},

            # 6) El descuadre del cuadro de pago, que hasta ahora se miraba a ojo.
            {"tipo": "umbral",
             "desc": "el descuadre del cuadro de pago (CPRT!I3) es chico",
             "clave": "DESCUADRE_CPRT", "maximo": UMBRAL_DESCUADRE_CPRT},

            # 7) La matriz esta armada con los datos de ahora, y el nombre
            #    definido la cubre justo. Caza que falte «Cuadro de pagos» o
            #    que falte «Actualiza Rango».
            {"tipo": "matriz_al_dia",
             "desc": "la matriz del cuadro de pagos está al día (y CPTEE la cubre)",
             "archivo": "a_0_cuadros", "hoja": "01.SSCC_Recurso_Técnico",
             "fila_tabla": 9, "nombre_definido": "CPTEE"},

            # 8) El CPRT corresponde a esa matriz. Caza que falte refrescar la
            #    tabla dinámica.
            # 9) Las formulas de la H del CPRT tienen que llegar a todas las filas:
            #    el csv se arma con la H, y donde no llegue el monto sale vacio.
            {"tipo": "formulas_cubren",
             "desc": "CPRT: las fórmulas de la H llegan a todas las filas",
             "archivo": "a_0_cuadros", "hoja": "CPRT",
             # solo_faltan: que la formula siga mas abajo esta bien; la dinamica
             # cambia de largo y el exportador corta por la columna A.
             "cols": ["H"], "fila_inicio": 7, "solo_faltan": True,
             "referencia": {"archivo": "a_0_cuadros", "hoja": "CPRT",
                            "col": "A", "fila_inicio": 7}},

            # 10) H solo puede ser igual a G o cero. Y avisa cuanto se retiene.
            {"tipo": "retencion_coherente",
             "desc": "CPRT: la H es igual a la G o es 0 (retención)",
             "archivo": "a_0_cuadros", "hoja": "CPRT", "fila_inicio": 7,
             "col_g": "G", "col_h": "H", "col_referencia": "A"},

            {"tipo": "cprt_al_dia",
             "desc": "el CPRT corresponde a la matriz (dinámica refrescada)",
             "archivo": "a_0_cuadros",
             "hoja_matriz": "01.SSCC_Recurso_Técnico",
             "hoja_cprt": "CPRT", "fila_cprt": 7},
        ],
    },

    # -------------------------------------------------------------------------
    # V17 - EL CONSOLIDADO TABULADO, que es el origen de casi todo lo demas.
    #
    # El sobrecosto llega ya calculado en la columna E, pero tambien llegan sus
    # componentes por separado. Recalcularlo es la unica verificacion de la cadena
    # que NO depende de que el archivo se copie bien: comprueba que el numero sea
    # correcto, no que coincida con otra copia de si mismo.
    #
    # Todas las columnas tienen el encabezado en la fila 2, asi que los datos
    # arrancan en la 3.
    # -------------------------------------------------------------------------
    "V17": {
        "archivo": "a_cons_tab",
        "titulo": "02 Consolidado_Tabulado: el sobrecosto recalculado desde sus componentes",
        "depende": [],
        "verif_previas": [],
        "comprobaciones": [
            {"tipo": "sobrecosto_por_fila",
             "desc": "(CV − CMg) × Generación × USD = Sobrecosto (col. E)",
             "archivo": "a_cons_tab", "hoja": "Sobrecostos", "fila_inicio": 3,
             "columnas": {"cv": "I", "cmg": "J", "gen": "G", "usd": "W",
                          "resultado": "E"},
             "tolerancia_fila": TOL_SOBRECOSTO_FILA},
        ],
    },
}

# =============================================================================
#  Definicion del arbol de carpetas / archivos
#  tipo: "carpeta" | "archivo" | "diarios"
#  espejo: id del archivo maestro con el que se compara la fecha de modificacion
# =============================================================================

XL = (".xlsm", ".xlsx", ".xlsb")
DB = (".mdb", ".accdb")

# La carpeta de detalles se llama distinto segun el modulo ("Detalles diarios"
# en Sobrecostos, "Detalle diario" en CO y CCA). Se prueban todas las variantes.
DD = ("Detalles diarios", "Detalle diario", "Detalle diarios", "Detalles diario")

NODOS = [
    # La carpeta FD NO es parte de 02 CASO RELIQUIDACION: esta un nivel arriba.
    # Es el origen de todas las hojas FD, asi que se muestra para avisar si
    # falta antes de lanzar una actualizacion, y para ver si es mas nueva que
    # el archivo que se actualizo con ella. No se verifica: es solo origen.
    # "sube": la carpeta se resuelve desde base.parent, no desde base.
    dict(id="c_fd", tipo="carpeta", pref="",
         texto="../FD/          « ORIGEN — fuera de 02 CASO RELIQUIDACION »"),
    dict(id="a_sscc_desempeno", tipo="archivo", pref="    └── ",
         texto="SSCC_Desempeno*.xlsx|xlsm   (origen de las hojas FD)",
         carpeta=["FD"], sube=True, solo_info=True,
         patron=r"^sscc_desempeno", ext=XL, espejo=None),

    dict(id="c_ent", tipo="carpeta", pref="├── ", texto="00 Entregables/"),

    dict(id="c_ent_sob", tipo="carpeta", pref="│   ├── ", texto="01 Sobrecostos/"),
    dict(id="d_ent_sob", tipo="diarios", pref="│   │   ├── ", texto="Detalles diarios/  (Detalle Sobrecostos AAAAMMDD)",
         carpeta=["00 Entregables", "01 Sobrecostos", DD],
         patron=r"^detalle sobrecostos (\d{8})", ext=XL, espejo="d_sob"),
    dict(id="a_calc_sscc_ent", tipo="archivo", pref="│   │   └── ", texto="Cálculo_SobrecostosSSCC_AAMM_*.xlsm",
         carpeta=["00 Entregables", "01 Sobrecostos"],
         patron=r"^calculo_sobrecostossscc_", ext=XL, espejo="a_calc_sscc_01"),

    dict(id="c_ent_co", tipo="carpeta", pref="│   ├── ", texto="02 Costo de Oportunidad/"),
    dict(id="d_ent_co", tipo="diarios", pref="│   │   ├── ", texto="Detalle diario/  (Detalle Costo de Oportunidad AAAAMMDD)",
         carpeta=["00 Entregables", "02 Costo de Oportunidad", DD],
         patron=r"^detalle costo de oportunidad (\d{8})", ext=XL, espejo=None),
    dict(id="a_calc_co", tipo="archivo", pref="│   │   └── ", texto="Cálculo_CO_AAMM_*.xlsm",
         carpeta=["00 Entregables", "02 Costo de Oportunidad"],
         patron=r"^calculo_co_", ext=XL, espejo=None),

    dict(id="c_ent_cca", tipo="carpeta", pref="│   ├── ", texto="03 Costo de Combustible Adicional/"),
    dict(id="d_ent_cca", tipo="diarios", pref="│   │   ├── ", texto="Detalle diario/  (Detalle CCA AAAAMMDD)",
         carpeta=["00 Entregables", "03 Costo de Combustible Adicional", DD],
         patron=r"^detalle cca (\d{8})", ext=XL, espejo=None),
    dict(id="a_cons_cca", tipo="archivo", pref="│   │   └── ", texto="Consolidado_CCA_AAMM_*.xlsm",
         carpeta=["00 Entregables", "03 Costo de Combustible Adicional"],
         patron=r"^consolidado_cca_", ext=XL, espejo=None),

    dict(id="a_0_cuadros", tipo="archivo", pref="│   ├── ", texto="0_CUADROS_RELIQUIDACIÓN SSCC_AAMM_*.xlsm",
         carpeta=["00 Entregables"], patron=r"^0_cuadros", ext=XL, espejo=None),
    dict(id="a_1_cuadros", tipo="archivo", pref="│   ├── ", texto="1_CUADROS_PAGO_SSCC_AAMM_*.xlsm",
         carpeta=["00 Entregables"], patron=r"^1_cuadros", ext=XL, espejo=None),
    dict(id="a_3_ent", tipo="archivo", pref="│   ├── ", texto="3_REMUNERACIÓN_SUBASTAS_E_ID_AAMM_*.xlsm",
         carpeta=["00 Entregables"], patron=r"^3_remuneracion_subastas", ext=XL, espejo="a_3_p9"),
    dict(id="a_4_rem", tipo="archivo", pref="│   ├── ", texto="4_REMUNERACIÓN_SC_CO_CCA_Y_Pagos_Retiros_AAMM_*.xlsx",
         carpeta=["00 Entregables"], patron=r"^4_remuneracion_sc", ext=XL, espejo=None),
    dict(id="a_5_ent", tipo="archivo", pref="│   ├── ", texto="5_REMUNERACIÓN_CRA_AAMM_*.xlsx",
         carpeta=["00 Entregables"], patron=r"^5_remuneracion_cra", ext=XL, espejo="a_5_p9"),
    dict(id="a_6_ent", tipo="archivo", pref="│   ├── ", texto="6_REMUNERACIÓN_REA_Y_CO_ERNC_AAMM_*.xlsx",
         carpeta=["00 Entregables"], patron=r"^6_remuneracion_rea", ext=XL, espejo="a_6_p9"),
    dict(id="a_9_pagos", tipo="archivo", pref="│   └── ", texto="9_Pagos_Retiros_CRA_REA_CO_ERNC_Subastas_AAMM_*.xlsx",
         carpeta=["00 Entregables"], patron=r"^9_pagos_retiros", ext=XL, espejo=None),

    dict(id="c_sob", tipo="carpeta", pref="├── ", texto="01 Sobrecostos/          « MAESTRO »"),
    dict(id="d_sob", tipo="diarios", pref="│   ├── ", texto="Detalles diarios/  (maestro)",
         carpeta=["01 Sobrecostos", DD],
         patron=r"^detalle sobrecostos (\d{8})", ext=XL, espejo=None),
    dict(id="a_cons_tab", tipo="archivo", pref="│   │   └── ", texto="02 Consolidado_Tabulado_AAMM_*.xlsm   (en Detalles diarios)",
         carpeta=["01 Sobrecostos", DD], patron=r"^02 consolidado_tabulado", ext=XL, espejo=None),
    dict(id="a_mdb_sscc", tipo="archivo", pref="│   ├── ", texto="03b ENTRADA_SOB_SSCC_AAMM_*.mdb",
         carpeta=["01 Sobrecostos"], patron=r"^03b entrada_sob_sscc", ext=DB, espejo=None),
    dict(id="a_calc_sscc_01", tipo="archivo", pref="│   └── ", texto="Cálculo_SobrecostosSSCC_AAMM_*.xlsm  (maestro)",
         carpeta=["01 Sobrecostos"], patron=r"^calculo_sobrecostossscc_", ext=XL, espejo=None),

    dict(id="c_sobe", tipo="carpeta", pref="├── ", texto="01.a Sobrecostos de Energia/"),
    dict(id="d_sobe", tipo="diarios", pref="│   ├── ", texto="Detalles diarios/  (Detalle Sobrecostos AAAAMMDD)",
         carpeta=["01.a Sobrecostos de Energia", DD],
         patron=r"^detalle sobrecostos (\d{8})", ext=XL, espejo="d_sob"),
    dict(id="a_mdb_sob", tipo="archivo", pref="│   ├── ", texto="03b ENTRADA_SOB_AAMM_*.mdb",
         carpeta=["01.a Sobrecostos de Energia"], patron=r"^03b entrada_sob_(?!sscc)", ext=DB, espejo=None),
    dict(id="a_consolidado", tipo="archivo", pref="│   ├── ", texto="Consolidado_AAMM_*.xlsm",
         carpeta=["01.a Sobrecostos de Energia"], patron=r"^consolidado_(?!cca)", ext=XL, espejo=None),
    dict(id="a_pago", tipo="archivo", pref="│   └── ", texto="Pago_Sobrecostos_AAMM_*.xlsx",
         carpeta=["01.a Sobrecostos de Energia"], patron=r"^pago_sobrecostos", ext=XL, espejo=None),

    dict(id="c_p9", tipo="carpeta", pref="└── ", texto="04 Planilla 9/          « MAESTRO de 3_ / 5_ / 6_ »"),
    dict(id="a_3_p9", tipo="archivo", pref="    ├── ", texto="3_REMUNERACIÓN_SUBASTAS_E_ID_AAMM_*.xlsm  (maestro)",
         carpeta=["04 Planilla 9"], patron=r"^3_remuneracion_subastas", ext=XL, espejo=None),
    dict(id="a_5_p9", tipo="archivo", pref="    ├── ", texto="5_REMUNERACIÓN_CRA_AAMM_*.xlsx  (maestro)",
         carpeta=["04 Planilla 9"], patron=r"^5_remuneracion_cra", ext=XL, espejo=None),
    dict(id="a_6_p9", tipo="archivo", pref="    ├── ", texto="6_REMUNERACIÓN_REA_Y_CO_ERNC_AAMM_*.xlsx  (maestro)",
         carpeta=["04 Planilla 9"], patron=r"^6_remuneracion_rea", ext=XL, espejo=None),
    dict(id="a_ocupar", tipo="archivo", pref="    ├── ", texto="Ocupar_este_para_Reliquidacion_AAMM_*.mdb",
         carpeta=["04 Planilla 9"], patron=r"^ocupar_este", ext=DB, espejo=None),
    dict(id="a_prorrata", tipo="archivo", pref="    ├── ", texto="Prorrata_Retiros_AAMM_*.xlsx",
         carpeta=["04 Planilla 9"], patron=r"^prorrata_retiros", ext=XL, espejo=None),
    # Origen de la carga a SQL Server. No se verifica su contenido: es un parquet,
    # no un Excel. Solo se muestra el nombre y la fecha, y desde aca se lanza la
    # carga. Lo que valida la carga es el propio script, contra el servidor.
    dict(id="a_retiros_parq", tipo="archivo", pref="    └── ",
         texto="Retiros_h.parquet   (se carga a SQL Server)",
         carpeta=["04 Planilla 9"], patron=r"^retiros", ext=(".parquet",),
         espejo=None, solo_info=True,
         estado_info="origen de la carga a SQL"),
]

NODO_POR_ID = {n["id"]: n for n in NODOS}


def _base_de(base, nodo):
    """Carpeta desde la que se resuelve un nodo. Casi todos bajan desde la raiz
    del caso; los marcados con "sube" (la carpeta FD) cuelgan un nivel arriba."""
    return Path(base).parent if nodo.get("sube") else Path(base)


def _texto_carpeta(nodo):
    """Ruta del nodo en texto, para la bitacora. Cada parte puede ser una tupla
    de alternativas (por ejemplo "Detalles diarios" / "Detalle diario"), asi que
    no se puede hacer un join directo."""
    partes = []
    for p in nodo["carpeta"]:
        partes.append(p if isinstance(p, str) else " | ".join(p))
    if nodo.get("sube"):
        partes.insert(0, "..")
    return "/".join(partes)


# =============================================================================
#  Botón "Actualizar data": qué script lanza cada archivo maestro
# =============================================================================
# Solo va en los MAESTROS, que son los que el usuario edita. Las copias de
# "00 Entregables" se actualizan copiando el maestro, no corriendo un script.
#   script   : ruta relativa a la carpeta del revisor. Puede traer subcarpeta.
#   planilla : para Actualiza_datos.py, cual radio preseleccionar.
#   texto    : rotulo del boton (por omision "Actualizar data").
# El valor es una LISTA porque una fila puede tener mas de un actualizador: el
# cuadro cero tiene dos, los reemplazos y los pasos del cuadro.
ACTUALIZADORES = {
    "a_calc_sscc_01": [dict(script="actualizadores/Actualiza_datos.py", planilla="sc")],
    "a_3_p9":         [dict(script="actualizadores/Actualiza_datos.py", planilla="p3")],
    # Escribe en SQL Server, no en un Excel. Va como actualizador igual: abre su
    # ventana, recibe la ruta por el JSON y el revisor no queda dueno del proceso.
    "a_retiros_parq": [dict(script="actualizadores/Carga_Retiros.py", texto="Cargar retiros")],

    # El .mdb de la planilla 9 se arma desde las planillas 3, 5, 6 (y 11).
    "a_ocupar":       [dict(script="actualizadores/Actualiza_Access_P9.py",
                            texto="Actualizar data"),
                       # Prorratear va en los TRES .mdb: sirve cualquiera.
                       dict(script="actualizadores/Prorratear.py", texto="Prorratear")],
    "a_5_p9": [
        dict(script="actualizadores/Actualiza_datos.py", planilla="p5"),
        # La hoja "SC y CO" no la toca Actualiza_datos.py: va aparte.
        dict(script="actualizadores/Actualiza_SC_CO.py", texto='Actualizar "SC y CO"'),
    ],
    "a_6_p9":         [dict(script="actualizadores/Actualiza_datos.py", planilla="p6")],
    "a_mdb_sscc":     [dict(script="actualizadores/Actualiza_Data_Access.py"),
                       dict(script="actualizadores/Prorratear.py", texto="Prorratear")],
    # Los dos de Energia abren la MISMA ventana, con los dos casilleros en
    # blanco: desde ahi se elige actualizar el Access, el Consolidado o los dos.
    "a_mdb_sob":      [dict(script="actualizadores/Actualiza_Energia.py"),
                       dict(script="actualizadores/Prorratear.py", texto="Prorratear")],
    "a_consolidado":  [dict(script="actualizadores/Actualiza_Energia.py")],
    "a_0_cuadros": [
        # Reemplazos REUC vive en su propia subcarpeta y tiene su propio
        # reemplazos_reuc.json (dentro de __config__), no el compartido.
        dict(script="Reemplazos REUC/ActualizaRemplazos.py",
             texto="Actualizar reemplazos"),
        # Los pasos del propio cuadro: tabla, tasa, formulas, macros, dinamica.
        dict(script="actualizadores/Actualiza_Cuadro0.py", texto="Actualizar cuadro 0"),
    ],
}

# Traduccion de los id del arbol a las claves del JSON de traspaso, que son las
# que esperan los actualizadores.
CLAVES_TRASPASO = {
    "a_sscc_desempeno": "sscc_desempeno",
    "a_cons_tab":       "consolidado_tabulado",
    "a_prorrata":       "prorrata_retiros",
    "a_calc_sscc_01":   "calculo_sscc_maestro",
    "a_calc_co":        "calculo_co",
    "a_cons_cca":       "consolidado_cca",
    "a_3_p9":           "p3",
    "a_5_p9":           "p5",
    "a_6_p9":           "p6",
    "a_mdb_sscc":       "mdb_sscc",
    "a_mdb_sob":        "mdb_sob",
    "a_consolidado":    "consolidado_energia",
    "a_0_cuadros":      "cuadro_0",
    "a_1_cuadros":      "cuadro_1",
    "a_retiros_parq":   "retiros_parquet",
    "a_ocupar":         "mdb_ocupar",
}

ARCHIVO_TRASPASO = "_traspaso_actualizador.json"
TRASPASO_VERSION = 1

# Acciones que corren DENTRO del revisor, sin lanzar otro proceso. Se distinguen
# de ACTUALIZADORES porque no abren ventana ni tocan Excel: leen y escriben
# nomas, asi que no vale la pena un .py aparte.
#   {id del arbol: [(rotulo del boton, nombre del metodo), ...]}
ACCIONES_INTERNAS = {
    "a_0_cuadros": [("Exportar CPRT", "_exportar_cprt")],
}

# colores
C_OK        = "#1a7f1a"
C_FALTA     = "#c00000"
C_AMARILLO  = "#ffe600"
C_VENCIDA   = "#ff9900"
C_GRIS      = "#777777"
C_NEUTRO    = "SystemButtonFace"

DIR_SCRIPT = Path(__file__).resolve().parent
DIR_RAIZ = DIR_SCRIPT.parent
CONFIG_PATH = DIR_RAIZ / "__config__" / "config.json"
DIR_CONFIG = DIR_RAIZ / "__config__"
DIR_SALIDAS = DIR_RAIZ / "00_Salidas"
ARCHIVO_ESTADO = "_revisor_verificaciones.json"


def dir_mes(aamm, crear=False):
    """00_Salidas/AAAA/MM Mes, hermana de Revisor_Relq.

    La logica vive en __comun__/salidas.py porque los comparadores tienen que armar
    exactamente la misma ruta; si se separan, uno lee donde el otro no escribe.
    """
    return _sal.carpeta_mes(DIR_SALIDAS, aamm, crear=crear)


def dir_config_mes(aamm, crear=False):
    """__config__/AAAA/MM Mes para estado, cache y traspasos internos."""
    return _sal.carpeta_mes(DIR_CONFIG, aamm, crear=crear)


def escribir_json(ruta, data):
    """Escritura atomica: primero un .tmp y despues os.replace.
    Evita dejar el archivo truncado si algo falla a medio camino."""
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    tmp = ruta.with_suffix(ruta.suffix + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, ruta)


# =============================================================================
#  Utilidades
# =============================================================================

def get_usuario():
    usuario = os.environ.get("USERNAME") or os.environ.get("USER") or "desconocido"
    return f"{socket.gethostname()}_{usuario}"


def leer_config():
    try:
        if CONFIG_PATH.exists():
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                return json.load(f).get(get_usuario(), {})
    except Exception:
        pass
    return {}


def _modificar_config(mutador):
    """Lee config.json completo, lo modifica con `mutador` y lo reescribe.
    IMPORTANTE: config.json lo comparten otros scripts. Solo se agregan o
    actualizan claves, nunca se borra nada, y si el archivo existe pero no se
    puede interpretar NO se escribe (mejor perder un ajuste que el archivo)."""
    todo = {}
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                todo = json.load(f)
            if not isinstance(todo, dict):
                return False
        except Exception:
            return False
    try:
        mutador(todo)
        escribir_json(CONFIG_PATH, todo)
        return True
    except Exception:
        return False


def guardar_config(data):
    return _modificar_config(
        lambda todo: todo.setdefault(get_usuario(), {}).update(data))


def leer_valores_cfg():
    """La ubicacion de los valores es estructural, no por usuario: va en la
    clave compartida '_valores' de config.json."""
    try:
        if CONFIG_PATH.exists():
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                return json.load(f).get("_valores", {})
    except Exception:
        pass
    return {}


def guardar_valores_cfg(clave, campos):
    return _modificar_config(
        lambda todo: todo.setdefault("_valores", {}).setdefault(clave, {}).update(campos))


def aplicar_valores_cfg():
    guardado = leer_valores_cfg()
    permitidos = ("hoja", "celda", "tabla", "columna", "where", "columna_tipo",
                  "fila_inicio", "columna_filtro", "valores_filtro",
                  "columna_etiqueta", "texto_fila", "columna_valor")
    for clave, campos in guardado.items():
        if clave not in VALORES:
            continue
        for k, v in campos.items():
            if k not in permitidos:
                continue
            if k == "fila_inicio":
                try:
                    v = int(str(v).strip())
                except Exception:
                    continue
            elif k == "valores_filtro" and isinstance(v, str):
                v = [x.strip() for x in v.split(",") if x.strip()]
            VALORES[clave][k] = v


def abrir_en_explorador(ruta, es_archivo=False):
    if not ruta:
        return
    p = Path(ruta)
    if not p.exists():
        return
    carpeta = p.parent if es_archivo else p
    try:
        if sys.platform == "win32":
            subprocess.Popen(["explorer", str(carpeta)])
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(carpeta)])
        else:
            subprocess.Popen(["xdg-open", str(carpeta)])
    except Exception:
        pass


def normalizar(texto):
    nfkd = unicodedata.normalize("NFKD", str(texto))
    limpio = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", limpio).strip().lower()


# =============================================================================
#  Cache de directorios
# =============================================================================
# Una relectura completa hacia 68 recorridos de carpeta para 13 carpetas
# distintas: cada nodo del arbol recorria la carpeta entera de nuevo, y encima
# resolver_carpeta recorria la raiz una vez por nodo. En un disco local no se
# nota; en la T: cada recorrido es un viaje de red y ahi esta el tiempo.
#
# Este cache guarda el listado de cada carpeta MIENTRAS dura una relectura, asi
# cada carpeta se recorre UNA vez. Se usa os.scandir en vez de iterdir porque
# trae la fecha y el tamano en el mismo recorrido: con Path.iterdir + .stat()
# cada archivo cuesta un viaje aparte.
#
# El cache esta apagado por omision y solo se enciende dentro de
# "with cache_directorios():". Fuera de ahi todo lee del disco como siempre, que
# es lo que hace falta para que mtime() no devuelva datos viejos cuando se
# comprueba si una verificacion vencio.
_DIR_CACHE = {"on": False, "datos": {}, "hits": 0, "scans": 0}


def _escanear(carpeta):
    """(subcarpetas, archivos) de una carpeta, en UN solo recorrido.
    subcarpetas: {nombre_normalizado: Path}
    archivos   : {nombre: (Path, mtime, tamano)}
    """
    subs, archs = {}, {}
    _DIR_CACHE["scans"] += 1
    try:
        with os.scandir(str(carpeta)) as it:
            for e in it:
                try:
                    if e.is_dir():
                        subs[normalizar(e.name)] = Path(e.path)
                    elif e.is_file():
                        st = e.stat()
                        archs[e.name] = (Path(e.path), st.st_mtime, st.st_size)
                except OSError:
                    continue
    except OSError:
        pass
    return subs, archs


def leer_dir(carpeta):
    """El listado de una carpeta, del cache si esta encendido."""
    if carpeta is None:
        return {}, {}
    if _DIR_CACHE["on"]:
        clave = str(carpeta)
        if clave in _DIR_CACHE["datos"]:
            _DIR_CACHE["hits"] += 1
            return _DIR_CACHE["datos"][clave]
        datos = _escanear(carpeta)
        _DIR_CACHE["datos"][clave] = datos
        return datos
    return _escanear(carpeta)


class cache_directorios:
    """Enciende el cache mientras dura el bloque. Devuelve (scans, hits) al salir
    en self.stats, para poder decir en la bitacora cuanto se ahorro."""

    def __enter__(self):
        _DIR_CACHE.update(on=True, datos={}, hits=0, scans=0)
        return self

    def __exit__(self, *_):
        self.stats = (_DIR_CACHE["scans"], _DIR_CACHE["hits"])
        _DIR_CACHE.update(on=False, datos={}, hits=0, scans=0)
        return False


def buscar_carpeta(base, nombre):
    """Busca subcarpeta tolerando tildes, mayusculas y espacios extra."""
    if not base:
        return None
    objetivo = normalizar(nombre)
    mapa, _ = leer_dir(base)
    if objetivo in mapa:
        return mapa[objetivo]
    subs = list(mapa.values())
    cands = [d for d in subs if normalizar(d.name).startswith(objetivo)]
    if cands:
        return sorted(cands, key=lambda d: len(d.name))[0]
    cands = [d for d in subs if objetivo in normalizar(d.name)]
    if cands:
        return sorted(cands, key=lambda d: len(d.name))[0]
    return None


def resolver_carpeta(base, partes):
    """Cada parte puede ser un nombre o una tupla de nombres alternativos."""
    actual = base
    for p in partes:
        opciones = p if isinstance(p, (list, tuple)) else [p]
        siguiente = None
        for o in opciones:
            siguiente = buscar_carpeta(actual, o)
            if siguiente is not None:
                break
        if siguiente is None:
            return None
        actual = siguiente
    return actual


def es_temporal(nombre):
    return nombre.startswith("~$") or nombre.startswith(".")


# Sufijos que deja Windows al copiar: "archivo - copia.mdb",
# "archivo - copia (2).mdb", "archivo - Copy.xlsm".
RE_COPIA = re.compile(r"(-\s*cop(?:ia|y)(?:\s*\(\d+\))?|\(\d+\))\s*$")


def es_copia(nombre_sin_extension):
    return bool(RE_COPIA.search(normalizar(nombre_sin_extension)))


def buscar_archivo(carpeta, patron_regex, extensiones):
    """Devuelve el archivo mas reciente que calza el patron (sobre nombre normalizado)."""
    if not carpeta:
        return None
    patron = re.compile(patron_regex)
    _, archivos = leer_dir(carpeta)
    cands, fechas = [], {}
    for nombre, (f, mt, _sz) in archivos.items():
        if es_temporal(nombre):
            continue
        if f.suffix.lower() not in extensiones:
            continue
        if patron.search(normalizar(f.stem)):
            cands.append(f)
            fechas[f] = mt
    if not cands:
        return None
    # Si hay varios, se descartan las copias ("- copia", "(2)"...). Solo se usan
    # si no queda ninguna otra opcion.
    sin_copias = [f for f in cands if not es_copia(f.stem)]
    if sin_copias:
        cands = sin_copias
    # La fecha ya vino del recorrido de la carpeta: no hace falta un stat por
    # archivo, que en la T: es un viaje de red cada uno.
    cands.sort(key=lambda p: fechas.get(p) or 0, reverse=True)
    return cands[0]


def listar_diarios(carpeta, patron_regex, extensiones):
    """{fecha_AAAAMMDD: Path} para las planillas diarias de una carpeta."""
    out = {}
    if not carpeta:
        return out
    patron = re.compile(patron_regex)
    _, archivos = leer_dir(carpeta)
    for nombre, (f, _mt, _sz) in archivos.items():
        if es_temporal(nombre):
            continue
        if f.suffix.lower() not in extensiones:
            continue
        m = patron.search(normalizar(f.stem))
        if m:
            fecha = m.group(1)
            if es_copia(f.stem) and fecha in out:
                continue
            if fecha in out and es_copia(out[fecha].stem):
                out[fecha] = f
            else:
                out.setdefault(fecha, f)
    return out


def mtime(p):
    """Fecha de modificacion. Con el cache encendido sale del recorrido de la
    carpeta, sin un stat por archivo."""
    if p is None:
        return None
    if _DIR_CACHE["on"]:
        _, archivos = leer_dir(Path(p).parent)
        dato = archivos.get(Path(p).name)
        if dato is not None:
            return dato[1]
    try:
        return p.stat().st_mtime
    except Exception:
        return None


def tamano(p):
    if p is None:
        return None
    if _DIR_CACHE["on"]:
        _, archivos = leer_dir(Path(p).parent)
        dato = archivos.get(Path(p).name)
        if dato is not None:
            return dato[2]
    try:
        return p.stat().st_size
    except Exception:
        return None


def fmt_fecha(ts):
    if ts is None:
        return "—"
    return datetime.fromtimestamp(ts).strftime("%d-%m-%Y %H:%M:%S")


def iguales_mtime(a, b):
    if a is None or b is None:
        return False
    return abs(a - b) <= TOL_MTIME


def fmt_monto(v):
    if v is None:
        return "—"
    try:
        return f"{float(v):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    except Exception:
        return str(v)


# =============================================================================
#  Estado de verificaciones (se guarda dentro de la carpeta del caso)
# =============================================================================

def detectar_aamm(rutas, diarios):
    """Deduce el AAMM reliquidado. Primero del sufijo _AAMM_ de los nombres de
    archivo (se queda con el mas repetido); si no, de las fechas AAAAMMDD de
    los detalles diarios. Devuelve (aamm, motivo) o (None, motivo)."""
    votos = {}
    pat = re.compile(r"_(\d{4})_")
    for r in rutas.values():
        if not r or not hasattr(r, "stem"):
            continue
        for m in pat.finditer(r.stem):
            votos[m.group(1)] = votos.get(m.group(1), 0) + 1
    if votos:
        top = max(votos.items(), key=lambda kv: (kv[1], kv[0]))
        detalle = ", ".join(f"{k}×{v}" for k, v in sorted(votos.items(),
                                                          key=lambda kv: -kv[1]))
        return top[0], f"del nombre de los archivos ({detalle})"

    votos = {}
    for mapa in diarios.values():
        for fecha in mapa:                      # AAAAMMDD
            if len(fecha) == 8:
                k = fecha[2:6]                  # AAMM
                votos[k] = votos.get(k, 0) + 1
    if votos:
        top = max(votos.items(), key=lambda kv: kv[1])
        return top[0], "de las fechas de los detalles diarios"
    return None, "no se pudo deducir"


def firma_verificador(vid):
    """Huella corta y estable de COMO esta definida una verificacion.

    Existe para que un resultado guardado con una definicion vieja no se muestre
    nunca como si fuera del chequeo actual. Si se cambian las hojas, columnas,
    filas o comprobaciones de un verificador, la huella cambia y el registro
    anterior se descarta en vez de quedar mostrando datos de rangos que ya no se
    leen (por ejemplo columnas que ni existen en la definicion nueva).
    """
    import hashlib
    v = VERIFICADORES.get(vid) or {}
    # Solo lo que define QUE se lee y QUE se compara. El titulo no entra: cambiar
    # una palabra del rotulo no deberia invalidar un resultado bueno.
    relevante = {
        "archivo": v.get("archivo"),
        "depende": list(v.get("depende") or []),
        "previas": list(v.get("verif_previas") or []),
        "comprobaciones": v.get("comprobaciones") or [],
    }
    crudo = json.dumps(relevante, sort_keys=True, ensure_ascii=False,
                       default=str)
    return hashlib.sha1(crudo.encode("utf-8")).hexdigest()[:12]


class Estado:
    """Verificaciones de un mes. Se guardan en __config__/AAAA/MM Mes."""

    def __init__(self):
        self.ruta = None
        self.aamm = None
        self.data = {}

    def cargar(self, aamm):
        self.aamm = str(aamm).strip() if aamm else None
        self.data = {}
        self.ruta = None
        if not self.aamm:
            return False
        self.ruta = dir_config_mes(self.aamm) / ARCHIVO_ESTADO
        if self.ruta.exists():
            try:
                with open(self.ruta, "r", encoding="utf-8") as f:
                    self.data = json.load(f)
                return True
            except Exception:
                self.data = {}
        return False

    def existe(self):
        return bool(self.ruta and self.ruta.exists())

    def guardar(self):
        if not self.aamm:
            return False
        try:
            dir_config_mes(self.aamm, crear=True)
            escribir_json(dir_config_mes(self.aamm) / ARCHIVO_ESTADO, self.data)
            return True
        except Exception:
            return False

    def get(self, vid):
        return self.data.get(vid)

    def vigente(self, vid):
        """El registro guardado, si hay alguno.

        No se mira la firma a proposito. Un resultado verificado vale como
        verificado, sin importar con que version de la definicion se corrio: al
        usuario le da lo mismo esa distincion y solo le ensuciaba la fila. La
        diferencia de definicion se avisa en la ventana de detalle, que es donde
        importa, porque ahi si se muestran hojas y rangos concretos.
        """
        return self.data.get(vid) or None

    def firma_guardada_distinta(self, vid):
        """True si hay un registro pero es de una definicion anterior.
        Solo para avisarlo en el detalle, no para invalidar nada."""
        reg = self.data.get(vid)
        return bool(reg) and reg.get("firma") != firma_verificador(vid)

    def set(self, vid, registro):
        registro = dict(registro)
        registro["firma"] = firma_verificador(vid)
        self.data[vid] = registro
        return self.guardar()


ARCHIVO_CACHE = "_revisor_cache_valores.json"


def partir_signo(clave):
    """En las listas de comprobaciones una clave puede llevar '-' delante para
    entrar con signo cambiado.  '-P4_L6' -> ('P4_L6', -1.0)"""
    c = str(clave)
    if c.startswith("-"):
        return c[1:], -1.0
    return c, 1.0


def huella_spec(spec):
    """Firma de QUE se lee de un archivo. Si cambia la hoja, la celda, la
    columna o el filtro, la huella cambia y el cache no aplica."""
    t = spec.get("tipo")
    if t == "excel_col":
        return "|".join(["col", str(spec.get("hoja", "")), str(spec.get("columna", "")),
                         str(spec.get("fila_inicio", "")),
                         str(spec.get("columna_filtro", "")),
                         ",".join(spec.get("valores_filtro") or [])])
    if t == "excel":
        return "|".join(["celda", str(spec.get("hoja", "")), str(spec.get("celda", ""))])
    if t == "excel_etiqueta":
        return "|".join(["etq", str(spec.get("hoja", "")),
                         str(spec.get("columna_etiqueta", "")),
                         str(spec.get("texto_fila", "")),
                         str(spec.get("columna_valor", "")),
                         str(spec.get("fila_inicio", ""))])
    if t == "mdb":
        return "|".join(["mdb", str(spec.get("tabla", "")), str(spec.get("columna", "")),
                         str(spec.get("where", ""))])
    return str(t)


class CacheValores:
    """Guarda el valor ya leido de cada origen junto con la ruta y la fecha de
    modificacion del archivo. Si el archivo no cambio y se pide lo mismo, no se
    vuelve a abrir. Se guarda en __config__/AAAA/MM Mes entre ejecuciones."""

    def __init__(self):
        self.aamm = None
        self.data = {}

    def cargar(self, aamm):
        self.aamm = str(aamm).strip() if aamm else None
        self.data = {}
        if not self.aamm:
            return
        ruta = dir_config_mes(self.aamm) / ARCHIVO_CACHE
        if ruta.exists():
            try:
                with open(ruta, "r", encoding="utf-8") as f:
                    self.data = json.load(f)
            except Exception:
                self.data = {}

    def guardar(self):
        if not self.aamm:
            return False
        try:
            dir_config_mes(self.aamm, crear=True)
            escribir_json(dir_config_mes(self.aamm) / ARCHIVO_CACHE, self.data)
            return True
        except Exception:
            return False

    def obtener(self, clave, ruta, huella):
        reg = self.data.get(clave)
        if not reg or ruta is None:
            return None
        if reg.get("archivo") != ruta.name or reg.get("huella") != huella:
            return None
        # OJO: aca la comparacion es EXACTA, sin la tolerancia de TOL_MTIME.
        # Esa tolerancia sirve para comparar copias entre discos distintos, pero
        # si se usara aca un archivo guardado 1 segundo despues de leerlo
        # devolveria el valor viejo. Se compara ademas el tamaño.
        ts_ahora, tam_ahora = mtime(ruta), tamano(ruta)
        if ts_ahora is None or reg.get("mtime") is None:
            return None
        if abs(reg["mtime"] - ts_ahora) > 1e-6:
            return None
        if reg.get("tamano") is not None and reg["tamano"] != tam_ahora:
            return None
        if not isinstance(reg.get("valor"), (int, float)):
            return None
        return reg

    def poner(self, clave, ruta, huella, valor, filas=None):
        if ruta is None:
            return
        self.data[clave] = {
            "archivo": ruta.name,
            "ruta": str(ruta),
            "mtime": mtime(ruta),
            "mtime_texto": fmt_fecha(mtime(ruta)),
            "tamano": tamano(ruta),
            "huella": huella,
            "valor": valor,
            "filas": filas,
            "leido": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        self.guardar()

    def descartar(self, claves=None):
        if claves is None:
            self.data = {}
        else:
            for c in claves:
                self.data.pop(c, None)
        self.guardar()


CACHE = CacheValores()


def leer_estado_mes(aamm):
    """Lee el estado de cualquier mes sin tocar el estado en uso."""
    ruta = dir_config_mes(aamm) / ARCHIVO_ESTADO
    if not ruta.exists():
        return None
    try:
        with open(ruta, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


ESTADO = Estado()


# =============================================================================
#  Lectura de valores
# =============================================================================

def _suma_rango_openpyxl(ws, ref):
    total = 0.0
    hubo = False
    for fila in ws[ref]:
        celdas = fila if isinstance(fila, tuple) else (fila,)
        for c in celdas:
            if isinstance(c.value, (int, float)):
                total += float(c.value)
                hubo = True
    return total if hubo else None


CACHE_COLUMNAS = {}     # (ruta, mtime, hoja, col, col_filtro, fila) -> {clave: [suma, n]}


def col_letra(n):
    """4 -> "D".  Al reves de col_letra_a_num."""
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def col_letra_a_num(letra):
    n = 0
    for c in str(letra).upper():
        if "A" <= c <= "Z":
            n = n * 26 + (ord(c) - ord("A") + 1)
    return n


def _es_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def leer_columna_excel(ruta, hoja, columna, fila_inicio, col_filtro, valores_filtro, log):
    """Suma una columna completa desde fila_inicio hacia abajo.
    Si se indica col_filtro, suma solo las filas cuyo valor de esa columna esta
    en valores_filtro (comparacion sin tildes ni mayusculas).
    Devuelve (suma, n_filas, {valor_filtro: suma}) o (None, 0, {})."""
    ci = col_letra_a_num(columna)
    cf = col_letra_a_num(col_filtro) if col_filtro else 0
    objetivo = {normalizar(v) for v in (valores_filtro or [])}

    # Cache: la misma columna se lee una sola vez por corrida aunque varios
    # valores la usen con distinto filtro (p.ej. SCMT y SCPC).
    ck = (str(ruta), mtime(ruta), hoja, ci, cf, int(fila_inicio))

    def desde_desglose(bruto):
        """bruto: {clave: [suma, n]} -> (total, n, {clave: suma})"""
        if cf and objetivo:
            sel = [v for k, v in bruto.items() if k in objetivo]
        else:
            sel = list(bruto.values())
        total = sum(v[0] for v in sel)
        n = sum(v[1] for v in sel)
        return total, n, {k: v[0] for k, v in bruto.items()}

    if ck in CACHE_COLUMNAS:
        return desde_desglose(CACHE_COLUMNAS[ck])

    def acumular(pares):
        """pares: iterable de (monto, tipo)"""
        bruto = {}
        for monto, tipo in pares:
            if not _es_num(monto):
                continue
            clave = "(todo)" if not cf else normalizar(tipo)
            reg = bruto.setdefault(clave, [0.0, 0])
            reg[0] += float(monto)
            reg[1] += 1
        CACHE_COLUMNAS[ck] = bruto
        return desde_desglose(bruto)

    # --- openpyxl (valores cacheados) --------------------------------------
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(ruta), data_only=True, read_only=True)
        try:
            real = resolver_hoja(wb.sheetnames, hoja)
            if real is None:
                log(f"    ! La hoja '{hoja}' no existe. Hojas: {', '.join(wb.sheetnames)}")
                return None, 0, {}
            ws = wb[real]
            ancho = max(ci, cf)

            def gen():
                for fila in ws.iter_rows(min_row=fila_inicio, max_col=ancho,
                                         values_only=True):
                    m = fila[ci - 1] if len(fila) >= ci else None
                    t = fila[cf - 1] if cf and len(fila) >= cf else None
                    yield m, t

            total, n, desg = acumular(gen())
            if desg:
                return total, n, desg
        finally:
            wb.close()
        log("    · sin valores cacheados, reintentando con Excel...")
    except Exception as e:
        log(f"    · openpyxl no pudo leer ({e}); reintentando con Excel...")

    # --- xlwings ----------------------------------------------------------
    app = wb2 = None
    try:
        import xlwings as xw
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        wb2 = app.books.open(str(ruta), read_only=True, update_links=False)
        nombres = [s.name for s in wb2.sheets]
        real = resolver_hoja(nombres, hoja)
        if real is None:
            log(f"    ! La hoja '{hoja}' no existe. Hojas: {', '.join(nombres)}")
            return None, 0, {}
        sh = wb2.sheets[real]
        ult = sh.used_range.last_cell.row
        if ult < fila_inicio:
            return 0.0, 0, {}
        montos = sh.range((fila_inicio, ci), (ult, ci)).value
        if not isinstance(montos, list):
            montos = [montos]
        if cf:
            tipos = sh.range((fila_inicio, cf), (ult, cf)).value
            if not isinstance(tipos, list):
                tipos = [tipos]
        else:
            tipos = [None] * len(montos)
        return acumular(zip(montos, tipos))
    except Exception as e:
        log(f"    ! Error leyendo con Excel: {e}")
        return None, 0, {}
    finally:
        try:
            if wb2 is not None:
                wb2.close()
            if app is not None:
                app.quit()
        except Exception:
            pass


NS_XL = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
NS_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"


def es_zip_excel(ruta):
    return Path(ruta).suffix.lower() in (".xlsx", ".xlsm", ".xltx", ".xltm")


def ubicar_hoja_xml(z, hoja):
    """Dentro del zip de un .xlsx/.xlsm, devuelve (ruta_del_xml, lista_de_hojas).
    ruta_del_xml es None si la hoja no existe."""
    import xml.etree.ElementTree as ET
    nombres = set(z.namelist())
    if "xl/workbook.xml" not in nombres:
        return None, []
    wb = ET.fromstring(z.read("xl/workbook.xml"))
    lista = list(wb.iter(f"{NS_XL}sheet"))
    hojas = [sh.get("name", "") for sh in lista]
    rid = None
    # "#1", "#2", ... permiten pedir la hoja por posicion cuando no se sabe el
    # nombre o cuando cambia de un mes a otro.
    m_pos = re.fullmatch(r"#(\d+)", str(hoja).strip())
    if m_pos:
        i = int(m_pos.group(1)) - 1
        if 0 <= i < len(lista):
            rid = lista[i].get(f"{NS_REL}id")
    else:
        for sh in lista:
            if normalizar(sh.get("name", "")) == normalizar(hoja):
                rid = sh.get(f"{NS_REL}id")
                break
    if rid is None or "xl/_rels/workbook.xml.rels" not in nombres:
        return None, hojas
    destino = None
    for r in ET.fromstring(z.read("xl/_rels/workbook.xml.rels")):
        if r.get("Id") == rid:
            destino = r.get("Target")
            break
    if not destino:
        return None, hojas
    ruta_hoja = destino[1:] if destino.startswith("/") else "xl/" + destino
    ruta_hoja = ruta_hoja.replace("xl/xl/", "xl/")
    return (ruta_hoja if ruta_hoja in nombres else None), hojas


def expandir_columnas(rango):
    """'CF:CI' -> ['CF','CG','CH','CI'].  'CD' -> ['CD']."""
    partes = str(rango).replace(" ", "").replace("$", "").upper().split(":")
    a = col_letra_a_num(partes[0])
    b = col_letra_a_num(partes[-1]) if len(partes) > 1 else a
    letras = []
    for n in range(min(a, b), max(a, b) + 1):
        s, x = "", n
        while x > 0:
            x, r = divmod(x - 1, 26)
            s = chr(ord("A") + r) + s
        letras.append(s)
    return letras


def buscar_marcas_rapido(ruta, hoja, fila_inicio, reglas, log, tope_detalle=30):
    """Busca errores de fórmula y textos prohibidos en columnas puntuales,
    escaneando el XML de la hoja por trozos en vez de cargar el libro.

    reglas: [{"rangos": ["CF:CI"], "errores": True, "textos": ["REVISAR"]}, ...]
    Devuelve {"conteo": {motivo: n}, "marcas": [(celda, motivo, valor)]} o None."""
    import zipfile
    import xml.etree.ElementTree as ET

    if not es_zip_excel(ruta):
        return None

    # columna -> indice de regla
    de_columna = {}
    for i, regla in enumerate(reglas or []):
        for r in regla.get("rangos", []):
            for letra in expandir_columnas(r):
                de_columna[letra.encode()] = i
    if not de_columna:
        return {"conteo": {}, "marcas": []}

    alternativas = b"|".join(sorted(de_columna, key=len, reverse=True))
    patron = re.compile(rb'<c r="(' + alternativas + rb')(\d+)"([^>]*?)(?:/>|>(.*?)</c>)',
                        re.S)
    fila_inicio = int(fila_inicio)
    conteo, marcas = {}, []

    def anotar(celda, motivo, valor):
        conteo[motivo] = conteo.get(motivo, 0) + 1
        if len(marcas) < tope_detalle:
            marcas.append((celda, motivo, str(valor)[:40]))

    try:
        with zipfile.ZipFile(str(ruta)) as z:
            ruta_hoja, hojas = ubicar_hoja_xml(z, hoja)
            if ruta_hoja is None:
                log(f"    ! La hoja '{hoja}' no existe. Hojas: {', '.join(hojas)}")
                return None

            # cadenas compartidas: se resuelven una vez y solo si hacen falta
            compartidas, malos = [], {}
            if any(regla.get("textos") for regla in reglas):
                if "xl/sharedStrings.xml" in set(z.namelist()):
                    ss = ET.fromstring(z.read("xl/sharedStrings.xml"))
                    compartidas = ["".join(si.itertext()) for si in ss.iter(f"{NS_XL}si")]
                # Una misma cadena puede estar prohibida por varias reglas, asi
                # que se guarda POR REGLA: {indice: {n_regla: texto}}
                for i, texto in enumerate(compartidas):
                    tn = normalizar(texto)
                    for idx, regla in enumerate(reglas):
                        for t in regla.get("textos", []):
                            if normalizar(t) in tn:
                                malos.setdefault(str(i).encode(), {})[idx] = t
                                break

            with z.open(ruta_hoja) as f:
                cola = b""
                while True:
                    trozo = f.read(1 << 20)
                    if not trozo:
                        break
                    buf = cola + trozo
                    fin = 0
                    for m in patron.finditer(buf):
                        fin = m.end()
                        col, nfila, attrs, cuerpo = m.group(1), m.group(2), m.group(3), m.group(4) or b""
                        if int(nfila) < fila_inicio:
                            continue
                        regla = reglas[de_columna[col]]
                        celda = (col + nfila).decode()
                        if b't="e"' in attrs:
                            if regla.get("errores"):
                                mv = re.search(rb"<v>([^<]*)</v>", cuerpo)
                                anotar(celda, "valor de error",
                                       mv.group(1).decode() if mv else "#?")
                            continue
                        if not regla.get("textos"):
                            continue
                        if b't="s"' in attrs:
                            mv = re.search(rb"<v>([^<]*)</v>", cuerpo)
                            porregla = malos.get(mv.group(1)) if mv else None
                            if porregla:
                                t = porregla.get(de_columna[col])
                                if t:
                                    anotar(celda, f"texto '{t}'",
                                           compartidas[int(mv.group(1))])
                        elif b'inlineStr' in attrs or b't="str"' in attrs:
                            # OJO: hay que leer SOLO el resultado, nunca el nodo
                            # <f> de la formula. Muchas formulas llevan la
                            # palabra buscada adentro (IFERROR(...,"REVISAR"))
                            # y mirar la formula marcaba toda la columna.
                            if b'inlineStr' in attrs:
                                mv = re.search(rb"<is>(.*?)</is>", cuerpo, re.S)
                                bruto = re.sub(rb"<[^>]+>", b"", mv.group(1)) if mv else b""
                            else:
                                mv = re.search(rb"<v>(.*?)</v>", cuerpo, re.S)
                                bruto = mv.group(1) if mv else b""
                            if not bruto:
                                continue
                            texto = bruto.decode("utf-8", "ignore")
                            texto = (texto.replace("&amp;", "&").replace("&lt;", "<")
                                     .replace("&gt;", ">").replace("&quot;", '"')
                                     .replace("&apos;", "'"))
                            tn = normalizar(texto)
                            for t in regla["textos"]:
                                if normalizar(t) in tn:
                                    anotar(celda, f"texto '{t}'", texto.strip())
                                    break
                    cola = buf[fin:] if fin else buf[-8192:]
    except Exception as e:
        log(f"    ! No se pudo escanear {Path(ruta).name}: {e}")
        return None
    return {"conteo": conteo, "marcas": marcas}


def leer_celdas_rapido(ruta, hoja, celdas):
    """Lee celdas puntuales de un .xlsx/.xlsm sin cargar el libro completo.

    Un .xlsm es un ZIP con XML adentro. Se recorre el XML de la hoja en
    streaming y se corta en cuanto se pasa de la ultima fila pedida, asi que
    para celdas de las primeras filas (H1, EE6) casi no se lee nada, aunque el
    archivo tenga miles de filas y millones de formulas.

    Devuelve {celda: valor} con los valores YA CALCULADOS que Excel dejo
    guardados. Si el archivo nunca fue calculado y guardado, no habra valores.
    Devuelve None si no se pudo (formato distinto, hoja inexistente, etc.)."""
    import zipfile
    import xml.etree.ElementTree as ET

    if not es_zip_excel(ruta):
        return None

    pedidas = {c.upper().replace("$", "") for c in celdas}
    if not pedidas:
        return {}
    try:
        fila_tope = max(int(re.sub(r"[^0-9]", "", c)) for c in pedidas)
    except ValueError:
        return None

    NS = NS_XL
    try:
        with zipfile.ZipFile(str(ruta)) as z:
            nombres = set(z.namelist())
            ruta_hoja, _ = ubicar_hoja_xml(z, hoja)
            if ruta_hoja is None:
                return None

            # 2) recorrer la hoja y cortar al pasar la ultima fila pedida
            compartidas = None
            out = {}
            with z.open(ruta_hoja) as f:
                for evento, el in ET.iterparse(f, events=("end",)):
                    if el.tag == f"{NS}row":
                        try:
                            if int(el.get("r", 0)) > fila_tope:
                                el.clear()
                                break
                        except ValueError:
                            pass
                        el.clear()
                        continue
                    if el.tag != f"{NS}c":
                        continue
                    ref = (el.get("r") or "").upper()
                    if ref not in pedidas:
                        el.clear()
                        continue
                    t = el.get("t")
                    nodo_v = el.find(f"{NS}v")
                    valor = None
                    if t == "inlineStr":
                        nodo_is = el.find(f"{NS}is")
                        if nodo_is is not None:
                            valor = "".join(nodo_is.itertext())
                    elif nodo_v is not None and nodo_v.text is not None:
                        bruto = nodo_v.text
                        if t == "s":                      # texto compartido
                            if compartidas is None:
                                compartidas = []
                                if "xl/sharedStrings.xml" in nombres:
                                    ss = ET.fromstring(z.read("xl/sharedStrings.xml"))
                                    compartidas = ["".join(si.itertext())
                                                   for si in ss.iter(f"{NS}si")]
                            try:
                                valor = compartidas[int(bruto)]
                            except Exception:
                                valor = None
                        elif t == "e":                    # error de formula
                            valor = bruto
                        elif t == "b":
                            valor = bool(int(bruto))
                        elif t == "str":
                            valor = bruto
                        else:                             # numero
                            try:
                                valor = float(bruto)
                            except ValueError:
                                valor = bruto
                    out[ref] = valor
                    el.clear()
                    if len(out) == len(pedidas):
                        break
            return out
    except Exception:
        return None


def diagnosticar_celda(ruta, hoja, celda, log):
    """Cuando una celda no entrega valor, explica POR QUE mirando el XML crudo:
    si esta vacia, si tiene formula sin resultado guardado, si trae texto, si es
    un error, o si es parte de una celda combinada."""
    import zipfile
    import xml.etree.ElementTree as ET

    if not es_zip_excel(ruta):
        log(f"      (no se puede diagnosticar un {Path(ruta).suffix})")
        return
    NS = NS_XL
    ref = celda.upper().replace("$", "")
    try:
        fila_tope = int(re.sub(r"[^0-9]", "", ref))
    except ValueError:
        return
    try:
        with zipfile.ZipFile(str(ruta)) as z:
            ruta_hoja, hojas = ubicar_hoja_xml(z, hoja)
            if ruta_hoja is None:
                log(f"      la hoja '{hoja}' no existe. Hojas: {', '.join(hojas)}")
                return

            encontrada, tiene_formula, formula, tipo, valor = False, False, None, None, None
            with z.open(ruta_hoja) as f:
                for _, el in ET.iterparse(f, events=("end",)):
                    if el.tag == f"{NS}row":
                        try:
                            if int(el.get("r", 0)) > fila_tope:
                                break
                        except ValueError:
                            pass
                        el.clear()
                        continue
                    if el.tag == f"{NS}c" and (el.get("r") or "").upper() == ref:
                        encontrada = True
                        tipo = el.get("t")
                        nf = el.find(f"{NS}f")
                        nv = el.find(f"{NS}v")
                        tiene_formula = nf is not None
                        formula = (nf.text or "")[:70] if nf is not None else None
                        valor = nv.text if nv is not None else None
                        break

            if not encontrada:
                log(f"      {hoja}!{ref} está vacía (la celda no existe en el archivo).")
                # ¿es parte de una celda combinada?
                try:
                    xml = z.read(ruta_hoja).decode("utf-8", "ignore")
                    i = xml.find("<mergeCells")
                    if i > -1:
                        for mr in re.findall(r'ref="([A-Z]+\d+:[A-Z]+\d+)"',
                                             xml[i:xml.find("</mergeCells>", i)]):
                            a, b = mr.split(":")
                            ca, fa = re.match(r"([A-Z]+)(\d+)", a).groups()
                            cb, fb = re.match(r"([A-Z]+)(\d+)", b).groups()
                            cc, fc = re.match(r"([A-Z]+)(\d+)", ref).groups()
                            if (col_letra_a_num(ca) <= col_letra_a_num(cc) <= col_letra_a_num(cb)
                                    and int(fa) <= int(fc) <= int(fb)):
                                log(f"      OJO: está dentro de la celda combinada {mr}; "
                                    f"el valor vive en {a}.")
                                break
                except Exception:
                    pass
                log("      Revisa la hoja y la celda en «Configurar valores...».")
                return

            if tiene_formula and valor is None:
                log(f"      {hoja}!{ref} tiene fórmula pero sin resultado guardado.")
                log(f"      fórmula: ={formula}")
                log("      Pasa cuando el archivo se guardó con cálculo en manual o lo "
                    "guardó otro programa.")
                log("      Solución: ábrelo en Excel, presiona F9 y guárdalo.")
            elif valor is None:
                log(f"      {hoja}!{ref} existe pero no tiene valor guardado.")
            elif tipo == "e":
                log(f"      {hoja}!{ref} tiene un error de fórmula: {valor}")
            else:
                log(f"      {hoja}!{ref} no es un número, es texto: {str(valor)[:50]!r}")
    except Exception as e:
        log(f"      (no se pudo diagnosticar: {e})")


def resolver_hoja(nombres, hoja):
    """Traduce lo pedido al nombre real de la hoja: acepta '#1' (por posicion) y
    tolera tildes y mayusculas. Devuelve None si no calza ninguna."""
    m = re.fullmatch(r"#(\d+)", str(hoja).strip())
    if m:
        i = int(m.group(1)) - 1
        return nombres[i] if 0 <= i < len(nombres) else None
    for n in nombres:
        if n == hoja:
            return n
    for n in nombres:
        if normalizar(n) == normalizar(hoja):
            return n
    return None


def es_significativo(v):
    """Sirve para decidir si una celda 'cuenta' al buscar el ultimo dato de una
    columna: se omiten vacios, ceros y errores de formula (#REF!, #N/D...)."""
    if v is None:
        return False
    if isinstance(v, str):
        t = v.strip()
        return bool(t) and not t.startswith("#")
    if isinstance(v, (int, float)):
        return abs(float(v)) > 0
    return True


def leer_columnas_rapido(ruta, hoja, columnas, fila_inicio, log):
    """Lee columnas completas de un .xlsx/.xlsm escaneando el XML por trozos.
    Devuelve {"COL": {fila: valor}} con los valores ya calculados, o None.
    Igual que en el resto, se lee SOLO el resultado y nunca el nodo <f>."""
    import zipfile
    import xml.etree.ElementTree as ET

    if not es_zip_excel(ruta):
        log(f"    ! {Path(ruta).name}: formato no soportado para lectura rápida")
        return None
    objetivo = [c.upper() for c in columnas]
    if not objetivo:
        return {}
    alternativas = b"|".join(sorted((c.encode() for c in objetivo),
                                    key=len, reverse=True))
    patron = re.compile(rb'<c r="(' + alternativas + rb')(\d+)"([^>]*?)(?:/>|>(.*?)</c>)',
                        re.S)
    fila_inicio = int(fila_inicio)
    datos = {c: {} for c in objetivo}

    def desescapar(b):
        # Delegado al desescapador de verdad: maneja tambien &#243; y &#x00F3;.
        return desescapar_xml(b)

    try:
        with zipfile.ZipFile(str(ruta)) as z:
            ruta_hoja, hojas = ubicar_hoja_xml(z, hoja)
            if ruta_hoja is None:
                log(f"    ! La hoja '{hoja}' no existe. Hojas: {', '.join(hojas)}")
                return None
            compartidas = []
            if "xl/sharedStrings.xml" in set(z.namelist()):
                ss = ET.fromstring(z.read("xl/sharedStrings.xml"))
                compartidas = ["".join(si.itertext()) for si in ss.iter(f"{NS_XL}si")]

            with z.open(ruta_hoja) as f:
                cola = b""
                while True:
                    trozo = f.read(1 << 20)
                    if not trozo:
                        break
                    buf = cola + trozo
                    fin = 0
                    for m in patron.finditer(buf):
                        fin = m.end()
                        col, nfila = m.group(1).decode(), int(m.group(2))
                        if nfila < fila_inicio:
                            continue
                        attrs, cuerpo = m.group(3), m.group(4) or b""
                        valor = None
                        if b'inlineStr' in attrs:
                            mv = re.search(rb"<is>(.*?)</is>", cuerpo, re.S)
                            if mv:
                                valor = desescapar(re.sub(rb"<[^>]+>", b"", mv.group(1)))
                        else:
                            mv = re.search(rb"<v>(.*?)</v>", cuerpo, re.S)
                            if mv:
                                bruto = mv.group(1)
                                if b't="s"' in attrs:
                                    try:
                                        valor = compartidas[int(bruto)]
                                    except Exception:
                                        valor = None
                                elif b't="e"' in attrs or b't="str"' in attrs:
                                    valor = desescapar(bruto)
                                elif b't="b"' in attrs:
                                    valor = bool(int(bruto))
                                else:
                                    try:
                                        valor = float(bruto)
                                    except ValueError:
                                        valor = desescapar(bruto)
                        if valor is not None and not (isinstance(valor, str)
                                                      and not valor.strip()):
                            datos[col][nfila] = valor
                    cola = buf[fin:] if fin else buf[-8192:]
    except Exception as e:
        log(f"    ! No se pudo leer {Path(ruta).name}: {e}")
        return None
    return datos


_ENT_XML = {"lt": "<", "gt": ">", "quot": '"', "apos": "'", "amp": "&"}
_RE_ENT = re.compile(r"&(?:#(\d+)|#[xX]([0-9a-fA-F]+)|(lt|gt|quot|apos|amp));")


def desescapar_xml(b):
    """Convierte el texto crudo del XML de Excel a texto de verdad.

    Hay que manejar las referencias NUMERICAS (&#243; = o con tilde), no solo las
    cinco entidades con nombre: los nombres de empresa chilenos vienen llenos de
    tildes y ñ, y algunos escritores de Excel las guardan asi. Si no se
    desescapan, "Enel Generaci&#243;n" y "Enel Generación" no se parecen en nada
    al comparar, y el cuadro de pago reporta un descuadre que no existe.

    Se resuelve en UNA pasada a proposito. Reemplazar "&amp;" primero y despues
    "&lt;" convertiria "&amp;lt;" (un literal "&lt;") en "<", que es otra cosa.
    """
    if isinstance(b, bytes):
        b = b.decode("utf-8", "ignore")

    def uno(m):
        dec, hexa, nombre = m.group(1), m.group(2), m.group(3)
        try:
            if dec is not None:
                return chr(int(dec))
            if hexa is not None:
                return chr(int(hexa, 16))
        except (ValueError, OverflowError):
            return m.group(0)
        return _ENT_XML[nombre]

    return _RE_ENT.sub(uno, b)


def leer_formulas_rapido(ruta, hoja, columnas, fila_inicio, log):
    """Devuelve {"COL": set(filas_que_TIENEN_formula)}.

    A diferencia de leer_columnas_rapido, que lee el resultado, esto detecta la
    PRESENCIA del nodo <f>, o sea si la celda es una formula o un valor escrito
    a mano. Sirve para saber hasta donde se arrastro una formula.

    Ojo con las formulas compartidas: la primera celda trae la formula completa
    (<f t="shared" ref="L5:L120" si="0">...) y las siguientes solo <f t="shared"
    si="0"/> sin texto. Como aca solo importa que exista un <f>, las dos formas
    cuentan igual.
    """
    import zipfile

    if not es_zip_excel(ruta):
        log(f"    ! {Path(ruta).name}: formato no soportado para leer fórmulas")
        return None
    objetivo = [c.upper() for c in columnas]
    if not objetivo:
        return {}
    alternativas = b"|".join(sorted((c.encode() for c in objetivo),
                                    key=len, reverse=True))
    patron = re.compile(rb'<c r="(' + alternativas + rb')(\d+)"([^>]*?)(?:/>|>(.*?)</c>)',
                        re.S)
    fila_inicio = int(fila_inicio)
    formulas = {c: set() for c in objetivo}
    try:
        with zipfile.ZipFile(str(ruta)) as z:
            ruta_hoja, hojas = ubicar_hoja_xml(z, hoja)
            if ruta_hoja is None:
                log(f"    ! La hoja '{hoja}' no existe. Hojas: {', '.join(hojas)}")
                return None
            with z.open(ruta_hoja) as f:
                cola = b""
                while True:
                    trozo = f.read(1 << 20)
                    if not trozo:
                        break
                    buf = cola + trozo
                    fin = 0
                    for m in patron.finditer(buf):
                        fin = m.end()
                        col, nfila = m.group(1).decode(), int(m.group(2))
                        if nfila < fila_inicio:
                            continue
                        cuerpo = m.group(4) or b""
                        if b"<f" in cuerpo:
                            formulas[col].add(nfila)
                    cola = buf[fin:] if fin else buf[-8192:]
    except Exception as e:
        log(f"    ! No se pudieron leer las fórmulas de {Path(ruta).name}: {e}")
        return None
    return formulas


def armar_tabla(datos, col_clave, cols_valor, log, etiqueta="",
                excluir=("0",), info=None):
    """Convierte {"A":{fila:val}} en {clave_normalizada: (nombre, [valores])}.
    Se salta las filas sin clave, que es lo que aparece como vacios al final.

    excluir: claves normalizadas que NO son empresas y hay que descartar. Por
             defecto el "0", porque ninguna empresa se llama asi y se cuela
             cuando sobran formulas arrastrando ceros.
    info:    dict opcional que se rellena con {"duplicadas", "excluidas",
             "vacias"} para que quien llame decida si eso es un fallo.
    """
    tabla, vacias, duplicadas, excluidas = {}, 0, [], []
    fuera = {normalizar(e) for e in (excluir or ())}
    filas = sorted(datos.get(col_clave, {}))
    for f in filas:
        nombre = datos[col_clave].get(f)
        # Un 0 numerico tampoco es empresa: se descarta igual que el "0" texto.
        if isinstance(nombre, (int, float)) and not isinstance(nombre, bool):
            excluidas.append((f, nombre))
            continue
        if not isinstance(nombre, str) or not nombre.strip():
            vacias += 1
            continue
        clave = normalizar(nombre)
        if clave in fuera:
            excluidas.append((f, nombre.strip()))
            continue
        vals = []
        for c in cols_valor:
            v = datos.get(c, {}).get(f)
            vals.append(float(v) if isinstance(v, (int, float)) else v)
        if clave in tabla:
            duplicadas.append((nombre.strip(), f))
        tabla[clave] = (nombre.strip(), vals)
    if info is not None:
        info.update(duplicadas=duplicadas, excluidas=excluidas, vacias=vacias)
    partes = [f"{len(tabla)} empresa(s)"]
    if vacias:
        partes.append(f"{vacias} fila(s) sin empresa omitidas")
    if excluidas:
        partes.append(f"{len(excluidas)} fila(s) con 0 o vacío descartadas")
    log(f"        {etiqueta}: " + ", ".join(partes))
    if duplicadas:
        log(f"        {etiqueta}: EMPRESAS REPETIDAS -> "
            + ", ".join(f"{n} (fila {f})" for n, f in duplicadas[:8]))
        if len(duplicadas) > 8:
            log(f"        {etiqueta}: ... y {len(duplicadas) - 8} repetición(es) más")
    return tabla


# =============================================================================
#  Coloreado del log y del detalle
# =============================================================================
# Los mensajes ya vienen rotulados: ">>" es fallo, "OK" es bien, "?" es sin
# datos y ".." es "trabajando". Pero las lineas de detalle van indentadas y sin
# rotulo, asi que el color se HEREDA del bloque al que pertenecen: todo lo que
# cuelga de un ">>" se pinta igual que el ">>".
C_LOG_MALO = "#c00000"      # rojo: algo no cuadra
C_LOG_DUDA = "#b45309"      # ambar: aviso, o no se pudo comprobar
C_LOG_BIEN = "#1d6b1d"      # verde: cuadra

# Palabras que pintan la linea sin importar el bloque en que caiga.
_PALABRAS_MALO = (
    "FALTA ", "FALTA EL ARCHIVO", "NO CUADRA", "NO SE LANZÓ", "NO SE LANZO",
    "ERROR", "REPETIDA", "VENCIDA", "SOBRAN", "FALTAN", "descuadrada",
    "no encontrada:", "sin ninguna fórmula",
)
# Ojo: aca NO van "omitidas" ni "descartadas". Aparecen en toda corrida sana
# (siempre hay una fila vacia al final de una tabla), y pintar de ambar algo que
# sale siempre le ensena al usuario a ignorar el ambar.
_PALABRAS_DUDA = (
    "OJO", "AMARILLO", "ADVERTENCIA", "SIN DATOS", "desactualizado",
    "no se pudo", "No se pudo", "ya no está vigente",
    "Problemas al leer", "sin verificar",
)


def configurar_tags_log(widget):
    """Deja el widget listo para recibir lineas con color."""
    widget.tag_config("malo", foreground=C_LOG_MALO)
    widget.tag_config("duda", foreground=C_LOG_DUDA)
    widget.tag_config("bien", foreground=C_LOG_BIEN)


def clasificar_linea(linea, bloque=None):
    """(tag, bloque_nuevo) para una linea del log.

    bloque recuerda si venimos de un ">>" o de un "OK", para que las lineas
    indentadas que siguen hereden el color. Devolver el bloque permite pintar
    de a una linea (log en vivo) o una lista entera (ventana de detalle).
    """
    txt = str(linea)
    pelado = txt.strip()
    if not pelado:
        return None, bloque

    # Separadores y encabezados cierran el bloque anterior.
    if set(pelado) <= set("=-_") and len(pelado) > 3:
        return None, None

    if pelado.startswith(">>"):
        return "malo", "malo"
    if pelado.startswith("OK"):
        return "bien", "bien"
    if pelado.startswith("?"):
        return "duda", "duda"
    if pelado.startswith(".."):
        # "trabajando en esto": todavia no se sabe, no cambia el bloque.
        return None, bloque

    indentada = txt[:1].isspace()

    # "Resumen: 3 archivo(s) faltante(s)" tiene que decidir por si misma y no
    # heredar el color del bloque anterior, que puede ser cualquiera.
    m_res = re.search(r"Resumen:\s*(\d+)\s*archivo", txt)
    if m_res:
        return ("malo" if int(m_res.group(1)) else None), None

    # Una palabra fuerte manda por sobre la herencia.
    for p in _PALABRAS_MALO:
        if p in txt:
            return "malo", (bloque if indentada else "malo")
    for p in _PALABRAS_DUDA:
        if p in txt:
            return "duda", (bloque if indentada else "duda")

    if indentada and bloque in ("malo", "duda"):
        # Detalle de un bloque que fallo: se pinta igual que el bloque.
        return bloque, bloque
    if indentada:
        # Detalle de un bloque bueno: sin color, para que el rojo resalte.
        return None, bloque
    return None, None


def insertar_con_color(widget, lineas, bloque=None):
    """Inserta lineas ya coloreadas. Devuelve el bloque en que quedo."""
    for l in lineas:
        tag, bloque = clasificar_linea(l, bloque)
        widget.insert("end", str(l) + "\n", (tag,) if tag else ())
    return bloque


# =============================================================================
#  Exportar el CPRT a csv
# =============================================================================
# El cuadro cero trae en su hoja "CPRT" el cuadro de pago ya armado. Hay que
# sacarlo tal cual a un .csv que se manda a pago.
#
# Detalles que importan y no son obvios:
#  - La codificacion es cp1252, NO utf-8. Y la hoja trae "Fecha Generaci„n:" con
#    el caracter roto de fabrica; se copia ASI COMO ESTA a proposito, sin
#    arreglarlo, porque el archivo que espera el destinatario lo trae asi.
#  - Fin de linea CRLF, y todas las lineas con el mismo ancho (las de cabecera
#    se rellenan con comas vacias).
#  - Debajo de los datos hay filas con formulas que arrastran 0. Se cortan por
#    la columna A: si A no dice "Fila", la fila no va.
HOJA_CPRT = "CPRT"
CPRT_FILAS_META = 5          # las 5 lineas de cabecera (Coordinador, Concepto, ...)
CPRT_FILA_ENCABEZADO = 6     # la fila de nombres de columna
# El csv lleva 7 campos: A:F mas UNA columna de monto. Esa columna es la H
# ("Monto retenido"), NO la G ("Monto"):
#     H = G puesta en 0 cuando el acreedor esta en la lista de retenciones
#         =SI(CONTAR.SI('Retenciones Empresas FMCP'!I:I;E7)<>0;0;1)*G7
# O sea que el csv paga el monto NETO DE RETENCIONES, igual que hace la macro
# CPRT_csv del libro (su columna M concatena hasta la H).
CPRT_N_CAMPOS = 7
CPRT_COL_MONTO = 8           # 1-based: 8 = H. Poner 7 para volver a la G.
# El ROTULO de esa columna se sigue tomando de la G, o sea que el encabezado dice
# "Monto" y no "Monto retenido". Es a proposito: el unico csv que sabemos que el
# destinatario acepto (CPRT_2312_R01D.csv) dice "Monto", y si el sistema que lo
# lee mira el nombre de la columna, cambiarlo lo rompe. Los DATOS si salen de la
# H. Poner False para que el rotulo tambien salga de la H.
CPRT_ROTULO_DESDE_G = True
CPRT_CODIF = "cp1252"


def _celda_csv(v):
    """Convierte un valor de celda al texto que va en el csv."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        # Los montos van enteros y sin separador de miles.
        return str(int(round(v))) if abs(v - round(v)) < 1e-9 else repr(v)
    if isinstance(v, datetime):
        return f"{v.month}/{v.day}/{v.year} {v.hour}:{v.minute:02d}"
    return str(v).strip()


def exportar_cprt(ruta_xlsm, ruta_csv, log):
    """Escribe el csv del CPRT. Devuelve (n_filas_datos, n_con_retencion)."""
    import openpyxl
    wb = openpyxl.load_workbook(str(ruta_xlsm), data_only=True, read_only=True)
    try:
        hoja = next((n for n in wb.sheetnames
                     if normalizar(n) == normalizar(HOJA_CPRT)), None)
        if hoja is None:
            raise RuntimeError(
                f"No está la hoja '{HOJA_CPRT}'. Hojas: {', '.join(wb.sheetnames)}")
        sh = wb[hoja]
        im = CPRT_COL_MONTO - 1           # indice 0-based de la columna de monto
        salida, n_datos, n_retenidas, retenido = [], 0, 0, 0.0
        raras = []
        for i, f in enumerate(sh.iter_rows(min_row=1, max_row=sh.max_row,
                                           min_col=1, max_col=8,
                                           values_only=True), 1):
            # Las 6 primeras son A:F; la septima es la columna de monto elegida.
            fila = list(f[:CPRT_N_CAMPOS - 1]) + [f[im]]
            if i <= CPRT_FILAS_META or i == CPRT_FILA_ENCABEZADO:
                if i == CPRT_FILA_ENCABEZADO and CPRT_ROTULO_DESDE_G:
                    fila[-1] = f[6]       # el rotulo de la G ("Monto")
                salida.append([_celda_csv(v) for v in fila])
                continue
            if f[0] is None or not str(f[0]).strip():
                continue          # cola de formulas arrastrando 0
            g, h = f[6], f[7]
            if isinstance(g, (int, float)) and isinstance(h, (int, float)):
                if abs(h - g) > TOLERANCIA:
                    if abs(h) <= TOLERANCIA:
                        n_retenidas += 1
                        retenido += float(g)
                    else:
                        # H no es G ni 0: la formula de la H esta mal.
                        raras.append((i, g, h))
            salida.append([_celda_csv(v) for v in fila])
            n_datos += 1
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if not n_datos:
        raise RuntimeError("La hoja CPRT no tiene filas de datos: no se escribe nada.")

    tmp = Path(ruta_csv).with_suffix(".csv.tmp")
    with open(tmp, "w", newline="", encoding=CPRT_CODIF, errors="replace") as fh:
        csv.writer(fh, delimiter=",", lineterminator="\r\n",
                   quoting=csv.QUOTE_MINIMAL).writerows(salida)
    os.replace(tmp, ruta_csv)

    letra = col_letra(CPRT_COL_MONTO)
    log(f"        {n_datos} fila(s) de datos, {len(salida)} línea(s) en total")
    log(f"        {CPRT_N_CAMPOS} campos (A:F + {letra}), "
        f"codificación {CPRT_CODIF}, fin de línea CRLF")
    if CPRT_ROTULO_DESDE_G and CPRT_COL_MONTO != 7:
        log(f"        los datos salen de la {letra}, pero el rótulo sigue diciendo "
            f'"{_celda_csv(salida[CPRT_FILA_ENCABEZADO - 1][-1])}"')
    if n_retenidas:
        log(f"        OJO: {n_retenidas} fila(s) van en 0 por retención, o sea "
            f"{fmt_monto(retenido)} que NO se paga.")
        log(f"        El csv lleva la columna {letra} (Monto retenido), así que eso "
            f"es a propósito. Confírmalo antes de mandarlo.")
    if raras:
        log(f"        >> {len(raras)} fila(s) con H que no es G ni 0: la fórmula "
            f"de la H está mal.")
        for i, g, h in raras[:10]:
            log(f"             fila {i}: G={fmt_monto(g)}  H={fmt_monto(h)}")
    return n_datos, n_retenidas


# =============================================================================
#  El cuadro de pagos: la matriz y la tabla dinamica
# =============================================================================
# CuadroPago arma en 01.SSCC_Recurso_Tecnico una matriz cruzada:
#     N8  = "Pagan"            O8:..8 = las empresas que RECIBEN
#     N9:N.. = las que PAGAN   el interior = el reparto proporcional
# Despues Actualiza_rango_1 apunta el nombre CPTEE a esa matriz, y de ahi la
# Power Query alimenta la dinamica de CPRT, que es de donde sale el csv de pago.
#
# La cadena tiene tres puntos donde se puede quedar desincronizada, y el usuario
# no se entera porque cada pieza sigue mostrando numeros del paso anterior:
#   1. se actualizaron los datos pero no se corrio CuadroPago  -> matriz vieja
#   2. se corrio CuadroPago pero no Actualiza Rango            -> CPTEE viejo
#   3. no se refresco la dinamica                              -> CPRT viejo
UMBRAL_PAR_SEGURO = 100.0     # ver comentario en comparar_cprt_con_matriz


def columnas_de_fila(ruta, hoja, fila, log):
    """Letras de las columnas que tienen algo en esa fila, en orden.

    Existe para no poner un tope fijo de columnas. La matriz del cuadro de pagos
    tiene una columna por cada empresa que RECIBE, y eso cambia todos los meses:
    con un tope escrito a mano, el mes que se pase se pierden receptores en
    silencio y las verificaciones reportan faltantes que en realidad estan ahi.
    """
    import zipfile
    if not es_zip_excel(ruta):
        log(f"    ! {Path(ruta).name}: formato no soportado")
        return None
    fila = int(fila)
    # El \D del final evita que la fila 8 matchee con la 80 o la 815.
    patron = re.compile(rb'<c r="([A-Z]{1,3})' + str(fila).encode() + rb'"')
    encontradas = set()
    try:
        with zipfile.ZipFile(str(ruta)) as z:
            ruta_hoja, hojas = ubicar_hoja_xml(z, hoja)
            if ruta_hoja is None:
                log(f"    ! La hoja '{hoja}' no existe. Hojas: {', '.join(hojas)}")
                return None
            with z.open(ruta_hoja) as f:
                cola = b""
                while True:
                    trozo = f.read(1 << 20)
                    if not trozo:
                        break
                    buf = cola + trozo
                    fin = 0
                    for m in patron.finditer(buf):
                        fin = m.end()
                        encontradas.add(m.group(1).decode())
                    cola = buf[fin:] if fin else buf[-64:]
    except Exception as e:
        log(f"    ! No se pudieron leer las columnas de la fila {fila}: {e}")
        return None
    return sorted(encontradas, key=col_letra_a_num)


def leer_matriz_pago(ruta, hoja, log, fila_enc=8, col_ini="N"):
    """Lee la matriz cruzada. Devuelve dict con:
        pagan     : {nombre: fila}
        reciben   : {nombre: columna}
        montos    : {(pagador, receptor): monto}
        ultima_fila / ultima_col  (numero de columna)
    o None si no se pudo leer.

    Se descartan la fila y la columna de totales: CuadroPago escribe un "Total"
    al final de cada lado, y esos no son pares de pago.
    """
    # Hasta donde llega el encabezado de la matriz lo dice el propio archivo.
    presentes = columnas_de_fila(ruta, hoja, fila_enc, log)
    if presentes is None:
        return None
    n_ini = col_letra_a_num(col_ini)
    ultima = max((col_letra_a_num(c) for c in presentes if
                  col_letra_a_num(c) >= n_ini), default=n_ini)
    cols = [col_letra(n) for n in range(n_ini, ultima + 1)]
    log(f"        encabezado de la matriz: {col_ini}{fila_enc} a "
        f"{col_letra(ultima)}{fila_enc}  ({len(cols)} columna(s))")
    datos = leer_columnas_rapido(ruta, hoja, cols, fila_enc, log)
    if datos is None:
        return None

    def es_nombre(v):
        return (isinstance(v, str) and v.strip()
                and not v.startswith("#")
                and normalizar(v) not in ("", "0", "total"))

    # Receptores: la fila del encabezado, de la segunda columna en adelante.
    reciben = {}
    for c in cols[1:]:
        v = datos.get(c, {}).get(fila_enc)
        if v is None or (isinstance(v, str) and not v.strip()):
            break                      # la matriz termina en el primer hueco
        if es_nombre(v):
            reciben[str(v).strip()] = c
    # Pagadores: la primera columna, de la fila siguiente hacia abajo.
    pagan = {}
    for f in sorted(datos.get(cols[0], {})):
        if f <= fila_enc:
            continue
        v = datos[cols[0]][f]
        if es_nombre(v):
            pagan[str(v).strip()] = f
    montos = {}
    for p, f in pagan.items():
        for r, c in reciben.items():
            v = datos.get(c, {}).get(f)
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                montos[(normalizar(p), normalizar(r))] = float(v)
    return {"pagan": pagan, "reciben": reciben, "montos": montos,
            "ultima_fila": max(pagan.values()) if pagan else fila_enc,
            "ultima_col": max((col_letra_a_num(c) for c in reciben.values()),
                              default=col_letra_a_num(col_ini))}


def leer_nombre_definido(ruta, nombre):
    """El rango al que apunta un nombre definido, tal como esta guardado.
    Puede venir en A1 ($N$8:$T$89) o en R1C1 localizado (F8C14:F84C22)."""
    import zipfile
    try:
        with zipfile.ZipFile(str(ruta)) as z:
            wx = z.read("xl/workbook.xml").decode("utf-8", "ignore")
    except Exception:
        return None
    m = re.search(r'<definedName name="' + re.escape(nombre)
                  + r'"[^>]*>([^<]*)</definedName>', wx)
    return m.group(1) if m else None


def celdas_de_rango(texto):
    """(fila_fin, col_fin) del final de un rango, sea A1 o R1C1 localizado.
    Devuelve None si no se entiende."""
    if not texto:
        return None
    t = texto.replace("=", "").strip()
    t = t.split("!")[-1]
    # R1C1 localizado: F8C14:F84C22  (Fila/Columna, en Excel en espanol)
    m = re.search(r"[FR](\d+)[CL](\d+)\s*$", t)
    if m:
        return int(m.group(1)), int(m.group(2))
    # A1: $N$8:$T$89
    m = re.search(r"\$?([A-Z]{1,3})\$?(\d+)\s*$", t)
    if m:
        return int(m.group(2)), col_letra_a_num(m.group(1))
    return None


# =============================================================================
#  "Traer maestro": rehacer una copia a partir de su maestro
# =============================================================================
# Las copias de "00 Entregables" (y los diarios de Sobrecostos de Energia) no se
# generan con un script: se copian del maestro a mano. Este boton hace eso mismo.
#
# CLAVE: se copia con shutil.copy2, que CONSERVA la fecha de modificacion. Si se
# copiara con copy() a secas, la copia quedaria con la fecha de ahora y el revisor
# la seguiria marcando en amarillo, porque compara justamente por fecha.


def plan_traer_maestro(nodo, ruta_copia, ruta_maestro, diarios_copia, diarios_maestro,
                       carpeta_copia):
    """Arma la lista de operaciones, sin tocar el disco.

    Devuelve (acciones, avisos). Cada accion es
        ("copiar"|"reemplazar"|"borrar", origen_o_None, destino, etiqueta)

    Se separa del ejecutor a proposito: asi se puede mostrar al usuario exactamente
    que se va a hacer ANTES de hacerlo, y se puede probar sin borrar nada.
    """
    acciones, avisos = [], []

    if nodo["tipo"] == "diarios":
        if carpeta_copia is None:
            avisos.append("No existe la carpeta de destino.")
            return [], avisos
        if not diarios_maestro:
            avisos.append("El maestro no tiene ningun archivo: no hay nada que traer.")
            return [], avisos
        for fecha in sorted(set(diarios_maestro) | set(diarios_copia)):
            pm = diarios_maestro.get(fecha)
            pp = diarios_copia.get(fecha)
            if pm and not pp:
                acciones.append(("copiar", pm, carpeta_copia / pm.name, fecha))
            elif pm and pp:
                if iguales_mtime(mtime(pm), mtime(pp)):
                    continue                      # ya esta al dia
                # Si el nombre cambio, primero se borra el viejo: si no quedarian
                # los dos y el revisor veria un archivo que sobra.
                if pp.name != pm.name:
                    acciones.append(("borrar", None, pp, fecha))
                acciones.append(("reemplazar", pm, carpeta_copia / pm.name, fecha))
            elif pp and not pm:
                acciones.append(("borrar", None, pp, fecha))
        return acciones, avisos

    # --- archivo suelto ---
    if ruta_maestro is None:
        avisos.append("No se encontro el archivo maestro.")
        return [], avisos
    if carpeta_copia is None:
        avisos.append("No existe la carpeta de destino.")
        return [], avisos
    destino = carpeta_copia / ruta_maestro.name
    if ruta_copia is None:
        acciones.append(("copiar", ruta_maestro, destino, ruta_maestro.name))
    elif iguales_mtime(mtime(ruta_maestro), mtime(ruta_copia)):
        pass                                       # ya esta al dia
    else:
        if ruta_copia.name != ruta_maestro.name:
            acciones.append(("borrar", None, ruta_copia, ruta_copia.name))
        acciones.append(("reemplazar", ruta_maestro, destino, ruta_maestro.name))
    return acciones, avisos


def ids_de_verificacion(vid):
    """Nodos del arbol que un verificador necesita leer.

    Se usa para releer del disco SOLO lo que hace falta antes de verificar. La
    lectura completa recorre todas las carpetas de diarios, que en un disco de
    red es lo que se lleva el tiempo, y para verificar un archivo no aporta nada.

    Se juntan de todos lados para no dejar ninguno afuera: el archivo propio, los
    "depende", los archivos que nombra cada comprobacion (incluidos los anidados
    en referencia/lado_a/lado_b) y los de los VALORES que use.
    """
    v = VERIFICADORES.get(vid)
    if not v:
        return set()
    ids = {v["archivo"]}
    ids.update(v.get("depende", []))
    for c in v.get("comprobaciones", []):
        for k in ("archivo", "archivo_a", "archivo_b"):
            if c.get(k):
                ids.add(c[k])
        for anidado in ("referencia", "lado_a", "lado_b"):
            sub = c.get(anidado)
            if isinstance(sub, dict) and sub.get("archivo"):
                ids.add(sub["archivo"])
        claves = []
        if c["tipo"] == "igualdad":
            claves = list(c.get("izq", [])) + list(c.get("der", []))
        elif c["tipo"] == "cero":
            claves = list(c.get("claves", []))
        elif c["tipo"] == "umbral":
            claves = [c["clave"]]
        for k in claves:
            spec = VALORES.get(partir_signo(k)[0])
            if spec and spec.get("archivo"):
                ids.add(spec["archivo"])
    return {i for i in ids if i}


def ids_de_verificaciones(vids):
    """Union de los nodos que necesitan varios verificadores, con sus previas."""
    ids, pendientes, vistos = set(), list(vids), set()
    while pendientes:
        vid = pendientes.pop()
        if vid in vistos:
            continue
        vistos.add(vid)
        ids |= ids_de_verificacion(vid)
        # Las previas se evaluan tambien, asi que sus archivos entran igual.
        pendientes += VERIFICADORES.get(vid, {}).get("verif_previas", [])
    return ids


def ultimo_significativo(datos, columna):
    """(fila, valor) del ultimo dato util de una columna, omitiendo vacios,
    ceros y errores. (None, None) si no hay ninguno."""
    for f in sorted(datos.get(columna, {}), reverse=True):
        v = datos[columna][f]
        if es_significativo(v):
            return f, v
    return None, None


def leer_valor_por_etiqueta(ruta, hoja, col_etiqueta, texto, col_valor,
                            fila_inicio, log):
    """Busca la fila cuya col_etiqueta diga `texto` y devuelve el numero que hay
    en col_valor de esa misma fila. Sirve para tablas que cambian de largo: no
    importa en que fila quede el "Total general", se lo busca por el rotulo.
    Si aparece mas de una vez se usa la ULTIMA."""
    datos = leer_columnas_rapido(ruta, hoja, [col_etiqueta, col_valor],
                                 fila_inicio, log)
    if datos is None:
        return None
    objetivo = normalizar(texto)
    filas = []
    for f, v in datos.get(col_etiqueta, {}).items():
        if isinstance(v, str) and normalizar(v) == objetivo:
            filas.append(f)
    if not filas:
        parecidas = [f"{col_etiqueta}{f}={str(v)[:24]!r}"
                     for f, v in sorted(datos.get(col_etiqueta, {}).items())[-6:]]
        log(f"    ! No encontré la fila '{texto}' en la columna {col_etiqueta}.")
        if parecidas:
            log(f"      últimos rótulos leídos: {', '.join(parecidas)}")
        return None
    fila = max(filas)
    if len(filas) > 1:
        log(f"      '{texto}' aparece {len(filas)} veces; se usa la fila {fila}")
    v = datos.get(col_valor, {}).get(fila)
    if not isinstance(v, (int, float)):
        log(f"    ! {col_valor}{fila} no es un número: {str(v)[:40]!r}")
        return None
    log(f"      fila '{texto}' = {col_etiqueta}{fila}, valor en {col_valor}{fila}")
    return float(v)


def leer_valor_excel(ruta, hoja, celda, log):
    """Lee una celda o la suma de un rango. Tres caminos, del mas rapido al mas
    lento: streaming del XML, openpyxl, y por ultimo Excel via xlwings."""
    # --- 1) streaming del XML: no carga el libro, corta al pasar la fila -----
    if ":" not in celda:
        rapido = leer_celdas_rapido(ruta, hoja, [celda])
        if rapido:
            v = rapido.get(celda.upper())
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return float(v)
            if isinstance(v, str) and v.startswith("#"):
                log(f"    ! {hoja}!{celda} tiene un error de fórmula: {v}")
                return None

    # --- 2) openpyxl --------------------------------------------------------
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(ruta), data_only=True, read_only=True)
        try:
            real = resolver_hoja(wb.sheetnames, hoja)
            if real is None:
                log(f"    ! La hoja '{hoja}' no existe. Hojas: {', '.join(wb.sheetnames)}")
                return None
            ws = wb[real]
            if ":" in celda:
                v = _suma_rango_openpyxl(ws, celda)
            else:
                v = ws[celda].value
                if not isinstance(v, (int, float)):
                    v = None
            if v is not None:
                return float(v)
            log(f"    · {hoja}!{celda} sin valor guardado en el archivo:")
            if ":" not in celda:
                diagnosticar_celda(ruta, hoja, celda, log)
            log("    · reintentando abriendo Excel...")
        finally:
            wb.close()
    except Exception as e:
        log(f"    · openpyxl no pudo leer ({e}); reintentando con Excel...")

    # --- xlwings ------------------------------------------------------------
    app = wb2 = None
    try:
        import xlwings as xw
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        wb2 = app.books.open(str(ruta), read_only=True, update_links=False)
        nombres = [s.name for s in wb2.sheets]
        real = resolver_hoja(nombres, hoja)
        if real is None:
            log(f"    ! La hoja '{hoja}' no existe. Hojas: {', '.join(nombres)}")
            return None
        sh = wb2.sheets[real]
        val = sh.range(celda).value
        if isinstance(val, list):
            total, hubo = 0.0, False
            pila = [val]
            while pila:
                x = pila.pop()
                if isinstance(x, list):
                    pila.extend(x)
                elif isinstance(x, (int, float)):
                    total += float(x)
                    hubo = True
            return total if hubo else None
        return float(val) if isinstance(val, (int, float)) else None
    except Exception as e:
        log(f"    ! Error leyendo con Excel: {e}")
        return None
    finally:
        try:
            if wb2 is not None:
                wb2.close()
            if app is not None:
                app.quit()
        except Exception:
            pass


def obtener_hojas(ruta):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(ruta), data_only=True, read_only=True)
        nombres = list(wb.sheetnames)
        wb.close()
        return nombres
    except Exception:
        pass
    app = wb2 = None
    try:
        import xlwings as xw
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        wb2 = app.books.open(str(ruta), read_only=True, update_links=False)
        return [s.name for s in wb2.sheets]
    except Exception:
        return []
    finally:
        try:
            if wb2 is not None:
                wb2.close()
            if app is not None:
                app.quit()
        except Exception:
            pass


def listar_hojas(ruta, log):
    hojas = obtener_hojas(ruta)
    if hojas:
        log(f"    Hojas de {ruta.name}: {', '.join(hojas)}")
    else:
        log(f"    No se pudieron listar las hojas de {ruta.name}")


def conexion_mdb(ruta):
    import pyodbc
    drivers = [d for d in pyodbc.drivers() if "Microsoft Access Driver" in d]
    if not drivers:
        raise RuntimeError(
            "No hay driver de Access instalado (o es de otra arquitectura que este "
            "Python). Instala 'Microsoft Access Database Engine' de la misma "
            f"arquitectura ({8 * 8 if sys.maxsize > 2**32 else 32} bits)."
        )
    cs = f"DRIVER={{{drivers[0]}}};DBQ={ruta};"
    return pyodbc.connect(cs, autocommit=True)


def listar_tablas_mdb(ruta, log):
    try:
        cn = conexion_mdb(ruta)
        cur = cn.cursor()
        tablas = [r.table_name for r in cur.tables(tableType="TABLE")]
        log(f"    Tablas de {ruta.name}: {', '.join(tablas) if tablas else '(ninguna)'}")
        for t in tablas:
            cols = [r.column_name for r in cur.columns(table=t)]
            log(f"      · {t}: {', '.join(cols)}")
        cn.close()
    except Exception as e:
        log(f"    No se pudo inspeccionar {ruta.name}: {e}")


def obtener_tablas_columnas(ruta):
    """{tabla: [columnas]} de una base Access. {} si no se pudo abrir."""
    out = {}
    cn = None
    try:
        cn = conexion_mdb(ruta)
        cur = cn.cursor()
        for t in [r.table_name for r in cur.tables(tableType="TABLE")]:
            out[t] = [r.column_name for r in cur.columns(table=t)]
    except Exception:
        pass
    finally:
        try:
            if cn is not None:
                cn.close()
        except Exception:
            pass
    return out


def desglose_por_tipo(ruta, tabla, columna, columna_tipo, where, log):
    """Escribe en el log la suma agrupada por tipo. Solo informativo."""
    cn = None
    try:
        cn = conexion_mdb(ruta)
        cur = cn.cursor()
        sql = (f"SELECT [{columna_tipo}], SUM([{columna}]), COUNT(*) "
               f"FROM [{tabla}]")
        if where.strip():
            sql += f" WHERE {where}"
        sql += f" GROUP BY [{columna_tipo}] ORDER BY [{columna_tipo}]"
        cur.execute(sql)
        filas = cur.fetchall()
        if filas:
            log(f"      desglose por {columna_tipo}:")
            for f in filas:
                log(f"        {str(f[0]):<12} {fmt_monto(f[1]):>22}   ({f[2]} filas)")
    except Exception as e:
        log(f"      (no se pudo desglosar por tipo: {e})")
    finally:
        try:
            if cn is not None:
                cn.close()
        except Exception:
            pass


def leer_valor_mdb(ruta, tabla, columna, where, log):
    cn = None
    try:
        cn = conexion_mdb(ruta)
        cur = cn.cursor()
        sql = f"SELECT SUM([{columna}]) FROM [{tabla}]"
        if where.strip():
            sql += f" WHERE {where}"
        cur.execute(sql)
        fila = cur.fetchone()
        return float(fila[0]) if fila and fila[0] is not None else None
    except Exception as e:
        log(f"    ! Error consultando {ruta.name}: {e}")
        return None
    finally:
        try:
            if cn is not None:
                cn.close()
        except Exception:
            pass


def obtener_valor(clave, rutas, log, usar_cache=True):
    """Devuelve (valor, mensaje_error_o_None).
    Si el archivo no cambio desde la ultima lectura y se pide exactamente lo
    mismo, devuelve el valor guardado sin abrir el archivo."""
    spec = VALORES[clave]
    ruta = rutas.get(spec["archivo"])
    if ruta is None:
        return None, f"falta el archivo de origen ({spec['archivo']})"

    huella = huella_spec(spec)
    if usar_cache:
        reg = CACHE.obtener(clave, ruta, huella)
        if reg is not None:
            extra = f", {reg['filas']} fila(s)" if reg.get("filas") else ""
            log(f"    {spec['etiqueta']}")
            log(f"      = {fmt_monto(reg['valor'])}   [ya calculado el "
                f"{reg['leido']}{extra}; {ruta.name} sin cambios, no se abrió]")
            return reg["valor"], None

    if spec["tipo"] == "excel_col":
        if not spec.get("hoja") or not spec.get("columna"):
            listar_hojas(ruta, log)
            return None, f"sin configurar hoja/columna para {clave}"
        filtro = ""
        if spec.get("columna_filtro"):
            filtro = (f"  filtrando {spec['columna_filtro']} = "
                      f"{'/'.join(spec.get('valores_filtro') or ['(nada)'])}")
        log(f"    {spec['etiqueta']}  <-  {spec['hoja']}!{spec['columna']}"
            f"{spec.get('fila_inicio', 2)} hacia abajo{filtro}")
        v, n, desg = leer_columna_excel(
            ruta, spec["hoja"], spec["columna"], int(spec.get("fila_inicio", 2)),
            spec.get("columna_filtro", ""), spec.get("valores_filtro") or [], log)
        if v is None:
            return None, f"no se pudo leer {clave}"
        log(f"      {n} fila(s) sumada(s)")
        if spec.get("columna_filtro") and len(desg) > 1:
            buscados = {normalizar(x) for x in (spec.get("valores_filtro") or [])}
            log(f"      valores presentes en la columna {spec['columna_filtro']}:")
            for k in sorted(desg):
                marca = "<--" if k in buscados else "   "
                log(f"        {marca} {k:<14} {fmt_monto(desg[k]):>22}")
        CACHE.poner(clave, ruta, huella, v, n)
        return v, None

    if spec["tipo"] == "excel_etiqueta":
        if not (spec.get("hoja") and spec.get("columna_etiqueta")
                and spec.get("texto_fila") and spec.get("columna_valor")):
            listar_hojas(ruta, log)
            return None, f"sin configurar la búsqueda por rótulo para {clave}"
        log(f"    {spec['etiqueta']}  <-  {spec['hoja']}: fila donde "
            f"{spec['columna_etiqueta']} = '{spec['texto_fila']}', "
            f"valor en {spec['columna_valor']}")
        v = leer_valor_por_etiqueta(
            ruta, spec["hoja"], spec["columna_etiqueta"], spec["texto_fila"],
            spec["columna_valor"], int(spec.get("fila_inicio", 1)), log)
        if v is not None:
            log(f"      = {fmt_monto(v)}")
            CACHE.poner(clave, ruta, huella, v)
        return v, None if v is not None else f"no se pudo leer {clave}"

    if spec["tipo"] == "excel":
        if not spec.get("hoja") or not spec.get("celda"):
            listar_hojas(ruta, log)
            return None, f"sin configurar hoja/celda para {clave}"
        log(f"    {spec['etiqueta']}  <-  {spec['hoja']}!{spec['celda']}")
        v = leer_valor_excel(ruta, spec["hoja"], spec["celda"], log)
        if v is not None:
            log(f"      = {fmt_monto(v)}")
            CACHE.poner(clave, ruta, huella, v)
        return v, None if v is not None else f"no se pudo leer {clave}"

    if spec["tipo"] == "mdb":
        if not spec.get("tabla") or not spec.get("columna"):
            listar_tablas_mdb(ruta, log)
            return None, f"sin configurar tabla/columna para {clave}"
        wh = spec.get("where", "")
        log(f"    {spec['etiqueta']}  <-  SUM([{spec['columna']}]) de [{spec['tabla']}]"
            + (f" WHERE {wh}" if wh.strip() else ""))
        v = leer_valor_mdb(ruta, spec["tabla"], spec["columna"], wh, log)
        if spec.get("columna_tipo"):
            desglose_por_tipo(ruta, spec["tabla"], spec["columna"],
                              spec["columna_tipo"], wh, log)
        if v is not None:
            CACHE.poner(clave, ruta, huella, v)
        return v, None if v is not None else f"no se pudo leer {clave}"

    return None, f"tipo de origen desconocido en {clave}"


# =============================================================================
#  Aplicacion
# =============================================================================

class Revisor:
    def __init__(self, root):
        self.root = root
        self.cfg = leer_config()
        self.rutas = {}          # id -> Path | None
        self.diarios = {}        # id -> {fecha: Path}
        self.detalle_diarios = {}  # id -> lista de strings con el detalle
        self.filas = {}          # id -> dict de widgets
        self.cola = queue.Queue()
        self.trabajando = False
        self.timer = {"on": False, "t0": 0.0}

        root.title("Revisor de entregables — CASO RELIQUIDACION")
        # Arranca MAXIMIZADA, aprovechando toda la pantalla: las filas del arbol
        # son anchas (nombre + fecha + estado + hasta cuatro botones) y en un
        # monitor chico se cortaban.
        # El geometry va igual como respaldo: es el tamano que toma la ventana
        # cuando el usuario la desmaximiza.
        root.geometry("1500x820")
        try:
            root.state("zoomed")                    # Windows
        except tk.TclError:
            try:
                root.attributes("-zoomed", True)    # varios Linux
            except tk.TclError:
                # Ultimo recurso: del tamano de la pantalla, pegada arriba a la
                # izquierda. No se usa -fullscreen a proposito, porque esconde la
                # barra de titulo y deja sin como cerrar o minimizar.
                root.geometry(f"{root.winfo_screenwidth()}x"
                              f"{root.winfo_screenheight()}+0+0")

        self.var_base = tk.StringVar(value=self.cfg.get("carpeta_base", "[sin seleccionar]"))
        self.var_estado = tk.StringVar(value="Listo")
        self.var_tiempo = tk.StringVar(value="00:00:00")
        self.var_aamm = tk.StringVar(value=self.cfg.get("ultimo_mes", ""))
        self.var_releer = tk.BooleanVar(value=False)
        self.var_aamm_info = tk.StringVar(
            value="último mes visto" if self.cfg.get("ultimo_mes") else "se detecta al actualizar")

        self._construir()
        if self.var_base.get() and Path(self.var_base.get()).is_dir():
            self.root.after(200, self.actualizar)
        self.root.after(300, self._bombear_cola)

    # ------------------------------------------------------------------ UI --
    def _construir(self):
        root = self.root

        # 1) barra inferior fija
        barra = tk.Frame(root)
        barra.pack(side="bottom", fill="x", pady=6)

        self.btn_act = tk.Button(barra, text="ACTUALIZAR", bg="#1f5fa8", fg="white",
                                 font=("Segoe UI", 10, "bold"), width=16,
                                 command=self.actualizar)
        self.btn_act.pack(side="left", padx=(16, 6))

        self.btn_todo = tk.Button(barra, text="Verificar todo", width=16,
                                  command=self.verificar_todo)
        self.btn_todo.pack(side="left", padx=6)

        self.btn_cfg = tk.Button(barra, text="Configurar valores...", width=20,
                                 command=self.configurar_valores)
        self.btn_cfg.pack(side="left", padx=6)

        self.chk_releer = tk.Checkbutton(
            barra, text="Releer todo (ignorar lo ya calculado)",
            variable=self.var_releer, font=("Segoe UI", 8))
        self.chk_releer.pack(side="left", padx=6)

        tk.Label(barra, textvariable=self.var_estado, anchor="w").pack(side="left", padx=14)
        tk.Label(barra, textvariable=self.var_tiempo, font=("Consolas", 10, "bold"),
                 fg="#2d7a2d").pack(side="right", padx=16)

        # 2) log fijo abajo
        marco_log = tk.LabelFrame(root, text="Bitácora")
        marco_log.pack(side="bottom", fill="x", padx=12, pady=(0, 4))
        self.txt_log = tk.Text(marco_log, height=9, font=("Consolas", 9), wrap="none")
        configurar_tags_log(self.txt_log)
        self._bloque_log = None
        sb_log = tk.Scrollbar(marco_log, command=self.txt_log.yview)
        self.txt_log.configure(yscrollcommand=sb_log.set)
        sb_log.pack(side="right", fill="y")
        self.txt_log.pack(side="left", fill="both", expand=True)

        self.progress = ttk.Progressbar(root, mode="determinate", length=400)
        self.progress.pack(side="bottom", fill="x", padx=12, pady=(0, 4))

        # 3) canvas con scroll
        canvas = tk.Canvas(root, borderwidth=0, highlightthickness=0)
        scroll = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        contenedor = tk.Frame(canvas)
        win = canvas.create_window((0, 0), window=contenedor, anchor="nw")

        def ajustar(_e=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(win, width=canvas.winfo_width())
        contenedor.bind("<Configure>", ajustar)
        canvas.bind("<Configure>", ajustar)
        # La rueda del mouse: el canvas del arbol solo scrollea si el puntero NO
        # esta sobre un widget que scrollea solo (la bitacora, o el texto de
        # cualquier ventana de detalle).
        #
        # bind_all captura la rueda GLOBALMENTE. Sin este filtro, al intentar
        # subir en la bitacora se movia la ventana entera: los dos respondian a
        # la misma rueda.
        def _rueda(e):
            w = e.widget.winfo_toplevel().winfo_containing(e.x_root, e.y_root)
            pasos = int(-e.delta / 120)
            # Se sube por los padres: el puntero puede estar sobre un hijo.
            while w is not None:
                if isinstance(w, (tk.Text, tk.Listbox)):
                    w.yview_scroll(pasos, "units")
                    return "break"
                if w is canvas:
                    break
                w = getattr(w, "master", None)
            canvas.yview_scroll(pasos, "units")

        canvas.bind_all("<MouseWheel>", _rueda)

        # --- selector de carpeta base
        f = tk.LabelFrame(contenedor, text="Carpeta  02 CASO RELIQUIDACION", padx=10, pady=6)
        f.pack(fill="x", padx=14, pady=(10, 4))
        self.lbl_base = tk.Label(f, textvariable=self.var_base, wraplength=980,
                                 justify="left", cursor="hand2", font=("Segoe UI", 9))
        self.lbl_base.pack(anchor="w")
        self.lbl_base.bind("<Button-1>", lambda e: abrir_en_explorador(self.var_base.get()))
        tk.Button(f, text="Examinar", command=self.examinar).pack(anchor="w", pady=(4, 0))
        self._color_base()

        # --- mes reliquidado y estado guardado
        fm = tk.LabelFrame(contenedor, text="Mes reliquidado (AAMM)", padx=10, pady=6)
        fm.pack(fill="x", padx=14, pady=4)
        linea = tk.Frame(fm)
        linea.pack(fill="x")
        ent = tk.Entry(linea, textvariable=self.var_aamm, width=8,
                       font=("Consolas", 11, "bold"), justify="center")
        ent.pack(side="left")
        ent.bind("<Return>", lambda e: self.ir_a_mes())
        tk.Button(linea, text="Ir a este mes", bg="#1f5fa8", fg="white",
                  font=("Segoe UI", 9, "bold"),
                  command=self.ir_a_mes).pack(side="left", padx=8)
        tk.Button(linea, text="Ver estado guardado",
                  command=self.ver_estado_mes).pack(side="left", padx=4)
        tk.Button(linea, text="Reiniciar este mes", fg="#a00000",
                  command=self.reiniciar_mes).pack(side="left", padx=12)
        tk.Label(linea, textvariable=self.var_aamm_info, font=("Segoe UI", 8),
                 fg="#444444", anchor="w").pack(side="left", padx=6)
        self.lbl_salidas = tk.Label(fm, text=f"Salidas: {DIR_SALIDAS}",
                                    font=("Consolas", 8), fg="blue", anchor="w",
                                    cursor="hand2")
        self.lbl_salidas.pack(fill="x", pady=(3, 0))
        self.lbl_salidas.bind("<Button-1>",
                              lambda e: abrir_en_explorador(dir_mes(self.var_aamm.get())
                                                            if dir_mes(self.var_aamm.get()).is_dir()
                                                            else DIR_SALIDAS))

        # --- leyenda
        leyenda = tk.Frame(contenedor)
        leyenda.pack(fill="x", padx=18, pady=(2, 6))
        for txt, bg, fg in [("desactualizado (copia ≠ maestro)", C_AMARILLO, "black"),
                            ("falta el archivo", C_NEUTRO, C_FALTA),
                            ("verificación vigente", C_NEUTRO, C_OK),
                            ("verificación vencida", C_VENCIDA, "black"),
                            ("sin verificar", C_NEUTRO, C_GRIS)]:
            tk.Label(leyenda, text=" " + txt + " ", bg=bg, fg=fg,
                     font=("Segoe UI", 8), relief="groove", bd=1).pack(side="left", padx=3)

        # --- arbol
        arbol = tk.LabelFrame(contenedor,
                              text="02 CASO RELIQUIDACION/   (+ la carpeta FD, que está un nivel arriba)",
                              padx=6, pady=6)
        arbol.pack(fill="both", expand=True, padx=14, pady=4)

        enc = tk.Frame(arbol)
        enc.pack(fill="x")
        tk.Label(enc, text="Estructura (nombre real del archivo)",
                 font=("Segoe UI", 8, "bold"), width=74, anchor="w").pack(side="left")
        tk.Label(enc, text="Fecha de modificación", font=("Segoe UI", 8, "bold"),
                 width=20, anchor="w").pack(side="left")
        tk.Label(enc, text="Estado", font=("Segoe UI", 8, "bold"), width=26,
                 anchor="w").pack(side="left")

        for nodo in NODOS:
            self._fila(arbol, nodo)

    def _fila(self, parent, nodo):
        fr = tk.Frame(parent)
        fr.pack(fill="x")
        widgets = {"frame": fr, "nodo": nodo}

        if nodo["tipo"] == "carpeta":
            lbl = tk.Label(fr, text=nodo["pref"] + nodo["texto"], font=("Consolas", 9, "bold"),
                           width=74, anchor="w")
            lbl.pack(side="left")
            widgets["lbl"] = lbl
            self.filas[nodo["id"]] = widgets
            return

        lbl = tk.Label(fr, text=nodo["pref"] + nodo["texto"], font=("Consolas", 9),
                       width=74, anchor="w", cursor="hand2")
        lbl.pack(side="left")
        lbl.bind("<Button-1>", lambda e, i=nodo["id"]: self._abrir(i))
        widgets["lbl"] = lbl

        lbl_fecha = tk.Label(fr, text="—", font=("Consolas", 8), width=20, anchor="w",
                             fg=C_GRIS)
        lbl_fecha.pack(side="left")
        widgets["fecha"] = lbl_fecha

        lbl_est = tk.Label(fr, text="", font=("Segoe UI", 8), width=26, anchor="w")
        lbl_est.pack(side="left")
        widgets["estado"] = lbl_est

        if nodo["tipo"] == "diarios":
            b = tk.Button(fr, text="detalle", font=("Segoe UI", 7),
                          command=lambda i=nodo["id"]: self._ver_detalle(i))
            b.pack(side="left", padx=3)
            widgets["btn_detalle"] = b

        vid = next((k for k, v in VERIFICADORES.items() if v["archivo"] == nodo["id"]), None)

        # El resultado de la verificacion va ANTES de los botones a proposito.
        # Es informacion que hay que poder leer siempre; los botones, si la fila
        # se queda corta, se corren a la derecha. Al final se cortaba justo esto
        # en la fila del cuadro cero, que es la que mas botones tiene.
        if vid:
            widgets["vid"] = vid
            b = tk.Button(fr, text="Verificar", font=("Segoe UI", 8), bg="#e8eef7",
                          command=lambda v=vid: self.verificar(v))
            b.pack(side="left", padx=4)
            widgets["btn_ver"] = b
            bd = tk.Button(fr, text="detalle", font=("Segoe UI", 7),
                           command=lambda v=vid: self.ver_detalle_verificacion(v))
            bd.pack(side="left", padx=(0, 4))
            widgets["btn_det_ver"] = bd

        # Los botones de actualizar y de exportar van UNO SOBRE OTRO, en una
        # columna. Es el unico motivo de este Frame: en la fila del cuadro cero
        # son tres y en una sola linea la fila se salia de la pantalla. Las filas
        # que tienen un solo boton se ven igual que antes.
        col = tk.Frame(fr)
        col.pack(side="left", padx=(0, 4))
        # Botones bajitos, para que la columna de tres no estire la fila:
        #   pady=0  quita el relleno interno vertical (por omision son 3px arriba
        #           y 3 abajo, o sea 6px por boton)
        #   bd=1    borde de 1px en vez de 2 (otros 2px)
        # Con los dos, cada boton pasa de ~23px de alto a ~15px.
        BAJO = dict(pady=0, bd=1)

        # Boton de actualizacion: solo en los maestros. Lanza el actualizador
        # que corresponde en una ventana aparte, ya apuntado al mes en curso.
        for k, spec in enumerate(ACTUALIZADORES.get(nodo["id"], [])):
            ba = tk.Button(col, text=spec.get("texto", "Actualizar data"),
                           font=("Segoe UI", 8), bg="#fdf0d5", **BAJO,
                           command=lambda i=nodo["id"], j=k:
                               self._lanzar_actualizador(i, j))
            ba.pack(fill="x")
            widgets.setdefault("btn_act_data", []).append(ba)

        # "Traer maestro": va solo en las filas que SON copia de otra, o sea las
        # que declaran espejo. No hay que listarlas a mano: el espejo ya dice
        # cuales son. Va en la misma columna que los otros botones.
        if nodo.get("espejo"):
            bm = tk.Button(col, text="Traer maestro", font=("Segoe UI", 8),
                           bg="#e8eef7", **BAJO,
                           command=lambda i=nodo["id"]: self._traer_maestro(i))
            bm.pack(fill="x")
            widgets.setdefault("btn_acciones", []).append(bm)

        # Acciones que corren dentro del revisor (por ejemplo exportar el csv).
        for rotulo, metodo in ACCIONES_INTERNAS.get(nodo["id"], []):
            bx = tk.Button(col, text=rotulo, font=("Segoe UI", 8), bg="#e6f0e6",
                           **BAJO,
                           command=lambda m=metodo, i=nodo["id"]:
                               getattr(self, m)(i))
            bx.pack(fill="x")
            widgets.setdefault("btn_acciones", []).append(bx)

        if vid:
            lv = tk.Label(fr, text="sin verificar", font=("Segoe UI", 8), fg=C_GRIS,
                          anchor="w", width=30)
            lv.pack(side="left")
            widgets["lbl_ver"] = lv

        self.filas[nodo["id"]] = widgets

    # ------------------------------------------------------------- helpers --
    def log(self, msg=""):
        # El bloque se guarda entre llamadas porque el log llega de a una linea:
        # asi las lineas indentadas heredan el color del ">>" que las encabeza.
        tag, self._bloque_log = clasificar_linea(msg, getattr(self, "_bloque_log", None))
        self.txt_log.insert("end", str(msg) + "\n", (tag,) if tag else ())
        self.txt_log.see("end")
        self.root.update_idletasks()

    def _color_base(self):
        v = self.var_base.get()
        self.lbl_base.config(fg="blue" if v and Path(v).is_dir() else C_FALTA)

    def _abrir(self, nid):
        r = self.rutas.get(nid)
        if r:
            abrir_en_explorador(r, es_archivo=True)
            return
        nodo = NODO_POR_ID[nid]
        base = self.var_base.get()
        if base and Path(base).is_dir():
            carp = resolver_carpeta(_base_de(base, nodo), nodo["carpeta"])
            if carp:
                abrir_en_explorador(carp)

    def examinar(self):
        ini = self.cfg.get("carpeta_base", "")
        ini = ini if ini and Path(ini).is_dir() else ""
        r = filedialog.askdirectory(title="Selecciona la carpeta 02 CASO RELIQUIDACION",
                                    initialdir=ini)
        if r:
            self.var_base.set(r)
            self.cfg["carpeta_base"] = r
            guardar_config({"carpeta_base": r})
            self._color_base()
            self.actualizar()

    def _fmt_tiempo(self, seg):
        m, s = divmod(int(seg), 60)
        h, m = divmod(m, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"

    def _tick(self):
        if self.timer["on"]:
            self.var_tiempo.set(self._fmt_tiempo(time.time() - self.timer["t0"]))
            self.root.after(500, self._tick)

    def _bloquear(self, on):
        est = "disabled" if on else "normal"
        self.btn_act.config(state=est, bg="#aaaaaa" if on else "#1f5fa8")
        self.btn_todo.config(state=est)
        self.btn_cfg.config(state=est)
        self.chk_releer.config(state=est)
        for w in self.filas.values():
            if "btn_ver" in w:
                w["btn_ver"].config(state=est)
            # Mientras corren verificaciones no se lanza ningun actualizador: si
            # el archivo cambia a media verificacion el resultado no sirve.
            for b in w.get("btn_act_data", []) + w.get("btn_acciones", []):
                b.config(state=est)

    # ---------------------------------------------------------- ACTUALIZAR --
    def actualizar(self, motivo=None, solo_ids=None):
        """Relee la carpeta: reubica los archivos, repinta fechas y recalcula la
        vigencia de las verificaciones. Devuelve True si pudo.

        solo_ids: si viene, se releen SOLO esos nodos del arbol y se salta el
        listado de las carpetas de diarios, que es lo que se lleva el tiempo en
        un disco de red. El resto de las rutas queda como estaba en memoria. Con
        None se relee todo, que es lo que hace el boton ACTUALIZAR.

        Es sincronico y no toca self.trabajando, asi que se puede llamar justo
        antes de verificar o de lanzar un actualizador para no trabajar sobre
        una foto vieja del disco."""
        base = self.var_base.get()
        if not base or not Path(base).is_dir():
            messagebox.showwarning("Falta la carpeta",
                                   "Selecciona primero la carpeta 02 CASO RELIQUIDACION.")
            return False
        base = Path(base)

        # La primera vez no hay nada en memoria, asi que no se puede releer a
        # medias: se relee todo aunque pidan lo contrario.
        parcial = bool(solo_ids) and bool(self.rutas)
        if parcial:
            solo_ids = set(solo_ids)
            # Los espejos de los nodos pedidos entran igual: se comparan fechas
            # contra el maestro y hace falta tener resuelta su ruta.
            for nid in list(solo_ids):
                esp = NODO_POR_ID.get(nid, {}).get("espejo")
                if esp:
                    solo_ids.add(esp)
            for nodo in NODOS:
                if nodo.get("espejo") in solo_ids:
                    solo_ids.add(nodo["id"])

        self.log("=" * 96)
        self.log(f"ACTUALIZAR  {datetime.now():%d-%m-%Y %H:%M:%S}"
                 + (f"   ({motivo})" if motivo else "")
                 + f"   ->  {base}")
        if parcial:
            self.log(f"  Relectura parcial: {len(solo_ids)} archivo(s), "
                     "sin recorrer las carpetas de diarios.")

        # Todo el trabajo de disco va dentro del cache de directorios: cada
        # carpeta se recorre UNA vez en vez de una por nodo. Se apaga al salir del
        # with, asi que lo que venga despues (comprobar si una verificacion vencio)
        # vuelve a leer del disco y no arrastra datos viejos.
        cd = cache_directorios()
        with cd:
            faltantes = self._releer(base, parcial, solo_ids)
        n_scans, n_hits = cd.stats
        if n_hits:
            self.log(f"  {n_scans} carpeta(s) recorrida(s), "
                     f"{n_hits} lectura(s) servida(s) del caché "
                     f"(antes serían {n_scans + n_hits})")

        self._informar_cache()
        if parcial:
            self.log(f"  Releídos {len(solo_ids)} archivo(s); "
                     f"{faltantes} faltante(s) entre ellos.")
        else:
            self.log(f"  Resumen: {faltantes} archivo(s) faltante(s).")
        self.var_estado.set(f"Actualizado {datetime.now():%H:%M:%S}")
        return True

    def _releer(self, base, parcial, solo_ids):
        """El trabajo de disco de actualizar(): ubicar los archivos y pintar el
        arbol. Separado para poder envolverlo en el cache de directorios.
        Devuelve la cantidad de archivos faltantes."""
        # 1) ubicar archivos y carpetas
        if not parcial:
            self.rutas, self.diarios = {}, {}
        faltantes = 0
        for nodo in NODOS:
            if nodo["tipo"] == "carpeta":
                continue
            if parcial and nodo["id"] not in solo_ids:
                continue
            carp = resolver_carpeta(_base_de(base, nodo), nodo["carpeta"])
            if nodo["tipo"] == "diarios":
                self.diarios[nodo["id"]] = listar_diarios(carp, nodo["patron"], nodo["ext"])
                self.rutas[nodo["id"]] = carp
            else:
                self.rutas[nodo["id"]] = buscar_archivo(carp, nodo["patron"], nodo["ext"])

        # 1b) deducir el mes reliquidado y cargar el estado de ese mes.
        # En una relectura PARCIAL todo esto se salta: el mes no cambia por
        # releer un archivo, y recargar el estado desde el disco no aporta nada.
        if parcial:
            aamm = (self.var_aamm.get() or "").strip()
        else:
            aamm, det = detectar_aamm(self.rutas, self.diarios)
            manual = self.var_aamm.get().strip()
            if aamm:
                if manual and manual != aamm:
                    self.log(f"  OJO: el AAMM escrito ({manual}) no coincide con el "
                             f"detectado ({aamm}). Se usa el detectado.")
                self.var_aamm.set(aamm)
                self.var_aamm_info.set(f"detectado {det}")
                # Se recuerda la carpeta de cada mes para poder volver a el despues.
                carpetas = dict(self.cfg.get("carpetas_por_mes") or {})
                carpetas[aamm] = str(base)
                self.cfg["carpetas_por_mes"] = carpetas
                self.cfg["ultimo_mes"] = aamm
                guardar_config({"ultimo_mes": aamm, "carpetas_por_mes": carpetas})
            elif manual:
                aamm = manual
                self.var_aamm_info.set("escrito a mano (no se pudo detectar)")
            else:
                self.var_aamm_info.set("no se pudo detectar: escríbelo a mano")
            if aamm:
                habia = ESTADO.cargar(aamm)
                CACHE.cargar(aamm)
                self.log(f"  Mes {aamm} — estado en {dir_config_mes(aamm) / ARCHIVO_ESTADO}"
                         + ("" if habia else "   (aún no existe, se crea al verificar)"))
            else:
                ESTADO.cargar(None)
                CACHE.cargar(None)
                self.log("  Sin AAMM: las verificaciones no se podrán guardar.")

        # 2) pintar archivos y comparar espejos
        for nodo in NODOS:
            if nodo["tipo"] == "carpeta":
                continue
            if parcial and nodo["id"] not in solo_ids:
                continue          # no se releyo: se deja como estaba
            w = self.filas[nodo["id"]]
            w["lbl"].config(bg=C_NEUTRO, fg="black")
            w["estado"].config(bg=C_NEUTRO, fg="black", text="")

            if nodo["tipo"] == "diarios":
                self._actualizar_diarios(nodo, w)
                continue

            r = self.rutas[nodo["id"]]
            if r is None:
                w["fecha"].config(text="—", fg=C_GRIS)
                w["lbl"].config(text=nodo["pref"] + nodo["texto"], fg=C_FALTA)
                w["estado"].config(text="FALTA EL ARCHIVO", fg=C_FALTA)
                self.log(f"  FALTA  {nodo['texto']}  ({_texto_carpeta(nodo)})")
                faltantes += 1
                for b in w.get("btn_act_data", []):
                    b.config(state="disabled")
                continue

            for b in w.get("btn_act_data", []):
                b.config(state="normal")

            w["lbl"].config(text=nodo["pref"] + r.name)
            w["fecha"].config(text=fmt_fecha(mtime(r)), fg="black")
            if nodo.get("solo_info"):
                # Nodos que son solo origen: no tienen maestro ni verificacion de
                # valores (la carpeta FD, el Retiros_h.parquet). El texto se puede
                # cambiar por nodo con "estado_info".
                w["estado"].config(text=nodo.get("estado_info",
                                                 "origen (no se verifica)"),
                                   fg=C_GRIS)
                continue
            esp = nodo.get("espejo")
            if not esp:
                w["estado"].config(text="ok", fg=C_OK)
                continue

            maestro = self.rutas.get(esp)
            if maestro is None:
                w["estado"].config(text="falta el maestro", fg=C_FALTA)
                continue
            if iguales_mtime(mtime(r), mtime(maestro)):
                w["estado"].config(text="igual al maestro", fg=C_OK)
            else:
                w["lbl"].config(bg=C_AMARILLO)
                w["estado"].config(text="DESACTUALIZADO", bg=C_AMARILLO, fg="black")
                self.log(f"  AMARILLO  {r.name}")
                self.log(f"            copia   : {fmt_fecha(mtime(r))}   ({r.parent})")
                self.log(f"            maestro : {fmt_fecha(mtime(maestro))}   ({maestro.parent})")

        # 3) vigencia de verificaciones. Es en memoria, no toca el disco.
        self._actualizar_verificaciones()
        return faltantes

    def _actualizar_diarios(self, nodo, w):
        carp = self.rutas.get(nodo["id"])
        propios = self.diarios.get(nodo["id"], {})
        detalle = []
        if carp is None:
            w["fecha"].config(text="—", fg=C_GRIS)
            w["lbl"].config(fg=C_FALTA)
            w["estado"].config(text="FALTA LA CARPETA", fg=C_FALTA)
            self.detalle_diarios[nodo["id"]] = ["No se encontró la carpeta."]
            return

        etiqueta = nodo["texto"].split("/")[0]
        w["lbl"].config(text=f"{nodo['pref']}{carp.name}/   ({etiqueta})"
                        if normalizar(carp.name) != normalizar(etiqueta)
                        else f"{nodo['pref']}{carp.name}/")
        w["fecha"].config(text=f"{len(propios)} archivo(s)", fg="black")
        esp = nodo.get("espejo")
        if not esp:
            w["estado"].config(text="maestro", fg=C_OK)
            detalle = [f"{f}  {fmt_fecha(mtime(p))}" for f, p in sorted(propios.items())]
            self.detalle_diarios[nodo["id"]] = detalle
            return

        maestros = self.diarios.get(esp, {})
        problemas = 0
        for fecha in sorted(set(maestros) | set(propios)):
            pm, pp = maestros.get(fecha), propios.get(fecha)
            if pm and not pp:
                detalle.append(f"{fecha}  FALTA en esta carpeta   (maestro {fmt_fecha(mtime(pm))})")
                problemas += 1
            elif pp and not pm:
                detalle.append(f"{fecha}  SOBRA: no existe en el maestro")
                problemas += 1
            elif iguales_mtime(mtime(pm), mtime(pp)):
                detalle.append(f"{fecha}  ok   {fmt_fecha(mtime(pp))}")
            else:
                detalle.append(f"{fecha}  DISTINTO   copia {fmt_fecha(mtime(pp))} | "
                               f"maestro {fmt_fecha(mtime(pm))}")
                problemas += 1
        self.detalle_diarios[nodo["id"]] = detalle or ["Sin archivos."]

        if problemas:
            w["lbl"].config(bg=C_AMARILLO)
            w["estado"].config(text=f"DESACTUALIZADO ({problemas})", bg=C_AMARILLO, fg="black")
            self.log(f"  AMARILLO  {_texto_carpeta(nodo)}: {problemas} día(s) con diferencia.")
            for d in detalle:
                if " ok " not in d:
                    self.log(f"            {d}")
        else:
            w["estado"].config(text=f"{len(propios)} día(s) al día", fg=C_OK)

    # -------------------------------------------------- ACTUALIZAR DATA --
    def _bloqueado_para_escritura(self, ruta):
        """True si otro proceso tiene el archivo tomado y no se va a poder
        guardar. Es la prueba DIRECTA de lo que importa: se pide permiso de
        escritura y se cierra al instante, sin escribir ni un byte.

        Es mejor que mirar el "~$": Excel deja ese archivo huérfano cuando se
        cae o lo matan, y entonces el "~$" existe para siempre aunque el libro
        esté cerrado. Esta prueba, en cambio, dice la verdad de ahora."""
        try:
            with open(ruta, "r+b"):
                return False
        except PermissionError:
            return True
        except OSError:
            # No existe, o algo raro con la red. No se bloquea aca: si de verdad
            # hay un problema, el actualizador lo va a decir con mejor detalle.
            return False

    def _lock_excel(self, ruta):
        """Ruta del "~$" que Excel deja al lado del libro abierto, si existe."""
        try:
            p = Path(ruta)
            lock = p.parent / ("~$" + p.name)
            return lock if lock.exists() else None
        except Exception:
            return None

    def _dueno_del_lock(self, lock):
        """Nombre de quien tiene el libro abierto. Excel lo guarda dentro del
        "~$": un byte con el largo y despues el nombre, a veces en ANSI y a
        veces en UTF-16. El formato no esta documentado, asi que esto es al
        mejor esfuerzo: si no se entiende se devuelve None y listo."""
        try:
            crudo = Path(lock).read_bytes()[:200]
        except Exception:
            return None
        if not crudo:
            return None
        candidatos = []
        # Se prueban las dos codificaciones y ambos desfases, porque el primer
        # byte es el largo y descuadra el UTF-16 si no se saltea.
        for datos in (crudo, crudo[1:]):
            for codec in ("utf-16-le", "latin-1"):
                try:
                    candidatos.append(datos.decode(codec, errors="ignore"))
                except Exception:
                    pass
        for texto in candidatos:
            # Los \x00 sobrantes cortarian el nombre en pedazos de 1 letra.
            limpio = texto.replace("\x00", "")
            m = re.search(r"[A-Za-z0-9._\-]{2,}(?:[ ][A-Za-z0-9._\-]+)*", limpio)
            if m:
                nombre = m.group(0).strip()
                if len(nombre) >= 3:
                    return nombre
        return None

    def _armar_traspaso(self, aamm, planilla, nid=None):
        """JSON que reciben los actualizadores. Solo se escriben las rutas que
        el revisor pudo resolver; el actualizador tolera que falte alguna.

        nid: la fila DESDE LA QUE se apreto el boton. Hace falta cuando el mismo
        script cuelga de varias filas (Prorratear esta en los tres .mdb): sin
        esto el script no puede saber a cual le dieron y tiene que adivinar.
        Se manda el id del nodo y tambien la clave de su ruta, ya resuelta.
        """
        rutas = {}
        for nid_, clave in CLAVES_TRASPASO.items():
            r = self.rutas.get(nid_)
            if r is not None:
                rutas[clave] = str(r)
        d = {"origen": "Revisor_Reliquidacion",
             "version": TRASPASO_VERSION,
             "aamm": aamm,
             "carpeta_reliq": self.var_base.get(),
             "rutas": rutas}
        if planilla:
            d["planilla"] = planilla
        if nid:
            d["nodo"] = nid
            # La ruta de ESA fila, sin que el script tenga que saber que clave
            # le corresponde.
            r = self.rutas.get(nid)
            if r is not None:
                d["ruta_nodo"] = str(r)
            clave = CLAVES_TRASPASO.get(nid)
            if clave:
                d["clave_nodo"] = clave
        return d

    def _lanzar_actualizador(self, nid, indice=0):
        spec = ACTUALIZADORES[nid][indice]
        # El script puede estar en una subcarpeta ("Reemplazos REUC/..."), asi
        # que se arma con Path para que la barra funcione igual en Windows.
        script_rel = Path(spec["script"])
        script_nombre = script_rel.name
        planilla = spec.get("planilla")
        nodo = NODO_POR_ID[nid]

        # Se relee el disco ANTES de leer la ruta del destino: si apareció una
        # revisión nueva (R01E donde antes había R01D), la ruta que tenía en
        # memoria apunta al archivo viejo y le mandaríamos ese al actualizador.
        # Se releen el destino y todo lo que va en el JSON de traspaso, que es
        # lo unico que el actualizador va a usar.
        if not self.actualizar(motivo=f"previo a actualizar data de {nid}",
                               solo_ids=set(CLAVES_TRASPASO) | {nid}):
            return
        destino = self.rutas.get(nid)

        if destino is None:
            messagebox.showwarning(
                "Falta el archivo",
                f"No se encontró el archivo de esta fila.\n\n{nodo['texto']}\n\n"
                "Aprieta ACTUALIZAR para volver a buscarlo.")
            return

        script = DIR_SCRIPT / script_rel
        if not script.is_file():
            messagebox.showerror(
                "Falta el actualizador",
                f"No se encontró:\n\n{script_rel}\n\nSe buscó en:\n{DIR_SCRIPT}")
            self.log(f"  ERROR: no se encontró {script_rel} en {DIR_SCRIPT}")
            return

        # Aviso obligatorio cuando el destino es un Excel que se va a escribir:
        # xlwings necesita poder guardarlo. Se prueba la escritura de verdad; el
        # "~$" solo sirve como pista de quien lo tiene, porque queda huerfano
        # cuando Excel se cae. Un .mdb no tiene "~$", asi que no aplica.
        destino_es_excel = destino.suffix.lower() in XL
        if destino_es_excel:
            lock = self._lock_excel(destino)
            if self._bloqueado_para_escritura(destino):
                dueno = self._dueno_del_lock(lock) if lock else None
                self.log(f"  NO SE LANZÓ {script_nombre}: el destino está tomado.")
                self.log(f"     {destino}")
                if dueno:
                    self.log(f"     lo tiene abierto: {dueno}")
                messagebox.showwarning(
                    "El archivo está en uso",
                    f"No se puede escribir en:\n\n{destino.name}\n\n"
                    + (f"Lo tiene abierto: {dueno}\n\n" if dueno else "")
                    + "Ciérralo (o pídele que lo cierre) y volvé a intentar.")
                return
            if lock is not None:
                # Se puede escribir, así que el "~$" sobró: Excel lo dejó tirado.
                dueno = self._dueno_del_lock(lock)
                self.log(f"  OJO: hay un '~$' huérfano al lado del destino.")
                self.log(f"     {lock}")
                if dueno:
                    self.log(f"     quedó a nombre de: {dueno}")
                self.log("     El archivo SÍ se puede escribir, así que se puede seguir.")
                if not messagebox.askyesno(
                        "Quedó un archivo de bloqueo",
                        f"Hay un archivo de bloqueo de Excel al lado del destino:\n\n"
                        f"{lock.name}\n\n"
                        + (f"Quedó a nombre de: {dueno}\n\n" if dueno else "")
                        + "Pero el archivo se puede escribir sin problema, así que "
                          "lo más probable es que sea basura de un Excel que se cerró mal.\n\n"
                          "Podés borrar ese archivo con tranquilidad (está oculto).\n\n"
                          "¿Actualizar igual?"):
                    return

        aamm = (self.var_aamm.get() or "").strip()
        if not aamm:
            if not messagebox.askyesno(
                    "Sin mes",
                    "No hay AAMM definido, así que no se sabe a qué mes pertenece "
                    "esto.\n\nEl traspaso se guardará en __config__/sin_mes.\n\n¿Lanzar igual?"):
                return

        # Se deja el mes en curso en config.json para que los actualizadores
        # abiertos a mano después arranquen apuntando al mes correcto.
        cambios = {"carpeta_reliq": self.var_base.get()}
        mdb = self.rutas.get("a_mdb_sscc")
        if mdb is not None:
            cambios["mdb"] = str(mdb)
        guardar_config(cambios)
        self.cfg.update(cambios)

        traspaso = self._armar_traspaso(aamm, planilla, nid)
        carpeta_salida = dir_config_mes(aamm or "sin_mes", crear=True)
        ruta_traspaso = carpeta_salida / ARCHIVO_TRASPASO
        try:
            escribir_json(ruta_traspaso, traspaso)
        except Exception as e:
            messagebox.showerror("No se pudo escribir el traspaso", str(e))
            self.log(f"  ERROR al escribir {ruta_traspaso}: {e}")
            return

        try:
            # cwd = carpeta del propio script, no la del revisor: el de
            # Reemplazos vive en su subcarpeta y resuelve cosas relativas a ella.
            subprocess.Popen([sys.executable, str(script), str(ruta_traspaso)],
                             cwd=str(script.parent))
        except Exception as e:
            messagebox.showerror("No se pudo lanzar el actualizador", str(e))
            self.log(f"  ERROR al lanzar {script_nombre}: {e}")
            return

        # Bitácora: si después algo no cuadra, esto permite reconstruir qué pasó.
        self.log("-" * 96)
        self.log(f"ACTUALIZAR DATA  {datetime.now():%d-%m-%Y %H:%M:%S}")
        self.log(f"  script   : {script_nombre}"
                 + (f"   (planilla {planilla})" if planilla else ""))
        self.log(f"  mes      : {aamm or 'sin_mes'}")
        self.log(f"  destino  : {destino}")
        self.log(f"  traspaso : {ruta_traspaso}")
        for clave in sorted(traspaso["rutas"]):
            self.log(f"     {clave:22s} {traspaso['rutas'][clave]}")
        faltan = [c for c in CLAVES_TRASPASO.values() if c not in traspaso["rutas"]]
        if faltan:
            self.log(f"  sin resolver: {', '.join(sorted(faltan))}")
        if destino_es_excel:
            self.log("  OJO: el actualizador deja el archivo GUARDADO y ABIERTO en Excel.")
            self.log("       El revisor lee la versión de disco, así que si sigues editando")
            self.log("       sin guardar, lo que ves y lo que verifica el revisor difieren.")
        self.log("  Cuando termine, aprieta ACTUALIZAR para releer fechas y vencer "
                 "las verificaciones que dependan del archivo.")

    def _traer_maestro(self, nid):
        """Rehace esta copia a partir de su maestro."""
        nodo = NODO_POR_ID[nid]
        esp = nodo.get("espejo")
        if not esp:
            return
        maestro = NODO_POR_ID[esp]

        # Se relee primero: si el maestro cambio recien, hay que copiar el de
        # ahora y no el que estaba en memoria.
        if not self.actualizar(motivo=f"previo a traer el maestro de {nid}",
                               solo_ids={nid, esp}):
            return

        base = Path(self.var_base.get())
        carpeta_copia = resolver_carpeta(_base_de(base, nodo), nodo["carpeta"])
        acciones, avisos = plan_traer_maestro(
            nodo,
            self.rutas.get(nid), self.rutas.get(esp),
            self.diarios.get(nid, {}), self.diarios.get(esp, {}),
            carpeta_copia)

        if avisos:
            messagebox.showwarning("No se puede traer el maestro",
                                   "\n".join(avisos))
            for a in avisos:
                self.log(f"  {a}")
            return
        if not acciones:
            messagebox.showinfo("Ya está al día",
                                f"{nodo['texto']}\n\nYa coincide con el maestro. "
                                "No hay nada que copiar.")
            return

        n_copiar = sum(1 for a in acciones if a[0] == "copiar")
        n_reemp = sum(1 for a in acciones if a[0] == "reemplazar")
        n_borrar = sum(1 for a in acciones if a[0] == "borrar")

        # Nada abierto en Excel: si un destino esta tomado, la copia falla a medias
        # y quedaria la carpeta con una parte vieja y otra nueva.
        tomados = []
        for tipo, _, destino, _ in acciones:
            if destino.exists() and self._bloqueado_para_escritura(destino):
                tomados.append(destino)
        if tomados:
            nombres = "\n".join(f"  • {t.name}" for t in tomados[:10])
            messagebox.showwarning(
                "Hay archivos en uso",
                "No se puede escribir en:\n\n" + nombres +
                "\n\nCiérralos y volvé a intentar.")
            self.log(f"  NO se trajo el maestro: {len(tomados)} archivo(s) en uso")
            for t in tomados:
                self.log(f"     en uso: {t}")
            return

        detalle = []
        for tipo, origen, destino, etiqueta in acciones[:14]:
            verbo = {"copiar": "copiar", "reemplazar": "REEMPLAZAR",
                     "borrar": "BORRAR"}[tipo]
            detalle.append(f"  {verbo}  {etiqueta}")
        if len(acciones) > 14:
            detalle.append(f"  ... y {len(acciones) - 14} más")

        resumen = ", ".join(p for p in (
            f"{n_copiar} a copiar" if n_copiar else "",
            f"{n_reemp} a reemplazar" if n_reemp else "",
            f"{n_borrar} a BORRAR" if n_borrar else "") if p)

        if not messagebox.askyesno(
                "Traer el maestro",
                f"Destino:\n{carpeta_copia}\n\n"
                f"Maestro:\n{maestro['texto']}\n\n"
                f"{resumen}\n\n" + "\n".join(detalle) +
                "\n\nLo que haya en el destino se pierde. ¿Seguir?"):
            return

        self.log("-" * 96)
        self.log(f"TRAER MAESTRO  {datetime.now():%d-%m-%Y %H:%M:%S}")
        self.log(f"  copia   : {nodo['texto']}")
        self.log(f"  maestro : {maestro['texto']}")
        self.log(f"  destino : {carpeta_copia}")

        ok, fallos = 0, []
        for tipo, origen, destino, etiqueta in acciones:
            try:
                if tipo == "borrar":
                    destino.unlink()
                    self.log(f"     borrado    {destino.name}")
                else:
                    destino.parent.mkdir(parents=True, exist_ok=True)
                    # copy2 conserva la fecha de modificacion, que es lo que el
                    # revisor compara. Con copy() a secas quedaria en amarillo.
                    shutil.copy2(str(origen), str(destino))
                    self.log(f"     {'copiado ' if tipo == 'copiar' else 'reemplazado'}"
                             f"  {destino.name}   ({fmt_fecha(mtime(destino))})")
                ok += 1
            except Exception as e:
                fallos.append((etiqueta, e))
                self.log(f"     ERROR en {etiqueta}: {e}")

        self.log(f"  {ok} de {len(acciones)} operación(es) sin problemas.")
        # Se relee para que el árbol muestre el resultado al toque.
        self.actualizar(motivo="después de traer el maestro", solo_ids={nid, esp})

        if fallos:
            messagebox.showerror(
                "Terminó con errores",
                f"{ok} de {len(acciones)} operaciones salieron bien.\n\n"
                f"Falló en {len(fallos)}; el detalle está en la bitácora.")
        else:
            messagebox.showinfo("Listo",
                                f"{ok} operación(es).\n\nLa copia quedó igual "
                                "que el maestro.")

    def _exportar_cprt(self, nid="a_0_cuadros"):
        """Saca la hoja CPRT del cuadro cero a un .csv listo para mandar a pago."""
        # Se relee primero, por lo mismo que el boton de actualizar data: si
        # apareció una revisión nueva, la ruta en memoria es la vieja.
        if not self.actualizar(motivo="previo a exportar el CPRT",
                               solo_ids={nid}):
            return
        origen = self.rutas.get(nid)
        if origen is None:
            messagebox.showwarning("Falta el archivo",
                                   "No se encontró el 0_CUADROS_RELIQUIDACIÓN.")
            return

        aamm = (self.var_aamm.get() or "").strip()
        # El sufijo de revisión sale del nombre del origen, no se inventa: el csv
        # tiene que quedar rotulado con la misma revisión de la que salió.
        m = re.search(r"_(R\d{1,2}[A-Za-z])", origen.stem)
        sufijo = m.group(1).upper() if m else None
        partes = ["CPRT"] + [p for p in (aamm or None, sufijo) if p]
        destino = origen.parent / ("_".join(partes) + ".csv")
        if not m:
            self.log("  OJO: no se pudo leer la revisión del nombre del cuadro cero; "
                     f"el csv sale como {destino.name}")

        if destino.exists():
            if not messagebox.askyesno(
                    "Ya existe",
                    f"Ya hay un archivo:\n\n{destino.name}\n\n"
                    f"en {destino.parent}\n\n¿Reemplazarlo?"):
                return

        self.log("-" * 96)
        self.log(f"EXPORTAR CPRT  {datetime.now():%d-%m-%Y %H:%M:%S}")
        self.log(f"  origen  : {origen}")
        self.log(f"  destino : {destino}")
        try:
            n, con_ret = exportar_cprt(origen, destino, self.log)
        except Exception as e:
            self.log(f"  ERROR al exportar el CPRT: {e}")
            self.log(traceback.format_exc())
            messagebox.showerror("No se pudo exportar", str(e))
            return

        aviso = ""
        if con_ret:
            aviso = (f"\n\nOJO: {con_ret} fila(s) van en 0 por retención.\n"
                     "El csv lleva la columna H (Monto retenido), así que esas "
                     "empresas NO reciben pago. Confírmalo antes de mandarlo.")
        self.log(f"  Listo: {destino.name}")
        if messagebox.askyesno(
                "CPRT exportado",
                f"{n} fila(s) escritas en:\n\n{destino.name}\n\n"
                f"Carpeta: {destino.parent}{aviso}\n\n¿Abrir la carpeta?"):
            abrir_en_explorador(destino, es_archivo=True)

    def _ver_detalle(self, nid):
        top = tk.Toplevel(self.root)
        top.title("Detalle diario — " + NODO_POR_ID[nid]["texto"])
        top.geometry("640x480")
        t = tk.Text(top, font=("Consolas", 9))
        sb = tk.Scrollbar(top, command=t.yview)
        t.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        t.pack(fill="both", expand=True)
        for linea in self.detalle_diarios.get(nid, ["(sin datos: presiona ACTUALIZAR)"]):
            t.insert("end", linea + "\n")
        t.config(state="disabled")

    def estado_verificador(self, vid, _visitados=None):
        """OK | VENCIDA | NO CUADRA | NO SE PUDO | SIN VERIFICAR.
        Propaga hacia arriba: si una verificacion previa no esta OK, esta queda
        VENCIDA aunque sus propios archivos no hayan cambiado."""
        # vigente() y no get(): si la definicion de la verificacion cambio, el
        # resultado guardado no dice nada sobre el chequeo que corre ahora.
        reg = ESTADO.vigente(vid)
        if not reg:
            return "SIN VERIFICAR"
        if reg.get("resultado") == "NO SE PUDO":
            return "NO SE PUDO"
        if self._motivos_archivos(vid, reg):
            return "VENCIDA"

        _visitados = set() if _visitados is None else _visitados
        if vid in _visitados:          # corta cualquier ciclo
            return "OK"
        _visitados.add(vid)
        for prev in VERIFICADORES[vid].get("verif_previas", []):
            if self.estado_verificador(prev, _visitados) != "OK":
                return "VENCIDA"

        return "OK" if reg.get("resultado") == "OK" else "NO CUADRA"

    def _motivos_archivos(self, vid, reg):
        """Razones de archivo: algo se modifico despues de haber verificado."""
        vspec = VERIFICADORES[vid]
        fecha_v = reg.get("fecha_ts", 0)
        motivos = []
        for aid in [vspec["archivo"]] + list(vspec["depende"]):
            r = self.rutas.get(aid)
            if r is None:
                motivos.append(f"falta {aid}")
                continue
            ts_ahora = mtime(r)
            ts_guard = (reg.get("mtimes") or {}).get(aid)
            if ts_guard is None:
                motivos.append(f"{r.name}: no registrado")
            elif not iguales_mtime(ts_ahora, ts_guard):
                motivos.append(f"{r.name}: modificado después")
            elif ts_ahora > fecha_v + TOL_MTIME:
                motivos.append(f"{r.name}: posterior a la verificación")
        return motivos

    def _motivos_vencida(self, vid, reg):
        """Todos los motivos: archivos modificados y verificaciones previas que
        no esten vigentes. Se usa el estado ACTUAL de la previa, no el que tenia
        guardado, para que la cadena se explique completa."""
        motivos = self._motivos_archivos(vid, reg)
        for prev in VERIFICADORES[vid].get("verif_previas", []):
            est = self.estado_verificador(prev)
            if est != "OK":
                motivos.append(f"la previa {prev} está {est}")
        return motivos

    def _actualizar_verificaciones(self):
        for vid, vspec in VERIFICADORES.items():
            w = self.filas.get(vspec["archivo"])
            if not w or "lbl_ver" not in w:
                continue
            reg = ESTADO.vigente(vid)
            lv = w["lbl_ver"]

            if not reg:
                lv.config(text=f"{vid}: SIN VERIFICAR", fg=C_GRIS, bg=C_NEUTRO)
                continue

            marca = datetime.fromtimestamp(reg.get("fecha_ts", 0)).strftime("%d-%m %H:%M")
            est = self.estado_verificador(vid)
            if est == "OK":
                lv.config(text=f"✔ {vid} OK  {marca}", fg=C_OK, bg=C_NEUTRO)
            elif est == "VENCIDA":
                lv.config(text=f"{vid} VENCIDA ({marca})", bg=C_VENCIDA, fg="black")
                self.log(f"  VERIFICACIÓN VENCIDA  {vid}: "
                         + "; ".join(self._motivos_vencida(vid, reg)))
            elif est == "NO SE PUDO":
                lv.config(text=f"✘ {vid} NO SE PUDO  {marca}", fg=C_FALTA, bg=C_NEUTRO)
            else:
                lv.config(text=f"✘ {vid} NO CUADRA  {marca}", fg=C_FALTA, bg=C_NEUTRO)

    def _orden_con_previas(self, vids):
        """Expande la lista con las verificaciones previas que hagan falta y
        devuelve (orden_topologico, faltantes) donde faltantes son las previas
        que no estan OK."""
        orden, faltantes, visto = [], [], set()

        def visitar(vid, pila):
            if vid in orden or vid in pila:
                return
            pila.add(vid)
            for prev in VERIFICADORES[vid].get("verif_previas", []):
                est = self.estado_verificador(prev)
                if est != "OK":
                    if prev not in faltantes:
                        faltantes.append(prev)
                    visitar(prev, pila)
            pila.discard(vid)
            if vid not in orden:
                orden.append(vid)

        for vid in vids:
            visitar(vid, set())
        return orden, faltantes

    # ----------------------------------------------------------- VERIFICAR --
    def configurar_valores(self):
        """Ventana para indicar donde esta cada valor, sin editar el codigo.
        Se guarda en config.json bajo '_valores' (compartido, no por usuario)."""
        if not self.rutas:
            self.actualizar()

        top = tk.Toplevel(self.root)
        top.title("Configurar ubicación de los valores")
        top.geometry("1080x700")

        tk.Label(top, justify="left", anchor="w", font=("Segoe UI", 8), fg="#444444",
                 text="La celda puede ser una sola (H120) o un rango a sumar (H10:H500).\n"
                      "Los combos se llenan con «Cargar hojas y tablas»; también puedes "
                      "escribir a mano.\nSe guarda en config.json y aplica para todos los "
                      "meses y usuarios."
                 ).pack(fill="x", padx=12, pady=(8, 4))

        cv = tk.Canvas(top, borderwidth=0, highlightthickness=0)
        sb = tk.Scrollbar(top, orient="vertical", command=cv.yview)
        cv.configure(yscrollcommand=sb.set)
        pie = tk.Frame(top)
        pie.pack(side="bottom", fill="x", pady=8)
        sb.pack(side="right", fill="y")
        cv.pack(side="left", fill="both", expand=True)
        cont = tk.Frame(cv)
        w_id = cv.create_window((0, 0), window=cont, anchor="nw")
        cont.bind("<Configure>", lambda e: (cv.configure(scrollregion=cv.bbox("all")),
                                            cv.itemconfig(w_id, width=cv.winfo_width())))

        campos = {}   # clave -> {campo: StringVar}
        combos = {}   # clave -> {"hoja": widget} / {"tabla":.., "columna":..}

        for clave, spec in VALORES.items():
            ruta = self.rutas.get(spec["archivo"])
            f = tk.LabelFrame(cont, text=f"{clave}   —   {spec['etiqueta']}",
                              padx=8, pady=5)
            f.pack(fill="x", padx=12, pady=4)
            tk.Label(f, text=(ruta.name if ruta else "[archivo no encontrado]"),
                     font=("Consolas", 8), fg="blue" if ruta else C_FALTA,
                     anchor="w").pack(fill="x")

            fila = tk.Frame(f)
            fila.pack(fill="x", pady=(3, 0))
            campos[clave], combos[clave] = {}, {}

            def add_combo(etq, campo, ancho=26):
                tk.Label(fila, text=etq, font=("Segoe UI", 8)).pack(side="left")
                v = tk.StringVar(value=spec.get(campo, "") or "")
                c = ttk.Combobox(fila, textvariable=v, width=ancho, font=("Consolas", 8))
                c.pack(side="left", padx=(2, 12))
                campos[clave][campo] = v
                combos[clave][campo] = c

            def add_entry(etq, campo, ancho=16):
                tk.Label(fila, text=etq, font=("Segoe UI", 8)).pack(side="left")
                bruto = spec.get(campo, "")
                if isinstance(bruto, (list, tuple)):
                    bruto = ", ".join(str(x) for x in bruto)
                v = tk.StringVar(value="" if bruto in (None, "") else str(bruto))
                tk.Entry(fila, textvariable=v, width=ancho,
                         font=("Consolas", 8)).pack(side="left", padx=(2, 12))
                campos[clave][campo] = v

            if spec["tipo"] == "excel_col":
                add_combo("Hoja:", "hoja", 26)
                add_entry("Col. monto:", "columna", 6)
                add_entry("1ª fila datos:", "fila_inicio", 6)
                add_entry("Col. filtro:", "columna_filtro", 6)
                add_entry("Valores (coma):", "valores_filtro", 18)
            elif spec["tipo"] == "excel_etiqueta":
                add_combo("Hoja:", "hoja", 24)
                add_entry("Col. rótulo:", "columna_etiqueta", 5)
                add_entry("Texto de la fila:", "texto_fila", 16)
                add_entry("Col. valor:", "columna_valor", 5)
                add_entry("1ª fila:", "fila_inicio", 5)
            elif spec["tipo"] == "excel":
                add_combo("Hoja:", "hoja", 30)
                add_entry("Celda o rango:", "celda", 16)
            else:
                add_combo("Tabla:", "tabla", 24)
                add_combo("Columna monto:", "columna", 20)
                add_entry("WHERE:", "where", 26)
                add_entry("Col. tipo:", "columna_tipo", 16)

        def cargar_listas():
            btn_cargar.config(state="disabled", text="Cargando...")
            top.update_idletasks()
            cache_x, cache_m = {}, {}
            for clave, spec in VALORES.items():
                ruta = self.rutas.get(spec["archivo"])
                if not ruta:
                    continue
                try:
                    if spec["tipo"] in ("excel", "excel_col", "excel_etiqueta"):
                        if ruta not in cache_x:
                            cache_x[ruta] = obtener_hojas(ruta)
                        combos[clave]["hoja"]["values"] = cache_x[ruta]
                    else:
                        if ruta not in cache_m:
                            cache_m[ruta] = obtener_tablas_columnas(ruta)
                        tc = cache_m[ruta]
                        combos[clave]["tabla"]["values"] = list(tc.keys())
                        t = campos[clave]["tabla"].get()
                        combos[clave]["columna"]["values"] = tc.get(
                            t, sorted({c for cols in tc.values() for c in cols}))
                except Exception as e:
                    self.log(f"  No se pudo leer la estructura de {ruta.name}: {e}")
            btn_cargar.config(state="normal", text="Cargar hojas y tablas")

        def guardar():
            n = 0
            for clave, cs in campos.items():
                datos = {k: v.get().strip() for k, v in cs.items()}
                guardar_valores_cfg(clave, datos)
                aplicados = dict(datos)
                if "fila_inicio" in aplicados:
                    try:
                        aplicados["fila_inicio"] = int(aplicados["fila_inicio"])
                    except Exception:
                        aplicados.pop("fila_inicio")
                if "valores_filtro" in aplicados:
                    aplicados["valores_filtro"] = [
                        x.strip() for x in aplicados["valores_filtro"].split(",") if x.strip()]
                VALORES[clave].update(aplicados)
                n += 1
            self.log(f"  Configuración de valores guardada ({n} orígenes).")
            top.destroy()
            self._actualizar_verificaciones()

        btn_cargar = tk.Button(pie, text="Cargar hojas y tablas", command=cargar_listas)
        btn_cargar.pack(side="left", padx=12)
        tk.Button(pie, text="Guardar", bg="#1a7f1a", fg="white", width=14,
                  font=("Segoe UI", 9, "bold"), command=guardar).pack(side="right", padx=12)
        tk.Button(pie, text="Cancelar", width=12,
                  command=top.destroy).pack(side="right")

    def _informar_cache(self):
        """Dice que valores siguen vigentes y cuales habra que volver a leer."""
        if not CACHE.aamm:
            return
        vigentes, caducos = [], []
        for clave, spec in VALORES.items():
            ruta = self.rutas.get(spec["archivo"])
            if ruta is None:
                continue
            if CACHE.obtener(clave, ruta, huella_spec(spec)) is not None:
                vigentes.append(clave)
            elif clave in CACHE.data:
                caducos.append(clave)
        if vigentes:
            self.log(f"  Ya calculado y vigente ({len(vigentes)}): "
                     + ", ".join(vigentes))
        if caducos:
            self.log(f"  Se volverá a leer ({len(caducos)}): " + ", ".join(caducos))

    def ir_a_mes(self):
        """Lleva TODA la ventana al mes escrito: cambia la carpeta del caso a la
        que se usó ese mes, recarga el árbol y su estado guardado."""
        aamm = self.var_aamm.get().strip()
        if not re.fullmatch(r"\d{4}", aamm):
            messagebox.showwarning("AAMM inválido",
                                   "Escribe el mes con 4 dígitos, por ejemplo 2401.")
            return
        carpetas = self.cfg.get("carpetas_por_mes") or {}
        destino = carpetas.get(aamm)

        if not destino or not Path(destino).is_dir():
            faltante = ("La carpeta que se usó ese mes ya no existe:\n" + destino
                        if destino else
                        f"Todavía no he visto el mes {aamm} en este equipo.")
            if not messagebox.askyesno(
                    f"Mes {aamm}",
                    f"{faltante}\n\n¿Buscar ahora la carpeta 02 CASO RELIQUIDACION "
                    f"del mes {aamm}?"):
                return
            r = filedialog.askdirectory(
                title=f"Carpeta 02 CASO RELIQUIDACION del mes {aamm}",
                initialdir=self.cfg.get("carpeta_base", ""))
            if not r:
                return
            destino = r

        self.var_base.set(destino)
        self.cfg["carpeta_base"] = destino
        guardar_config({"carpeta_base": destino})
        self._color_base()
        self.log("=" * 96)
        self.log(f"Cambiando la ventana al mes {aamm}")
        self.var_aamm.set(aamm)
        self.actualizar()
        # Si los archivos de esa carpeta son de otro mes, actualizar() lo detecta
        # y avisa; el cuadro queda con el mes real.
        real = self.var_aamm.get().strip()
        if real != aamm:
            messagebox.showwarning(
                "El mes no coincide",
                f"Pediste el mes {aamm}, pero los archivos de esa carpeta son "
                f"del mes {real}.\n\nLa ventana quedó en {real}.")

    def reiniciar_mes(self):
        """Borra el estado y los valores guardados de un mes, para partir limpio."""
        aamm = self.var_aamm.get().strip()
        if not re.fullmatch(r"\d{4}", aamm):
            messagebox.showwarning("AAMM inválido",
                                   "Escribe el mes con 4 dígitos, por ejemplo 2407.")
            return
        carpeta = dir_config_mes(aamm)
        archivos = [carpeta / ARCHIVO_ESTADO, carpeta / ARCHIVO_CACHE]
        existen = [a for a in archivos if a.exists()]
        if not existen:
            messagebox.showinfo("Nada que borrar",
                                f"El mes {aamm} no tiene información guardada.\n\n"
                                f"Se buscó en:\n{carpeta}")
            return
        if not messagebox.askyesno(
                f"Reiniciar el mes {aamm}",
                "Se van a borrar las verificaciones y los valores guardados de "
                f"este mes:\n\n" + "\n".join(f"  · {a.name}" for a in existen) +
                f"\n\nEn:\n{carpeta}\n\n"
                "Los archivos de la reliquidación NO se tocan.\n"
                "Después habrá que volver a correr las verificaciones.\n\n"
                "¿Continuar?"):
            return
        borrados, fallidos = [], []
        for a in existen:
            try:
                a.unlink()
                borrados.append(a.name)
            except Exception as e:
                fallidos.append(f"{a.name}: {e}")
        ESTADO.cargar(aamm)
        CACHE.cargar(aamm)
        CACHE_COLUMNAS.clear()
        for w in self.filas.values():
            if "lbl_ver" in w:
                w["lbl_ver"].config(text="SIN VERIFICAR", fg=C_GRIS, bg=C_NEUTRO)
        self.log(f"  Mes {aamm} reiniciado: borrado {', '.join(borrados)}")
        for f in fallidos:
            self.log(f"  ! No se pudo borrar {f}")
        self._actualizar_verificaciones()
        self.var_estado.set(f"Mes {aamm} reiniciado")
        if fallidos:
            messagebox.showerror("Quedó algo sin borrar", "\n".join(fallidos))

    def _ventana_texto(self, titulo, lineas, ancho="900x620"):
        top = tk.Toplevel(self.root)
        top.title(titulo)
        top.geometry(ancho)
        t = tk.Text(top, font=("Consolas", 9), wrap="none")
        configurar_tags_log(t)
        sy = tk.Scrollbar(top, command=t.yview)
        sx = tk.Scrollbar(top, orient="horizontal", command=t.xview)
        t.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        sx.pack(side="bottom", fill="x")
        sy.pack(side="right", fill="y")
        t.pack(fill="both", expand=True)
        insertar_con_color(t, lineas)
        t.config(state="disabled")
        return top

    def ver_detalle_verificacion(self, vid):
        """Totales, comprobaciones y bitácora de la última corrida."""
        v = VERIFICADORES[vid]
        reg = ESTADO.vigente(vid)
        out = [f"{vid}   {v['titulo']}", "=" * 92]
        if not reg:
            out += ["", "Este verificador todavía no se ha corrido en el mes "
                        f"{ESTADO.aamm or '(sin mes)'}.",
                    "Presiona «Verificar» para generarlo."]
            self._ventana_texto(f"Detalle {vid}", out)
            return

        out += [f"Mes            : {ESTADO.aamm}",
                f"Resultado      : {reg.get('resultado')}",
                f"Estado actual  : {self.estado_verificador(vid)}",
                f"Verificado el  : {reg.get('fecha')}",
                f"Por            : {reg.get('usuario')}", ""]

        # Se avisa aca y no en la fila: en la fila daba lo mismo y estorbaba,
        # pero aca abajo se listan hojas y rangos concretos, y si la definicion
        # cambio esos rangos pueden no ser los que se leen hoy.
        if ESTADO.firma_guardada_distinta(vid):
            out += ["Nota: este resultado se generó con una definición anterior "
                    "de este verificador,",
                    "así que las hojas y rangos de abajo pueden no ser los que se "
                    "leen ahora.",
                    "Para tenerlo con las comprobaciones actuales, aprieta "
                    "«Verificar».", ""]

        motivos = self._motivos_vencida(vid, reg)
        if motivos:
            out += ["Por qué ya no está vigente:"] + [f"  - {m}" for m in motivos] + [""]

        previas = reg.get("previas") or {}
        if previas:
            out += ["Verificaciones previas al momento de correr esta:"]
            for k, e in previas.items():
                out.append(f"  {k:<5} {e:<14} {VERIFICADORES.get(k, {}).get('titulo', '')[:60]}")
            out.append("")

        comps = reg.get("comprobaciones") or []
        if comps:
            out += ["Comprobaciones:", "-" * 92]
            for c in comps:
                simbolo = {"OK": "OK  ", "NO CUADRA": ">>> ", "SIN DATOS": " ?  ",
                           "DESACTIVADA": "--  "}.get(c.get("estado"), "    ")
                out.append(f"{simbolo}{c.get('desc')}"
                           + ("   (en valor absoluto)" if c.get("absoluto") else ""))
                if c["tipo"] == "igualdad" and c.get("izquierda") is not None:
                    for lado, lista in (("izq", c.get("izq_claves") or []),
                                        ("der", c.get("der_claves") or [])):
                        for k in lista:
                            bk, sg = partir_signo(k)
                            etq = VALORES.get(bk, {}).get("etiqueta", bk)
                            if sg < 0:
                                etq = "(con signo cambiado) " + etq
                            val = (reg.get("valores") or {}).get(bk)
                            if val is not None:
                                val = val * sg
                            out.append(f"        {lado}  {etq:<52}{fmt_monto(val):>20}")
                    out.append(f"        {'total izquierda':<57}{fmt_monto(c['izquierda']):>20}")
                    out.append(f"        {'total derecha':<57}{fmt_monto(c['derecha']):>20}")
                    out.append(f"        {'diferencia':<57}{fmt_monto(c['diferencia']):>20}")
                elif c["tipo"] == "cero":
                    for k, val in (c.get("valores") or {}).items():
                        marca = "  <-- no es 0" if k in (c.get("fuera") or []) else ""
                        etq = VALORES.get(partir_signo(k)[0], {}).get(
                            "etiqueta", partir_signo(k)[0])
                        out.append(f"        {etq:<52}{fmt_monto(val):>20}{marca}")
                elif c["tipo"] == "tabla":
                    out.append(f"        {c.get('filas_a')} fila(s) en un lado, "
                               f"{c.get('filas_b')} en el otro")
                    for e in c.get("solo_a") or []:
                        out.append(f"        solo en el primero : {e}")
                    for e in c.get("solo_b") or []:
                        out.append(f"        solo en el segundo : {e}")
                    for emp, i, va, vb in c.get("difs") or []:
                        out.append(f"        {emp[:30]:<32} col {i + 2}  "
                                   f"{fmt_monto(va):>18} vs {fmt_monto(vb):>18}")
                    if c.get("n_difs", 0) > len(c.get("difs") or []):
                        out.append(f"        ... y {c['n_difs'] - len(c['difs'])} "
                                   "diferencia(s) más")
                    for et, n, f in c.get("duplicadas") or []:
                        out.append(f"        REPETIDA en {et}: {n} (fila {f})")
                elif c["tipo"] == "formulas_cubren":
                    out.append(f"        {c.get('n_empresas')} empresa(s); "
                               f"las fórmulas deberían llegar a la fila "
                               f"{c.get('esperada')}")
                    for col, ult, n in c.get("faltan") or []:
                        out.append(f"        {col}: llega a {ult}, FALTAN {n} "
                                   "fila(s) (empresas sin calcular)")
                    for col, ult, n in c.get("sobran") or []:
                        out.append(f"        {col}: llega a {ult}, SOBRAN {n} "
                                   "fila(s) (arrastra 0 o vacío)")
                    if c.get("sin_formula"):
                        out.append("        sin ninguna fórmula: "
                                   + ", ".join(c["sin_formula"]))
                elif c["tipo"] == "suma_por_empresa":
                    out.append(f"        {c.get('n_detalle')} empresa(s) en el "
                               f"detalle, {c.get('n_resumen')} en el resumen")
                    for n in c.get("solo_detalle") or []:
                        out.append(f"        solo en el detalle: {n}")
                    for n in c.get("solo_resumen") or []:
                        out.append(f"        solo en el resumen: {n}")
                    for n in c.get("duplicadas") or []:
                        out.append(f"        REPETIDA en el resumen: {n}")
                    for nom, suma, total, dif, nfilas in c.get("difs") or []:
                        out.append(f"        {nom[:26]:<28} {fmt_monto(suma):>16} vs "
                                   f"{fmt_monto(total):>16}  dif {fmt_monto(dif):>14}"
                                   f"  ({nfilas} fila(s))")
                    if c.get("n_difs", 0) > len(c.get("difs") or []):
                        out.append(f"        ... y {c['n_difs'] - len(c['difs'])} "
                                   "empresa(s) descuadrada(s) más")
                elif c["tipo"] == "pago_por_empresa":
                    out.append(f"        {c.get('n_ok')} par(es) cuadran, "
                               f"{c.get('n_difs')} con diferencia")
                    if c.get("n_signo"):
                        out.append(f"        {c['n_signo']} con signo opuesto "
                                   "(convención de la planilla 1)")
                    if c.get("comunes"):
                        out.append("        conceptos comparados: "
                                   + ", ".join(c["comunes"]))
                    if c.get("solo_resumen"):
                        out.append("        solo en la planilla 1: "
                                   + ", ".join(c["solo_resumen"]))
                    if c.get("solo_detalle"):
                        out.append("        solo en el detalle: "
                                   + ", ".join(c["solo_detalle"]))
                    for etq, vd, vr, d in c.get("difs") or []:
                        out.append(f"        {etq}: {fmt_monto(vd)} vs "
                                   f"{fmt_monto(vr)}   dif {fmt_monto(d)}")
                    for x in c.get("faltan_resumen") or []:
                        out.append(f"        falta en la planilla 1: {x}")
                    for x in c.get("faltan_detalle") or []:
                        out.append(f"        falta en el detalle: {x}")
                elif c["tipo"] == "bloque_contra_origen":
                    out.append(f"        bloque {c.get('bloque')}: "
                               f"destino {c.get('n_destino')} filas / "
                               f"{fmt_monto(c.get('suma_destino'))}")
                    out.append(f"        origen : {c.get('n_origen')} filas / "
                               f"{fmt_monto(c.get('suma_origen'))}")
                    out.append(f"        diferencia: "
                               f"{fmt_monto(c.get('diferencia'))}")
                    for k in c.get("solo_origen") or []:
                        out.append(f"        solo en el origen: {k}")
                    for k in c.get("solo_destino") or []:
                        out.append(f"        solo en el destino: {k}")
                elif c["tipo"] == "retencion_coherente":
                    out.append(f"        {c.get('n_filas')} filas: "
                               f"{c.get('n_iguales')} con H=G, "
                               f"{c.get('n_retenidas')} retenidas")
                    if c.get("total_retenido"):
                        out.append(f"        retenido: "
                                   f"{fmt_monto(c['total_retenido'])}")
                    for f, g, h in c.get("raras") or []:
                        out.append(f"        fila {f}: G={fmt_monto(g)} "
                                   f"H={fmt_monto(h)}  (ni igual ni 0)")
                    for f in c.get("sin_h") or []:
                        out.append(f"        fila {f}: sin valor en H")
                    for f, g in c.get("retenidas") or []:
                        out.append(f"        fila {f}: retenido {fmt_monto(g)}")
                elif c["tipo"] == "prorrata_al_dia":
                    out.append(f"        origen  : {c.get('n_origen')} "
                               f"suministradores / {c.get('horas_origen')} horas")
                    out.append(f"        planilla: {c.get('n_destino')} "
                               f"suministradores / {c.get('horas_destino')} horas")
                    out.append(f"        {c.get('n_ok')} cuadran")
                    for k in c.get("solo_origen") or []:
                        out.append(f"        falta en la planilla: {k}")
                    for k in c.get("solo_destino") or []:
                        out.append(f"        columna que no es del origen: {k}")
                    if c.get("sobran_en_cero"):
                        out.append(f"        {c['sobran_en_cero']} columna(s) en 0 "
                                   "que no están en el origen (normal)")
                    for k, vd, vo, d in c.get("difs") or []:
                        out.append(f"        {k}: planilla {vd:.6f} vs "
                                   f"origen {vo:.6f}")
                elif c["tipo"] == "suma_fila":
                    out.append(f"        {c.get('n_filas')} fila(s)")
                    for esc, n in sorted((c.get("escalas") or {}).items()):
                        out.append(f"        {n} fila(s) suman {float(esc):g}")
                    if c.get("mezcla"):
                        out.append("        MEZCLA filas que suman 1 con filas que "
                                   "suman 100")
                    for fila_mala in c.get("malas") or []:
                        f, t = fila_mala[0], fila_mala[1]
                        bl = fila_mala[2] if len(fila_mala) > 2 else "?"
                        out.append(f"        fila {f} ({bl}): suma {t}")
                    if c.get("por_bloque"):
                        out.append("        filas malas por bloque: "
                                   + ", ".join(f"{k}: {v}" for k, v in
                                               sorted(c["por_bloque"].items())))
                    if c.get("ceros"):
                        out.append(f"        {len(c['ceros'])} fila(s) suman 0 "
                                   "(permitido): "
                                   + ", ".join(str(f) for f in c["ceros"][:20]))
                elif c["tipo"] == "centrales_en_lista":
                    out.append(f"        lista de {c.get('n_lista')}; "
                               f"{c.get('n_vistas')} vistas, "
                               f"{c.get('n_fuera')} fuera de la lista")
                    for x in c.get("fuera") or []:
                        out.append(f"        no es embalse: {x}")
                    for x in c.get("sufijo") or []:
                        out.append(f"        «-número» fuera de la lista: {x}")
                    for x in c.get("faltan") or []:
                        out.append(f"        no apareció: {x}")
                elif c["tipo"] == "sobrecosto_por_fila":
                    out.append(f"        {c.get('n_filas')} fila(s), "
                               f"{c.get('n_ok')} cuadran fila a fila")
                    out.append(f"        recalculado: {fmt_monto(c.get('suma_calculada'))}")
                    out.append(f"        columna E  : {fmt_monto(c.get('suma_esperada'))}")
                    out.append(f"        diferencia : {fmt_monto(c.get('diferencia'))}")
                    for f, ca, es, d in c.get("difs") or []:
                        out.append(f"        fila {f}: recalc {fmt_monto(ca)} vs "
                                   f"{fmt_monto(es)}  dif {fmt_monto(d)}")
                    if c.get("n_difs", 0) > len(c.get("difs") or []):
                        out.append(f"        ... y {c['n_difs'] - len(c['difs'])} "
                                   "fila(s) más")
                    for f, e, x in c.get("incompletas") or []:
                        out.append(f"        fila {f}: falta {x}"
                                   + (f" (trae {fmt_monto(e)})" if e is not None else ""))
                elif c["tipo"] == "matriz_al_dia":
                    out.append(f"        matriz: {c.get('n_pagan')} pagan x "
                               f"{c.get('n_reciben')} reciben")
                    for et, k in (("PAGAN falta", "faltan_pagan"),
                                  ("PAGAN sobra", "sobran_pagan"),
                                  ("RECIBEN falta", "faltan_reciben"),
                                  ("RECIBEN sobra", "sobran_reciben")):
                        for n in c.get(k) or []:
                            out.append(f"        {et}: {n}")
                    if c.get("aviso_rango"):
                        out.append(f"        {c['aviso_rango']}")
                elif c["tipo"] == "cprt_al_dia":
                    out.append(f"        CPRT: {c.get('n_cprt')} pares | "
                               f"matriz: {c.get('n_matriz')} pares")
                    for n in c.get("fantasmas") or []:
                        out.append(f"        en el CPRT y no en la matriz: {n}")
                    for n in c.get("perdidos") or []:
                        out.append(f"        en la matriz y no en el CPRT: {n}")
                    for a, b, x, y in c.get("difs") or []:
                        out.append(f"        {a} -> {b}: CPRT {fmt_monto(x)} "
                                   f"vs matriz {fmt_monto(y)}")
                elif c["tipo"] == "mismas_empresas":
                    out.append(f"        {c.get('n_a')} en {c.get('nombre_a')}, "
                               f"{c.get('n_b')} en {c.get('nombre_b')}")
                    for n in c.get("solo_a") or []:
                        out.append(f"        FALTA en {c.get('nombre_b')}: {n}")
                    for n in c.get("solo_b") or []:
                        out.append(f"        SOBRA en {c.get('nombre_b')}: {n}")
                    for et, n, f in c.get("duplicadas") or []:
                        out.append(f"        REPETIDA en {et}: {n} (fila {f})")
                elif c["tipo"] == "pertenencia":
                    out.append(f"        {c.get('n_origen')} en origen, "
                               f"{c.get('n_destino')} en destino")
                    for e in c.get("faltan") or []:
                        out.append(f"        no encontrada: {e}")
                    for e in c.get("duplicadas") or []:
                        out.append(f"        REPETIDA en el destino: {e}")
                elif c["tipo"] == "ultimo_igual":
                    out.append(f"        {c.get('a')} = {str(c.get('valor_a'))[:40]!r}")
                    out.append(f"        {c.get('b')} = {str(c.get('valor_b'))[:40]!r}")
                elif c["tipo"] == "marcas":
                    if not c.get("conteo"):
                        out.append("        sin errores ni textos prohibidos")
                    for motivo, n in (c.get("conteo") or {}).items():
                        out.append(f"        {n} celda(s) con {motivo}")
                    for celda, motivo, val in (c.get("marcas") or []):
                        out.append(f"          {celda:<9} {motivo:<18} {val}")
                out.append("")

        if reg.get("errores"):
            out += ["Problemas al leer los datos:"]
            out += [f"  - {e}" for e in reg["errores"]] + [""]

        mt = reg.get("mtimes") or {}
        if mt:
            out += ["Fechas de modificación al momento de verificar:", "-" * 92]
            for aid, ts in mt.items():
                nodo = NODO_POR_ID.get(aid, {})
                actual = mtime(self.rutas.get(aid)) if self.rutas.get(aid) else None
                estado = "sin cambios" if iguales_mtime(ts, actual) else ">>> CAMBIÓ DESPUÉS"
                out.append(f"  {nodo.get('texto', aid)[:44]:<46}{fmt_fecha(ts)}   {estado}")
            out.append("")

        out += ["Bitácora de la corrida:", "-" * 92] + (reg.get("log") or ["(no guardada)"])
        self._ventana_texto(f"Detalle {vid} — {reg.get('resultado')}", out)

    def ver_estado_mes(self):
        """Abre el estado guardado del AAMM escrito en el cuadro."""
        aamm = self.var_aamm.get().strip()
        if not re.fullmatch(r"\d{4}", aamm):
            messagebox.showwarning("AAMM inválido",
                                   "Escribe el mes con 4 dígitos, por ejemplo 2407.")
            return
        carpeta = dir_config_mes(aamm)
        data = leer_estado_mes(aamm)
        if data is None:
            messagebox.showinfo(
                "No existe",
                f"No existe estado guardado para el mes {aamm}.\n\n"
                f"Se buscó en:\n{carpeta / ARCHIVO_ESTADO}\n\n"
                "Se creará la primera vez que corras una verificación de ese mes.")
            self.log(f"  No existe estado para el mes {aamm}.")
            return

        out = [f"Estado guardado del mes {aamm}",
               f"{carpeta / ARCHIVO_ESTADO}", "=" * 92, ""]
        if not data:
            out.append("El archivo existe pero está vacío o ilegible.")
        for vid in VERIFICADORES:
            reg = data.get(vid)
            titulo = VERIFICADORES[vid]["titulo"]
            if not reg:
                out += [f"{vid}  SIN VERIFICAR", f"      {titulo}", ""]
                continue
            out += [f"{vid}  {reg.get('resultado')}   ({reg.get('fecha')}"
                    f"  por {reg.get('usuario')})",
                    f"      {titulo}"]
            if reg.get("izquierda") is not None:
                out.append(f"      izq {fmt_monto(reg['izquierda'])}   "
                           f"der {fmt_monto(reg['derecha'])}   "
                           f"dif {fmt_monto(reg['diferencia'])}")
            for c in reg.get("comprobaciones") or []:
                simbolo = {"OK": "ok ", "NO CUADRA": ">> ",
                           "SIN DATOS": " ? "}.get(c.get("estado"), "   ")
                linea = f"      {simbolo}{c.get('desc')}"
                if c.get("diferencia") is not None:
                    linea += f"   (dif {fmt_monto(c['diferencia'])})"
                out.append(linea)
            for e in reg.get("errores") or []:
                out.append(f"      ! {e}")
            out.append("")
        extra = [k for k in data if k not in VERIFICADORES]
        if extra:
            out += ["Verificadores guardados que ya no existen en el script: "
                    + ", ".join(extra), ""]
        self._ventana_texto(f"Estado del mes {aamm}", out)

    def verificar(self, vid):
        self._lanzar([vid])

    def verificar_todo(self):
        self._lanzar(list(VERIFICADORES.keys()))

    def _lanzar(self, vids, preguntar_previas=True):
        if self.trabajando:
            return

        # Se relee el disco ANTES de nada. Va primero que el calculo de las
        # previas a proposito: si un archivo cambio desde la ultima vez, recien
        # aca las verificaciones que dependian de el pasan a VENCIDA, y entonces
        # el aviso de "faltan previas" se arma con la realidad de ahora y no con
        # una foto vieja.
        etiqueta = vids[0] if len(vids) == 1 else f"{len(vids)} verificaciones"
        # Solo se relee lo que estas verificaciones leen. Verificar todo termina
        # pidiendo casi todo, y ahi la relectura completa sale igual de rapida.
        necesarios = ids_de_verificaciones(vids)
        if not self.actualizar(motivo=f"previo a verificar {etiqueta}",
                               solo_ids=necesarios):
            return

        # --- verificaciones previas: si faltan, ofrecer correrlas antes
        if preguntar_previas:
            orden, faltantes = self._orden_con_previas(vids)
            if faltantes:
                pedidos = [v for v in vids]
                detalle = "\n".join(
                    f"   · {f}  ({self.estado_verificador(f)})\n"
                    f"       {VERIFICADORES[f]['titulo']}" for f in faltantes)
                r = messagebox.askyesnocancel(
                    "Faltan verificaciones previas",
                    f"{', '.join(pedidos)} depende de verificaciones que no están "
                    f"al día:\n\n{detalle}\n\n"
                    "¿Correr primero las que faltan y después las que pediste?\n\n"
                    "Sí = correr todo en orden\n"
                    "No = correr solo lo que pediste\n"
                    "Cancelar = no hacer nada")
                if r is None:
                    return
                if r:
                    vids = orden
                    self.log("  Orden de verificación: " + " -> ".join(vids))
                else:
                    self.log("  Se corre sin las verificaciones previas: "
                             + ", ".join(faltantes))
        if not ESTADO.aamm:
            if not messagebox.askyesno(
                    "Sin mes definido",
                    "No hay un AAMM definido, así que el resultado no se va a "
                    "guardar en 00_Salidas.\n\n¿Verificar igual?"):
                return
        self.trabajando = True
        self._bloquear(True)
        self.timer["on"] = True
        self.timer["t0"] = time.time()
        self._tick()
        self.progress.config(maximum=len(vids), value=0)
        self.var_estado.set("Verificando...")
        threading.Thread(target=self._worker, args=(vids,), daemon=True).start()

    def _worker(self, vids):
        usar_cache = not self.var_releer.get()
        if not usar_cache:
            CACHE_COLUMNAS.clear()
            CACHE.descartar()
            self.cola.put(("log", "  (releyendo todo: se descartó lo ya calculado)"))
        try:
            try:
                import pythoncom
                pythoncom.CoInitialize()
            except Exception:
                pass
            for i, vid in enumerate(vids, start=1):
                try:
                    self._correr_verificacion(vid, usar_cache)
                except Exception as e:
                    self.cola.put(("log", f"  ERROR en {vid}: {e}"))
                    self.cola.put(("log", traceback.format_exc()))
                self.cola.put(("prog", i))
        finally:
            try:
                import pythoncom
                pythoncom.CoUninitialize()
            except Exception:
                pass
            self.cola.put(("fin", None))

    def _correr_verificacion(self, vid, usar_cache=True):
        v = VERIFICADORES[vid]
        lineas = []

        def L(m=""):
            lineas.append(str(m))
            self.cola.put(("log", m))

        L("-" * 96)
        L(f"{vid}  {v['titulo']}")
        ts_inicio = time.time()

        # --- estado de los verificadores previos (informativo en el registro)
        previas = {}
        for prev in v.get("verif_previas", []):
            previas[prev] = self.estado_verificador(prev)
        pendientes = [k for k, e in previas.items() if e != "OK"]
        if pendientes:
            L("  Aviso: verificaciones previas no vigentes -> "
              + ", ".join(f"{k}: {previas[k]}" for k in pendientes))

        # --- leer todos los valores que se necesitan
        claves = []
        for c in v["comprobaciones"]:
            if not c.get("activa", True):
                continue
            if c["tipo"] == "igualdad":
                claves += [partir_signo(k)[0] for k in c["izq"] + c["der"]]
            elif c["tipo"] == "cero":
                claves += [partir_signo(k)[0] for k in c["claves"]]
            elif c["tipo"] == "umbral":
                claves.append(c["clave"])
        valores, errores = {}, []
        for clave in claves:
            if clave in valores:
                continue
            val, err = obtener_valor(clave, self.rutas, L, usar_cache=usar_cache)
            valores[clave] = val
            if err:
                errores.append(err)


        # --- correr las comprobaciones
        ok_total = True
        detalle = []
        for c in v["comprobaciones"]:
            if not c.get("activa", True):
                L(f"  -- {c.get('desc')}   (desactivada)")
                detalle.append({"tipo": c["tipo"], "desc": c.get("desc"),
                                "estado": "DESACTIVADA"})
                continue
            r = self._comprobar(c, valores, L)
            detalle.append(r)
            if r["estado"] not in ("OK", "DESACTIVADA"):
                ok_total = False

        if any(r["estado"] == "SIN DATOS" for r in detalle) or errores:
            for e in errores:
                L(f"  ! {e}")
            resultado = "NO SE PUDO"
        else:
            resultado = "OK" if ok_total else "NO CUADRA"

        L(f"  RESULTADO {vid}: " + {"OK": "CUADRA", "NO CUADRA": "NO CUADRA",
                                    "NO SE PUDO": "NO SE PUDO VERIFICAR"}[resultado])

        mtimes = {}
        for aid in [v["archivo"]] + list(v["depende"]):
            r = self.rutas.get(aid)
            mtimes[aid] = mtime(r) if r else None

        registro = {
            "fecha_ts": ts_inicio,
            "fecha": datetime.fromtimestamp(ts_inicio).strftime("%Y-%m-%d %H:%M:%S"),
            "usuario": get_usuario(),
            "resultado": resultado,
            "errores": errores,
            "valores": valores,
            "mtimes": mtimes,
            "comprobaciones": detalle,
            "previas": previas,
            "log": lineas,
        }
        self.cola.put(("ver_guardar", (vid, registro)))

    def _comprobar(self, c, valores, L):
        """Corre una comprobacion y devuelve un dict con su resultado."""
        tipo = c["tipo"]
        base = {"tipo": tipo, "desc": c.get("desc", tipo)}

        if tipo == "igualdad":
            faltan = [k for k in c["izq"] + c["der"]
                      if valores.get(partir_signo(k)[0]) is None]
            if faltan:
                L(f"  ? {c['desc']}: sin datos ({', '.join(faltan)})")
                return dict(base, estado="SIN DATOS")
            izq = sum(valores[k] * sg for k, sg in map(partir_signo, c["izq"]))
            der = sum(valores[k] * sg for k, sg in map(partir_signo, c["der"]))
            absoluto = bool(c.get("absoluto"))
            if absoluto:
                izq, der = abs(izq), abs(der)
            dif = izq - der
            ok = abs(dif) <= TOLERANCIA
            L(f"  {'OK ' if ok else '>> '}{c['desc']}"
              + ("   (en valor absoluto)" if absoluto else ""))
            L(f"        {fmt_monto(izq)}  vs  {fmt_monto(der)}"
              f"   dif {fmt_monto(dif)}")
            # Si al cambiarle el signo a un lado la diferencia se hace mucho mas
            # chica, lo que falla es el signo y no los montos. Se avisa para no
            # dejar a la vista un descuadre gigante que en realidad no lo es.
            suma = izq + der
            if not ok and not absoluto and abs(suma) < abs(dif) / 2:
                L(f"        OJO: parece un problema de signo. Con un lado "
                  f"invertido la diferencia sería {fmt_monto(suma)}.")
                if abs(suma) <= TOLERANCIA:
                    L("        Los dos montos son iguales y de signo opuesto: "
                      "seguramente falta un '-' delante de una clave.")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        izquierda=izq, derecha=der, diferencia=dif,
                        dif_invertida=suma, absoluto=absoluto,
                        izq_claves=list(c["izq"]), der_claves=list(c["der"]))

        if tipo == "umbral":
            # |valor| <= maximo. Para residuos de redondeo, donde exigir 0 seria
            # irreal pero un numero grande delata un problema de verdad.
            clave = c["clave"]
            val = valores.get(clave)
            if val is None:
                L(f"  ? {c['desc']}: sin datos ({clave})")
                return dict(base, estado="SIN DATOS")
            tope = c.get("maximo", UMBRAL_DESCUADRE_CPRT)
            ok = abs(val) <= tope
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        {VALORES[clave]['etiqueta']}")
            L(f"        vale {fmt_monto(val)}   (máximo aceptado {fmt_monto(tope)})")
            if not ok:
                L("        Un descuadre grande suele ser que el Cuadro de pagos se armó")
                L("        con una tabla distinta de la que quedó, o que faltó apretar")
                L("        «Actualiza Rango» antes de refrescar la tabla dinámica.")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        valor=val, maximo=tope)

        if tipo == "cero":
            faltan = [k for k in c["claves"]
                      if valores.get(partir_signo(k)[0]) is None]
            if faltan:
                L(f"  ? {c['desc']}: sin datos ({', '.join(faltan)})")
                return dict(base, estado="SIN DATOS")
            malos = {}
            for k in c["claves"]:
                bk, sg = partir_signo(k)
                if abs(valores[bk] * sg) > TOLERANCIA:
                    malos[k] = valores[bk] * sg
            ok = not malos
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            for k, val in malos.items():
                L(f"        {VALORES[partir_signo(k)[0]]['etiqueta']} debería ser 0 "
                  f"y vale {fmt_monto(val)}")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        valores={k: valores[partir_signo(k)[0]] * partir_signo(k)[1]
                                 for k in c["claves"]},
                        fuera=list(malos))

        if tipo == "marcas":
            ruta = self.rutas.get(c["archivo"])
            if ruta is None:
                L(f"  ? {c['desc']}: falta el archivo")
                return dict(base, estado="SIN DATOS")
            res = buscar_marcas_rapido(ruta, c["hoja"], c["fila_inicio"],
                                       c["reglas"], L)
            if res is None:
                return dict(base, estado="SIN DATOS")
            ok = not res["conteo"]
            rangos = ", ".join(r for regla in c["reglas"] for r in regla["rangos"])
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        {rangos}, desde la fila {c['fila_inicio']}")
            if ok:
                L("        sin errores ni textos prohibidos")
            for motivo, n in res["conteo"].items():
                L(f"        {n} celda(s) con {motivo}")
            for celda, motivo, val in res["marcas"]:
                L(f"          {celda:<9} {motivo:<18} {val}")
            total = sum(res["conteo"].values())
            if total > len(res["marcas"]):
                L(f"          ... y {total - len(res['marcas'])} más")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        conteo=res["conteo"], marcas=res["marcas"])

        if tipo == "tabla":
            ra = self.rutas.get(c["archivo_a"])
            rb = self.rutas.get(c["archivo_b"])
            if ra is None or rb is None:
                L(f"  ? {c['desc']}: falta uno de los archivos")
                return dict(base, estado="SIN DATOS")
            L(f"  .. {c['desc']}")
            da = leer_columnas_rapido(ra, c["hoja_a"], c["cols_a"], c["fila_a"], L)
            db = leer_columnas_rapido(rb, c["hoja_b"], c["cols_b"], c["fila_b"], L)
            if da is None or db is None:
                return dict(base, estado="SIN DATOS")
            info_a, info_b = {}, {}
            ta = armar_tabla(da, c["cols_a"][0], c["cols_a"][1:], L,
                             c.get("nombre_a", "A"), info=info_a)
            tb = armar_tabla(db, c["cols_b"][0], c["cols_b"][1:], L,
                             c.get("nombre_b", "B"), info=info_b)
            solo_a = [ta[k][0] for k in ta if k not in tb]
            solo_b = [tb[k][0] for k in tb if k not in ta]
            difs = []
            for k in ta:
                if k not in tb:
                    continue
                for i, (va, vb) in enumerate(zip(ta[k][1], tb[k][1])):
                    na = va if isinstance(va, (int, float)) else None
                    nb = vb if isinstance(vb, (int, float)) else None
                    if na is None and nb is None:
                        continue
                    if na is None or nb is None or abs(na - nb) > TOLERANCIA:
                        difs.append((ta[k][0], i, va, vb))
            # Las empresas repetidas son un fallo, no un aviso: la tabla se pega
            # copiada y una empresa dos veces significa que a alguien se le paga
            # o se le cobra dos veces.
            dups = []
            if c.get("sin_duplicados"):
                dups = ([(c.get("nombre_a", "A"), n, f) for n, f in info_a["duplicadas"]]
                        + [(c.get("nombre_b", "B"), n, f) for n, f in info_b["duplicadas"]])
            ok = not (solo_a or solo_b or difs or dups)
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            for et, n, f in dups[:10]:
                L(f"        REPETIDA en {et}: {n[:34]} (fila {f})")
            if len(dups) > 10:
                L(f"        ... y {len(dups) - 10} repetición(es) más")
            if solo_a:
                L(f"        {len(solo_a)} empresa(s) solo en {c.get('nombre_a')}: "
                  f"{', '.join(sorted(solo_a)[:10])}")
            if solo_b:
                L(f"        {len(solo_b)} empresa(s) solo en {c.get('nombre_b')}: "
                  f"{', '.join(sorted(solo_b)[:10])}")
            if difs:
                nombres = c.get("nombres_valor") or ["valor 1", "valor 2"]
                L(f"        {len(difs)} diferencia(s) de monto:")
                for emp, i, va, vb in difs[:20]:
                    et = nombres[i] if i < len(nombres) else f"col {i + 2}"
                    dif = (va - vb) if isinstance(va, (int, float)) and isinstance(vb, (int, float)) else None
                    L(f"          {emp[:28]:<30} {et:<12} {fmt_monto(va):>18} vs "
                      f"{fmt_monto(vb):>18}   dif {fmt_monto(dif)}")
                if len(difs) > 20:
                    L(f"          ... y {len(difs) - 20} más")
            if ok:
                L(f"        las {len(ta)} empresas y sus montos coinciden")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        solo_a=solo_a, solo_b=solo_b, n_difs=len(difs),
                        difs=[(e, i, va, vb) for e, i, va, vb in difs[:40]],
                        duplicadas=[(et, n, f) for et, n, f in dups[:40]],
                        filas_a=len(ta), filas_b=len(tb))

        if tipo == "formulas_cubren":
            # Las formulas de unas columnas tienen que llegar exactamente hasta
            # la ultima empresa: ni cortarse antes (empresa sin calcular) ni
            # seguir despues (arrastra 0 o vacio y ensucia los totales).
            ruta = self.rutas.get(c["archivo"])
            ref = c["referencia"]
            ruta_ref = self.rutas.get(ref["archivo"])
            if ruta is None or ruta_ref is None:
                L(f"  ? {c['desc']}: falta uno de los archivos")
                return dict(base, estado="SIN DATOS")
            cols = []
            for r in c["cols"]:
                cols.extend(expandir_columnas(r))
            f_ini = int(c["fila_inicio"])
            f_ini_ref = int(ref.get("fila_inicio", 1))

            datos_ref = leer_columnas_rapido(ruta_ref, ref["hoja"], [ref["col"]],
                                             f_ini_ref, L)
            if datos_ref is None:
                return dict(base, estado="SIN DATOS")
            # Ultima fila con una empresa de verdad en la columna de referencia.
            filas_emp = [f for f, v in datos_ref.get(ref["col"], {}).items()
                         if isinstance(v, str) and v.strip()
                         and normalizar(v) not in ("", "0")
                         and not v.startswith("#")]
            if not filas_emp:
                L(f"  ? {c['desc']}: no hay empresas en "
                  f"{ref['col']}{f_ini_ref} hacia abajo")
                return dict(base, estado="SIN DATOS")
            ultima_ref = max(filas_emp)
            n_emp = len(filas_emp)
            # Las filas pueden estar corridas entre hojas (K desde 5, A desde 9).
            desfase = f_ini - f_ini_ref
            esperada = ultima_ref + desfase

            formulas = leer_formulas_rapido(ruta, c["hoja"], cols, f_ini, L)
            if formulas is None:
                return dict(base, estado="SIN DATOS")
            # Tambien se mira hasta donde hay VALOR, no solo formula. Motivos:
            #  - un 0 arrastrado por una formula que sobra es un valor, y hay que
            #    cazarlo aunque la formula este "bien" puesta;
            #  - si la columna trae valores pegados a mano en vez de formulas,
            #    igual hay que revisar que no sobren filas.
            valores_col = leer_columnas_rapido(ruta, c["hoja"], cols, f_ini, L)
            if valores_col is None:
                valores_col = {}

            faltan, sobran, sin_formula = [], [], []
            for col in cols:
                fs = formulas.get(col, set())
                vs = set(valores_col.get(col, {}))
                if not fs:
                    sin_formula.append(col)
                alcance = max(fs | vs) if (fs or vs) else None
                if alcance is None:
                    faltan.append((col, 0, esperada - f_ini + 1))
                    continue
                if alcance < esperada:
                    faltan.append((col, alcance, esperada - alcance))
                elif alcance > esperada:
                    sobran.append((col, alcance, alcance - esperada))
            # Que una columna traiga valores pegados en vez de formulas no es un
            # error por si mismo: se avisa, pero lo que decide es el alcance.
            # Con solo_faltan, que las formulas sigan mas abajo NO es un error.
            # Hace falta para la H del CPRT: ahi el bloque de datos lo genera una
            # tabla dinamica que crece y se encoge, la formula esta arrastrada bien
            # abajo a proposito, y el exportador corta por la columna A. Lo unico
            # que importa es que no se quede corta.
            if c.get("solo_faltan"):
                sobran = []
            ok = not (faltan or sobran)
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        {n_emp} empresa(s) en {ref['col']}{f_ini_ref}:"
              f"{ref['col']}{ultima_ref}"
              + (f"   (las fórmulas van corridas {desfase:+d} fila(s))" if desfase else ""))
            L(f"        {'+'.join(c['cols'])} debería llegar hasta la fila {esperada}")
            for col, ult, n in faltan:
                L(f"          {col}: llega a la fila {ult}, FALTAN {n} "
                  f"-> hay empresas sin calcular")
            for col, ult, n in sobran:
                L(f"          {col}: llega a la fila {ult}, SOBRAN {n} "
                  f"-> arrastra 0 o vacío por debajo de la última empresa")
            if sin_formula:
                L(f"        (sin fórmulas, con valores escritos: "
                  f"{', '.join(sin_formula)})")
            if ok:
                L(f"        las {len(cols)} columna(s) llegan justo a la fila {esperada}")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        esperada=esperada, n_empresas=n_emp,
                        faltan=faltan[:40], sobran=sobran[:40],
                        sin_formula=sin_formula)

        if tipo == "suma_por_empresa":
            # La columna de detalle trae la empresa repetida (una fila por
            # reemplazo). Se suma por empresa y se compara contra el total de la
            # tabla resumen. Todo dentro del mismo archivo y la misma hoja.
            ruta = self.rutas.get(c["archivo"])
            if ruta is None:
                L(f"  ? {c['desc']}: falta el archivo")
                return dict(base, estado="SIN DATOS")
            cols_res = list(c["cols_resumen"])          # [empresa, val1, val2, ...]
            necesarias = [c["col_empresa"], c["col_monto"]] + cols_res
            datos = leer_columnas_rapido(ruta, c["hoja"], necesarias,
                                         c["fila_inicio"], L)
            if datos is None:
                return dict(base, estado="SIN DATOS")

            # 1) detalle: sumar el monto agrupando por empresa
            detalle, sin_empresa = {}, 0
            col_e, col_m = c["col_empresa"], c["col_monto"]
            for f in sorted(datos.get(col_e, {})):
                nombre = datos[col_e][f]
                if not isinstance(nombre, str) or not nombre.strip():
                    continue
                k = normalizar(nombre)
                if k in ("", "0"):
                    continue
                v = datos.get(col_m, {}).get(f)
                if not isinstance(v, (int, float)):
                    if v is not None:
                        sin_empresa += 1
                    continue
                acum, nom, filas = detalle.get(k, (0.0, nombre.strip(), 0))
                detalle[k] = (acum + float(v), nom, filas + 1)

            # 2) resumen: la tabla de totales
            info = {}
            resumen = armar_tabla(datos, cols_res[0], cols_res[1:], L,
                                  c.get("nombre_resumen", "resumen"), info=info)

            difs, solo_det, solo_res = [], [], []
            for k, (suma, nom, nfilas) in detalle.items():
                if k not in resumen:
                    solo_det.append(nom)
                    continue
                vals = [v for v in resumen[k][1] if isinstance(v, (int, float))]
                total = sum(vals)
                if abs(suma - total) > TOLERANCIA:
                    difs.append((nom, suma, total, suma - total, nfilas))
            for k in resumen:
                if k not in detalle:
                    solo_res.append(resumen[k][0])

            dups = [n for n, _ in info.get("duplicadas", [])]
            ok = not (difs or solo_det or solo_res or dups)
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        {len(detalle)} empresa(s) distintas en {col_e} "
              f"(sumando {col_m}), {len(resumen)} en la tabla resumen")
            if sin_empresa:
                L(f"        {sin_empresa} fila(s) con empresa pero sin monto numérico")
            for n in solo_det[:15]:
                L(f"          en el detalle y NO en el resumen: {n[:40]}")
            for n in solo_res[:15]:
                L(f"          en el resumen y NO en el detalle: {n[:40]}")
            for n in dups[:10]:
                L(f"          REPETIDA en el resumen: {n[:40]}")
            if difs:
                etiqueta = " + ".join(cols_res[1:])
                L(f"        {len(difs)} empresa(s) descuadrada(s) "
                  f"(suma de {col_m}  vs  {etiqueta}):")
                for nom, suma, total, dif, nfilas in sorted(
                        difs, key=lambda x: -abs(x[3]))[:20]:
                    L(f"          {nom[:26]:<28} {fmt_monto(suma):>16} vs "
                      f"{fmt_monto(total):>16}  dif {fmt_monto(dif):>14}"
                      f"  ({nfilas} fila(s))")
                if len(difs) > 20:
                    L(f"          ... y {len(difs) - 20} más")
                # Si invirtiendo el signo cuadra, el problema es de convencion.
                invertidos = sum(1 for _, s, t, _, _ in difs
                                 if abs(s + t) <= TOLERANCIA)
                if invertidos:
                    L(f"        OJO: en {invertidos} de esas empresas cuadraría "
                      f"con el signo invertido. Puede que el neto se calcule "
                      f"como recibe - paga y no como paga + recibe.")
            if ok:
                L(f"        las {len(detalle)} empresas cuadran una por una")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        n_detalle=len(detalle), n_resumen=len(resumen),
                        difs=[(n, s, t, d, nf) for n, s, t, d, nf in
                              sorted(difs, key=lambda x: -abs(x[3]))[:40]],
                        n_difs=len(difs), solo_detalle=solo_det[:40],
                        solo_resumen=solo_res[:40], duplicadas=dups[:40])

        if tipo == "suma_calculada":
            # Recalcula un monto FILA POR FILA a partir de sus componentes y lo
            # compara contra la columna que ya lo trae calculado.
            #   valor_fila = (suma de "terminos", con su signo) * (los "factores")
            # En el tabulado: (CV - CMg) * Generacion * USD  contra la columna E.
            #
            # Se compara el TOTAL, que es lo que se pidio, pero tambien se cuenta
            # cuantas filas fallan por su cuenta: en un total, dos errores de
            # signo contrario se cancelan y no se ven. Y saber si falla UNA fila o
            # TODAS distingue un dato malo de un problema de redondeo.
            ruta = self.rutas.get(c["archivo"])
            if ruta is None:
                L(f"  ? {c['desc']}: falta el archivo")
                return dict(base, estado="SIN DATOS")
            L(f"  .. {c['desc']}")

            terminos = [(t.lstrip("+-").upper(), -1.0 if t.startswith("-") else 1.0)
                        for t in c["terminos"]]
            factores = [f.upper() for f in c.get("factores", [])]
            contra = c["contra"].upper()
            cols = sorted({t for t, _ in terminos} | set(factores) | {contra})
            f_ini = int(c["fila_inicio"])

            datos = leer_columnas_rapido(ruta, c["hoja"], cols, f_ini, L)
            if datos is None:
                return dict(base, estado="SIN DATOS")

            def num(col, fila):
                v = datos.get(col, {}).get(fila)
                return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None

            filas = sorted({f for col in cols for f in datos.get(col, {})
                            if f >= f_ini})
            tol_fila = float(c.get("tolerancia_fila", 0.5))
            suma_calc = suma_real = 0.0
            n_usadas = n_sin_datos = 0
            malas = []
            for f in filas:
                comps = {col: num(col, f) for col in cols}
                real = comps[contra]
                # Una fila sin ningun dato es el relleno del final: se ignora.
                if real is None and all(comps[col] is None for col, _ in terminos):
                    continue
                if any(comps[col] is None for col, _ in terminos) or \
                   any(comps[col] is None for col in factores):
                    n_sin_datos += 1
                    continue
                calc = sum(comps[col] * sg for col, sg in terminos)
                for col in factores:
                    calc *= comps[col]
                real = real if real is not None else 0.0
                suma_calc += calc
                suma_real += real
                n_usadas += 1
                if abs(calc - real) > tol_fila:
                    malas.append((f, calc, real, calc - real))

            if not n_usadas:
                L(f"  ? {c['desc']}: no hubo ninguna fila con todos los datos")
                return dict(base, estado="SIN DATOS")

            formula = " ".join(
                (("- " if sg < 0 else ("+ " if i else "")) + col)
                for i, (col, sg) in enumerate(terminos))
            if factores:
                formula = f"({formula}) * " + " * ".join(factores)
            dif = suma_calc - suma_real
            tol = float(c.get("tolerancia", TOLERANCIA))
            ok = abs(dif) <= tol
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        {n_usadas} fila(s) usadas, desde la {f_ini}"
              + (f"; {n_sin_datos} omitida(s) por falta de algún componente"
                 if n_sin_datos else ""))
            L(f"        {formula}  =  {fmt_monto(suma_calc)}")
            L(f"        columna {contra}          =  {fmt_monto(suma_real)}")
            L(f"        diferencia            =  {fmt_monto(dif)}"
              f"   (máximo aceptado {fmt_monto(tol)})")
            if suma_real:
                L(f"        en proporción         =  {dif / suma_real:.2%}")
            if malas:
                L(f"        {len(malas)} de {n_usadas} fila(s) no cuadran por su "
                  f"cuenta (más de {fmt_monto(tol_fila)} de diferencia):")
                for f, calc, real, d in sorted(malas, key=lambda x: -abs(x[3]))[:12]:
                    L(f"          fila {f}: calculado {fmt_monto(calc)} vs "
                      f"{contra}{f} {fmt_monto(real)}   dif {fmt_monto(d)}")
                if len(malas) > 12:
                    L(f"          ... y {len(malas) - 12} fila(s) más")
                if len(malas) == n_usadas:
                    L("        Fallan TODAS las filas: mirá si es redondeo o si "
                      "alguna columna no es la que se cree.")
            elif ok:
                L(f"        las {n_usadas} filas cuadran una por una")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        suma_calculada=suma_calc, suma_real=suma_real,
                        diferencia=dif, tolerancia=tol, formula=formula,
                        n_filas=n_usadas, n_sin_datos=n_sin_datos,
                        n_malas=len(malas),
                        malas=[(f, a, b, d) for f, a, b, d in
                               sorted(malas, key=lambda x: -abs(x[3]))[:40]])

        if tipo == "retencion_coherente":
            # La H del CPRT es la G con las retenciones puestas en 0:
            #     H = SI(CONTAR.SI(retenciones; acreedor)<>0; 0; 1) * G
            # O sea que en cada fila la H solo puede ser IGUAL a la G, o CERO.
            # Cualquier otro valor significa que la formula de la H esta mal, y eso
            # SI es un error, porque el csv se arma con la H.
            # Que haya filas en 0 no es un error: es una retencion de verdad. Pero
            # se avisa con el monto, porque es plata que no se paga.
            ruta = self.rutas.get(c["archivo"])
            if ruta is None:
                L(f"  ? {c['desc']}: falta el archivo")
                return dict(base, estado="SIN DATOS")
            L(f"  .. {c['desc']}")
            cg, ch, cref = c["col_g"], c["col_h"], c["col_referencia"]
            d = leer_columnas_rapido(ruta, c["hoja"], [cg, ch, cref],
                                     c["fila_inicio"], L)
            if d is None:
                return dict(base, estado="SIN DATOS")
            filas = [f for f, v in d.get(cref.upper(), {}).items()
                     if isinstance(v, str) and v.strip()]
            iguales, retenidas, raras, sin_h = 0, [], [], []
            total_ret = 0.0
            for f in sorted(filas):
                g = d.get(cg.upper(), {}).get(f)
                h = d.get(ch.upper(), {}).get(f)
                if not isinstance(g, (int, float)) or isinstance(g, bool):
                    continue
                if not isinstance(h, (int, float)) or isinstance(h, bool):
                    sin_h.append(f)      # la formula de la H no llega hasta aca
                    continue
                if abs(h - g) <= TOLERANCIA:
                    iguales += 1
                elif abs(h) <= TOLERANCIA:
                    retenidas.append((f, float(g)))
                    total_ret += float(g)
                else:
                    raras.append((f, float(g), float(h)))
            ok = not raras and not sin_h
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        {len(filas)} fila(s): {iguales} con {ch} = {cg}, "
              f"{len(retenidas)} retenida(s) en 0")
            if sin_h:
                L(f"        {len(sin_h)} fila(s) SIN valor en {ch}: la fórmula no "
                  f"llega hasta abajo. El csv saldría con el monto vacío.")
                for f in sin_h[:10]:
                    L(f"             fila {f}")
            if raras:
                L(f"        {len(raras)} fila(s) con {ch} que no es {cg} ni 0: "
                  f"la fórmula de la {ch} está mal.")
                for f, g, h in sorted(raras, key=lambda x: -abs(x[1] - x[2]))[:10]:
                    L(f"             fila {f}: {cg}={fmt_monto(g)}  "
                      f"{ch}={fmt_monto(h)}")
            if retenidas:
                L(f"        OJO: {fmt_monto(total_ret)} retenido en total. El csv "
                  f"se arma con la {ch}, así que esas empresas NO reciben pago:")
                for f, g in sorted(retenidas, key=lambda x: -x[1])[:10]:
                    L(f"             fila {f}: {fmt_monto(g)}")
                if len(retenidas) > 10:
                    L(f"             ... y {len(retenidas) - 10} más")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        n_filas=len(filas), n_iguales=iguales,
                        n_retenidas=len(retenidas), total_retenido=total_ret,
                        retenidas=[(f, g) for f, g in retenidas[:40]],
                        raras=[(f, g, h) for f, g, h in raras[:40]],
                        sin_h=sin_h[:40])

        if tipo == "prorrata_al_dia":
            # La hoja PRORRATA_RETIROS de las planillas 3, 5 y 6 es el pivote de
            # la tabla larga del Prorrata_Retiros: mismos numeros, otra forma.
            #     origen  (PRORRATA_HORARIA_TABULAR, desde la fila 2):
            #             A Hora | B Suministrador | C Prorrata_horaria
            #     destino (PRORRATA_RETIROS): B8="Hora", C8.. suministradores,
            #             B9.. las horas, el resto los valores (0 si falta)
            #
            # Como son los MISMOS numeros reordenados, la suma por suministrador
            # tiene que coincidir exactamente. Eso caza el olvido de actualizar:
            # otro mes trae otros totales y casi siempre otros suministradores.
            r_dest = self.rutas.get(c["archivo"])
            r_orig = self.rutas.get(c["origen"]["archivo"])
            if r_dest is None or r_orig is None:
                L(f"  ? {c['desc']}: falta uno de los archivos")
                return dict(base, estado="SIN DATOS")
            L(f"  .. {c['desc']}")

            # --- origen: tabla larga -> suma por suministrador ---
            o = c["origen"]
            do = leer_columnas_rapido(
                r_orig, o["hoja"], [o["col_hora"], o["col_suministrador"],
                                    o["col_valor"]], o["fila_inicio"], L)
            if do is None:
                return dict(base, estado="SIN DATOS")
            ch, cs, cv = (x.upper() for x in (o["col_hora"],
                                              o["col_suministrador"],
                                              o["col_valor"]))
            suma_orig, horas_orig, n_orig = {}, set(), 0
            for f in sorted(do.get(cs, {})):
                sumi = do[cs].get(f)
                if not isinstance(sumi, str) or not sumi.strip():
                    continue
                v = do.get(cv, {}).get(f)
                if not isinstance(v, (int, float)) or isinstance(v, bool):
                    continue
                k = normalizar(sumi)
                suma_orig[k] = suma_orig.get(k, 0.0) + float(v)
                h = do.get(ch, {}).get(f)
                if h is not None:
                    horas_orig.add(str(h).strip())
                n_orig += 1
            L(f"        origen : {n_orig} fila(s), {len(suma_orig)} "
              f"suministrador(es), {len(horas_orig)} hora(s)")

            # --- destino: la matriz ---
            fila_enc = int(c["fila_encabezado"])
            cols = columnas_de_fila(r_dest, c["hoja"], fila_enc, L)
            if cols is None:
                return dict(base, estado="SIN DATOS")
            n_ini = col_letra_a_num(c["col_inicio"])
            cols = [x for x in cols if col_letra_a_num(x) >= n_ini]
            if not cols:
                L(f"  >> {c['desc']}")
                L(f"        la fila {fila_enc} de {c['hoja']} está VACÍA desde "
                  f"{c['col_inicio']}: la prorrata nunca se pegó.")
                return dict(base, estado="NO CUADRA", n_origen=len(suma_orig),
                            n_destino=0, difs=[], solo_origen=[], solo_destino=[])
            datos = leer_columnas_rapido(r_dest, c["hoja"], cols, fila_enc, L)
            if datos is None:
                return dict(base, estado="SIN DATOS")
            # La primera columna es "Hora"; de la segunda en adelante, un
            # suministrador por columna.
            #
            # OJO con dos cosas de estas hojas:
            #  - A la DERECHA del bloque de prorrata puede haber mas columnas con
            #    los MISMOS nombres de suministrador pero con MONTOS (las
            #    planillas 5 y 6 los tienen), y columnas de total. Por eso se toma
            #    solo la PRIMERA columna de cada nombre: el bloque pegado empieza
            #    en B8, asi que es el de mas a la izquierda. Sumar las dos daba
            #    numeros de millones contra prorratas de dos digitos.
            #  - En la planilla 3 puede haber centrales pegadas a mano que quedan
            #    en 0%. Esas no estan en el origen y no son un error.
            col_hora = cols[0]
            filas_dato = sorted(f for f in datos.get(col_hora, {})
                                if f > fila_enc)
            n_filas_dest = len(filas_dato)
            suma_dest, vistos, repetidas = {}, set(), []
            for col in cols[1:]:
                nom = datos.get(col, {}).get(fila_enc)
                if not isinstance(nom, str) or not nom.strip():
                    continue
                k = normalizar(nom)
                if k in vistos:
                    repetidas.append(f"{nom.strip()} ({col})")
                    continue
                vistos.add(k)
                tot = 0.0
                for f in filas_dato:
                    v = datos.get(col, {}).get(f)
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        tot += float(v)
                suma_dest[k] = tot
            L(f"        destino: {n_filas_dest} fila(s) de hora, "
              f"{len(suma_dest)} columna(s) con nombre")
            if repetidas:
                L(f"        {len(repetidas)} columna(s) con un nombre ya visto "
                  f"(bloques de montos a la derecha): se usa la primera de cada "
                  f"una")

            # --- comparacion ---
            # La exigencia es ASIMETRICA a proposito: todo suministrador del
            # ORIGEN tiene que estar en la planilla con la misma suma. Lo que
            # sobra en la planilla solo se informa, porque hay dos motivos
            # legitimos: centrales pegadas a mano que quedan en 0% (planilla 3) y
            # columnas de totales o de montos que son parte de la hoja
            # (planillas 5 y 6).
            tol = c.get("tolerancia", TOL_PRORRATA_SUMA)
            solo_o = sorted(k for k in suma_orig if k not in suma_dest)
            sobran = sorted(k for k in suma_dest if k not in suma_orig)
            # De lo que sobra, solo se nombra lo que tiene algo distinto de cero:
            # una central en 0% no aporta nada y no vale la pena listarla.
            sobran_con_valor = [k for k in sobran if abs(suma_dest[k]) > tol]
            sobran_en_cero = [k for k in sobran if abs(suma_dest[k]) <= tol]
            difs = []
            n_ok = 0
            for k, v in suma_orig.items():
                if k not in suma_dest:
                    continue
                d = suma_dest[k] - v
                if abs(d) > tol:
                    difs.append((k, suma_dest[k], v, d))
                else:
                    n_ok += 1
            # Las horas: solo se avisa. La hoja puede tener filas de mas abajo
            # que no son parte de la matriz.
            dif_horas = bool(horas_orig) and len(horas_orig) != n_filas_dest
            ok = not (difs or solo_o)
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        {n_ok} de {len(suma_orig)} suministrador(es) del origen "
              f"cuadran")
            if dif_horas:
                L(f"        (las horas no coinciden: {len(horas_orig)} en el "
                  f"origen y {n_filas_dest} filas en la planilla)")
            for k in solo_o[:10]:
                L(f"          FALTA en la planilla: {k[:40]} "
                  f"(origen {suma_orig[k]:.6f})")
            if len(solo_o) > 10:
                L(f"          ... y {len(solo_o) - 10} más")
            for k, vd, vo, d in sorted(difs, key=lambda x: -abs(x[3]))[:10]:
                L(f"          {k[:34]:<36} planilla {vd:.6f} vs origen {vo:.6f}")
            if len(difs) > 10:
                L(f"          ... y {len(difs) - 10} más")
            if sobran_en_cero:
                L(f"        {len(sobran_en_cero)} columna(s) de la planilla en 0 "
                  f"que no están en el origen (normal): "
                  + ", ".join(x[:20] for x in sobran_en_cero[:6])
                  + (" ..." if len(sobran_en_cero) > 6 else ""))
            if sobran_con_valor:
                L(f"        {len(sobran_con_valor)} columna(s) de la planilla que "
                  f"no son del origen (totales, montos): "
                  + ", ".join(x[:22] for x in sobran_con_valor[:6])
                  + (" ..." if len(sobran_con_valor) > 6 else ""))
            if not ok:
                L("        La prorrata de esta planilla NO es la del "
                  "Prorrata_Retiros de este mes.")
                L("        Hay que correr «Actualizar data» y marcar Prorrata.")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        n_origen=len(suma_orig), n_destino=len(suma_dest),
                        n_ok=n_ok, horas_origen=len(horas_orig),
                        horas_destino=n_filas_dest,
                        solo_origen=solo_o[:40],
                        solo_destino=sobran_con_valor[:40],
                        sobran_en_cero=len(sobran_en_cero),
                        difs=[(k, vd, vo, d) for k, vd, vo, d in
                              sorted(difs, key=lambda x: -abs(x[3]))[:40]])

        if tipo == "centrales_sin_dueno":
            # Cruza dos tablas del mismo Access: toda central con MONTO en
            # Sobrecostos tiene que tener dueño en Central_Empresa.
            #
            # Una central sin dueño NO es un error por si misma: en
            # CONSUMOS_PROPIOS de la planilla 6 hay centrales sin propietario a
            # proposito. Lo que no puede pasar es que una central con plata no
            # tenga a quien pagarle o a quien cobrarle.
            ruta = self.rutas.get(c["archivo"])
            if ruta is None:
                L(f"  ? {c['desc']}: falta el archivo")
                return dict(base, estado="SIN DATOS")
            L(f"  .. {c['desc']}")
            t_sob = c.get("tabla_montos", "Sobrecostos")
            t_ce = c.get("tabla_duenos", "Central_Empresa")
            cn = None
            try:
                cn = conexion_mdb(ruta)
                cur = cn.cursor()
                tablas = {r.table_name for r in cur.tables(tableType="TABLE")}
                for t in (t_sob, t_ce):
                    if t not in tablas:
                        L(f"  ? {c['desc']}: no está la tabla [{t}]. "
                          f"Hay: {', '.join(sorted(tablas))}")
                        return dict(base, estado="SIN DATOS")

                def col(tabla, objetivo):
                    cols = [r.column_name for r in cur.columns(table=tabla)]
                    for x in cols:
                        if normalizar(x) == normalizar(objetivo):
                            return x
                    for x in cols:
                        if normalizar(objetivo) in normalizar(x):
                            return x
                    return None

                c_cen_s = col(t_sob, "Central")
                c_mon = col(t_sob, "Sobrecosto")
                c_cen_e = col(t_ce, "Central")
                c_emp = col(t_ce, "Empresa")
                if not all((c_cen_s, c_mon, c_cen_e, c_emp)):
                    L(f"  ? {c['desc']}: no se encontraron las columnas "
                      f"necesarias en las dos tablas.")
                    return dict(base, estado="SIN DATOS")

                # Dueños: por central, si tiene una empresa de verdad.
                duenos = {}
                cur.execute(f"SELECT [{c_cen_e}], [{c_emp}] FROM [{t_ce}]")
                for cen, emp in cur.fetchall():
                    if cen is None or not str(cen).strip():
                        continue
                    k = clave_central(cen)
                    tiene = (emp is not None and str(emp).strip() != ""
                             and str(emp).strip() != "0")
                    duenos[k] = duenos.get(k, False) or tiene

                cur.execute(f"SELECT [{c_cen_s}], SUM([{c_mon}]) FROM [{t_sob}] "
                            f"GROUP BY [{c_cen_s}]")
                montos = []
                for cen, suma in cur.fetchall():
                    if cen is None or not str(cen).strip():
                        montos.append((None, "(central vacía)", float(suma or 0)))
                    else:
                        montos.append((clave_central(cen), str(cen).strip(),
                                       float(suma or 0)))
            except Exception as e:
                L(f"  ? {c['desc']}: no se pudo leer el Access: {e}")
                return dict(base, estado="SIN DATOS")
            finally:
                try:
                    if cn is not None:
                        cn.close()
                except Exception:
                    pass

            tol = c.get("tolerancia", TOLERANCIA)
            sin_dueno, sin_fila, con_dueno, sin_plata = [], [], 0, 0
            for k, nombre, suma in montos:
                if abs(suma) <= tol:
                    sin_plata += 1
                elif k is None or k not in duenos:
                    sin_fila.append((nombre, suma))
                elif not duenos[k]:
                    sin_dueno.append((nombre, suma))
                else:
                    con_dueno += 1

            ok = not (sin_dueno or sin_fila)
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        {len(montos)} central(es) en [{t_sob}], "
              f"{len(duenos)} en [{t_ce}]")
            L(f"        {con_dueno} con monto y con dueño; {sin_plata} sin monto "
              f"(esas pueden no tener dueño)")
            if sin_dueno:
                L(f"        {len(sin_dueno)} central(es) con MONTO y el dueño "
                  f"VACÍO o en 0:")
                for n, v in sorted(sin_dueno, key=lambda x: -abs(x[1]))[:15]:
                    L(f"          {n[:34]:<36} {fmt_monto(v)}")
                if len(sin_dueno) > 15:
                    L(f"          ... y {len(sin_dueno) - 15} más")
            if sin_fila:
                L(f"        {len(sin_fila)} central(es) con MONTO que NO están en "
                  f"[{t_ce}]:")
                for n, v in sorted(sin_fila, key=lambda x: -abs(x[1]))[:15]:
                    L(f"          {n[:34]:<36} {fmt_monto(v)}")
                if len(sin_fila) > 15:
                    L(f"          ... y {len(sin_fila) - 15} más")
            if not ok:
                L("        Esa plata no tiene a quién pagarse ni a quién cobrarse.")
                L("        Hay que completar el propietario en la planilla de")
                L("        origen y volver a actualizar el Access.")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        n_centrales=len(montos), n_duenos=len(duenos),
                        con_dueno=con_dueno, sin_plata=sin_plata,
                        sin_dueno=[(n, v) for n, v in
                                   sorted(sin_dueno, key=lambda x: -abs(x[1]))[:40]],
                        sin_fila=[(n, v) for n, v in
                                  sorted(sin_fila, key=lambda x: -abs(x[1]))[:40]])

        if tipo == "pago_por_empresa":
            # La planilla 1 y la 9 (o la 4) son dos calculos PARALELOS del mismo
            # pago: la 1 lo saca por empresa y concepto directo, la 9 lo saca
            # prorrateando por retiros y despues lo agrupa. Tienen que dar lo
            # mismo, y hasta ahora nadie lo comparaba.
            #
            # Detalles que importan:
            #  - Los nombres de concepto NO se escriben igual: la 1 dice
            #    "CO ERNC" y la 9 "CO_ERNC". Por eso clave_concepto().
            #  - El monto de la 9 corresponde a la columna PAGA de la 1, no a
            #    RECIBE. Verificado en 2409: contra PAGA los 664 pares cuadran
            #    (peor diferencia 58 pesos); contra RECIBE solo 41 de 664.
            #  - La 1 tiene conceptos que la 9 no (los "ID", y CCA/CO/SC_SSCC que
            #    viven en la planilla 4). Solo se comparan los que estan en las
            #    DOS; los que sobran de un lado se informan, no fallan.
            det, res = c["detalle"], c["resumen"]
            r_det = self.rutas.get(det["archivo"])
            r_res = self.rutas.get(res["archivo"])
            if r_det is None or r_res is None:
                L(f"  ? {c['desc']}: falta uno de los archivos")
                return dict(base, estado="SIN DATOS")
            L(f"  .. {c['desc']}")

            # --- lado detalle (la planilla 9 o 4: una fila por retiro) ---
            dd = leer_columnas_rapido(
                r_det, det["hoja"],
                [det["col_concepto"], det["col_empresa"], det["col_monto"]],
                det["fila_inicio"], L)
            if dd is None:
                return dict(base, estado="SIN DATOS")
            cc, ce, cm = (x.upper() for x in
                          (det["col_concepto"], det["col_empresa"], det["col_monto"]))
            pdet, n_det = {}, 0
            for f in sorted(dd.get(cc, {})):
                con, emp = dd[cc].get(f), dd.get(ce, {}).get(f)
                if not isinstance(con, str) or not con.strip():
                    continue
                if not isinstance(emp, str) or not emp.strip():
                    continue
                v = dd.get(cm, {}).get(f)
                if not isinstance(v, (int, float)) or isinstance(v, bool):
                    continue
                k = (clave_concepto(con), clave_concepto(emp))
                pdet[k] = pdet.get(k, 0.0) + float(v)
                n_det += 1
            L(f"        {det['nombre']}: {n_det} fila(s) -> {len(pdet)} par(es) "
              f"(concepto, empresa)")

            # --- lado resumen (la planilla 1: ya viene por empresa y concepto) ---
            dr = leer_columnas_rapido(
                r_res, res["hoja"],
                [res["col_concepto"], res["col_empresa"], res["col_monto"]],
                res["fila_inicio"], L)
            if dr is None:
                return dict(base, estado="SIN DATOS")
            rc, re_, rm = (x.upper() for x in
                           (res["col_concepto"], res["col_empresa"], res["col_monto"]))
            pres, n_res = {}, 0
            for f in sorted(dr.get(rc, {})):
                con, emp = dr[rc].get(f), dr.get(re_, {}).get(f)
                if not isinstance(con, str) or not con.strip():
                    continue
                if not isinstance(emp, str) or not emp.strip():
                    continue
                v = dr.get(rm, {}).get(f)
                if not isinstance(v, (int, float)) or isinstance(v, bool):
                    continue
                k = (clave_concepto(con), clave_concepto(emp))
                pres[k] = pres.get(k, 0.0) + float(v)
                n_res += 1
            L(f"        {res['nombre']}: {n_res} fila(s) -> {len(pres)} par(es)")

            # Solo los conceptos que estan en los dos lados.
            con_det = {k[0] for k in pdet}
            con_res = {k[0] for k in pres}
            comunes = con_det & con_res
            solo_det = sorted(con_det - con_res)
            solo_res = sorted(con_res - con_det)
            L(f"        conceptos en común: {len(comunes)}  "
              f"({', '.join(sorted(comunes))})")
            if solo_res:
                L(f"        solo en {res['nombre']}: {', '.join(solo_res)}"
                  f"   (no se comparan)")
            if solo_det:
                L(f"        solo en {det['nombre']}: {', '.join(solo_det)}"
                  f"   (no se comparan)")

            tol = c.get("tolerancia", TOL_PAGO_EMPRESA)
            # El signo de la columna PAGA de la planilla 1 NO es consistente: es
            # positivo para los conceptos de la planilla 9 (CPF, CSF, CTF, CRA,
            # REA, CO ERNC) y NEGATIVO para los de la planilla 4 (CCA, CO,
            # SC_SSCC). Verificado en 2409, en la misma columna E de la misma hoja.
            # Como la convencion cambia dentro del mismo archivo, no se puede
            # distinguir "convencion" de "error de signo", asi que se compara la
            # MAGNITUD y se informa aparte cuantos pares venian con signo opuesto.
            absoluto = c.get("absoluto", True)
            difs, faltan_res, faltan_det = [], [], []
            n_ok, n_signo = 0, 0
            for k, v in pdet.items():
                if k[0] not in comunes:
                    continue
                if k not in pres:
                    # Un par que falta con monto ~0 no es un problema: la
                    # planilla 1 lista la empresa con 0 y la 9 simplemente no
                    # tiene filas para ella en ese concepto.
                    if abs(v) > tol:
                        faltan_res.append((k, v))
                    continue
                vr = pres[k]
                if absoluto:
                    if v * vr < 0:
                        n_signo += 1
                    d = abs(v) - abs(vr)
                else:
                    d = v - vr
                if abs(d) > tol:
                    difs.append((k, v, vr, d))
                else:
                    n_ok += 1
            for k, v in pres.items():
                if k[0] in comunes and k not in pdet and abs(v) > tol:
                    faltan_det.append((k, v))

            ok = not (difs or faltan_res or faltan_det)
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        {n_ok} par(es) cuadran dentro de {fmt_monto(tol)}"
              + ("  (se compara la magnitud)" if absoluto else ""))
            if n_signo:
                L(f"        {n_signo} par(es) venían con el signo opuesto, que es "
                  f"convención de la planilla 1 y no un error.")
            if difs:
                L(f"        {len(difs)} par(es) con diferencia mayor:")
                for (co, em), vd, vr, d in sorted(difs, key=lambda x: -abs(x[3]))[:15]:
                    L(f"          {co:<12} {em[:24]:<26} "
                      f"{det['nombre']} {fmt_monto(vd):>16}  vs  "
                      f"{res['nombre']} {fmt_monto(vr):>16}   dif {fmt_monto(d)}")
                if len(difs) > 15:
                    L(f"          ... y {len(difs) - 15} más")
            for (co, em), v in faltan_res[:10]:
                L(f"          {co} / {em[:26]}: está en {det['nombre']} "
                  f"({fmt_monto(v)}) y NO en {res['nombre']}")
            if len(faltan_res) > 10:
                L(f"          ... y {len(faltan_res) - 10} más")
            for (co, em), v in faltan_det[:10]:
                L(f"          {co} / {em[:26]}: está en {res['nombre']} "
                  f"({fmt_monto(v)}) y NO en {det['nombre']}")
            if len(faltan_det) > 10:
                L(f"          ... y {len(faltan_det) - 10} más")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        n_ok=n_ok, n_difs=len(difs), n_signo=n_signo,
                        comunes=sorted(comunes),
                        solo_detalle=solo_det, solo_resumen=solo_res,
                        difs=[(f"{co} / {em}", vd, vr, d) for (co, em), vd, vr, d in
                              sorted(difs, key=lambda x: -abs(x[3]))[:40]],
                        faltan_resumen=[f"{co} / {em}" for (co, em), _ in faltan_res[:40]],
                        faltan_detalle=[f"{co} / {em}" for (co, em), _ in faltan_det[:40]])

        if tipo == "bloque_contra_origen":
            # Compara UN bloque del destino (SC o CO) contra su origen filtrado por
            # embalses: cantidad de filas y suma del monto. Es lo unico que dice si
            # los datos pegados son los de este mes; el resto de V10 solo mira que
            # el destino sea coherente consigo mismo.
            # Se hace por bloque separado a proposito, para poder decir CUAL de los
            # dos hay que volver a traer y no actualizar los dos al azar.
            de, org = c["destino"], c["origen"]
            r_dest = self.rutas.get(de["archivo"])
            r_orig = self.rutas.get(org["archivo"])
            if r_dest is None or r_orig is None:
                L(f"  ? {c['desc']}: falta uno de los archivos")
                return dict(base, estado="SIN DATOS")
            L(f"  .. {c['desc']}")
            permitidas = {clave_central(x) for x in CENTRALES_EMBALSE}

            # --- destino: solo las filas de este tipo ---
            dd = leer_columnas_rapido(
                r_dest, de["hoja"],
                [de["col_tipo"], de["col_monto"], de["col_central"]],
                de["fila_inicio"], L)
            if dd is None:
                return dict(base, estado="SIN DATOS")
            objetivo = normalizar(de["tipo"])
            n_dest, suma_dest, cent_dest = 0, 0.0, set()
            for f in sorted(dd.get(de["col_tipo"].upper(), {})):
                if normalizar(dd[de["col_tipo"].upper()][f]) != objetivo:
                    continue
                n_dest += 1
                v = dd.get(de["col_monto"].upper(), {}).get(f)
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    suma_dest += float(v)
                nom = dd.get(de["col_central"].upper(), {}).get(f)
                if isinstance(nom, str) and nom.strip():
                    cent_dest.add(clave_central(nom))

            # --- origen: solo los embalses ---
            do = leer_columnas_rapido(
                r_orig, org["hoja"], [org["col_central"], org["col_monto"]],
                org["fila_inicio"], L)
            if do is None:
                return dict(base, estado="SIN DATOS")
            n_orig, suma_orig, cent_orig = 0, 0.0, set()
            for f in sorted(do.get(org["col_central"].upper(), {})):
                nom = do[org["col_central"].upper()][f]
                if not isinstance(nom, str) or not nom.strip():
                    continue
                k = clave_central(nom)
                if k not in permitidas:
                    continue
                n_orig += 1
                cent_orig.add(k)
                v = do.get(org["col_monto"].upper(), {}).get(f)
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    suma_orig += float(v)

            dif = suma_dest - suma_orig
            ok_filas = n_dest == n_orig
            ok_monto = abs(dif) <= TOLERANCIA
            solo_d = sorted(cent_dest - cent_orig)
            solo_o = sorted(cent_orig - cent_dest)
            ok = ok_filas and ok_monto and not solo_d and not solo_o
            etq = de["tipo"]
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        destino ({etq}): {n_dest} fila(s), "
              f"{fmt_monto(suma_dest)}   [{de['hoja']} col {de['col_monto']}]")
            L(f"        origen  : {n_orig} fila(s), {fmt_monto(suma_orig)}   "
              f"[{Path(r_orig).name}, {org['hoja']} col {org['col_monto']}]")
            if not ok_filas:
                L(f"        FALTAN o SOBRAN filas: {n_dest - n_orig:+d}")
            if not ok_monto:
                L(f"        diferencia de monto: {fmt_monto(dif)}")
            for k in solo_o[:10]:
                L(f"          está en el origen y NO en el destino: {k}")
            for k in solo_d[:10]:
                L(f"          está en el destino y NO en el origen: {k}")
            if not ok:
                L(f"        >>> HAY QUE VOLVER A TRAER LOS {etq} <<<")
                L(f"            (no hace falta tocar el otro bloque)")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        bloque=etq, n_destino=n_dest, n_origen=n_orig,
                        suma_destino=suma_dest, suma_origen=suma_orig,
                        diferencia=dif, solo_destino=solo_d[:40],
                        solo_origen=solo_o[:40])

        if tipo == "suma_fila":
            # Cada fila del bloque de prorrata tiene que sumar el 100%: escrito
            # como 1 o como 100, segun la planilla. Se aceptan los dos, pero se
            # avisa si un mismo archivo mezcla, porque eso ya es un error.
            ruta = self.rutas.get(c["archivo"])
            if ruta is None:
                L(f"  ? {c['desc']}: falta el archivo")
                return dict(base, estado="SIN DATOS")
            L(f"  .. {c['desc']}")
            cols = expandir_columnas(c["rango"])
            f_ini = int(c["fila_inicio"])
            datos = leer_columnas_rapido(ruta, c["hoja"], cols, f_ini, L)
            if datos is None:
                return dict(base, estado="SIN DATOS")
            # Se usa una columna de referencia para saber cuales filas EXISTEN:
            # una fila sin prorrata no es lo mismo que una fila que no existe.
            # Con col_tipo se puede decir si la fila mala es de SC o de CO, que es
            # lo que hace falta para saber cual bloque volver a traer.
            col_tipo = c.get("col_tipo")
            tipos = {}
            if col_tipo:
                dt = leer_columnas_rapido(ruta, c["hoja"], [col_tipo], f_ini, L)
                if dt is not None:
                    tipos = {f: (str(v).strip() if isinstance(v, str) else v)
                             for f, v in dt.get(col_tipo.upper(), {}).items()}

            def de_quien(f):
                t = tipos.get(f)
                return str(t) if t else "?"

            col_ref = c.get("col_referencia")
            filas_ref = None
            if col_ref:
                dref = leer_columnas_rapido(ruta, c["hoja"], [col_ref], f_ini, L)
                if dref is not None:
                    filas_ref = {f for f, v in dref.get(col_ref.upper(), {}).items()
                                 if isinstance(v, str) and v.strip()
                                 and clave_central(v) not in ("", "0")}
            if filas_ref is None:
                filas_ref = set()
                for col in cols:
                    filas_ref |= set(datos.get(col, {}))
                filas_ref = {f for f in filas_ref if f >= f_ini}

            malas, ceros, escalas = [], [], {}
            for f in sorted(filas_ref):
                vals = [datos.get(col, {}).get(f) for col in cols]
                nums = [float(v) for v in vals
                        if isinstance(v, (int, float)) and not isinstance(v, bool)]
                total = sum(nums)
                # Una fila que suma 0 es VALIDA: no reparte nada. Se cuenta para
                # dejarlo a la vista, pero no hace fallar. Y no entra en la
                # deteccion de escala, porque el 0 no dice si la planilla trabaja
                # en 1 o en 100.
                if not nums or abs(total) <= TOL_PRORRATA:
                    ceros.append(f)
                    continue
                # ¿1 o 100? se acepta el que quede mas cerca.
                cerca = min(TOTALES_PRORRATA, key=lambda t: abs(total - t))
                tol = TOL_PRORRATA * max(1.0, cerca)
                if abs(total - cerca) <= tol:
                    escalas[cerca] = escalas.get(cerca, 0) + 1
                else:
                    malas.append((f, total))

            mezcla = len(escalas) > 1
            ok = not malas and not mezcla
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        {len(filas_ref)} fila(s) en {c['rango']}{f_ini} hacia abajo")
            for esc, n in sorted(escalas.items()):
                L(f"        {n} fila(s) suman {esc:g}")
            if mezcla:
                L("        OJO: el archivo MEZCLA filas que suman 1 con filas que")
                L("             suman 100. Una de las dos está mal.")
            # Se agrupa por bloque para poder decir cual hay que volver a traer.
            por_bloque = {}
            for f, t in malas:
                por_bloque.setdefault(de_quien(f), []).append((f, t))
            for f, t in malas[:15]:
                L(f"          fila {f} ({de_quien(f)}): suma {t!r}")
            if len(malas) > 15:
                L(f"          ... y {len(malas) - 15} fila(s) más")
            if malas and col_tipo:
                resumen = ", ".join(f"{len(v)} de {k}"
                                    for k, v in sorted(por_bloque.items()))
                L(f"        las filas malas son: {resumen}")
                solo = [k for k in por_bloque if k != "?"]
                if len(solo) == 1:
                    L(f"        >>> HAY QUE VOLVER A TRAER LOS {solo[0]} <<<")
                    L(f"            (no hace falta tocar el otro bloque)")
            if ceros:
                L(f"        {len(ceros)} fila(s) suman 0 (no reparten nada, "
                  f"está permitido)")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        n_filas=len(filas_ref), escalas=dict(escalas),
                        malas=[(f, t, de_quien(f)) for f, t in malas[:40]],
                        ceros=ceros[:40], mezcla=mezcla,
                        por_bloque={k: len(v) for k, v in por_bloque.items()})

        if tipo == "centrales_en_lista":
            # Dos cosas, segun los flags:
            #   exigir  -> toda central de la columna tiene que estar en la lista
            #   sufijo  -> avisar de las "-numero" que NO estan en la lista, que es
            #              la senal de que apareció una unidad de embalse nueva
            lista = CENTRALES_EMBALSE
            ruta = self.rutas.get(c["archivo"])
            if ruta is None:
                L(f"  ? {c['desc']}: falta el archivo")
                return dict(base, estado="SIN DATOS")
            L(f"  .. {c['desc']}")
            permitidas = {clave_central(x) for x in lista}
            col = c["columna"].upper()
            datos = leer_columnas_rapido(ruta, c["hoja"], [col],
                                         c["fila_inicio"], L)
            if datos is None:
                return dict(base, estado="SIN DATOS")
            vistas, fuera, sospechosas = {}, {}, {}
            for f in sorted(datos.get(col, {})):
                v = datos[col][f]
                if not isinstance(v, str) or not v.strip() or v.startswith("#"):
                    continue
                nom = v.strip()
                k = clave_central(nom)
                if k in ("", "0"):
                    continue
                if k in permitidas:
                    vistas[k] = nom
                else:
                    fuera.setdefault(k, (nom, f))
                    if RE_UNIDAD_CENTRAL.search(nom):
                        sospechosas.setdefault(k, (nom, f))

            malos = []
            if c.get("exigir"):
                malos = list(fuera.values())
            avisar = list(sospechosas.values()) if c.get("avisar_sufijo") else []
            faltan = [x for x in lista if clave_central(x) not in vistas] \
                if c.get("exigir_todas") else []

            ok = not (malos or avisar or faltan)
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        lista: {len(lista)} central(es) de embalse")
            L(f"        en {col}{c['fila_inicio']} hacia abajo: "
              f"{len(vistas)} de la lista, {len(fuera)} fuera de la lista")
            for nom, f in malos[:15]:
                L(f"          {col}{f}: «{nom[:34]}» no es una central de embalse")
            if len(malos) > 15:
                L(f"          ... y {len(malos) - 15} más")
            for nom, f in avisar[:15]:
                L(f"          {col}{f}: «{nom[:34]}» termina en «-número» y NO está")
                L(f"                   en la lista. ¿Es una unidad nueva?")
            if len(avisar) > 15:
                L(f"          ... y {len(avisar) - 15} más")
            for x in faltan[:15]:
                L(f"          falta: {x} no aparece en ninguna fila")
            if len(faltan) > 15:
                L(f"          ... y {len(faltan) - 15} más")
            if ok:
                L("        todas las centrales cuadran con la lista")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        n_lista=len(lista), n_vistas=len(vistas), n_fuera=len(fuera),
                        fuera=[f"{n} ({col}{fi})" for n, fi in malos[:40]],
                        sufijo=[f"{n} ({col}{fi})" for n, fi in avisar[:40]],
                        faltan=faltan[:40])

        if tipo == "sobrecosto_por_fila":
            # Recalcula el sobrecosto desde sus componentes y lo compara con la
            # columna que lo trae ya calculado:
            #     sobrecosto = (CV - CMg) * Generacion * USD
            # Se compara el TOTAL, que es lo pedido, pero tambien fila por fila:
            # si dos filas se equivocan en sentidos opuestos el total cuadra y el
            # archivo igual esta mal, y sin el detalle por fila no habria como
            # saber DONDE mirar.
            ruta = self.rutas.get(c["archivo"])
            if ruta is None:
                L(f"  ? {c['desc']}: falta el archivo")
                return dict(base, estado="SIN DATOS")
            L(f"  .. {c['desc']}")
            cols = c["columnas"]
            orden = ["cv", "cmg", "gen", "usd", "resultado"]
            letras = [cols[k] for k in orden]
            datos = leer_columnas_rapido(ruta, c["hoja"], letras,
                                         c["fila_inicio"], L)
            if datos is None:
                return dict(base, estado="SIN DATOS")

            def num(letra, fila):
                v = datos.get(letra.upper(), {}).get(fila)
                return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None

            # Se recorren todas las filas donde haya ALGO, no solo las de la
            # columna del resultado: una fila con componentes y sin resultado es
            # justamente uno de los errores que hay que cazar.
            filas = set()
            for letra in letras:
                filas |= set(datos.get(letra.upper(), {}))
            filas = sorted(f for f in filas if f >= int(c["fila_inicio"]))

            tol_fila = c.get("tolerancia_fila", TOL_SOBRECOSTO_FILA)
            suma_calc = suma_esp = 0.0
            n_ok = 0
            difs, incompletas = [], []
            for f in filas:
                cv, cmg, gen = num(cols["cv"], f), num(cols["cmg"], f), num(cols["gen"], f)
                usd, esp = num(cols["usd"], f), num(cols["resultado"], f)
                if None in (cv, cmg, gen, usd):
                    # Sin componentes no se puede recalcular. Solo molesta si la
                    # fila SI trae un sobrecosto: ahi hay algo que no cuadra.
                    if esp is not None and abs(esp) > TOLERANCIA:
                        faltan = [cols[k] for k, v in
                                  (("cv", cv), ("cmg", cmg), ("gen", gen), ("usd", usd))
                                  if v is None]
                        incompletas.append((f, esp, "+".join(faltan)))
                    continue
                calc = (cv - cmg) * gen * usd
                suma_calc += calc
                if esp is None:
                    incompletas.append((f, None, cols["resultado"]))
                    continue
                suma_esp += esp
                d = calc - esp
                if abs(d) > tol_fila:
                    difs.append((f, calc, esp, d, cv, cmg, gen, usd))
                else:
                    n_ok += 1

            dif_total = suma_calc - suma_esp
            ok_total = abs(dif_total) <= max(TOLERANCIA, tol_fila)
            ok = ok_total and not difs and not incompletas
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        ({cols['cv']} − {cols['cmg']}) × {cols['gen']} × "
              f"{cols['usd']}   vs   {cols['resultado']}")
            L(f"        {len(filas)} fila(s) leídas, {n_ok} cuadran fila a fila")
            L(f"        recalculado : {fmt_monto(suma_calc)}")
            L(f"        columna {cols['resultado']}    : {fmt_monto(suma_esp)}")
            L(f"        diferencia  : {fmt_monto(dif_total)}"
              f"   (máximo aceptado {fmt_monto(max(TOLERANCIA, tol_fila))})")
            if difs:
                L(f"        {len(difs)} fila(s) no cuadran (tolerancia por fila "
                  f"{fmt_monto(tol_fila)}):")
                for f, calc, esp, d, cv, cmg, gen, usd in sorted(
                        difs, key=lambda x: -abs(x[3]))[:15]:
                    L(f"          fila {f:>5}: recalc {fmt_monto(calc):>16} vs "
                      f"{fmt_monto(esp):>16}  dif {fmt_monto(d):>14}")
                    L(f"                     {cols['cv']}={cv} {cols['cmg']}={cmg} "
                      f"{cols['gen']}={gen} {cols['usd']}={usd}")
                if len(difs) > 15:
                    L(f"          ... y {len(difs) - 15} fila(s) más")
                if ok_total:
                    L("        OJO: el TOTAL cuadra pero hay filas que no. Se están")
                    L("             compensando entre ellas, así que el archivo está")
                    L("             mal aunque la suma dé bien.")
            for f, esp, falta in incompletas[:15]:
                if esp is None:
                    L(f"          fila {f}: tiene componentes pero {falta} está vacía")
                else:
                    L(f"          fila {f}: trae {fmt_monto(esp)} pero falta {falta}")
            if len(incompletas) > 15:
                L(f"          ... y {len(incompletas) - 15} fila(s) incompleta(s) más")
            if ok:
                L(f"        las {n_ok} filas cuadran una por una")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        suma_calculada=suma_calc, suma_esperada=suma_esp,
                        diferencia=dif_total, n_filas=len(filas), n_ok=n_ok,
                        n_difs=len(difs),
                        difs=[(f, ca, es, d) for f, ca, es, d, *_ in
                              sorted(difs, key=lambda x: -abs(x[3]))[:40]],
                        incompletas=[(f, e, x) for f, e, x in incompletas[:40]])

        if tipo == "matriz_al_dia":
            # La matriz tiene que estar armada con las empresas de AHORA. Si se
            # actualizaron los datos y no se corrio CuadroPago, la matriz queda
            # con las del mes anterior y nadie lo nota: sigue mostrando numeros.
            ruta = self.rutas.get(c["archivo"])
            if ruta is None:
                L(f"  ? {c['desc']}: falta el archivo")
                return dict(base, estado="SIN DATOS")
            L(f"  .. {c['desc']}")
            m = leer_matriz_pago(ruta, c["hoja"], L)
            if m is None or not m["pagan"] or not m["reciben"]:
                L("  >> no se pudo leer la matriz, o esta vacia. "
                  "Seguramente falta correr Cuadro de pagos.")
                return dict(base, estado="NO CUADRA", n_pagan=0, n_reciben=0)
            L(f"        matriz: {len(m['pagan'])} que pagan x "
              f"{len(m['reciben'])} que reciben, "
              f"{len(m['montos'])} par(es) con monto")

            # La verdad de quien paga y quien recibe esta en la tabla I:K.
            datos = leer_columnas_rapido(ruta, c["hoja"], ["I", "J", "K"],
                                         c["fila_tabla"], L)
            if datos is None:
                return dict(base, estado="SIN DATOS")
            esp_pagan, esp_reciben = {}, {}
            for f in sorted(datos.get("I", {})):
                nom = datos["I"][f]
                if not isinstance(nom, str) or not nom.strip():
                    continue
                k = normalizar(nom)
                if k in ("", "0"):
                    continue
                j = datos.get("J", {}).get(f)
                kk = datos.get("K", {}).get(f)
                if isinstance(j, (int, float)) and not isinstance(j, bool):
                    esp_pagan[k] = nom.strip()
                if isinstance(kk, (int, float)) and not isinstance(kk, bool):
                    esp_reciben[k] = nom.strip()
            L(f"        tabla I:K: {len(esp_pagan)} con monto en J (pagan), "
              f"{len(esp_reciben)} con monto en K (reciben)")

            hay_pagan = {normalizar(x) for x in m["pagan"]}
            hay_reciben = {normalizar(x) for x in m["reciben"]}
            faltan_p = [esp_pagan[k] for k in esp_pagan if k not in hay_pagan]
            sobran_p = [x for x in m["pagan"] if normalizar(x) not in esp_pagan]
            faltan_r = [esp_reciben[k] for k in esp_reciben if k not in hay_reciben]
            sobran_r = [x for x in m["reciben"] if normalizar(x) not in esp_reciben]

            # Y el nombre definido tiene que cubrir justo esa matriz, si no la
            # consulta lee de mas o de menos (falto apretar Actualiza Rango).
            nombre = c.get("nombre_definido")
            aviso_rango, rango_ok = None, True
            if nombre:
                crudo = leer_nombre_definido(ruta, nombre)
                fin = celdas_de_rango(crudo)
                # El rango excluye la fila y columna de totales: por eso -1 no,
                # sino que termina justo en la ultima empresa.
                if fin is None:
                    aviso_rango = f"no se pudo leer el nombre '{nombre}': {crudo!r}"
                    rango_ok = False
                elif (fin[0], fin[1]) != (m["ultima_fila"], m["ultima_col"]):
                    aviso_rango = (
                        f"{nombre} termina en fila {fin[0]} columna {fin[1]}, "
                        f"y la matriz termina en fila {m['ultima_fila']} columna "
                        f"{m['ultima_col']}. Falta apretar «Actualiza Rango».")
                    rango_ok = False
                else:
                    aviso_rango = (f"{nombre} cubre justo la matriz "
                                   f"(hasta fila {fin[0]}, columna {fin[1]})")

            ok = not (faltan_p or sobran_p or faltan_r or sobran_r) and rango_ok
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            for et, lista in (("PAGAN, falta en la matriz", faltan_p),
                              ("PAGAN, sobra en la matriz", sobran_p),
                              ("RECIBEN, falta en la matriz", faltan_r),
                              ("RECIBEN, sobra en la matriz", sobran_r)):
                for n in sorted(lista)[:10]:
                    L(f"          {et}: {n[:40]}")
                if len(lista) > 10:
                    L(f"          ... y {len(lista) - 10} mas ({et})")
            if aviso_rango:
                L(f"        {aviso_rango}")
            if not ok and (faltan_p or sobran_p or faltan_r or sobran_r):
                L("        La matriz no corresponde a los datos de ahora: "
                  "falta correr «Cuadro de pagos».")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        n_pagan=len(m["pagan"]), n_reciben=len(m["reciben"]),
                        faltan_pagan=faltan_p[:40], sobran_pagan=sobran_p[:40],
                        faltan_reciben=faltan_r[:40], sobran_reciben=sobran_r[:40],
                        aviso_rango=aviso_rango)

        if tipo == "cprt_al_dia":
            # El CPRT sale de la dinamica. Si no se refresco, sus pares son los
            # del mes anterior aunque la matriz ya este bien.
            ruta = self.rutas.get(c["archivo"])
            if ruta is None:
                L(f"  ? {c['desc']}: falta el archivo")
                return dict(base, estado="SIN DATOS")
            L(f"  .. {c['desc']}")
            m = leer_matriz_pago(ruta, c["hoja_matriz"], L)
            if m is None or not m["montos"]:
                L("  >> no se pudo leer la matriz; sin ella no hay con que "
                  "comparar el CPRT.")
                return dict(base, estado="SIN DATOS")
            d = leer_columnas_rapido(ruta, c["hoja_cprt"],
                                     ["B", "E", "G"], c["fila_cprt"], L)
            if d is None:
                return dict(base, estado="SIN DATOS")
            pares_cprt = {}
            for f in sorted(d.get("B", {})):
                deu, acr = d["B"].get(f), d.get("E", {}).get(f)
                if not (isinstance(deu, str) and deu.strip()):
                    continue
                if not (isinstance(acr, str) and acr.strip()):
                    continue
                monto = d.get("G", {}).get(f)
                pares_cprt[(normalizar(deu), normalizar(acr))] = (
                    deu.strip(), acr.strip(),
                    float(monto) if isinstance(monto, (int, float)) else None)
            L(f"        CPRT: {len(pares_cprt)} par(es)   |   "
              f"matriz: {len(m['montos'])} par(es) con monto")

            # Direccion 1: todo par del CPRT tiene que existir en la matriz, y
            # con el mismo monto redondeado. Esto caza los pares viejos que
            # quedaron de un refresco anterior.
            fantasmas, difs = [], []
            for k, (deu, acr, monto) in pares_cprt.items():
                if k not in m["montos"]:
                    fantasmas.append(f"{deu} -> {acr}")
                    continue
                if monto is not None and abs(round(m["montos"][k]) - round(monto)) > 1:
                    difs.append((deu, acr, monto, m["montos"][k]))
            # Direccion 2: solo se exigen los pares CLARAMENTE grandes. El CPRT
            # descarta los montos chicos, y el corte exacto no esta documentado
            # (en 2312 los excluidos llegaban a 8,7 y el menor incluido era 22,3).
            # Pedir los grandes caza un refresco viejo sin inventar el umbral.
            perdidos = [f"{p} -> {r}" for (p, r), v in m["montos"].items()
                        if abs(v) >= UMBRAL_PAR_SEGURO and (p, r) not in pares_cprt]
            ok = not (fantasmas or difs or perdidos)
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            for n in sorted(fantasmas)[:10]:
                L(f"          en el CPRT y NO en la matriz: {n[:52]}")
            if len(fantasmas) > 10:
                L(f"          ... y {len(fantasmas) - 10} mas")
            for n in sorted(perdidos)[:10]:
                L(f"          en la matriz y NO en el CPRT: {n[:52]}")
            if len(perdidos) > 10:
                L(f"          ... y {len(perdidos) - 10} mas")
            for deu, acr, mc, mm in sorted(difs, key=lambda x: -abs(x[2] - x[3]))[:10]:
                L(f"          {deu[:18]} -> {acr[:18]}: CPRT {fmt_monto(mc)} "
                  f"vs matriz {fmt_monto(mm)}")
            if len(difs) > 10:
                L(f"          ... y {len(difs) - 10} diferencia(s) mas")
            if not ok:
                L("        El CPRT no corresponde a la matriz de ahora: "
                  "falta refrescar la tabla dinámica.")
            else:
                L(f"        los {len(pares_cprt)} pares del CPRT calzan con la matriz")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        n_cprt=len(pares_cprt), n_matriz=len(m["montos"]),
                        fantasmas=fantasmas[:40], perdidos=perdidos[:40],
                        difs=[(a, b, x, y) for a, b, x, y in difs[:40]])

        if tipo == "mismas_empresas":
            # Los dos lados tienen que traer EXACTAMENTE las mismas empresas: ni
            # una de mas ni una de menos. Los 0 y los vacios se descartan porque
            # son lo que arrastra una formula que sobra, y eso esta permitido.
            # Se compara el conjunto, no el largo ni la ultima: asi da igual el
            # orden y da igual que una columna tenga cola de ceros.
            la, lb = c["lado_a"], c["lado_b"]
            ra = self.rutas.get(la["archivo"])
            rb = self.rutas.get(lb["archivo"])
            if ra is None or rb is None:
                L(f"  ? {c['desc']}: falta uno de los archivos")
                return dict(base, estado="SIN DATOS")
            L(f"  .. {c['desc']}")

            def leer_lado(ruta, esp):
                d = leer_columnas_rapido(ruta, esp["hoja"], [esp["col"]],
                                         esp.get("fila_inicio", 1), L)
                if d is None:
                    return None
                col = esp["col"].upper()
                out, dups, basura = {}, [], 0
                for f in sorted(d.get(col, {})):
                    v = d[col][f]
                    if isinstance(v, str) and v.startswith("#"):
                        basura += 1          # error de formula
                        continue
                    if not isinstance(v, str) or not v.strip():
                        basura += 1          # numero (0) o vacio
                        continue
                    k = normalizar(v)
                    if k in ("", "0"):
                        basura += 1
                        continue
                    if k in out:
                        dups.append((v.strip(), f))
                        continue
                    out[k] = (v.strip(), f)
                nombre = esp.get("nombre", f"{esp['hoja']} {col}")
                L(f"        {nombre}: {len(out)} empresa(s)"
                  + (f", {basura} fila(s) con 0/vacío descartadas" if basura else ""))
                return {"emp": out, "dups": dups, "nombre": nombre}

            a = leer_lado(ra, la)
            b = leer_lado(rb, lb)
            if a is None or b is None:
                return dict(base, estado="SIN DATOS")

            solo_a = [a["emp"][k][0] for k in a["emp"] if k not in b["emp"]]
            solo_b = [b["emp"][k][0] for k in b["emp"] if k not in a["emp"]]
            dups = []
            if c.get("sin_duplicados", True):
                dups = ([(a["nombre"], n, f) for n, f in a["dups"]]
                        + [(b["nombre"], n, f) for n, f in b["dups"]])
            ok = not (solo_a or solo_b or dups)
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            if solo_a:
                L(f"        {len(solo_a)} empresa(s) en {a['nombre']} que FALTAN "
                  f"en {b['nombre']}:")
                for n in sorted(solo_a)[:15]:
                    L(f"          {n[:44]}")
                if len(solo_a) > 15:
                    L(f"          ... y {len(solo_a) - 15} más")
            if solo_b:
                L(f"        {len(solo_b)} empresa(s) en {b['nombre']} que SOBRAN "
                  f"(no están en {a['nombre']}):")
                for n in sorted(solo_b)[:15]:
                    L(f"          {n[:44]}")
                if len(solo_b) > 15:
                    L(f"          ... y {len(solo_b) - 15} más")
            for et, n, f in dups[:10]:
                L(f"        REPETIDA en {et}: {n[:36]} (fila {f})")
            if len(dups) > 10:
                L(f"        ... y {len(dups) - 10} repetición(es) más")
            if ok:
                L(f"        las {len(a['emp'])} empresas son las mismas en los dos lados")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        solo_a=solo_a[:40], solo_b=solo_b[:40],
                        duplicadas=[(et, n, f) for et, n, f in dups[:40]],
                        n_a=len(a["emp"]), n_b=len(b["emp"]),
                        nombre_a=a["nombre"], nombre_b=b["nombre"])

        if tipo == "pertenencia":
            ruta = self.rutas.get(c["archivo"])
            if ruta is None:
                L(f"  ? {c['desc']}: falta el archivo")
                return dict(base, estado="SIN DATOS")
            datos = leer_columnas_rapido(
                ruta, c["hoja"], list(c["origen"]) + list(c["destino"]),
                c.get("fila_inicio", 1), L)
            if datos is None:
                return dict(base, estado="SIN DATOS")

            def textos(cols, recolectar_dups=False):
                """{clave: (nombre, col, fila)}. Descarta vacios, errores y los
                "0" que aparecen cuando sobran formulas."""
                out, dups = {}, []
                for col in cols:
                    for f in sorted(datos.get(col, {})):
                        v = datos[col][f]
                        if not isinstance(v, str) or not v.strip() or v.startswith("#"):
                            continue
                        k = normalizar(v)
                        if k in ("", "0"):
                            continue
                        if k in out:
                            if recolectar_dups:
                                dups.append((v.strip(), col, f))
                            continue
                        out[k] = (v.strip(), col, f)
                return (out, dups) if recolectar_dups else out

            orig = textos(c["origen"])
            dest, dups = textos(c["destino"], recolectar_dups=True)
            if not c.get("sin_duplicados_destino"):
                dups = []
            faltan = [orig[k] for k in orig if k not in dest]
            ok = not (faltan or dups)
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        {len(orig)} empresa(s) distintas en {'+'.join(c['origen'])}, "
              f"{len(dest)} en {'+'.join(c['destino'])}")
            for nombre, col, f in faltan[:20]:
                L(f"          falta: {nombre[:34]:<36} (está en {col}{f})")
            if len(faltan) > 20:
                L(f"          ... y {len(faltan) - 20} más")
            for nombre, col, f in dups[:20]:
                L(f"          REPETIDA en {col}{f}: {nombre[:34]}")
            if len(dups) > 20:
                L(f"          ... y {len(dups) - 20} repetición(es) más")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        faltan=[n for n, _, _ in faltan[:40]],
                        duplicadas=[f"{n} ({col}{f})" for n, col, f in dups[:40]],
                        n_origen=len(orig), n_destino=len(dest))

        if tipo == "ultimo_igual":
            ra = self.rutas.get(c["archivo_a"])
            rb = self.rutas.get(c["archivo_b"])
            if ra is None or rb is None:
                L(f"  ? {c['desc']}: falta uno de los archivos")
                return dict(base, estado="SIN DATOS")
            da = leer_columnas_rapido(ra, c["hoja_a"], [c["col_a"]],
                                      c.get("fila_a", 1), L)
            db = (da if (ra == rb and c["hoja_a"] == c["hoja_b"])
                  else leer_columnas_rapido(rb, c["hoja_b"], [c["col_b"]],
                                            c.get("fila_b", 1), L))
            if da is None or db is None:
                return dict(base, estado="SIN DATOS")
            if ra == rb and c["hoja_a"] == c["hoja_b"] and c["col_b"] not in da:
                da2 = leer_columnas_rapido(ra, c["hoja_a"], [c["col_a"], c["col_b"]],
                                           c.get("fila_a", 1), L)
                da = db = da2 if da2 is not None else da
            fa, va = ultimo_significativo(da, c["col_a"])
            fb, vb = ultimo_significativo(db, c["col_b"])
            if va is None or vb is None:
                L(f"  ? {c['desc']}: una de las columnas no tiene datos útiles "
                  f"({c['col_a']}={va}, {c['col_b']}={vb})")
                return dict(base, estado="SIN DATOS")
            if isinstance(va, (int, float)) and isinstance(vb, (int, float)):
                ok = abs(va - vb) <= TOLERANCIA
            else:
                ok = normalizar(va) == normalizar(vb)
            L(f"  {'OK ' if ok else '>> '}{c['desc']}")
            L(f"        {c['col_a']}{fa} = {str(va)[:38]!r}")
            L(f"        {c['col_b']}{fb} = {str(vb)[:38]!r}")
            return dict(base, estado="OK" if ok else "NO CUADRA",
                        a=f"{c['col_a']}{fa}", b=f"{c['col_b']}{fb}",
                        valor_a=va, valor_b=vb)

        L(f"  ? comprobación de tipo desconocido: {tipo}")
        return dict(base, estado="SIN DATOS")

    def _bombear_cola(self):
        try:
            while True:
                tipo, dato = self.cola.get_nowait()
                if tipo == "log":
                    self.log(dato)
                elif tipo == "prog":
                    self.progress.config(value=dato)
                elif tipo == "ver_guardar":
                    vid, reg = dato
                    if not ESTADO.set(vid, reg):
                        self.log(f"  ! No se pudo guardar el estado de {vid} en "
                                 f"{DIR_SALIDAS}. Revisa permisos de escritura.")
                elif tipo == "ver_fallo":
                    w = self.filas.get(VERIFICADORES[dato]["archivo"])
                    if w and "lbl_ver" in w:
                        w["lbl_ver"].config(text="NO SE PUDO VERIFICAR", fg=C_FALTA,
                                            bg=C_NEUTRO)
                elif tipo == "fin":
                    self.trabajando = False
                    self.timer["on"] = False
                    self._bloquear(False)
                    self._actualizar_verificaciones()
                    self.var_estado.set(f"Verificación terminada {datetime.now():%H:%M:%S}")
        except queue.Empty:
            pass
        self.root.after(250, self._bombear_cola)


def main():
    aplicar_valores_cfg()          # sobrescribe VALORES con lo guardado en config.json
    root = tk.Tk()
    app = Revisor(root)
    app.log("Revisor de entregables — CASO RELIQUIDACION")
    app.log("1) Examinar -> carpeta 02 CASO RELIQUIDACION   2) ACTUALIZAR   3) Verificar")
    app.log("Si a un verificador le falta saber la hoja/celda, usa «Configurar valores...»:")
    app.log("ahí eliges la hoja de una lista y escribes la celda o el rango a sumar.")
    legado = _sal.carpetas_legado(DIR_SALIDAS)
    if legado:
        nombres = ", ".join(p.name for p in legado)
        app.log(f"OJO: hay {len(legado)} carpeta(s) con el formato viejo en "
                f"00_Salidas ({nombres}).")
        app.log("     La estructura nueva es 00_Salidas/AAAA/MM Mes "
                "(ej: 2024/07 Julio).")
        app.log("     Mové el contenido a mano; desde acá no se lee ni se escribe "
                "en ellas.")
    faltan = [k for k, v in VALORES.items()
              if (v["tipo"] == "excel" and not (v.get("hoja") and v.get("celda")))
              or (v["tipo"] == "excel_col" and not (v.get("hoja") and v.get("columna")))
              or (v["tipo"] == "excel_etiqueta" and not (v.get("hoja")
                  and v.get("texto_fila") and v.get("columna_valor")))
              or (v["tipo"] == "mdb" and not (v.get("tabla") and v.get("columna")))]
    if faltan:
        app.log("Pendientes de configurar: " + ", ".join(faltan))
    root.mainloop()


if __name__ == "__main__":
    main()
